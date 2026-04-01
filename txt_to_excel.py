import argparse
import json
import re
import unicodedata
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path

from isced_f import classify_bachelor_program, clean_program_title, split_bachelor_programs
from openpyxl import load_workbook

from whed_excel_export import export_txt_directory_outputs, is_noise_program_name, write_table_workbook


DB_UNIVERSITY_COLUMNS = [
    "id",
    "country_id",
    "city_id",
    "code",
    "name",
    "type",
    "created_at",
    "updated_at",
]

DB_PROGRAM_COLUMNS = [
    "id",
    "isced_code",
    "country_id",
    "holland_match_id",
    "name",
    "show_on_report",
    "map_point",
    "riasec_code",
    "value_code",
    "dignity",
    "year",
    "details",
    "alias",
    "provider_name",
    "provider_id",
    "created_at",
    "updated_at",
]

DB_UNIVERSITY_PROGRAM_COLUMNS = [
    "id",
    "country_id",
    "program_id",
    "university_id",
    "university_program_code",
    "name",
    "year",
    "conditions",
    "details",
    "created_at",
    "updated_at",
]

DB_UNIVERSITY_PLACEMENT_CONDITION_COLUMNS = [
    "id",
    "country_id",
    "code",
    "content",
    "provider",
    "created_at",
    "updated_at",
]

SELF_INFERRED_PROGRAM_ATTRIBUTES_BY_ISCED = {
    "0111": ("SA", "PVF", 12),
    "0113": ("SAI", "PVF", 12),
    "0114": ("SAI", "PVF", 12),
    "0200": ("SAI", "FPV", 11),
    "0211": ("ARE", "VFO", 10),
    "0212": ("AER", "VFP", 11),
    "0213": ("AIS", "VFO", 11),
    "0214": ("RAE", "VFO", 11),
    "0215": ("AEI", "VFO", 11),
    "0221": ("SE", "PVF", 11),
    "0222": ("SIA", "VFO", 12),
    "0223": ("IAS", "VFO", 11),
    "0231": ("SAI", "FPV", 11),
    "0232": ("SAI", "FPV", 11),
    "0300": ("IAS", "VFO", 12),
    "0311": ("ECI", "VWF", 2),
    "0312": ("EIS", "PVF", 12),
    "0313": ("ISA", "PFV", 10),
    "0314": ("IAS", "VFO", 10),
    "0321": ("AES", "VOF", 10),
    "0322": ("AIS", "VFO", 11),
    "0410": ("EC", "FWV", 2),
    "0411": ("CIE", "PHV", 2),
    "0412": ("ECI", "WFO", 2),
    "0413": ("EC", "FWV", 2),
    "0416": ("ECS", "FWV", 2),
    "0500": ("IR", "VFO", 8),
    "0511": ("IRS", "VFO", 8),
    "0512": ("IR", "VFO", 8),
    "0521": ("RCE", "VWH", 3),
    "0522": ("RIC", "HFW", 6),
    "0531": ("IRC", "VFW", 9),
    "0532": ("IR", "VFO", 8),
    "0533": ("IR", "VFO", 8),
    "0541": ("IR", "VFO", 8),
    "0542": ("IR", "VFO", 8),
    "0610": ("IRC", "VWF", 8),
    "0613": ("IRC", "VWF", 8),
    "0710": ("IRC", "VWF", 7),
    "0713": ("IR", "VWF", 8),
    "0714": ("IRC", "VWF", 8),
    "0715": ("IRC", "FWV", 7),
    "0716": ("RIC", "FWV", 7),
    "0719": ("IRC", "VWF", 8),
    "0721": ("RCA", "HFW", 6),
    "0722": ("RIC", "FWV", 7),
    "0731": ("ARI", "VFP", 9),
    "0732": ("RIC", "FWO", 7),
    "0811": ("RIC", "HFW", 6),
    "0821": ("RIC", "HFW", 6),
    "0831": ("RIC", "HFW", 6),
    "0841": ("IRS", "HFW", 9),
    "0900": ("SIR", "PFV", 12),
    "0910": ("SIR", "PFV", 12),
    "0911": ("IRS", "HFW", 9),
    "0912": ("IRS", "VOW", 12),
    "0913": ("SIR", "PFV", 12),
    "0915": ("SIR", "PFV", 12),
    "0923": ("SIA", "PVF", 12),
    "1000": ("ESC", "FWV", 2),
    "1014": ("SRE", "PVF", 1),
    "1015": ("ESC", "FWV", 2),
    "1032": ("RSE", "HPV", 3),
    "1041": ("ERC", "FWV", 6),
}

SELF_INFERRED_PROGRAM_ATTRIBUTES_BY_BROAD = {
    "00": ("ESC", "PVF", 12),
    "01": ("SAI", "PVF", 12),
    "02": ("SAI", "FPV", 11),
    "03": ("IAS", "VFO", 12),
    "04": ("EC", "FWV", 2),
    "05": ("IR", "VFO", 8),
    "06": ("IRC", "VWF", 8),
    "07": ("IRC", "VWF", 7),
    "08": ("RIC", "HFW", 6),
    "09": ("SIR", "PFV", 12),
    "10": ("ESC", "FWV", 2),
}

UNMATCHED_LOCATION_COLUMNS = [
    "university_name",
    "raw_country",
    "raw_city",
    "raw_province",
    "country_id",
    "country_name",
    "reason",
]

COUNTRY_ALIASES = {
    "czechia": "Czech Republic",
    "slovak republic": "Slovakia",
    "united states of america": "United States",
}

SPECIAL_CHAR_TRANSLITERATION = {
    "\u00c6": "AE",
    "\u00d0": "D",
    "\u00d8": "O",
    "\u00de": "TH",
    "\u00df": "ss",
    "\u00e6": "ae",
    "\u00f0": "d",
    "\u00f8": "o",
    "\u00fe": "th",
    "\u0110": "D",
    "\u0111": "d",
    "\u0130": "I",
    "\u0131": "i",
    "\u0141": "L",
    "\u0142": "l",
    "\u0152": "OE",
    "\u0153": "oe",
    "\u1e9e": "SS",
}

CITY_ALIASES_BY_COUNTRY = {
    "Austria": {
        "wien": "Vienna",
    },
    "Belgium": {
        "bruxelles": "Brussels",
        "ghent": "Gent",
    },
    "Cyprus": {
        "lefkosia": "Nicosia",
        "nicosia cyprus": "Nicosia",
    },
    "Denmark": {
        "kgs lyngby": "Kongens Lyngby",
    },
    "Germany": {
        "alfter bei bonn": "Alfter",
        "berlin karlshorst": "Berlin",
        "duesseldorf": "Dusseldorf",
        "esslingen am neckar": "Esslingen",
        "freiburg im breisgau": "Freiburg",
        "kassel bad wilhelmshohe": "Kassel",
        "muenchen": "Munich",
        "mulheim an der ruhr": "Mulheim",
        "munchen": "Munich",
        "potsdam babelsberg": "Potsdam",
        "rottenburg am neckar": "Rottenburg",
    },
    "Italy": {
        "firenze": "Florence",
        "milano": "Milan",
        "roma": "Rome",
        "venezia": "Venice",
    },
    "Netherlands": {
        "den haag": "The Hague",
        "hertogenbosch": "s Hertogenbosch",
    },
    "Portugal": {
        "lisboa": "Lisbon",
    },
    "Romania": {
        "bucuresti": "Bucharest",
    },
    "Sweden": {
        "gothenburg": "Goteborg",
    },
}

REGION_ALIASES_BY_COUNTRY = {
    "Germany": {
        "baden wuerttemberg": "BW",
        "baden wurttemberg": "BW",
        "bavaria": "BY",
        "berlin": "BE",
        "bayern": "BY",
        "brandenburg": "BB",
        "hesse": "HE",
        "hessen": "HE",
        "holstein": "SH",
        "lower saxony": "NI",
        "niedersachsen": "NI",
        "nordrhein westfalen": "NW",
        "north rhine westfalia": "NW",
        "north rhine westphalia": "NW",
        "northrhine westphalia": "NW",
        "rheinland pfaltz": "RP",
        "rheinland pfalz": "RP",
        "rhineland palatinate": "RP",
        "sachsen anhalt": "ST",
        "saxony anhalt": "ST",
        "schleswig holstein": "SH",
    },
    "Italy": {
        "emilia romagna": "Emilia-Romagna",
        "fi": "52",
        "lombardia": "Lombardy",
        "mi": "25",
        "pr": "33",
        "si": "52",
        "vc": "21",
    },
    "Netherlands": {
        "flevoland": "FL",
        "gelderland": "GE",
        "limburg": "LI",
        "noord brabant": "NB",
        "overijsel": "OV",
        "utrecht": "UT",
        "zuid holland": "ZH",
    },
    "United States": {
        "dc": "DC",
        "district of columbia": "DC",
    },
}

ADMIN_WORDS = {
    "autonomous",
    "capital",
    "canton",
    "city",
    "county",
    "department",
    "departement",
    "district",
    "federal",
    "governorate",
    "hlavni",
    "kerulet",
    "main",
    "mesto",
    "metropolitan",
    "municipality",
    "oblast",
    "okres",
    "prefecture",
    "province",
    "region",
    "republic",
    "special",
    "state",
    "territory",
    "voivodeship",
}

STOPWORDS = {"da", "de", "di", "do", "el", "la", "le", "of"}


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description="Convert WHED TXT files into Excel exports, including relational database-style tables."
    )
    parser.add_argument(
        "--input-dir",
        default="Data",
        help="Folder that contains institution TXT files.",
    )
    parser.add_argument(
        "--output-file",
        default="whed_data.xlsx",
        help="Legacy all-in-one Excel file to create.",
    )
    parser.add_argument(
        "--skip-full-workbook",
        action="store_true",
        help="Skip generating the legacy all-in-one workbook.",
    )
    parser.add_argument(
        "--universities-file",
        default="universities.xlsx",
        help="Excel file to create for university records.",
    )
    parser.add_argument(
        "--programs-file",
        default="programs.xlsx",
        help="Excel file to create for unique programs.",
    )
    parser.add_argument(
        "--university-programs-file",
        default="university_programs.xlsx",
        help="Excel file to create for university-program pivot rows.",
    )
    parser.add_argument(
        "--enrichment-file",
        default=None,
        help="Optional enrichment JSONL file. Defaults to whed_enrichment.jsonl next to the outputs if present.",
    )
    parser.add_argument(
        "--program-source",
        choices=("all-degree-fields", "bachelors"),
        default="bachelors",
        help="Which parsed degree fields should feed the programs and university_programs outputs.",
    )
    parser.add_argument(
        "--all-countries",
        action="store_true",
        help="Include every parsed country instead of the current allowed-country filter.",
    )
    parser.add_argument(
        "--merge-db-universities",
        action="store_true",
        help="Merge WHED workbook universities into the database universities workbook format.",
    )
    parser.add_argument(
        "--merge-db-programs",
        action="store_true",
        help="Merge WHED bachelor programs into the database programs workbook format.",
    )
    parser.add_argument(
        "--merge-db-university-programs",
        action="store_true",
        help="Merge WHED bachelor offerings into the database university_programs workbook format.",
    )
    parser.add_argument(
        "--merge-db-university-placement-conditions",
        action="store_true",
        help="Merge WHED admission requirement ids into the database university_placement_conditions workbook format.",
    )
    parser.add_argument(
        "--db-whed-file",
        default="whed_data.xlsx",
        help="WHED workbook that contains the Institutions sheet.",
    )
    parser.add_argument(
        "--db-universities-file",
        default=r"References/TercihAnalizi Database Tables/universities.xlsx",
        help="Existing database universities workbook whose ids should be continued.",
    )
    parser.add_argument(
        "--db-output-file",
        default=None,
        help="Merged database universities workbook to create. Defaults to the same path as --db-universities-file.",
    )
    parser.add_argument(
        "--db-countries-file",
        default=r"References/TercihAnalizi Database Tables/countries.xlsx",
        help="Countries reference workbook.",
    )
    parser.add_argument(
        "--db-cities-file",
        default=r"References/TercihAnalizi Database Tables/cities.xlsx",
        help="Cities reference workbook.",
    )
    parser.add_argument(
        "--db-districts-file",
        default=r"References/TercihAnalizi Database Tables/districts.xlsx",
        help="Districts reference workbook.",
    )
    parser.add_argument(
        "--db-unmatched-file",
        default=None,
        help="Optional workbook path for rows whose city_id could not be matched.",
    )
    parser.add_argument(
        "--db-programs-file",
        default=r"References/TercihAnalizi Database Tables/programs.xlsx",
        help="Existing database programs workbook whose ids should be continued.",
    )
    parser.add_argument(
        "--db-programs-output-file",
        default=None,
        help="Merged database programs workbook to create. Defaults to the same path as --db-programs-file.",
    )
    parser.add_argument(
        "--db-isced-codes-file",
        default=r"References/TercihAnalizi Database Tables/isced_codes.xlsx",
        help="ISCED codes reference workbook.",
    )
    parser.add_argument(
        "--db-holland-matches-file",
        default=r"References/TercihAnalizi Database Tables/holland_matches.xlsx",
        help="Holland matches reference workbook.",
    )
    parser.add_argument(
        "--db-university-programs-file",
        default=r"References/TercihAnalizi Database Tables/university_programs.xlsx",
        help="Existing database university_programs workbook whose ids should be continued.",
    )
    parser.add_argument(
        "--db-university-programs-output-file",
        default=None,
        help=(
            "Merged database university_programs workbook to create. "
            "Defaults to the same path as --db-university-programs-file."
        ),
    )
    parser.add_argument(
        "--db-university-placement-conditions-file",
        default=r"References/TercihAnalizi Database Tables/university_placement_conditions.xlsx",
        help="Existing database university_placement_conditions workbook whose ids should be continued.",
    )
    parser.add_argument(
        "--db-university-placement-conditions-output-file",
        default=None,
        help=(
            "Merged database university_placement_conditions workbook to create. "
            "Defaults to the same path as --db-university-placement-conditions-file."
        ),
    )
    return parser


def normalize_text(value: object) -> str:
    text = "" if value is None else str(value)
    for source, replacement in SPECIAL_CHAR_TRANSLITERATION.items():
        text = text.replace(source, replacement)
    text = text.strip().lower()
    text = unicodedata.normalize("NFKD", text)
    text = "".join(char for char in text if not unicodedata.combining(char))
    text = text.replace("&", " and ")
    text = text.replace("'", " ")
    text = re.sub(r"\([^)]*\)", " ", text)
    text = re.sub(r"[/,.-]+", " ", text)
    text = re.sub(r"[^a-z0-9\s]", " ", text)
    return re.sub(r"\s+", " ", text).strip()


def simplify_admin_name(value: object) -> str:
    normalized = normalize_text(value)
    words = [word for word in normalized.split() if word not in ADMIN_WORDS and word not in STOPWORDS]
    return " ".join(words).strip() or normalized


def parse_country_name(raw_country: object) -> str:
    base_country = re.split(r"\s+-\s+", str(raw_country or "").strip(), maxsplit=1)[0].strip()
    return COUNTRY_ALIASES.get(normalize_text(base_country), base_country)


def parse_region_value(raw_country: object, raw_province: object) -> str:
    province = str(raw_province or "").strip()
    if province:
        return province
    parts = re.split(r"\s+-\s+", str(raw_country or "").strip(), maxsplit=1)
    return parts[1].strip() if len(parts) == 2 else ""


def dedupe_preserving_order(values: list[str]) -> list[str]:
    ordered: list[str] = []
    seen: set[str] = set()
    for value in values:
        normalized = normalize_text(value)
        if not normalized or normalized in seen:
            continue
        seen.add(normalized)
        ordered.append(value)
    return ordered


def expand_location_candidates(country_name: str, value: object) -> list[str]:
    raw_value = str(value or "").strip()
    if not raw_value:
        return []

    candidates = [raw_value]
    country_aliases = CITY_ALIASES_BY_COUNTRY.get(country_name, {})
    alias = country_aliases.get(normalize_text(raw_value))
    if alias:
        candidates.append(alias)

    without_numeric_suffix = re.sub(r"\b\d+[a-z]*\b$", "", raw_value, flags=re.IGNORECASE).strip(" ,-/")
    if without_numeric_suffix and without_numeric_suffix != raw_value:
        candidates.append(without_numeric_suffix)

    without_mail_suffix = re.sub(r"\b(?:cedex|air)\b$", "", raw_value, flags=re.IGNORECASE).strip(" ,-/")
    if without_mail_suffix and without_mail_suffix != raw_value:
        candidates.append(without_mail_suffix)

    without_single_letter_suffix = re.sub(r"\b[A-Z]\b$", "", raw_value).strip(" ,-/")
    if without_single_letter_suffix and without_single_letter_suffix != raw_value:
        candidates.append(without_single_letter_suffix)

    expanded_abbreviation = re.sub(r"^Kgs\.\s+", "Kongens ", raw_value, flags=re.IGNORECASE)
    if expanded_abbreviation != raw_value:
        candidates.append(expanded_abbreviation)

    if "-" in raw_value:
        dash_parts = [part.strip() for part in raw_value.split("-") if part.strip()]
        if len(dash_parts) > 1:
            candidates.append(dash_parts[0])

    if "," in raw_value:
        candidates.append(raw_value.split(",", 1)[0].strip())

    return dedupe_preserving_order(candidates)


def choose_single(values: list[object]) -> object | None:
    unique_values: list[object] = []
    seen: set[object] = set()
    for value in values:
        if value in seen:
            continue
        seen.add(value)
        unique_values.append(value)
    return unique_values[0] if len(unique_values) == 1 else None


def first_sheet(path: Path):
    workbook = load_workbook(path, read_only=True, data_only=True)
    return workbook[workbook.sheetnames[0]]


def build_country_lookup(countries_file: Path) -> dict[str, tuple[object, str]]:
    sheet = first_sheet(countries_file)
    rows = sheet.iter_rows(values_only=True)
    header = list(next(rows))
    id_index = header.index("id")
    name_index = header.index("name")

    lookup: dict[str, tuple[object, str]] = {}
    for row in rows:
        country_id = row[id_index]
        country_name = str(row[name_index] or "").strip()
        lookup[normalize_text(country_name)] = (country_id, country_name)
    return lookup


def build_city_indexes(
    cities_file: Path,
) -> tuple[
    dict[tuple[object, str], list[object]],
    dict[tuple[object, str], list[object]],
    dict[tuple[object, str], object],
]:
    sheet = first_sheet(cities_file)
    rows = sheet.iter_rows(values_only=True)
    header = list(next(rows))
    id_index = header.index("id")
    country_id_index = header.index("country_id")
    name_index = header.index("name")
    iso2_index = header.index("iso2")

    exact_index: dict[tuple[object, str], list[object]] = defaultdict(list)
    simplified_index: dict[tuple[object, str], list[object]] = defaultdict(list)
    iso2_index_map: dict[tuple[object, str], object] = {}

    for row in rows:
        country_id = row[country_id_index]
        city_id = row[id_index]
        name = row[name_index]
        exact_index[(country_id, normalize_text(name))].append(city_id)
        simplified_index[(country_id, simplify_admin_name(name))].append(city_id)

        iso2_code = normalize_text(row[iso2_index])
        if iso2_code:
            iso2_index_map[(country_id, iso2_code)] = city_id

    return exact_index, simplified_index, iso2_index_map


def build_district_indexes(
    districts_file: Path,
) -> tuple[
    dict[tuple[object, str], list[tuple[object, str]]],
    dict[tuple[object, str], list[tuple[object, str]]],
]:
    sheet = first_sheet(districts_file)
    rows = sheet.iter_rows(values_only=True)
    header = list(next(rows))
    country_id_index = header.index("country_id")
    city_id_index = header.index("city_id")
    name_index = header.index("name")
    state_code_index = header.index("state_code")

    exact_index: dict[tuple[object, str], list[tuple[object, str]]] = defaultdict(list)
    simplified_index: dict[tuple[object, str], list[tuple[object, str]]] = defaultdict(list)

    for row in rows:
        country_id = row[country_id_index]
        city_id = row[city_id_index]
        district_name = row[name_index]
        state_code = normalize_text(row[state_code_index])

        exact_index[(country_id, normalize_text(district_name))].append((city_id, state_code))
        simplified_index[(country_id, simplify_admin_name(district_name))].append((city_id, state_code))

    return exact_index, simplified_index


def load_existing_db_universities(universities_file: Path) -> list[dict[str, object]]:
    sheet = first_sheet(universities_file)
    rows = sheet.iter_rows(values_only=True)
    header = list(next(rows))
    index_by_name = {str(column): position for position, column in enumerate(header)}

    existing_rows: list[dict[str, object]] = []
    for row in rows:
        existing_rows.append({column: row[index_by_name[column]] for column in DB_UNIVERSITY_COLUMNS})
    return existing_rows


def load_existing_db_programs(programs_file: Path) -> list[dict[str, object]]:
    sheet = first_sheet(programs_file)
    rows = sheet.iter_rows(values_only=True)
    header = list(next(rows))
    index_by_name = {str(column): position for position, column in enumerate(header)}

    existing_rows: list[dict[str, object]] = []
    for row in rows:
        existing_rows.append({column: row[index_by_name[column]] for column in DB_PROGRAM_COLUMNS})
    return existing_rows


def load_existing_db_university_programs(university_programs_file: Path) -> list[dict[str, object]]:
    sheet = first_sheet(university_programs_file)
    rows = sheet.iter_rows(values_only=True)
    header = list(next(rows))
    index_by_name = {str(column): position for position, column in enumerate(header)}

    existing_rows: list[dict[str, object]] = []
    for row in rows:
        existing_rows.append({column: row[index_by_name[column]] for column in DB_UNIVERSITY_PROGRAM_COLUMNS})
    return existing_rows


def load_existing_db_university_placement_conditions(
    university_placement_conditions_file: Path,
) -> list[dict[str, object]]:
    sheet = first_sheet(university_placement_conditions_file)
    rows = sheet.iter_rows(values_only=True)
    header = list(next(rows))
    index_by_name = {str(column): position for position, column in enumerate(header)}

    existing_rows: list[dict[str, object]] = []
    for row in rows:
        existing_rows.append(
            {column: row[index_by_name[column]] for column in DB_UNIVERSITY_PLACEMENT_CONDITION_COLUMNS}
        )
    return existing_rows


def build_university_country_index(universities_file: Path) -> dict[str, list[object]]:
    sheet = first_sheet(universities_file)
    rows = sheet.iter_rows(values_only=True)
    header = list(next(rows))
    country_id_index = header.index("country_id")
    name_index = header.index("name")

    index: dict[str, list[object]] = defaultdict(list)
    for row in rows:
        country_id = row[country_id_index]
        normalized_name = normalize_text(row[name_index])
        if country_id is None or not normalized_name:
            continue
        index[normalized_name].append(country_id)
    return index


def build_university_id_index(universities_file: Path) -> dict[tuple[object, str], list[object]]:
    index: dict[tuple[object, str], list[object]] = defaultdict(list)
    for row in load_existing_db_universities(universities_file):
        country_id = row.get("country_id")
        normalized_name = normalize_text(row.get("name"))
        university_id = row.get("id")
        if country_id is None or not normalized_name or university_id is None:
            continue
        index[(country_id, normalized_name)].append(university_id)
    return index


def build_program_row_index(programs_file: Path) -> dict[tuple[object, str, int], list[dict[str, object]]]:
    index: dict[tuple[object, str, int], list[dict[str, object]]] = defaultdict(list)
    for row in load_existing_db_programs(programs_file):
        normalized_name = normalize_text(row.get("name"))
        if not normalized_name:
            continue
        key = (row.get("country_id"), normalized_name, int(row.get("year") or 0))
        index[key].append(row)
    return index


def build_isced_code_lookup(isced_codes_file: Path) -> dict[str, object]:
    sheet = first_sheet(isced_codes_file)
    rows = sheet.iter_rows(values_only=True)
    header = list(next(rows))
    id_index = header.index("id")
    broad_index = header.index("broad_code")
    narrow_index = header.index("narrow_code")
    detailed_index = header.index("detailed_code")

    lookup: dict[str, object] = {}
    for row in rows:
        code_id = row[id_index]
        broad_code = str(row[broad_index] or "").strip()
        narrow_code = str(row[narrow_index] or "").strip()
        detailed_code = str(row[detailed_index] or "").strip()

        if detailed_code:
            lookup[detailed_code.zfill(4)] = code_id
            continue

        if narrow_code:
            lookup[f"{narrow_code.zfill(3)}0"] = code_id
            continue

        if broad_code:
            lookup[f"{broad_code.zfill(2)}00"] = code_id

    return lookup


def build_holland_match_index(
    holland_matches_file: Path,
) -> tuple[dict[str, dict[str, object]], dict[str, list[dict[str, object]]]]:
    sheet = first_sheet(holland_matches_file)
    rows = sheet.iter_rows(values_only=True)
    header = list(next(rows))
    id_index = header.index("id")
    riasec_index = header.index("riasec_code")
    map_point_index = header.index("map_point")
    isced_index = header.index("isced_code")
    value_index = header.index("value_code")

    grouped_rows: dict[str, list[dict[str, object]]] = defaultdict(list)
    for row in rows:
        normalized_isced = normalize_isced_code(row[isced_index])
        if not normalized_isced:
            continue
        grouped_rows[normalized_isced].append(
            {
                "id": row[id_index],
                "riasec_code": row[riasec_index],
                "map_point": row[map_point_index],
                "value_code": row[value_index],
            }
        )

    unique_matches: dict[str, dict[str, object]] = {}
    ambiguous_matches: dict[str, list[dict[str, object]]] = {}
    for isced_code, match_rows in grouped_rows.items():
        if len(match_rows) == 1:
            unique_matches[isced_code] = match_rows[0]
        else:
            ambiguous_matches[isced_code] = match_rows

    return unique_matches, ambiguous_matches


def load_whed_institutions(whed_file: Path) -> list[dict[str, object]]:
    workbook = load_workbook(whed_file, read_only=True, data_only=True)
    sheet = workbook["Institutions"]
    rows = sheet.iter_rows(values_only=True)
    header = list(next(rows))
    index_by_name = {str(column): position for position, column in enumerate(header)}

    records: list[dict[str, object]] = []
    for row in rows:
        records.append(
            {
                "University Name": row[index_by_name["University Name"]],
                "Country": row[index_by_name["Country"]],
                "City": row[index_by_name["City"]],
                "Province": row[index_by_name["Province"]],
                "Institution Funding": row[index_by_name["Institution Funding"]],
            }
        )
    return records


def load_whed_bachelor_program_records(whed_file: Path) -> list[dict[str, object]]:
    workbook = load_workbook(whed_file, read_only=True, data_only=True)
    sheet = workbook["Institutions"]
    rows = sheet.iter_rows(values_only=True)
    header = list(next(rows))
    index_by_name = {str(column): position for position, column in enumerate(header)}

    records: list[dict[str, object]] = []
    for row in rows:
        records.append(
            {
                "University Name": row[index_by_name["University Name"]],
                "Country": row[index_by_name["Country"]],
                "Bachelor's Degree": row[index_by_name["Bachelor's Degree"]],
                "ISCED-F": row[index_by_name["ISCED-F"]],
            }
        )
    return records


def load_whed_admission_requirement_records(whed_file: Path) -> list[dict[str, object]]:
    workbook = load_workbook(whed_file, read_only=True, data_only=True)
    sheet = workbook["Admission Requirement IDs"]
    rows = sheet.iter_rows(values_only=True)
    header = list(next(rows))
    index_by_name = {str(column): position for position, column in enumerate(header)}

    records: list[dict[str, object]] = []
    for row in rows:
        records.append(
            {
                "Country": row[index_by_name["Country"]],
                "Condition ID": row[index_by_name["Condition ID"]],
                "Condition": row[index_by_name["Condition"]],
                "Usage Count": row[index_by_name["Usage Count"]],
            }
        )
    return records


def build_whed_admission_requirement_usage_index(whed_file: Path) -> dict[tuple[str, str], int]:
    workbook = load_workbook(whed_file, read_only=True, data_only=True)
    sheet = workbook["Institutions"]
    rows = sheet.iter_rows(values_only=True)
    header = list(next(rows))
    index_by_name = {str(column): position for position, column in enumerate(header)}

    usage_index: Counter[tuple[str, str]] = Counter()
    for row in rows:
        raw_country = str(row[index_by_name["Country"]] or "").strip()
        raw_ids = str(row[index_by_name["Admission Requirement IDs"]] or "").strip()
        if not raw_country or not raw_ids:
            continue

        for condition_id in re.split(r"[;,]", raw_ids):
            normalized_condition_id = str(condition_id).strip()
            if normalized_condition_id:
                usage_index[(raw_country, normalized_condition_id)] += 1

    return dict(usage_index)


def choose_city_id_from_district_candidates(
    candidates: list[tuple[object, str]],
    preferred_region_codes: set[str],
) -> object | None:
    if not candidates:
        return None

    if preferred_region_codes:
        filtered_candidates = [candidate for candidate in candidates if candidate[1] in preferred_region_codes]
        filtered_city_id = choose_single([candidate[0] for candidate in filtered_candidates])
        if filtered_city_id is not None:
            return filtered_city_id

    return choose_single([candidate[0] for candidate in candidates])


def match_city_id(
    country_id: object,
    country_name: str,
    city_value: object,
    region_value: object,
    city_exact_index: dict[tuple[object, str], list[object]],
    city_simplified_index: dict[tuple[object, str], list[object]],
    city_iso2_index: dict[tuple[object, str], object],
    district_exact_index: dict[tuple[object, str], list[tuple[object, str]]],
    district_simplified_index: dict[tuple[object, str], list[tuple[object, str]]],
) -> object | None:
    region_candidates = expand_location_candidates(country_name, region_value)
    city_candidates = expand_location_candidates(country_name, city_value)

    if city_candidates and region_candidates:
        combined_candidate = f"{city_candidates[0]} {region_candidates[0]}".strip()
        city_candidates = dedupe_preserving_order([combined_candidate, *city_candidates])

    normalized_region_values: list[str] = []
    for region_candidate in region_candidates:
        normalized_region = normalize_text(region_candidate)
        if normalized_region:
            normalized_region_values.append(normalized_region)
        region_alias = REGION_ALIASES_BY_COUNTRY.get(country_name, {}).get(normalized_region)
        if region_alias:
            normalized_region_values.append(normalize_text(region_alias))

    normalized_region_values = list(dict.fromkeys(normalized_region_values))
    simplified_region_values = list(
        dict.fromkeys(
            simplify_admin_name(region_candidate)
            for region_candidate in region_candidates
            if simplify_admin_name(region_candidate)
        )
    )
    preferred_region_codes = set(normalized_region_values)

    for region_key in [*normalized_region_values, *simplified_region_values]:
        direct_city_id = choose_single(city_exact_index.get((country_id, region_key), []))
        if direct_city_id is not None:
            return direct_city_id

        direct_city_id = choose_single(city_simplified_index.get((country_id, region_key), []))
        if direct_city_id is not None:
            return direct_city_id

        iso2_city_id = city_iso2_index.get((country_id, region_key))
        if iso2_city_id is not None:
            return iso2_city_id

    for city_candidate in city_candidates:
        normalized_city = normalize_text(city_candidate)
        simplified_city = simplify_admin_name(city_candidate)

        direct_city_id = choose_single(city_exact_index.get((country_id, normalized_city), []))
        if direct_city_id is not None:
            return direct_city_id

        direct_city_id = choose_single(city_simplified_index.get((country_id, simplified_city), []))
        if direct_city_id is not None:
            return direct_city_id

    for city_candidate in city_candidates:
        normalized_city = normalize_text(city_candidate)
        simplified_city = simplify_admin_name(city_candidate)

        district_city_id = choose_city_id_from_district_candidates(
            district_exact_index.get((country_id, normalized_city), []),
            preferred_region_codes,
        )
        if district_city_id is not None:
            return district_city_id

        district_city_id = choose_city_id_from_district_candidates(
            district_simplified_index.get((country_id, simplified_city), []),
            preferred_region_codes,
        )
        if district_city_id is not None:
            return district_city_id

    for region_candidate in region_candidates:
        normalized_region = normalize_text(region_candidate)
        simplified_region = simplify_admin_name(region_candidate)

        district_city_id = choose_city_id_from_district_candidates(
            district_exact_index.get((country_id, normalized_region), []),
            preferred_region_codes,
        )
        if district_city_id is not None:
            return district_city_id

        district_city_id = choose_city_id_from_district_candidates(
            district_simplified_index.get((country_id, simplified_region), []),
            preferred_region_codes,
        )
        if district_city_id is not None:
            return district_city_id

    return None


def normalize_isced_code(value: object) -> str:
    raw_value = str(value or "").strip()
    if not raw_value:
        return ""

    digits = re.sub(r"\D", "", raw_value)
    if not digits:
        return ""

    if len(digits) == 1:
        return f"0{digits}00"
    if len(digits) == 2:
        return f"{digits}00"
    if len(digits) == 3:
        return digits.zfill(4)
    return digits[:4]


def split_isced_values(value: object) -> list[str]:
    codes: list[str] = []
    for part in re.split(r"[;,]", str(value or "")):
        normalized_code = normalize_isced_code(part)
        if normalized_code:
            codes.append(normalized_code)
    return codes


def isced_specificity(code: str) -> int:
    if not code:
        return -1
    if code.endswith("00"):
        return 0
    if code.endswith("0"):
        return 1
    return 2


def choose_preferred_isced_code(codes: Counter[str]) -> str:
    if not codes:
        return ""

    return max(
        codes.items(),
        key=lambda item: (item[1], isced_specificity(item[0]), len(item[0]), item[0]),
    )[0]


def choose_preferred_id(values: list[object]) -> object | None:
    unique_values: list[object] = []
    seen: set[object] = set()
    for value in values:
        if value is None or value in seen:
            continue
        seen.add(value)
        unique_values.append(value)

    if not unique_values:
        return None

    numeric_values: list[tuple[int, object]] = []
    for value in unique_values:
        try:
            numeric_values.append((int(value), value))
        except (TypeError, ValueError):
            continue

    if numeric_values:
        return max(numeric_values, key=lambda item: item[0])[1]

    return unique_values[-1]


def slugify_identifier(value: object) -> str:
    return normalize_text(value).replace(" ", "-")


def resolve_isced_code_id(isced_code: str, lookup: dict[str, object]) -> object | None:
    normalized_code = normalize_isced_code(isced_code)
    if not normalized_code:
        return None

    direct_match = lookup.get(normalized_code)
    if direct_match is not None:
        return direct_match

    narrow_fallback = lookup.get(f"{normalized_code[:3]}0")
    if narrow_fallback is not None:
        return narrow_fallback

    return lookup.get(f"{normalized_code[:2]}00")


def resolve_program_country_id(
    *,
    university_name: object,
    raw_country: object,
    university_country_index: dict[str, list[object]],
    country_lookup: dict[str, tuple[object, str]],
) -> object | None:
    normalized_university_name = normalize_text(university_name)
    university_country_ids = list(dict.fromkeys(university_country_index.get(normalized_university_name, [])))
    if len(university_country_ids) == 1:
        return university_country_ids[0]

    country_name = parse_country_name(raw_country)
    country_match = country_lookup.get(normalize_text(country_name))
    if country_match is None:
        return university_country_ids[0] if len(university_country_ids) == 1 else None

    country_id = country_match[0]
    if country_id in university_country_ids:
        return country_id

    if university_country_ids:
        return university_country_ids[0] if len(university_country_ids) == 1 else None

    return country_id


def extract_whed_bachelor_program_items(record: dict[str, object]) -> list[dict[str, str]]:
    raw_programs = split_bachelor_programs(str(record.get("Bachelor's Degree") or ""))
    if not raw_programs:
        return []

    cleaned_items: list[dict[str, object]] = []
    for raw_name in raw_programs:
        program_name = clean_program_title(raw_name)
        cleaned_items.append(
            {
                "raw_name": raw_name,
                "program_name": program_name,
                "keep": not is_noise_program_name(program_name),
            }
        )

    isced_codes = split_isced_values(record.get("ISCED-F"))
    aligned_codes = [""] * len(raw_programs)

    kept_indexes = [index for index, item in enumerate(cleaned_items) if item["keep"]]
    if len(isced_codes) == len(raw_programs):
        aligned_codes = list(isced_codes)
    elif len(isced_codes) == len(kept_indexes):
        for kept_index, isced_code in zip(kept_indexes, isced_codes, strict=False):
            aligned_codes[kept_index] = isced_code
    elif len(isced_codes) == 1 and len(kept_indexes) == 1:
        aligned_codes[kept_indexes[0]] = isced_codes[0]

    items_by_key: dict[str, dict[str, str]] = {}
    for index, item in enumerate(cleaned_items):
        if not item["keep"]:
            continue

        program_name = str(item["program_name"])
        normalized_name = normalize_text(program_name)
        if not normalized_name:
            continue

        resolved_isced = aligned_codes[index] or normalize_isced_code(classify_bachelor_program(str(item["raw_name"])))
        existing_item = items_by_key.get(normalized_name)
        if existing_item is None:
            items_by_key[normalized_name] = {"name": program_name, "isced_f": resolved_isced}
            continue

        if isced_specificity(resolved_isced) > isced_specificity(existing_item["isced_f"]):
            existing_item["isced_f"] = resolved_isced

    return list(items_by_key.values())


def is_noise_db_program_name(program_name: str, country_lookup: dict[str, tuple[object, str]]) -> bool:
    normalized_name = normalize_text(program_name)
    if not normalized_name:
        return True

    if normalized_name in country_lookup:
        return True

    country_parts = [part.strip() for part in re.split(r"\band\b|/|,", normalized_name) if part.strip()]
    if len(country_parts) > 1 and all(part in country_lookup for part in country_parts):
        return True

    if normalized_name in {"short", "some"}:
        return True

    if re.search(
        r"\b(universidade|universidad|universite|university|college|institute|institut|hochschule)\b",
        normalized_name,
    ):
        return True

    return bool(re.search(r"\b(diplom|magistr)\b", normalized_name))


def choose_ambiguous_holland_match_id(program_name: str, isced_f: str) -> int | None:
    normalized_name = normalize_text(program_name)
    if not normalized_name:
        return None

    def has_any(*parts: str) -> bool:
        return any(part in normalized_name for part in parts)

    def has_all(*parts: str) -> bool:
        return all(part in normalized_name for part in parts)

    if isced_f == "0414":
        if has_any("e commerce", "ecommerce", "electronic commerce"):
            return 82
        if "public relations" in normalized_name:
            return 134
        if "brand" in normalized_name:
            return 194
        if "medical" in normalized_name and has_any("marketing", "promotion"):
            return 286
        if "advertis" in normalized_name:
            return 135
        if "marketing" in normalized_name:
            return 229

    if isced_f == "0421":
        if has_any("court", "judicial", "paralegal", "clerk") or has_all("legal", "office"):
            return 145
        if "law" in normalized_name:
            return 2

    if isced_f == "0412":
        if has_all("actuar", "insur"):
            return 260
        if "actuar" in normalized_name:
            return 7
        if "public finance" in normalized_name or "fiscal" in normalized_name:
            return 191
        if "international finance" in normalized_name:
            return 295
        if has_any("capital market", "capital markets", "securit", "stock exchange"):
            return 316
        if has_any("bank", "insur"):
            return 29
        if "finance" in normalized_name:
            return 110

    if isced_f == "0222":
        if "archaeolog" in normalized_name:
            return 20 if "art history" in normalized_name else 19
        if "museum" in normalized_name:
            return 209
        if "folklore" in normalized_name:
            return 133
        if "textile" in normalized_name and has_any("restoration", "preservation", "conservation"):
            return 319
        if has_any("preservation", "restoration", "conservation", "heritage"):
            return 181
        if "history" in normalized_name:
            return 41

    if isced_f == "0533":
        if has_any("optic", "acoustic") or has_all("sound", "engineering"):
            return 218
        if "photon" in normalized_name:
            return 116
        if "engineering" in normalized_name:
            return 111
        if has_any("physics", "astronomy", "astrophysics", "nuclear", "space science"):
            return 21

    if isced_f == "0212":
        if has_all("stage", "decor") or has_all("set", "design"):
            return 255
        if has_any("lighting", "sound"):
            return 254
        if has_any("fashion", "textile", "interior", "industrial design", "furniture") or "design" in normalized_name:
            return 146

    if isced_f == "0731":
        if "restoration" in normalized_name:
            return 204
        if has_any("town planning", "urban", "regional planning") or (
            "planning" in normalized_name and "architecture" not in normalized_name
        ):
            return 274
        if "landscape" in normalized_name:
            return 172
        if has_any("drawing", "draft"):
            return 306
        if "architecture" in normalized_name:
            return 147

    if isced_f == "0732":
        if "geotech" in normalized_name:
            return 123
        if "civil engineering" in normalized_name or "construction engineering" in normalized_name:
            return 160
        if "smart infrastructure" in normalized_name:
            return 325
        if has_any("ecological building", "green building", "surface design"):
            return 334
        if has_any("building", "construction", "surveying", "mapping"):
            return 79

    if isced_f == "0915":
        if has_any("dietet", "nutrition"):
            return 33
        if has_any("speech therapy", "speech and language", "speech language", "audiology and speech"):
            return 77
        if has_any("occupational therapy", "ergotherapy"):
            return 102
        if has_any("physical therapy", "physiotherapy", "rehabilitation"):
            return 113

    if isced_f == "0914":
        if "audiology" in normalized_name:
            return 215
        if "audiometr" in normalized_name:
            return 216
        if has_any("orthotic", "prosthetic", "orthopedic prost"):
            return 222
        if has_any("imaging", "radiology", "radiotherapy"):
            return 285
        if has_any("telehealth", "medical data", "health information"):
            return 332
        if has_any("first aid", "emergency"):
            return 153
        if "electroneurophys" in normalized_name:
            return 91
        if has_all("operating", "room") or has_all("surgery", "services"):
            return 14
        if has_any("anesthesia", "dialysis", "perfusion", "nuclear medicine", "pathology", "optometry", "optician", "autopsy"):
            return 15
        if has_any("laboratory", "biomedical", "medical technology"):
            return 49

    if isced_f == "1014":
        if has_any("sports management", "sport management"):
            return 85
        if has_any("underwater", "diving"):
            return 271
        if has_any("sport", "sports", "physical education", "coach", "coaching", "athletic"):
            return 16

    if isced_f == "0413":
        if "human resource" in normalized_name:
            return 158
        if has_all("international", "entrepreneur"):
            return 296
        if "entrepreneur" in normalized_name:
            return 127
        if "health" in normalized_name and "management" in normalized_name:
            return 253
        if "health" in normalized_name and "administration" in normalized_name:
            return 251
        if has_any("rail", "transport management"):
            return 244
        if has_any(
            "business administration",
            "business management",
            "hotel management",
            "industrial management",
            "information management",
        ) or normalized_name == "management":
            return 168
        if "energy management" in normalized_name:
            return 98

    if isced_f == "0314":
        if "anthropolog" in normalized_name:
            return 17
        if "sociolog" in normalized_name:
            return 270
        if has_all("communication", "culture") or "cultural studies" in normalized_name:
            return 182
        if has_any("population", "citizenship", "demography"):
            return 212
        if has_any("family studies", "consumer studies"):
            return 6

    if isced_f == "0231":
        if has_any("teaching", "teacher education", "language education"):
            return 10

    if isced_f == "0232":
        if has_all("applied", "translation") or has_all("applied", "interpreting"):
            return 299
        if "translation studies" in normalized_name:
            return 60
        if has_any("translation", "interpretation", "interpreting", "translator"):
            return 9
        if has_all("english", "linguistics"):
            return 156
        if has_any("literature", "linguistics", "philology", "writing"):
            return 8

    if isced_f == "0213":
        if "art history" in normalized_name:
            return 256
        if has_all("art", "management") or has_all("culture", "management"):
            return 257

    if isced_f == "0321":
        if has_any("television", "broadcast"):
            return 283
        if has_any("new media", "media") and "journal" in normalized_name:
            return 311
        if "journal" in normalized_name:
            return 117

    if isced_f == "0221":
        if "islam" in normalized_name:
            return 161
        if has_any("theology", "religion", "religious", "bible", "catholic", "christian", "judaic", "jewish"):
            return 150

    return None


def choose_ambiguous_holland_match(
    *,
    program_name: str,
    isced_f: str,
    ambiguous_matches: list[dict[str, object]],
) -> dict[str, object] | None:
    preferred_id = choose_ambiguous_holland_match_id(program_name, isced_f)
    if preferred_id is None:
        return None

    for ambiguous_match in ambiguous_matches:
        if int(ambiguous_match["id"]) == preferred_id:
            return ambiguous_match

    return None


def infer_program_attributes(
    *,
    program_name: str,
    isced_f: str,
) -> tuple[str, str, int, str] | None:
    normalized_name = normalize_text(program_name)

    if normalized_name:
        if any(token in normalized_name for token in ("broadcast", "journal", "media")):
            return ("AES", "VOF", 10, "keyword_media")
        if any(token in normalized_name for token in ("book", "literature", "writing", "philology")):
            return ("SAI", "FPV", 11, "keyword_literature")
        if any(token in normalized_name for token in ("religion", "religious", "theology")):
            return ("SE", "PVF", 11, "keyword_religion")

    exact_match = SELF_INFERRED_PROGRAM_ATTRIBUTES_BY_ISCED.get(isced_f)
    if exact_match is not None:
        return (*exact_match, "isced_default")

    broad_match = SELF_INFERRED_PROGRAM_ATTRIBUTES_BY_BROAD.get(isced_f[:2])
    if broad_match is not None:
        return (*broad_match, "broad_default")

    if normalized_name:
        if any(token in normalized_name for token in ("teach", "education")):
            return ("SAI", "PVF", 12, "keyword_education")
        if any(token in normalized_name for token in ("management", "business", "marketing", "commerce")):
            return ("EC", "FWV", 2, "keyword_business")
        if any(token in normalized_name for token in ("engineering", "technology", "computer", "software", "data")):
            return ("IRC", "VWF", 8, "keyword_technology")
        if any(token in normalized_name for token in ("therapy", "health", "nursing", "medical")):
            return ("SIR", "PFV", 12, "keyword_health")
        if any(token in normalized_name for token in ("art", "music", "design", "theatre", "dance")):
            return ("AEI", "VFO", 11, "keyword_arts")

    return None


def parse_json_object(value: object) -> dict[str, object] | None:
    if isinstance(value, dict):
        return value

    raw_text = str(value or "").strip()
    if not raw_text:
        return None

    try:
        parsed = json.loads(raw_text)
    except (TypeError, ValueError, json.JSONDecodeError):
        return None
    return parsed if isinstance(parsed, dict) else None


def has_whed_source(details: object) -> bool:
    parsed = parse_json_object(details)
    if parsed is not None:
        return normalize_text(parsed.get("source")) == "whed"

    raw_text = str(details or "")
    return '"source"' in raw_text and "WHED" in raw_text.upper()


def build_whed_admission_requirement_code(
    *,
    raw_country: str,
    condition_id: str,
    country_variants_by_base_country: dict[str, set[str]],
) -> str:
    base_country = parse_country_name(raw_country)
    country_variants = country_variants_by_base_country.get(base_country, set())
    if len(country_variants) <= 1:
        return condition_id

    raw_country_parts = re.split(r"\s+-\s+", raw_country, maxsplit=1)
    country_scope = raw_country_parts[1] if len(raw_country_parts) == 2 else raw_country
    country_scope_slug = slugify_identifier(country_scope)
    if not country_scope_slug:
        country_scope_slug = slugify_identifier(raw_country) or "scope"
    return f"{country_scope_slug}:{condition_id}"


def normalize_university_type(raw_funding: object) -> str:
    normalized_funding = normalize_text(raw_funding)
    return "state" if normalized_funding == "public" else "foundation"


def renumber_codes_by_country(rows: list[dict[str, object]]) -> None:
    country_counters: Counter[object] = Counter()
    for row in rows:
        country_id = row.get("country_id")
        country_counters[country_id] += 1
        row["code"] = str(country_counters[country_id])


def merge_whed_universities_into_db(
    *,
    whed_file: Path,
    db_universities_file: Path,
    output_file: Path,
    countries_file: Path,
    cities_file: Path,
    districts_file: Path,
    unmatched_file: Path | None,
) -> dict[str, object]:
    existing_rows = load_existing_db_universities(db_universities_file)
    whed_records = load_whed_institutions(whed_file)

    country_lookup = build_country_lookup(countries_file)
    city_exact_index, city_simplified_index, city_iso2_index = build_city_indexes(cities_file)
    district_exact_index, district_simplified_index = build_district_indexes(districts_file)

    max_existing_id = max(int(row["id"]) for row in existing_rows if row.get("id") is not None)
    next_id = max_existing_id + 1
    now = datetime.now().replace(microsecond=0)

    appended_rows: list[dict[str, object]] = []
    unmatched_rows: list[dict[str, object]] = []
    unmatched_countries = 0

    for record in whed_records:
        university_name = str(record.get("University Name") or "").strip()
        raw_country = record.get("Country")
        raw_city = record.get("City")
        raw_province = record.get("Province")

        country_name = parse_country_name(raw_country)
        country_match = country_lookup.get(normalize_text(country_name))
        if country_match is None:
            unmatched_countries += 1
            unmatched_rows.append(
                {
                    "university_name": university_name,
                    "raw_country": raw_country,
                    "raw_city": raw_city,
                    "raw_province": raw_province,
                    "country_id": None,
                    "country_name": country_name,
                    "reason": "country_id_not_found",
                }
            )
            continue

        country_id, matched_country_name = country_match
        region_value = parse_region_value(raw_country, raw_province)
        city_id = match_city_id(
            country_id=country_id,
            country_name=matched_country_name,
            city_value=raw_city,
            region_value=region_value,
            city_exact_index=city_exact_index,
            city_simplified_index=city_simplified_index,
            city_iso2_index=city_iso2_index,
            district_exact_index=district_exact_index,
            district_simplified_index=district_simplified_index,
        )

        if city_id is None:
            unmatched_rows.append(
                {
                    "university_name": university_name,
                    "raw_country": raw_country,
                    "raw_city": raw_city,
                    "raw_province": raw_province,
                    "country_id": country_id,
                    "country_name": matched_country_name,
                    "reason": "city_id_not_found",
                }
            )

        appended_rows.append(
            {
                "id": next_id,
                "country_id": country_id,
                "city_id": city_id,
                "code": "",
                "name": university_name,
                "type": normalize_university_type(record.get("Institution Funding")),
                "created_at": now,
                "updated_at": now,
            }
        )
        next_id += 1

    merged_rows = [*existing_rows, *appended_rows]
    renumber_codes_by_country(merged_rows)

    write_table_workbook(
        output_file=output_file,
        sheet_name="universities",
        columns=DB_UNIVERSITY_COLUMNS,
        rows=merged_rows,
    )

    unmatched_output = unmatched_file
    if unmatched_output is not None:
        write_table_workbook(
            output_file=unmatched_output,
            sheet_name="unmatched_locations",
            columns=UNMATCHED_LOCATION_COLUMNS,
            rows=unmatched_rows,
        )

    return {
        "existing_rows": len(existing_rows),
        "appended_rows": len(appended_rows),
        "merged_rows": len(merged_rows),
        "unmatched_countries": unmatched_countries,
        "unmatched_cities": sum(1 for row in unmatched_rows if row["reason"] == "city_id_not_found"),
        "output_file": output_file,
        "unmatched_file": unmatched_output,
    }


def merge_whed_programs_into_db(
    *,
    whed_file: Path,
    db_programs_file: Path,
    output_file: Path,
    universities_file: Path,
    countries_file: Path,
    isced_codes_file: Path,
    holland_matches_file: Path,
) -> dict[str, object]:
    raw_existing_rows = load_existing_db_programs(db_programs_file)
    whed_records = load_whed_bachelor_program_records(whed_file)
    university_country_index = build_university_country_index(universities_file)
    country_lookup = build_country_lookup(countries_file)
    isced_code_lookup = build_isced_code_lookup(isced_codes_file)
    holland_unique_match_index, holland_ambiguous_match_index = build_holland_match_index(holland_matches_file)

    existing_rows: list[dict[str, object]] = []
    existing_keys: set[tuple[object, str, int]] = set()
    deduped_existing = 0
    replaced_existing_whed = 0
    for row in raw_existing_rows:
        if normalize_text(row.get("provider_name")) == "whed":
            replaced_existing_whed += 1
            continue
        key = (row.get("country_id"), normalize_text(row.get("name")), int(row.get("year") or 0))
        if key in existing_keys:
            deduped_existing += 1
            continue
        existing_keys.add(key)
        existing_rows.append(row)

    now = datetime.now().replace(microsecond=0)
    max_existing_id = max(int(row["id"]) for row in existing_rows if row.get("id") is not None)
    next_id = max_existing_id + 1

    aggregated_rows: dict[tuple[object, str, int], dict[str, object]] = {}
    missing_country = 0
    bachelor_offerings = 0

    for record in whed_records:
        country_id = resolve_program_country_id(
            university_name=record.get("University Name"),
            raw_country=record.get("Country"),
            university_country_index=university_country_index,
            country_lookup=country_lookup,
        )
        if country_id is None:
            missing_country += 1
            continue

        for item in extract_whed_bachelor_program_items(record):
            if is_noise_db_program_name(item["name"], country_lookup):
                continue

            bachelor_offerings += 1
            key = (country_id, normalize_text(item["name"]), 4)
            aggregate = aggregated_rows.setdefault(
                key,
                {
                    "country_id": country_id,
                    "name": item["name"],
                    "year": 4,
                    "offer_count": 0,
                    "isced_counts": Counter(),
                },
            )
            aggregate["offer_count"] += 1
            isced_f = item["isced_f"]
            if isced_f:
                aggregate["isced_counts"][isced_f] += 1

    appended_rows: list[dict[str, object]] = []
    missing_isced = 0
    matched_holland = 0
    heuristic_holland = 0
    self_inferred_holland = 0
    ambiguous_holland = 0
    missing_holland = 0
    skipped_existing = 0

    for key in sorted(aggregated_rows, key=lambda value: (value[0], value[1], value[2])):
        aggregate = aggregated_rows[key]
        if key in existing_keys:
            skipped_existing += 1
            continue

        selected_isced_f = choose_preferred_isced_code(aggregate["isced_counts"])
        selected_isced_id = resolve_isced_code_id(selected_isced_f, isced_code_lookup)
        if selected_isced_f and selected_isced_id is None:
            missing_isced += 1

        holland_match = holland_unique_match_index.get(selected_isced_f)
        holland_match_id = None
        riasec_code = ""
        value_code = ""
        map_point = None
        holland_status = "missing"
        self_inference_basis = ""

        if holland_match is not None:
            holland_match_id = holland_match["id"]
            riasec_code = str(holland_match.get("riasec_code") or "")
            value_code = str(holland_match.get("value_code") or "")
            map_point = holland_match.get("map_point")
            holland_status = "matched"
            matched_holland += 1
        elif selected_isced_f in holland_ambiguous_match_index:
            ambiguous_match = choose_ambiguous_holland_match(
                program_name=aggregate["name"],
                isced_f=selected_isced_f,
                ambiguous_matches=holland_ambiguous_match_index[selected_isced_f],
            )
            if ambiguous_match is not None:
                holland_match_id = ambiguous_match["id"]
                riasec_code = str(ambiguous_match.get("riasec_code") or "")
                value_code = str(ambiguous_match.get("value_code") or "")
                map_point = ambiguous_match.get("map_point")
                holland_status = "heuristic"
                heuristic_holland += 1
            else:
                holland_status = "ambiguous"
        if holland_match_id is None:
            inferred_attributes = infer_program_attributes(
                program_name=aggregate["name"],
                isced_f=selected_isced_f,
            )
            if inferred_attributes is not None:
                riasec_code, value_code, map_point, self_inference_basis = inferred_attributes
                holland_match_id = "?"
                holland_status = "self_inferred"
                self_inferred_holland += 1
            elif holland_status == "ambiguous":
                ambiguous_holland += 1
            else:
                missing_holland += 1

        details = {
            "source": "WHED",
            "degree_type": "Bachelor's Degree",
            "whed_offer_count": int(aggregate["offer_count"]),
            "isced_f": selected_isced_f,
            "holland_status": holland_status,
        }
        if self_inference_basis:
            details["self_inference_basis"] = self_inference_basis

        appended_rows.append(
            {
                "id": next_id,
                "isced_code": str(selected_isced_id) if selected_isced_id is not None else "",
                "country_id": aggregate["country_id"],
                "holland_match_id": holland_match_id,
                "name": aggregate["name"],
                "show_on_report": 1 if holland_match_id is not None else 0,
                "map_point": map_point,
                "riasec_code": riasec_code,
                "value_code": value_code,
                "dignity": 0,
                "year": 4,
                "details": json.dumps(details, ensure_ascii=True),
                "alias": json.dumps([aggregate["name"]], ensure_ascii=True),
                "provider_name": "whed",
                "provider_id": "",
                "created_at": now,
                "updated_at": now,
            }
        )
        next_id += 1

    merged_rows = [*existing_rows, *appended_rows]
    write_table_workbook(
        output_file=output_file,
        sheet_name="programs",
        columns=DB_PROGRAM_COLUMNS,
        rows=merged_rows,
    )

    return {
        "existing_rows": len(existing_rows),
        "deduped_existing": deduped_existing,
        "replaced_existing_whed": replaced_existing_whed,
        "bachelor_offerings": bachelor_offerings,
        "unique_candidate_rows": len(aggregated_rows),
        "appended_rows": len(appended_rows),
        "skipped_existing": skipped_existing,
        "merged_rows": len(merged_rows),
        "matched_holland": matched_holland,
        "heuristic_holland": heuristic_holland,
        "self_inferred_holland": self_inferred_holland,
        "ambiguous_holland": ambiguous_holland,
        "missing_holland": missing_holland,
        "missing_country": missing_country,
        "missing_isced": missing_isced,
        "output_file": output_file,
    }


def merge_whed_university_programs_into_db(
    *,
    whed_file: Path,
    db_university_programs_file: Path,
    output_file: Path,
    universities_file: Path,
    programs_file: Path,
    countries_file: Path,
) -> dict[str, object]:
    raw_existing_rows = load_existing_db_university_programs(db_university_programs_file)
    whed_records = load_whed_bachelor_program_records(whed_file)
    university_country_index = build_university_country_index(universities_file)
    university_id_index = build_university_id_index(universities_file)
    program_row_index = build_program_row_index(programs_file)
    country_lookup = build_country_lookup(countries_file)

    existing_rows: list[dict[str, object]] = []
    existing_keys: set[tuple[object, object, object, int]] = set()
    replaced_existing_whed = 0
    for row in raw_existing_rows:
        if has_whed_source(row.get("details")):
            replaced_existing_whed += 1
            continue

        existing_rows.append(row)
        existing_keys.add(
            (
                row.get("country_id"),
                row.get("program_id"),
                row.get("university_id"),
                int(row.get("year") or 0),
            )
        )

    now = datetime.now().replace(microsecond=0)
    max_existing_id = max(int(row["id"]) for row in existing_rows if row.get("id") is not None)
    next_id = max_existing_id + 1

    aggregated_rows: dict[tuple[object, object, object, int], dict[str, object]] = {}
    missing_country = 0
    missing_university = 0
    missing_program = 0
    bachelor_offerings = 0

    for record in whed_records:
        country_id = resolve_program_country_id(
            university_name=record.get("University Name"),
            raw_country=record.get("Country"),
            university_country_index=university_country_index,
            country_lookup=country_lookup,
        )
        if country_id is None:
            missing_country += 1
            continue

        university_key = (country_id, normalize_text(record.get("University Name")))
        university_id = choose_preferred_id(university_id_index.get(university_key, []))
        if university_id is None:
            missing_university += 1
            continue

        for item in extract_whed_bachelor_program_items(record):
            if is_noise_db_program_name(item["name"], country_lookup):
                continue

            bachelor_offerings += 1
            program_key = (country_id, normalize_text(item["name"]), 4)
            program_rows = program_row_index.get(program_key, [])
            program_id = choose_preferred_id([row.get("id") for row in program_rows])
            if program_id is None:
                missing_program += 1
                continue

            key = (country_id, program_id, university_id, 4)
            aggregate = aggregated_rows.setdefault(
                key,
                {
                    "country_id": country_id,
                    "program_id": program_id,
                    "university_id": university_id,
                    "name": item["name"],
                    "year": 4,
                    "offer_count": 0,
                },
            )
            aggregate["offer_count"] += 1

    appended_rows: list[dict[str, object]] = []
    skipped_existing = 0
    for key in sorted(aggregated_rows, key=lambda value: (value[0], value[1], value[2], value[3])):
        aggregate = aggregated_rows[key]
        if key in existing_keys:
            skipped_existing += 1
            continue

        details = {
            "source": "WHED",
            "degree_type": "Bachelor's Degree",
            "whed_offer_count": int(aggregate["offer_count"]),
        }

        appended_rows.append(
            {
                "id": next_id,
                "country_id": aggregate["country_id"],
                "program_id": aggregate["program_id"],
                "university_id": aggregate["university_id"],
                "university_program_code": "?",
                "name": aggregate["name"],
                "year": aggregate["year"],
                "conditions": "",
                "details": json.dumps(details, ensure_ascii=True),
                "created_at": now,
                "updated_at": now,
            }
        )
        next_id += 1

    merged_rows = [*existing_rows, *appended_rows]
    write_table_workbook(
        output_file=output_file,
        sheet_name="university_programs",
        columns=DB_UNIVERSITY_PROGRAM_COLUMNS,
        rows=merged_rows,
    )

    return {
        "existing_rows": len(existing_rows),
        "replaced_existing_whed": replaced_existing_whed,
        "bachelor_offerings": bachelor_offerings,
        "unique_candidate_rows": len(aggregated_rows),
        "appended_rows": len(appended_rows),
        "skipped_existing": skipped_existing,
        "merged_rows": len(merged_rows),
        "missing_country": missing_country,
        "missing_university": missing_university,
        "missing_program": missing_program,
        "output_file": output_file,
    }


def merge_whed_university_placement_conditions_into_db(
    *,
    whed_file: Path,
    db_university_placement_conditions_file: Path,
    output_file: Path,
    countries_file: Path,
) -> dict[str, object]:
    raw_existing_rows = load_existing_db_university_placement_conditions(db_university_placement_conditions_file)
    whed_condition_rows = load_whed_admission_requirement_records(whed_file)
    usage_index = build_whed_admission_requirement_usage_index(whed_file)
    country_lookup = build_country_lookup(countries_file)

    existing_rows: list[dict[str, object]] = []
    existing_keys: set[tuple[object, str, str]] = set()
    replaced_existing_whed = 0
    for row in raw_existing_rows:
        provider = str(row.get("provider") or "")
        if normalize_text(provider) == "whed":
            replaced_existing_whed += 1
            continue

        existing_rows.append(row)
        existing_keys.add((row.get("country_id"), str(row.get("code") or ""), normalize_text(provider)))

    max_existing_id = max((int(row["id"]) for row in existing_rows if row.get("id") is not None), default=0)
    next_id = max_existing_id + 1
    now = datetime.now().replace(microsecond=0)

    country_variants_by_base_country: dict[str, set[str]] = defaultdict(set)
    for raw_country, _condition_id in usage_index:
        country_variants_by_base_country[parse_country_name(raw_country)].add(raw_country)

    appended_rows: list[dict[str, object]] = []
    missing_country = 0
    skipped_unused = 0
    usage_mismatches = 0
    scoped_codes = 0

    for row in whed_condition_rows:
        raw_country = str(row.get("Country") or "").strip()
        condition_id = str(row.get("Condition ID") or "").strip()
        condition_content = str(row.get("Condition") or "").strip()
        if not raw_country or not condition_id or not condition_content:
            skipped_unused += 1
            continue

        usage_key = (raw_country, condition_id)
        if usage_key not in usage_index:
            skipped_unused += 1
            continue

        source_usage_count = int(row.get("Usage Count") or 0)
        if source_usage_count != usage_index[usage_key]:
            usage_mismatches += 1

        country_name = parse_country_name(raw_country)
        country_match = country_lookup.get(normalize_text(country_name))
        if country_match is None:
            missing_country += 1
            continue

        country_id, _matched_country_name = country_match
        code = build_whed_admission_requirement_code(
            raw_country=raw_country,
            condition_id=condition_id,
            country_variants_by_base_country=country_variants_by_base_country,
        )
        if code != condition_id:
            scoped_codes += 1

        key = (country_id, code, "whed")
        if key in existing_keys:
            continue

        appended_rows.append(
            {
                "id": next_id,
                "country_id": country_id,
                "code": code,
                "content": condition_content,
                "provider": "whed",
                "created_at": now,
                "updated_at": now,
            }
        )
        existing_keys.add(key)
        next_id += 1

    merged_rows = [*existing_rows, *appended_rows]
    write_table_workbook(
        output_file=output_file,
        sheet_name="university_placement_conditions",
        columns=DB_UNIVERSITY_PLACEMENT_CONDITION_COLUMNS,
        rows=merged_rows,
    )

    return {
        "existing_rows": len(existing_rows),
        "replaced_existing_whed": replaced_existing_whed,
        "source_rows": len(whed_condition_rows),
        "appended_rows": len(appended_rows),
        "merged_rows": len(merged_rows),
        "missing_country": missing_country,
        "skipped_unused": skipped_unused,
        "usage_mismatches": usage_mismatches,
        "scoped_codes": scoped_codes,
        "output_file": output_file,
    }


def main() -> int:
    args = build_parser().parse_args()

    if args.merge_db_universities:
        output_file = Path(args.db_output_file) if args.db_output_file else Path(args.db_universities_file)
        unmatched_file = (
            Path(args.db_unmatched_file)
            if args.db_unmatched_file
            else output_file.with_name(f"{output_file.stem}_unmatched_locations{output_file.suffix}")
        )

        results = merge_whed_universities_into_db(
            whed_file=Path(args.db_whed_file),
            db_universities_file=Path(args.db_universities_file),
            output_file=output_file,
            countries_file=Path(args.db_countries_file),
            cities_file=Path(args.db_cities_file),
            districts_file=Path(args.db_districts_file),
            unmatched_file=unmatched_file,
        )

        print(
            f"[done] Existing DB universities: {results['existing_rows']} -> {Path(args.db_universities_file).resolve()}",
            flush=True,
        )
        print(f"[done] Appended WHED universities: {results['appended_rows']}", flush=True)
        print(f"[done] Merged DB universities: {results['merged_rows']} -> {Path(output_file).resolve()}", flush=True)
        if results["unmatched_countries"]:
            print(f"[warn] Country matches missing: {results['unmatched_countries']}", flush=True)
        print(f"[warn] City matches missing: {results['unmatched_cities']}", flush=True)
        if results["unmatched_file"] is not None:
            print(f"[done] Unmatched location review file: {Path(results['unmatched_file']).resolve()}", flush=True)
        return 0

    if args.merge_db_programs:
        output_file = Path(args.db_programs_output_file) if args.db_programs_output_file else Path(args.db_programs_file)
        results = merge_whed_programs_into_db(
            whed_file=Path(args.db_whed_file),
            db_programs_file=Path(args.db_programs_file),
            output_file=output_file,
            universities_file=Path(args.db_universities_file),
            countries_file=Path(args.db_countries_file),
            isced_codes_file=Path(args.db_isced_codes_file),
            holland_matches_file=Path(args.db_holland_matches_file),
        )

        print(
            f"[done] Existing DB programs: {results['existing_rows']} -> {Path(args.db_programs_file).resolve()}",
            flush=True,
        )
        print(f"[done] Deduped existing exact program rows: {results['deduped_existing']}", flush=True)
        print(f"[done] Rebuilt prior WHED program rows: {results['replaced_existing_whed']}", flush=True)
        print(f"[done] Bachelor offerings scanned: {results['bachelor_offerings']}", flush=True)
        print(f"[done] Unique WHED candidate programs: {results['unique_candidate_rows']}", flush=True)
        print(f"[done] Appended WHED programs: {results['appended_rows']}", flush=True)
        print(f"[done] Skipped existing program keys: {results['skipped_existing']}", flush=True)
        print(f"[done] Merged DB programs: {results['merged_rows']} -> {Path(output_file).resolve()}", flush=True)
        print(f"[warn] Country matches missing: {results['missing_country']}", flush=True)
        print(f"[warn] ISCED ids missing: {results['missing_isced']}", flush=True)
        print(f"[warn] Holland matches ambiguous: {results['ambiguous_holland']}", flush=True)
        print(f"[warn] Holland matches missing: {results['missing_holland']}", flush=True)
        print(f"[done] Holland matches filled exactly: {results['matched_holland']}", flush=True)
        print(f"[done] Holland matches filled heuristically: {results['heuristic_holland']}", flush=True)
        print(f"[done] Holland-like attributes self-inferred: {results['self_inferred_holland']}", flush=True)
        return 0

    if args.merge_db_university_programs:
        output_file = (
            Path(args.db_university_programs_output_file)
            if args.db_university_programs_output_file
            else Path(args.db_university_programs_file)
        )
        results = merge_whed_university_programs_into_db(
            whed_file=Path(args.db_whed_file),
            db_university_programs_file=Path(args.db_university_programs_file),
            output_file=output_file,
            universities_file=Path(args.db_universities_file),
            programs_file=Path(args.db_programs_file),
            countries_file=Path(args.db_countries_file),
        )

        print(
            "[done] Existing DB university_programs: "
            f"{results['existing_rows']} -> {Path(args.db_university_programs_file).resolve()}",
            flush=True,
        )
        print(f"[done] Rebuilt prior WHED university_program rows: {results['replaced_existing_whed']}", flush=True)
        print(f"[done] Bachelor offerings scanned: {results['bachelor_offerings']}", flush=True)
        print(f"[done] Unique WHED candidate university_program rows: {results['unique_candidate_rows']}", flush=True)
        print(f"[done] Appended WHED university_program rows: {results['appended_rows']}", flush=True)
        print(f"[done] Skipped existing university_program keys: {results['skipped_existing']}", flush=True)
        print(
            f"[done] Merged DB university_programs: {results['merged_rows']} -> {Path(output_file).resolve()}",
            flush=True,
        )
        print(f"[warn] Country matches missing: {results['missing_country']}", flush=True)
        print(f"[warn] University matches missing: {results['missing_university']}", flush=True)
        print(f"[warn] Program matches missing: {results['missing_program']}", flush=True)
        return 0

    if args.merge_db_university_placement_conditions:
        output_file = (
            Path(args.db_university_placement_conditions_output_file)
            if args.db_university_placement_conditions_output_file
            else Path(args.db_university_placement_conditions_file)
        )
        results = merge_whed_university_placement_conditions_into_db(
            whed_file=Path(args.db_whed_file),
            db_university_placement_conditions_file=Path(args.db_university_placement_conditions_file),
            output_file=output_file,
            countries_file=Path(args.db_countries_file),
        )

        print(
            "[done] Existing DB university_placement_conditions: "
            f"{results['existing_rows']} -> {Path(args.db_university_placement_conditions_file).resolve()}",
            flush=True,
        )
        print(
            f"[done] Rebuilt prior WHED university_placement_condition rows: {results['replaced_existing_whed']}",
            flush=True,
        )
        print(f"[done] WHED admission requirement rows scanned: {results['source_rows']}", flush=True)
        print(f"[done] Appended WHED university_placement_condition rows: {results['appended_rows']}", flush=True)
        print(
            f"[done] Merged DB university_placement_conditions: {results['merged_rows']} -> {Path(output_file).resolve()}",
            flush=True,
        )
        print(f"[done] Scoped region-specific WHED codes: {results['scoped_codes']}", flush=True)
        print(f"[warn] Country matches missing: {results['missing_country']}", flush=True)
        print(f"[warn] Unused source rows skipped: {results['skipped_unused']}", flush=True)
        print(f"[warn] Usage mismatches between sheets: {results['usage_mismatches']}", flush=True)
        return 0

    results = export_txt_directory_outputs(
        input_dir=Path(args.input_dir),
        output_file=None if args.skip_full_workbook else Path(args.output_file),
        universities_output_file=Path(args.universities_file),
        programs_output_file=Path(args.programs_file),
        university_programs_output_file=Path(args.university_programs_file),
        enrichment_file=Path(args.enrichment_file) if args.enrichment_file else None,
        include_all_countries=args.all_countries,
        program_source=args.program_source,
    )

    if "full_workbook" in results:
        print(f"[done] Full workbook rows: {results['full_workbook']} -> {Path(args.output_file).resolve()}", flush=True)
    print(f"[done] Universities rows: {results['universities']} -> {Path(args.universities_file).resolve()}", flush=True)
    print(f"[done] Programs rows: {results['programs']} -> {Path(args.programs_file).resolve()}", flush=True)
    print(
        f"[done] University-program rows: {results['university_programs']} -> {Path(args.university_programs_file).resolve()}",
        flush=True,
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
