from __future__ import annotations

import argparse
import json
import re
import uuid
from collections import Counter, defaultdict
from pathlib import Path

from openpyxl import load_workbook

from txt_to_excel import (
    build_city_indexes,
    build_country_lookup,
    build_district_indexes,
    build_holland_match_index,
    build_whed_admission_requirement_code,
    build_whed_admission_requirement_usage_index,
    choose_ambiguous_holland_match,
    choose_preferred_isced_code,
    extract_whed_bachelor_program_items,
    infer_program_attributes,
    is_noise_db_program_name,
    load_whed_admission_requirement_records,
    match_city_id,
    normalize_text,
    normalize_university_type,
    parse_country_name,
    parse_region_value,
    simplify_admin_name,
    slugify_identifier,
)


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description="Export WHED workbook data into the requested country-scoped JSON format."
    )
    parser.add_argument(
        "--whed-file",
        default="whed_data.xlsx",
        help="WHED workbook that contains the Institutions and Admission Requirement IDs sheets.",
    )
    parser.add_argument(
        "--output-dir",
        default="DataJSON",
        help="Folder that will receive the generated JSON files.",
    )
    parser.add_argument(
        "--countries-file",
        default=r"References/TercihAnalizi Database Tables/countries.xlsx",
        help="Countries reference workbook.",
    )
    parser.add_argument(
        "--cities-file",
        default=r"References/TercihAnalizi Database Tables/cities.xlsx",
        help="Cities reference workbook.",
    )
    parser.add_argument(
        "--districts-file",
        default=r"References/TercihAnalizi Database Tables/districts.xlsx",
        help="Districts reference workbook.",
    )
    parser.add_argument(
        "--holland-matches-file",
        default=r"References/TercihAnalizi Database Tables/holland_matches.xlsx",
        help="Holland match reference workbook.",
    )
    return parser


def load_full_institution_rows(whed_file: Path) -> list[dict[str, object]]:
    workbook = load_workbook(whed_file, read_only=True, data_only=True)
    sheet = workbook["Institutions"]
    rows = sheet.iter_rows(values_only=True)
    header = list(next(rows))
    index_by_name = {str(column): position for position, column in enumerate(header)}

    selected_columns = [
        "University Name",
        "IAU Code",
        "Country",
        "City",
        "Province",
        "Institution Funding",
        "Language(s)",
        "Admission Language Requirement Level",
        "Bachelor's Degree",
        "ISCED-F",
        "Admission Requirement IDs",
    ]

    records: list[dict[str, object]] = []
    for row in rows:
        records.append({column: row[index_by_name[column]] for column in selected_columns})
    return records


def split_condition_ids(value: object) -> list[str]:
    condition_ids: list[str] = []
    for part in re.split(r"[;,]", str(value or "")):
        normalized = str(part).strip()
        if normalized:
            condition_ids.append(normalized)
    return condition_ids


def generate_program_provider_id(country_id: object, program_name: str) -> str:
    name_seed = f"namespace={country_id}&name={program_name.strip()}&version=1.0"
    return str(uuid.uuid5(uuid.NAMESPACE_URL, name_seed))


def ensure_provider_id_collisions_absent(
    programs_by_scope: dict[object, list[dict[str, object]]],
    *,
    scope_label: str,
) -> None:
    owner_by_provider_id: dict[str, tuple[str, str]] = {}
    duplicate_details: list[str] = []

    for scope_name, programs in programs_by_scope.items():
        normalized_scope_name = str(scope_name).strip()
        for program in programs:
            provider_id = str(program.get("provider_id") or "").strip()
            if not provider_id:
                continue

            owner = (normalized_scope_name, str(program.get("name") or "").strip())
            previous_owner = owner_by_provider_id.get(provider_id)
            if previous_owner is None:
                owner_by_provider_id[provider_id] = owner
                continue

            if previous_owner == owner:
                continue

            duplicate_details.append(
                f"{provider_id} ({previous_owner[0]} / {previous_owner[1]} -> {owner[0]} / {owner[1]})"
            )
            if len(duplicate_details) >= 10:
                break

        if len(duplicate_details) >= 10:
            break

    if duplicate_details:
        duplicates_text = "; ".join(duplicate_details)
        raise ValueError(f"{scope_label} provider_id collision detected: {duplicates_text}")


LANGUAGE_CODE_BY_NAME = {
    "Arabic": "ar",
    "Basque": "eu",
    "Belarusian": "be",
    "Bosnian": "bs",
    "Bulgarian": "bg",
    "Cambodian": "km",
    "Catalan": "ca",
    "Chinese": "zh",
    "Croatian": "hr",
    "Czech": "cs",
    "Danish": "da",
    "Dutch": "nl",
    "Dzongkha": "dz",
    "English": "en",
    "Estonian": "et",
    "Finnish": "fi",
    "French": "fr",
    "Galician": "gl",
    "German": "de",
    "Greek": "el",
    "Hebrew": "he",
    "Hungarian": "hu",
    "Italian": "it",
    "Korean": "ko",
    "Latin": "la",
    "Latvian": "lv",
    "Lithuanian": "lt",
    "Maltese": "mt",
    "Polish": "pl",
    "Portuguese": "pt",
    "Romanian": "ro",
    "Russian": "ru",
    "Scottish Gaelic": "gd",
    "Serbian": "sr",
    "Slovak": "sk",
    "Slovenian": "sl",
    "Spanish": "es",
    "Swedish": "sv",
    "Thai": "th",
    "Ukrainian": "uk",
}

LANGUAGE_AFFILIATION_SUFFIX_RE = re.compile(r"\s+Religious Affiliation\b.*$", flags=re.IGNORECASE)


def normalize_language_name(value: object) -> str:
    language_name = str(value or "").strip().strip(" .;,:")
    if not language_name:
        return ""

    language_name = LANGUAGE_AFFILIATION_SUFFIX_RE.sub("", language_name).strip()
    return language_name


def resolve_language_code(language_name: object) -> str:
    normalized_language_name = normalize_language_name(language_name)
    if not normalized_language_name:
        return ""

    direct_match = LANGUAGE_CODE_BY_NAME.get(normalized_language_name)
    if direct_match:
        return direct_match

    for known_language_name, language_code in sorted(
        LANGUAGE_CODE_BY_NAME.items(),
        key=lambda item: len(item[0]),
        reverse=True,
    ):
        if normalized_language_name == known_language_name:
            return language_code
        if normalized_language_name.startswith(f"{known_language_name} "):
            return language_code

    return ""


def normalize_proficiency_level(value: object) -> str | None:
    proficiency_level = str(value or "").strip().rstrip(".")
    if not proficiency_level:
        return None

    normalized_upper = proficiency_level.upper()
    if normalized_upper.startswith("CEFR "):
        cefr_value = proficiency_level[5:].strip()
        if all(token not in normalized_upper for token in ("IELTS", "TOEFL", "CAMBRIDGE", "DUOLINGO", "PTE")):
            return cefr_value

    return proficiency_level


def upsert_education_language(
    items: list[dict[str, object]],
    *,
    code: str,
    proficiency_level: str | None,
) -> None:
    if not code:
        return

    for item in items:
        if item["code"] != code:
            continue
        if item["proficiency_level"] is None and proficiency_level is not None:
            item["proficiency_level"] = proficiency_level
        return

    items.append(
        {
            "code": code,
            "proficiency_level": proficiency_level,
        }
    )


def parse_education_languages(
    *,
    admission_language_requirement: object,
    fallback_languages: object,
) -> list[dict[str, object]]:
    raw_requirement = str(admission_language_requirement or "").strip()
    education_languages: list[dict[str, object]] = []

    instruction_match = re.search(
        r"Instruction language\(s\):\s*(.+?)(?:\.\s|$)",
        raw_requirement,
        flags=re.IGNORECASE,
    )
    if instruction_match is not None:
        for part in instruction_match.group(1).split(";"):
            language_code = resolve_language_code(part)
            upsert_education_language(
                education_languages,
                code=language_code,
                proficiency_level=None,
            )

    proficiency_match = re.match(
        r"^(.+?)\s+proficiency required(?:(?:\s*:\s*(.+))|(?:\s*;\s*score/level not specified\.?))$",
        raw_requirement,
        flags=re.IGNORECASE,
    )
    if proficiency_match is not None:
        language_code = resolve_language_code(proficiency_match.group(1))
        proficiency_level = normalize_proficiency_level(proficiency_match.group(2))
        upsert_education_language(
            education_languages,
            code=language_code,
            proficiency_level=proficiency_level,
        )

    if not education_languages:
        for part in str(fallback_languages or "").split(";"):
            language_code = resolve_language_code(part)
            upsert_education_language(
                education_languages,
                code=language_code,
                proficiency_level=None,
            )

    return education_languages


def choose_canonical_name(
    name_counts: Counter[str],
    first_seen_order: dict[str, int],
) -> str:
    if not name_counts:
        return ""

    return min(
        name_counts,
        key=lambda name: (-name_counts[name], first_seen_order[name], normalize_text(name), name),
    )


def ordered_aliases(canonical_name: str, first_seen_order: dict[str, int]) -> list[str]:
    aliases = sorted(first_seen_order, key=lambda name: (name != canonical_name, first_seen_order[name], name))
    return aliases or ([canonical_name] if canonical_name else [])


def build_country_variants_by_base_country(
    institution_rows: list[dict[str, object]],
    usage_index: dict[tuple[str, str], int],
) -> dict[str, set[str]]:
    variants: dict[str, set[str]] = defaultdict(set)

    for raw_country, _condition_id in usage_index:
        if raw_country:
            variants[parse_country_name(raw_country)].add(raw_country)

    for row in institution_rows:
        raw_country = str(row.get("Country") or "").strip()
        raw_ids = split_condition_ids(row.get("Admission Requirement IDs"))
        if raw_country and raw_ids:
            variants[parse_country_name(raw_country)].add(raw_country)

    return variants


def build_placement_conditions(
    *,
    whed_file: Path,
    country_lookup: dict[str, tuple[object, str]],
    country_variants_by_base_country: dict[str, set[str]],
) -> tuple[dict[object, list[dict[str, str]]], dict[object, dict[str, int]], dict[str, int]]:
    usage_index = build_whed_admission_requirement_usage_index(whed_file)
    source_rows = load_whed_admission_requirement_records(whed_file)

    placement_conditions_by_country: dict[object, list[dict[str, str]]] = defaultdict(list)
    placement_condition_positions: dict[object, dict[str, int]] = defaultdict(dict)
    stats = {
        "condition_rows_scanned": len(source_rows),
        "condition_rows_used": 0,
        "condition_rows_skipped_unused": 0,
        "condition_country_misses": 0,
    }

    for row in source_rows:
        raw_country = str(row.get("Country") or "").strip()
        condition_id = str(row.get("Condition ID") or "").strip()
        condition = str(row.get("Condition") or "").strip()

        if not raw_country or not condition_id or not condition:
            stats["condition_rows_skipped_unused"] += 1
            continue

        if (raw_country, condition_id) not in usage_index:
            stats["condition_rows_skipped_unused"] += 1
            continue

        country_name = parse_country_name(raw_country)
        country_match = country_lookup.get(normalize_text(country_name))
        if country_match is None:
            stats["condition_country_misses"] += 1
            continue

        country_id = country_match[0]
        scoped_code = build_whed_admission_requirement_code(
            raw_country=raw_country,
            condition_id=condition_id,
            country_variants_by_base_country=country_variants_by_base_country,
        )
        if scoped_code in placement_condition_positions[country_id]:
            continue

        placement_conditions_by_country[country_id].append(
            {
                "code": scoped_code,
                "content": condition,
            }
        )
        placement_condition_positions[country_id][scoped_code] = len(placement_conditions_by_country[country_id])
        stats["condition_rows_used"] += 1

    return placement_conditions_by_country, placement_condition_positions, stats


def build_programs(
    *,
    institution_rows: list[dict[str, object]],
    country_lookup: dict[str, tuple[object, str]],
    holland_matches_file: Path,
) -> tuple[dict[object, list[dict[str, object]]], dict[tuple[object, str], str], dict[str, int]]:
    holland_unique_match_index, holland_ambiguous_match_index = build_holland_match_index(holland_matches_file)

    aggregates: dict[tuple[object, str], dict[str, object]] = {}
    stats = {
        "program_country_misses": 0,
        "program_offerings_scanned": 0,
        "program_unique_rows": 0,
        "program_holland_exact": 0,
        "program_holland_heuristic": 0,
        "program_holland_inferred": 0,
        "program_holland_missing": 0,
    }
    name_sequence = 0

    for row in institution_rows:
        raw_country = row.get("Country")
        country_name = parse_country_name(raw_country)
        country_match = country_lookup.get(normalize_text(country_name))
        if country_match is None:
            stats["program_country_misses"] += 1
            continue

        country_id = country_match[0]
        for item in extract_whed_bachelor_program_items(row):
            if is_noise_db_program_name(item["name"], country_lookup):
                continue

            stats["program_offerings_scanned"] += 1
            key = (country_id, normalize_text(item["name"]))
            aggregate = aggregates.setdefault(
                key,
                {
                    "country_id": country_id,
                    "name_counts": Counter(),
                    "first_seen_order": {},
                    "offer_count": 0,
                    "isced_counts": Counter(),
                },
            )

            aggregate["name_counts"][item["name"]] += 1
            if item["name"] not in aggregate["first_seen_order"]:
                aggregate["first_seen_order"][item["name"]] = name_sequence
                name_sequence += 1
            aggregate["offer_count"] += 1
            if item["isced_f"]:
                aggregate["isced_counts"][item["isced_f"]] += 1

    programs_by_country: dict[object, list[dict[str, object]]] = defaultdict(list)
    provider_id_by_program_key: dict[tuple[object, str], str] = {}

    for key in sorted(aggregates, key=lambda item: (item[0], item[1])):
        aggregate = aggregates[key]
        canonical_name = choose_canonical_name(aggregate["name_counts"], aggregate["first_seen_order"])
        alias_list = ordered_aliases(canonical_name, aggregate["first_seen_order"])
        selected_isced = choose_preferred_isced_code(aggregate["isced_counts"])

        holland_match = holland_unique_match_index.get(selected_isced)
        holland_match_id = None
        riasec_code = ""
        value_code = ""
        map_point = None
        holland_status = "missing"
        self_inference_basis = ""

        if holland_match is not None:
            holland_match_id = int(holland_match["id"])
            riasec_code = str(holland_match.get("riasec_code") or "")
            value_code = str(holland_match.get("value_code") or "")
            map_point = holland_match.get("map_point")
            holland_status = "matched"
            stats["program_holland_exact"] += 1
        elif selected_isced in holland_ambiguous_match_index:
            ambiguous_match = choose_ambiguous_holland_match(
                program_name=canonical_name,
                isced_f=selected_isced,
                ambiguous_matches=holland_ambiguous_match_index[selected_isced],
            )
            if ambiguous_match is not None:
                holland_match_id = int(ambiguous_match["id"])
                riasec_code = str(ambiguous_match.get("riasec_code") or "")
                value_code = str(ambiguous_match.get("value_code") or "")
                map_point = ambiguous_match.get("map_point")
                holland_status = "heuristic"
                stats["program_holland_heuristic"] += 1

        if holland_match_id is None:
            inferred_attributes = infer_program_attributes(
                program_name=canonical_name,
                isced_f=selected_isced,
            )
            if inferred_attributes is not None:
                riasec_code, value_code, map_point, self_inference_basis = inferred_attributes
                holland_status = "self_inferred"
                stats["program_holland_inferred"] += 1
            else:
                stats["program_holland_missing"] += 1

        provider_id = generate_program_provider_id(aggregate["country_id"], canonical_name)
        provider_id_by_program_key[key] = provider_id

        details: dict[str, object] = {
            "source": "WHED",
            "degree_type": "Bachelor's Degree",
            "whed_offer_count": int(aggregate["offer_count"]),
            "holland_status": holland_status,
        }
        if self_inference_basis:
            details["self_inference_basis"] = self_inference_basis

        programs_by_country[aggregate["country_id"]].append(
            {
                "isced_code": selected_isced,
                "holland_match_id": holland_match_id,
                "name": canonical_name,
                "map_point": map_point,
                "riasec_code": riasec_code,
                "value_code": value_code,
                "dignity": 0,
                "year": 4,
                "details": details,
                "alias": alias_list,
                "provider_name": "whed",
                "provider_id": provider_id,
            }
        )

    for country_id in programs_by_country:
        programs_by_country[country_id].sort(key=lambda item: (normalize_text(item["name"]), item["name"]))

    ensure_provider_id_collisions_absent(
        programs_by_scope=programs_by_country,
        scope_label="country program",
    )
    stats["program_unique_rows"] = len(aggregates)
    return programs_by_country, provider_id_by_program_key, stats


def build_universities(
    *,
    institution_rows: list[dict[str, object]],
    country_lookup: dict[str, tuple[object, str]],
    city_exact_index: dict[tuple[object, str], list[object]],
    city_simplified_index: dict[tuple[object, str], list[object]],
    city_iso2_index: dict[tuple[object, str], object],
    district_exact_index: dict[tuple[object, str], list[tuple[object, str]]],
    district_simplified_index: dict[tuple[object, str], list[tuple[object, str]]],
    placement_condition_positions: dict[object, dict[str, int]],
    country_variants_by_base_country: dict[str, set[str]],
    provider_id_by_program_key: dict[tuple[object, str], str],
) -> tuple[dict[object, list[dict[str, object]]], dict[str, int]]:
    universities_by_country: dict[object, list[dict[str, object]]] = defaultdict(list)
    stats = {
        "university_country_misses": 0,
        "university_city_misses": 0,
        "university_rows_used": 0,
        "university_program_rows": 0,
        "condition_refs_missing_from_array": 0,
    }
    fallback_code_counters: Counter[object] = Counter()

    for row in institution_rows:
        raw_country = row.get("Country")
        country_name = parse_country_name(raw_country)
        country_match = country_lookup.get(normalize_text(country_name))
        if country_match is None:
            stats["university_country_misses"] += 1
            continue

        country_id, matched_country_name = country_match
        region_value = parse_region_value(raw_country, row.get("Province"))
        city_id = match_city_id(
            country_id=country_id,
            country_name=matched_country_name,
            city_value=row.get("City"),
            region_value=region_value,
            city_exact_index=city_exact_index,
            city_simplified_index=city_simplified_index,
            city_iso2_index=city_iso2_index,
            district_exact_index=district_exact_index,
            district_simplified_index=district_simplified_index,
        )
        if city_id is None:
            stats["university_city_misses"] += 1

        raw_iau_code = str(row.get("IAU Code") or "").strip()
        if raw_iau_code:
            university_code = raw_iau_code
        else:
            fallback_code_counters[country_id] += 1
            university_code = f"WHED-{country_id}-{fallback_code_counters[country_id]}"

        education_languages = parse_education_languages(
            admission_language_requirement=row.get("Admission Language Requirement Level"),
            fallback_languages=row.get("Language(s)"),
        )

        condition_numbers: list[str] = []
        seen_condition_numbers: set[str] = set()
        for condition_id in split_condition_ids(row.get("Admission Requirement IDs")):
            scoped_code = build_whed_admission_requirement_code(
                raw_country=str(raw_country or "").strip(),
                condition_id=condition_id,
                country_variants_by_base_country=country_variants_by_base_country,
            )
            position = placement_condition_positions.get(country_id, {}).get(scoped_code)
            if position is None:
                stats["condition_refs_missing_from_array"] += 1
                continue

            position_text = str(position)
            if position_text in seen_condition_numbers:
                continue

            seen_condition_numbers.add(position_text)
            condition_numbers.append(position_text)

        university_programs: list[dict[str, object]] = []
        program_sequence = 1
        for item in extract_whed_bachelor_program_items(row):
            if is_noise_db_program_name(item["name"], country_lookup):
                continue

            program_key = (country_id, normalize_text(item["name"]))
            program_provider_id = provider_id_by_program_key.get(
                program_key,
                generate_program_provider_id(country_id, item["name"]),
            )

            details: dict[str, object] = {
                "source": "WHED",
            }
            if item["isced_f"]:
                details["isced_f"] = item["isced_f"]
            details["education_language"] = education_languages

            university_programs.append(
                {
                    "program_provider_id": program_provider_id,
                    "name": item["name"],
                    "university_program_code": f"{university_code}-{program_sequence}",
                    "year": 4,
                    "conditions": ", ".join(condition_numbers),
                    "details": details,
                }
            )
            program_sequence += 1
            stats["university_program_rows"] += 1

        universities_by_country[country_id].append(
            {
                "city_id": city_id,
                "name": str(row.get("University Name") or "").strip(),
                "code": university_code,
                "type": normalize_university_type(row.get("Institution Funding")),
                "university_programs": university_programs,
            }
        )
        stats["university_rows_used"] += 1

    return universities_by_country, stats


def write_country_json_files(
    *,
    output_dir: Path,
    country_names_by_id: dict[object, str],
    programs_by_country: dict[object, list[dict[str, object]]],
    universities_by_country: dict[object, list[dict[str, object]]],
    placement_conditions_by_country: dict[object, list[dict[str, str]]],
) -> dict[str, int]:
    output_dir.mkdir(parents=True, exist_ok=True)

    country_ids = sorted(
        {
            *programs_by_country.keys(),
            *universities_by_country.keys(),
            *placement_conditions_by_country.keys(),
        }
    )

    all_countries_payload: list[dict[str, object]] = []
    file_count = 0
    for country_id in country_ids:
        payload = {
            "country_id": country_id,
            "programs": programs_by_country.get(country_id, []),
            "universities": universities_by_country.get(country_id, []),
            "placement_conditions": placement_conditions_by_country.get(country_id, []),
        }
        all_countries_payload.append(payload)

        country_name = country_names_by_id.get(country_id, str(country_id))
        file_name = f"{country_id}-{slugify_identifier(country_name) or country_id}.json"
        output_path = output_dir / file_name
        output_path.write_text(
            json.dumps(payload, ensure_ascii=False, indent=2),
            encoding="utf-8",
        )
        file_count += 1

    (output_dir / "all_countries.json").write_text(
        json.dumps(all_countries_payload, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )

    return {
        "country_json_files": file_count,
        "all_countries_entries": len(all_countries_payload),
    }


def parse_us_state_name(raw_country: object) -> str:
    raw_country_text = str(raw_country or "").strip()
    if not raw_country_text.startswith("United States of America"):
        return ""

    parts = re.split(r"\s+-\s+", raw_country_text, maxsplit=1)
    return parts[1].strip() if len(parts) == 2 else ""


def build_us_state_placement_conditions(
    *,
    whed_file: Path,
    country_lookup: dict[str, tuple[object, str]],
    country_variants_by_base_country: dict[str, set[str]],
) -> tuple[dict[str, list[dict[str, str]]], dict[str, dict[str, int]], dict[str, int]]:
    usage_index = build_whed_admission_requirement_usage_index(whed_file)
    source_rows = load_whed_admission_requirement_records(whed_file)

    placement_conditions_by_state: dict[str, list[dict[str, str]]] = defaultdict(list)
    placement_condition_positions_by_state: dict[str, dict[str, int]] = defaultdict(dict)
    stats = {
        "us_state_condition_rows_used": 0,
        "us_state_condition_rows_skipped_unused": 0,
    }

    for row in source_rows:
        raw_country = str(row.get("Country") or "").strip()
        state_name = parse_us_state_name(raw_country)
        if not state_name:
            continue

        condition_id = str(row.get("Condition ID") or "").strip()
        condition = str(row.get("Condition") or "").strip()
        if not condition_id or not condition or (raw_country, condition_id) not in usage_index:
            stats["us_state_condition_rows_skipped_unused"] += 1
            continue

        country_name = parse_country_name(raw_country)
        country_match = country_lookup.get(normalize_text(country_name))
        if country_match is None:
            continue

        scoped_code = build_whed_admission_requirement_code(
            raw_country=raw_country,
            condition_id=condition_id,
            country_variants_by_base_country=country_variants_by_base_country,
        )
        if scoped_code in placement_condition_positions_by_state[state_name]:
            continue

        placement_conditions_by_state[state_name].append(
            {
                "code": scoped_code,
                "content": condition,
            }
        )
        placement_condition_positions_by_state[state_name][scoped_code] = len(placement_conditions_by_state[state_name])
        stats["us_state_condition_rows_used"] += 1

    return placement_conditions_by_state, placement_condition_positions_by_state, stats


def build_us_state_programs(
    *,
    institution_rows: list[dict[str, object]],
    country_lookup: dict[str, tuple[object, str]],
    holland_matches_file: Path,
) -> tuple[dict[str, list[dict[str, object]]], dict[tuple[str, str], str], dict[str, int]]:
    holland_unique_match_index, holland_ambiguous_match_index = build_holland_match_index(holland_matches_file)

    aggregates: dict[tuple[str, str], dict[str, object]] = {}
    stats = {
        "us_state_program_unique_rows": 0,
        "us_state_program_offerings_scanned": 0,
        "us_state_program_holland_exact": 0,
        "us_state_program_holland_heuristic": 0,
        "us_state_program_holland_inferred": 0,
        "us_state_program_holland_missing": 0,
    }
    name_sequence = 0

    for row in institution_rows:
        raw_country = row.get("Country")
        state_name = parse_us_state_name(raw_country)
        if not state_name:
            continue

        country_name = parse_country_name(raw_country)
        country_match = country_lookup.get(normalize_text(country_name))
        if country_match is None:
            continue

        country_id = country_match[0]
        for item in extract_whed_bachelor_program_items(row):
            if is_noise_db_program_name(item["name"], country_lookup):
                continue

            stats["us_state_program_offerings_scanned"] += 1
            key = (state_name, normalize_text(item["name"]))
            aggregate = aggregates.setdefault(
                key,
                {
                    "country_id": country_id,
                    "state_name": state_name,
                    "name_counts": Counter(),
                    "first_seen_order": {},
                    "offer_count": 0,
                    "isced_counts": Counter(),
                },
            )
            aggregate["name_counts"][item["name"]] += 1
            if item["name"] not in aggregate["first_seen_order"]:
                aggregate["first_seen_order"][item["name"]] = name_sequence
                name_sequence += 1
            aggregate["offer_count"] += 1
            if item["isced_f"]:
                aggregate["isced_counts"][item["isced_f"]] += 1

    programs_by_state: dict[str, list[dict[str, object]]] = defaultdict(list)
    provider_id_by_state_program_key: dict[tuple[str, str], str] = {}

    for key in sorted(aggregates, key=lambda item: (item[0], item[1])):
        aggregate = aggregates[key]
        canonical_name = choose_canonical_name(aggregate["name_counts"], aggregate["first_seen_order"])
        alias_list = ordered_aliases(canonical_name, aggregate["first_seen_order"])
        selected_isced = choose_preferred_isced_code(aggregate["isced_counts"])

        holland_match = holland_unique_match_index.get(selected_isced)
        holland_match_id = None
        riasec_code = ""
        value_code = ""
        map_point = None
        holland_status = "missing"
        self_inference_basis = ""

        if holland_match is not None:
            holland_match_id = int(holland_match["id"])
            riasec_code = str(holland_match.get("riasec_code") or "")
            value_code = str(holland_match.get("value_code") or "")
            map_point = holland_match.get("map_point")
            holland_status = "matched"
            stats["us_state_program_holland_exact"] += 1
        elif selected_isced in holland_ambiguous_match_index:
            ambiguous_match = choose_ambiguous_holland_match(
                program_name=canonical_name,
                isced_f=selected_isced,
                ambiguous_matches=holland_ambiguous_match_index[selected_isced],
            )
            if ambiguous_match is not None:
                holland_match_id = int(ambiguous_match["id"])
                riasec_code = str(ambiguous_match.get("riasec_code") or "")
                value_code = str(ambiguous_match.get("value_code") or "")
                map_point = ambiguous_match.get("map_point")
                holland_status = "heuristic"
                stats["us_state_program_holland_heuristic"] += 1

        if holland_match_id is None:
            inferred_attributes = infer_program_attributes(
                program_name=canonical_name,
                isced_f=selected_isced,
            )
            if inferred_attributes is not None:
                riasec_code, value_code, map_point, self_inference_basis = inferred_attributes
                holland_status = "self_inferred"
                stats["us_state_program_holland_inferred"] += 1
            else:
                stats["us_state_program_holland_missing"] += 1

        provider_id = generate_program_provider_id(aggregate["country_id"], canonical_name)
        provider_id_by_state_program_key[key] = provider_id

        details: dict[str, object] = {
            "source": "WHED",
            "degree_type": "Bachelor's Degree",
            "whed_offer_count": int(aggregate["offer_count"]),
            "holland_status": holland_status,
        }
        if self_inference_basis:
            details["self_inference_basis"] = self_inference_basis

        programs_by_state[aggregate["state_name"]].append(
            {
                "isced_code": selected_isced,
                "holland_match_id": holland_match_id,
                "name": canonical_name,
                "map_point": map_point,
                "riasec_code": riasec_code,
                "value_code": value_code,
                "dignity": 0,
                "year": 4,
                "details": details,
                "alias": alias_list,
                "provider_name": "whed",
                "provider_id": provider_id,
            }
        )

    for state_name in programs_by_state:
        programs_by_state[state_name].sort(key=lambda item: (normalize_text(item["name"]), item["name"]))

    stats["us_state_program_unique_rows"] = len(aggregates)
    return programs_by_state, provider_id_by_state_program_key, stats


def build_us_state_universities(
    *,
    institution_rows: list[dict[str, object]],
    country_lookup: dict[str, tuple[object, str]],
    city_exact_index: dict[tuple[object, str], list[object]],
    city_simplified_index: dict[tuple[object, str], list[object]],
    city_iso2_index: dict[tuple[object, str], object],
    district_exact_index: dict[tuple[object, str], list[tuple[object, str]]],
    district_simplified_index: dict[tuple[object, str], list[tuple[object, str]]],
    placement_condition_positions_by_state: dict[str, dict[str, int]],
    country_variants_by_base_country: dict[str, set[str]],
    provider_id_by_state_program_key: dict[tuple[str, str], str],
) -> tuple[dict[str, list[dict[str, object]]], dict[str, int]]:
    universities_by_state: dict[str, list[dict[str, object]]] = defaultdict(list)
    stats = {
        "us_state_university_rows_used": 0,
        "us_state_university_program_rows": 0,
        "us_state_city_misses": 0,
        "us_state_condition_refs_missing": 0,
    }

    for row in institution_rows:
        raw_country = row.get("Country")
        state_name = parse_us_state_name(raw_country)
        if not state_name:
            continue

        country_name = parse_country_name(raw_country)
        country_match = country_lookup.get(normalize_text(country_name))
        if country_match is None:
            continue

        country_id, matched_country_name = country_match
        region_value = parse_region_value(raw_country, row.get("Province"))
        city_id = match_city_id(
            country_id=country_id,
            country_name=matched_country_name,
            city_value=row.get("City"),
            region_value=region_value,
            city_exact_index=city_exact_index,
            city_simplified_index=city_simplified_index,
            city_iso2_index=city_iso2_index,
            district_exact_index=district_exact_index,
            district_simplified_index=district_simplified_index,
        )
        if city_id is None:
            stats["us_state_city_misses"] += 1

        university_code = str(row.get("IAU Code") or "").strip()
        if not university_code:
            university_code = f"WHED-229-{slugify_identifier(state_name)}-{stats['us_state_university_rows_used'] + 1}"

        education_languages = parse_education_languages(
            admission_language_requirement=row.get("Admission Language Requirement Level"),
            fallback_languages=row.get("Language(s)"),
        )

        condition_numbers: list[str] = []
        seen_condition_numbers: set[str] = set()
        for condition_id in split_condition_ids(row.get("Admission Requirement IDs")):
            scoped_code = build_whed_admission_requirement_code(
                raw_country=str(raw_country or "").strip(),
                condition_id=condition_id,
                country_variants_by_base_country=country_variants_by_base_country,
            )
            position = placement_condition_positions_by_state.get(state_name, {}).get(scoped_code)
            if position is None:
                stats["us_state_condition_refs_missing"] += 1
                continue

            position_text = str(position)
            if position_text in seen_condition_numbers:
                continue

            seen_condition_numbers.add(position_text)
            condition_numbers.append(position_text)

        university_programs: list[dict[str, object]] = []
        program_sequence = 1
        for item in extract_whed_bachelor_program_items(row):
            if is_noise_db_program_name(item["name"], country_lookup):
                continue

            program_key = (state_name, normalize_text(item["name"]))
            program_provider_id = provider_id_by_state_program_key.get(
                program_key,
                generate_program_provider_id(country_id, item["name"]),
            )

            details: dict[str, object] = {
                "source": "WHED",
            }
            if item["isced_f"]:
                details["isced_f"] = item["isced_f"]
            details["education_language"] = education_languages

            university_programs.append(
                {
                    "program_provider_id": program_provider_id,
                    "name": item["name"],
                    "university_program_code": f"{university_code}-{program_sequence}",
                    "year": 4,
                    "conditions": ", ".join(condition_numbers),
                    "details": details,
                }
            )
            program_sequence += 1
            stats["us_state_university_program_rows"] += 1

        universities_by_state[state_name].append(
            {
                "city_id": city_id,
                "name": str(row.get("University Name") or "").strip(),
                "code": university_code,
                "type": normalize_university_type(row.get("Institution Funding")),
                "university_programs": university_programs,
            }
        )
        stats["us_state_university_rows_used"] += 1

    return universities_by_state, stats


def write_us_state_json_files(
    *,
    output_dir: Path,
    country_id: int,
    programs_by_state: dict[str, list[dict[str, object]]],
    universities_by_state: dict[str, list[dict[str, object]]],
    placement_conditions_by_state: dict[str, list[dict[str, str]]],
) -> dict[str, int]:
    output_dir.mkdir(parents=True, exist_ok=True)

    state_names = sorted(
        {
            *programs_by_state.keys(),
            *universities_by_state.keys(),
            *placement_conditions_by_state.keys(),
        }
    )

    manifest: list[dict[str, object]] = []
    for state_name in state_names:
        payload = {
            "country_id": country_id,
            "programs": programs_by_state.get(state_name, []),
            "universities": universities_by_state.get(state_name, []),
            "placement_conditions": placement_conditions_by_state.get(state_name, []),
        }

        file_name = f"{country_id}-united-states-{slugify_identifier(state_name)}.json"
        output_path = output_dir / file_name
        output_path.write_text(
            json.dumps(payload, ensure_ascii=False, indent=2),
            encoding="utf-8",
        )
        manifest.append(
            {
                "state": state_name,
                "file": file_name,
                "programs": len(payload["programs"]),
                "universities": len(payload["universities"]),
                "placement_conditions": len(payload["placement_conditions"]),
            }
        )

    (output_dir / "manifest.json").write_text(
        json.dumps(manifest, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )

    return {
        "us_state_json_files": len(state_names),
    }


GERMANY_STATE_MANUAL_CITY_OVERRIDES = {
    "freiburg im breisgau": "Baden-Württemberg",
    "marburg": "Hesse",
    "marburg lahn": "Hesse",
    "wildau": "Brandenburg",
}


def build_germany_state_resolution_context(
    cities_file: Path,
    districts_file: Path,
) -> dict[str, object]:
    workbook = load_workbook(cities_file, read_only=True, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    rows = sheet.iter_rows(values_only=True)
    header = list(next(rows))
    index_by_name = {str(column): position for position, column in enumerate(header)}

    city_row_by_id: dict[object, dict[str, object]] = {}
    state_name_by_code: dict[str, str] = {}
    germany_state_ids: set[object] = set()
    for row in rows:
        if row[index_by_name["country_id"]] != 78:
            continue

        city_id = row[index_by_name["id"]]
        city_name = str(row[index_by_name["name"]] or "").strip()
        iso2 = str(row[index_by_name["iso2"]] or "").strip()
        city_row_by_id[city_id] = {
            "name": city_name,
            "iso2": iso2,
        }
        if iso2:
            state_name_by_code[iso2] = city_name
            germany_state_ids.add(city_id)

    workbook = load_workbook(districts_file, read_only=True, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    rows = sheet.iter_rows(values_only=True)
    header = list(next(rows))
    index_by_name = {str(column): position for position, column in enumerate(header)}

    state_code_by_district_name: dict[str, str] = {}
    state_code_by_district_simple_name: dict[str, str] = {}
    for row in rows:
        if row[index_by_name["country_id"]] != 78:
            continue

        district_name = str(row[index_by_name["name"]] or "").strip()
        state_code = str(row[index_by_name["state_code"]] or "").strip()
        if not district_name or not state_code:
            continue

        normalized_district_name = normalize_text(district_name)
        if normalized_district_name:
            state_code_by_district_name.setdefault(normalized_district_name, state_code)

        simplified_district_name = simplify_admin_name(district_name)
        if simplified_district_name:
            state_code_by_district_simple_name.setdefault(simplified_district_name, state_code)

    return {
        "city_row_by_id": city_row_by_id,
        "state_name_by_code": state_name_by_code,
        "germany_state_ids": germany_state_ids,
        "state_code_by_district_name": state_code_by_district_name,
        "state_code_by_district_simple_name": state_code_by_district_simple_name,
    }


def resolve_germany_state_name(
    *,
    city_id: object,
    context: dict[str, object],
) -> str:
    city_row_by_id = context["city_row_by_id"]
    state_name_by_code = context["state_name_by_code"]
    germany_state_ids = context["germany_state_ids"]
    state_code_by_district_name = context["state_code_by_district_name"]
    state_code_by_district_simple_name = context["state_code_by_district_simple_name"]

    city_row = city_row_by_id.get(city_id)
    if city_row is None:
        return ""

    if city_id in germany_state_ids:
        return str(city_row["name"] or "").strip()

    city_name = str(city_row["name"] or "").strip()
    candidates = [
        city_name,
        city_name.split("(", 1)[0].strip(),
        city_name.split("-", 1)[0].strip(),
        city_name.split("/", 1)[0].strip(),
        re.split(r"\bbei\b", city_name, maxsplit=1, flags=re.IGNORECASE)[0].strip(),
    ]

    normalized_city_name = normalize_text(city_name)
    if normalized_city_name in GERMANY_STATE_MANUAL_CITY_OVERRIDES:
        return GERMANY_STATE_MANUAL_CITY_OVERRIDES[normalized_city_name]

    for candidate in candidates:
        if not candidate:
            continue

        normalized_candidate = normalize_text(candidate)
        if normalized_candidate in GERMANY_STATE_MANUAL_CITY_OVERRIDES:
            return GERMANY_STATE_MANUAL_CITY_OVERRIDES[normalized_candidate]

        state_code = state_code_by_district_name.get(normalized_candidate)
        if state_code:
            return state_name_by_code.get(state_code, "")

        simplified_candidate = simplify_admin_name(candidate)
        if simplified_candidate in GERMANY_STATE_MANUAL_CITY_OVERRIDES:
            return GERMANY_STATE_MANUAL_CITY_OVERRIDES[simplified_candidate]

        state_code = state_code_by_district_simple_name.get(simplified_candidate)
        if state_code:
            return state_name_by_code.get(state_code, "")

    return ""


def remap_condition_numbers(
    condition_text: str,
    index_map: dict[int, int],
) -> str:
    remapped_numbers: list[str] = []
    seen_numbers: set[int] = set()
    for part in condition_text.split(","):
        normalized = str(part).strip()
        if not normalized:
            continue

        try:
            original_index = int(normalized)
        except ValueError:
            continue

        new_index = index_map.get(original_index)
        if new_index is None or new_index in seen_numbers:
            continue

        seen_numbers.add(new_index)
        remapped_numbers.append(str(new_index))

    return ", ".join(remapped_numbers)


def write_germany_state_json_files(
    *,
    output_dir: Path,
    programs: list[dict[str, object]],
    universities: list[dict[str, object]],
    placement_conditions: list[dict[str, str]],
    germany_state_context: dict[str, object],
) -> dict[str, object]:
    output_dir.mkdir(parents=True, exist_ok=True)

    universities_by_state: dict[str, list[dict[str, object]]] = defaultdict(list)
    unresolved_universities: list[str] = []
    for university in universities:
        state_name = resolve_germany_state_name(
            city_id=university.get("city_id"),
            context=germany_state_context,
        )
        if not state_name:
            unresolved_universities.append(str(university.get("name") or ""))
            continue
        universities_by_state[state_name].append(university)

    program_by_provider_id = {
        str(program.get("provider_id") or ""): program for program in programs if str(program.get("provider_id") or "").strip()
    }

    manifest: list[dict[str, object]] = []
    for state_name in sorted(universities_by_state):
        state_universities_payload: list[dict[str, object]] = []
        used_provider_ids: set[str] = set()
        used_condition_indexes: set[int] = set()

        for university in universities_by_state[state_name]:
            for university_program in university.get("university_programs", []):
                provider_id = str(university_program.get("program_provider_id") or "").strip()
                if provider_id:
                    used_provider_ids.add(provider_id)

                for part in str(university_program.get("conditions") or "").split(","):
                    normalized = part.strip()
                    if not normalized:
                        continue
                    try:
                        used_condition_indexes.add(int(normalized))
                    except ValueError:
                        continue

        sorted_condition_indexes = sorted(
            index for index in used_condition_indexes if 1 <= index <= len(placement_conditions)
        )
        condition_index_map = {
            original_index: new_index
            for new_index, original_index in enumerate(sorted_condition_indexes, start=1)
        }

        state_placement_conditions = [
            placement_conditions[original_index - 1] for original_index in sorted_condition_indexes
        ]
        state_programs = [
            program
            for program in programs
            if str(program.get("provider_id") or "").strip() in used_provider_ids
        ]

        for university in universities_by_state[state_name]:
            state_university_programs: list[dict[str, object]] = []
            for university_program in university.get("university_programs", []):
                state_university_programs.append(
                    {
                        "program_provider_id": university_program.get("program_provider_id"),
                        "name": university_program.get("name"),
                        "university_program_code": university_program.get("university_program_code"),
                        "year": university_program.get("year"),
                        "conditions": remap_condition_numbers(
                            str(university_program.get("conditions") or ""),
                            condition_index_map,
                        ),
                        "details": university_program.get("details"),
                    }
                )

            state_universities_payload.append(
                {
                    "city_id": university.get("city_id"),
                    "name": university.get("name"),
                    "code": university.get("code"),
                    "type": university.get("type"),
                    "university_programs": state_university_programs,
                }
            )

        payload = {
            "country_id": 78,
            "programs": state_programs,
            "universities": state_universities_payload,
            "placement_conditions": state_placement_conditions,
        }

        file_name = f"78-germany-{slugify_identifier(state_name)}.json"
        output_path = output_dir / file_name
        output_path.write_text(
            json.dumps(payload, ensure_ascii=False, indent=2),
            encoding="utf-8",
        )
        manifest.append(
            {
                "state": state_name,
                "file": file_name,
                "programs": len(state_programs),
                "universities": len(state_universities_payload),
                "placement_conditions": len(state_placement_conditions),
            }
        )

    (output_dir / "manifest.json").write_text(
        json.dumps(manifest, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )

    return {
        "germany_state_json_files": len(manifest),
        "germany_state_universities": sum(len(items) for items in universities_by_state.values()),
        "germany_state_unresolved": len(unresolved_universities),
        "germany_state_unresolved_names": unresolved_universities,
    }


def main() -> int:
    args = build_parser().parse_args()

    whed_file = Path(args.whed_file)
    output_dir = Path(args.output_dir)
    countries_file = Path(args.countries_file)
    cities_file = Path(args.cities_file)
    districts_file = Path(args.districts_file)
    holland_matches_file = Path(args.holland_matches_file)

    institution_rows = load_full_institution_rows(whed_file)
    country_lookup = build_country_lookup(countries_file)
    city_exact_index, city_simplified_index, city_iso2_index = build_city_indexes(cities_file)
    district_exact_index, district_simplified_index = build_district_indexes(districts_file)

    usage_index = build_whed_admission_requirement_usage_index(whed_file)
    country_variants_by_base_country = build_country_variants_by_base_country(institution_rows, usage_index)

    placement_conditions_by_country, placement_condition_positions, condition_stats = build_placement_conditions(
        whed_file=whed_file,
        country_lookup=country_lookup,
        country_variants_by_base_country=country_variants_by_base_country,
    )

    programs_by_country, provider_id_by_program_key, program_stats = build_programs(
        institution_rows=institution_rows,
        country_lookup=country_lookup,
        holland_matches_file=holland_matches_file,
    )

    universities_by_country, university_stats = build_universities(
        institution_rows=institution_rows,
        country_lookup=country_lookup,
        city_exact_index=city_exact_index,
        city_simplified_index=city_simplified_index,
        city_iso2_index=city_iso2_index,
        district_exact_index=district_exact_index,
        district_simplified_index=district_simplified_index,
        placement_condition_positions=placement_condition_positions,
        country_variants_by_base_country=country_variants_by_base_country,
        provider_id_by_program_key=provider_id_by_program_key,
    )

    country_names_by_id: dict[object, str] = {}
    for normalized_country_name, (country_id, country_name) in country_lookup.items():
        if country_id not in country_names_by_id:
            country_names_by_id[country_id] = country_name
        if normalized_country_name == normalize_text(country_name):
            country_names_by_id[country_id] = country_name

    write_stats = write_country_json_files(
        output_dir=output_dir,
        country_names_by_id=country_names_by_id,
        programs_by_country=programs_by_country,
        universities_by_country=universities_by_country,
        placement_conditions_by_country=placement_conditions_by_country,
    )

    us_state_placement_conditions, us_state_condition_positions, us_state_condition_stats = (
        build_us_state_placement_conditions(
            whed_file=whed_file,
            country_lookup=country_lookup,
            country_variants_by_base_country=country_variants_by_base_country,
        )
    )
    us_state_programs, us_state_provider_ids, us_state_program_stats = build_us_state_programs(
        institution_rows=institution_rows,
        country_lookup=country_lookup,
        holland_matches_file=holland_matches_file,
    )
    us_state_universities, us_state_university_stats = build_us_state_universities(
        institution_rows=institution_rows,
        country_lookup=country_lookup,
        city_exact_index=city_exact_index,
        city_simplified_index=city_simplified_index,
        city_iso2_index=city_iso2_index,
        district_exact_index=district_exact_index,
        district_simplified_index=district_simplified_index,
        placement_condition_positions_by_state=us_state_condition_positions,
        country_variants_by_base_country=country_variants_by_base_country,
        provider_id_by_state_program_key=us_state_provider_ids,
    )

    us_state_output_dir = output_dir / "UnitedStatesStates"
    us_state_write_stats = write_us_state_json_files(
        output_dir=us_state_output_dir,
        country_id=229,
        programs_by_state=us_state_programs,
        universities_by_state=us_state_universities,
        placement_conditions_by_state=us_state_placement_conditions,
    )

    germany_state_context = build_germany_state_resolution_context(
        cities_file=cities_file,
        districts_file=districts_file,
    )
    germany_state_output_dir = output_dir / "GermanyStates"
    germany_state_write_stats = write_germany_state_json_files(
        output_dir=germany_state_output_dir,
        programs=programs_by_country.get(78, []),
        universities=universities_by_country.get(78, []),
        placement_conditions=placement_conditions_by_country.get(78, []),
        germany_state_context=germany_state_context,
    )

    summary = {
        "source_file": str(whed_file.resolve()),
        "output_dir": str(output_dir.resolve()),
        "institution_rows_scanned": len(institution_rows),
        **condition_stats,
        **program_stats,
        **university_stats,
        **write_stats,
        **us_state_condition_stats,
        **us_state_program_stats,
        **us_state_university_stats,
        **us_state_write_stats,
        **germany_state_write_stats,
    }
    (output_dir / "summary.json").write_text(
        json.dumps(summary, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )

    print(f"[done] Institution rows scanned: {summary['institution_rows_scanned']}", flush=True)
    print(f"[done] Country JSON files written: {summary['country_json_files']}", flush=True)
    print(f"[done] Programs exported: {summary['program_unique_rows']}", flush=True)
    print(f"[done] Universities exported: {summary['university_rows_used']}", flush=True)
    print(f"[done] University programs exported: {summary['university_program_rows']}", flush=True)
    print(f"[done] Placement conditions exported: {summary['condition_rows_used']}", flush=True)
    print(f"[done] US state JSON files written: {summary['us_state_json_files']}", flush=True)
    print(f"[done] Germany state JSON files written: {summary['germany_state_json_files']}", flush=True)
    print(f"[warn] Country misses (programs): {summary['program_country_misses']}", flush=True)
    print(f"[warn] Country misses (universities): {summary['university_country_misses']}", flush=True)
    print(f"[warn] City misses: {summary['university_city_misses']}", flush=True)
    print(
        f"[warn] Missing condition references in university_programs: {summary['condition_refs_missing_from_array']}",
        flush=True,
    )
    print(f"[done] Summary file: {(output_dir / 'summary.json').resolve()}", flush=True)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
