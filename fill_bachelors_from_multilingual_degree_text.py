from __future__ import annotations

import argparse
import re
import shutil
import unicodedata
from pathlib import Path
import sys

from openpyxl import load_workbook

from isced_f import clean_program_title
from whed_excel_export import is_noise_program_name


WORKBOOK_PATH = Path("whed_data.xlsx")
BACKUP_SUFFIX = ".bachelors_multilingual_backup"
DEGREES_COLUMN = "Degrees"
BACHELORS_COLUMN = "Bachelor's Degree"
RAW_TEXT_COLUMN = "Raw Text"
UNIVERSITY_NAME_COLUMN = "University Name"
IAU_CODE_COLUMN = "IAU Code"
SECTION_TITLES = {
    "General Information",
    "Officers",
    "Divisions",
    "Degrees",
    "Academic Periodicals",
    "Student & Staff Numbers",
}
NOTE_LABELS = {"note", "notes", "remarks", "remark"}
FIELD_LABELS = {"fields of study"}
META_LABELS = {"more details", "duration", "main press"}
NOISE_ITEM_PATTERNS = (
    re.compile(r"^\d+\s*(?:yr|yrs|year|years)\b", re.IGNORECASE),
    re.compile(r"^(?:also|double degree|dual degree|joint degree|programme|programmes?)\b", re.IGNORECASE),
)

BACHELOR_TITLE_PATTERNS = (
    re.compile(r"\bbachelor(?:'s)?\b", re.IGNORECASE),
    re.compile(r"\bbachelier\b", re.IGNORECASE),
    re.compile(r"\bbaccalaure(?:at|us)\b", re.IGNORECASE),
    re.compile(r"\bbacharel(?:ado)?\b", re.IGNORECASE),
    re.compile(r"\bbakala[a-z]*\b", re.IGNORECASE),
    re.compile(r"\balapfokozat\b", re.IGNORECASE),
    re.compile(r"\balapkepz[a-z]*\b", re.IGNORECASE),
    re.compile(r"\bkandidaatin\b", re.IGNORECASE),
    re.compile(r"\bkandidatexamen\b", re.IGNORECASE),
    re.compile(r"\blicence\b", re.IGNORECASE),
    re.compile(r"\blicencia(?:do|tura)?\b", re.IGNORECASE),
    re.compile(r"\blicencjat\b", re.IGNORECASE),
    re.compile(r"\blicenta\b", re.IGNORECASE),
    re.compile(r"\blaurea\b(?!\s+magistrale)", re.IGNORECASE),
    re.compile(r"\bprimo livello\b", re.IGNORECASE),
    re.compile(r"\bundergraduate\b", re.IGNORECASE),
    re.compile(r"\bfirst cycle\b", re.IGNORECASE),
    re.compile(r"\b1\.\s*stopnje\b", re.IGNORECASE),
    re.compile(r"\bprvostup[a-z]*\b", re.IGNORECASE),
    re.compile(r"\bpreddiplom[a-z]*\b", re.IGNORECASE),
    re.compile(r"\bprofesinio bakalauro\b", re.IGNORECASE),
    re.compile(r"\bbenke\b", re.IGNORECASE),
    re.compile(r"\bdiploma accademico di primo livello\b", re.IGNORECASE),
)

MASTER_TITLE_PATTERNS = (
    re.compile(r"\bmaster\b", re.IGNORECASE),
    re.compile(r"\bmaitris[a-z]*\b", re.IGNORECASE),
    re.compile(r"\bmagister\b", re.IGNORECASE),
    re.compile(r"\bmagistr\b", re.IGNORECASE),
    re.compile(r"\bmaisterim\b", re.IGNORECASE),
    re.compile(r"\blaurea magistrale\b", re.IGNORECASE),
    re.compile(r"\bsecondo livello\b", re.IGNORECASE),
    re.compile(r"\b2\.\s*stopnje\b", re.IGNORECASE),
    re.compile(r"\bdiplomskog\b", re.IGNORECASE),
    re.compile(r"\bgraduate\b", re.IGNORECASE),
    re.compile(r"\bkandidatgrad\b", re.IGNORECASE),
    re.compile(r"\blicentiat\b", re.IGNORECASE),
    re.compile(r"\blisensiaatti\b", re.IGNORECASE),
    re.compile(r"\bmetaptychiako\b", re.IGNORECASE),
    re.compile(r"\bdiplom[a-z]* master\b", re.IGNORECASE),
)

DOCTOR_TITLE_PATTERNS = (
    re.compile(r"\bdoctor\b", re.IGNORECASE),
    re.compile(r"\bdoktor\b", re.IGNORECASE),
    re.compile(r"\bph\.?\s*d\b", re.IGNORECASE),
    re.compile(r"\bdoctorat\b", re.IGNORECASE),
    re.compile(r"\bdottorato\b", re.IGNORECASE),
    re.compile(r"\bdoctoral\b", re.IGNORECASE),
    re.compile(r"\bdidaktoriko\b", re.IGNORECASE),
    re.compile(r"\bdaktaro\b", re.IGNORECASE),
)

OTHER_DEGREE_TITLE_PATTERNS = (
    re.compile(r"\bdiploma\b", re.IGNORECASE),
    re.compile(r"\bdiplom[a-z-]*\b", re.IGNORECASE),
    re.compile(r"\bcertificate\b", re.IGNORECASE),
    re.compile(r"\bcertificat[a-z]*\b", re.IGNORECASE),
    re.compile(r"\bassociate degree\b", re.IGNORECASE),
    re.compile(r"\bhigher diploma\b", re.IGNORECASE),
    re.compile(r"\bundergraduate diploma\b", re.IGNORECASE),
    re.compile(r"\bexamen\b", re.IGNORECASE),
    re.compile(r"\btutkinto\b", re.IGNORECASE),
    re.compile(r"\bgraad\b", re.IGNORECASE),
    re.compile(r"\boklevel\b", re.IGNORECASE),
)

BACHELOR_NOTE_PATTERNS = (
    re.compile(r"\bBachelor\s+with\s+Honours\s+in\s+([^.;]+)", re.IGNORECASE),
    re.compile(r"\bBachelor(?:'s)?\s+degree(?:\s+\([^)]*\))?\s+in\s+([^.;]+)", re.IGNORECASE),
    re.compile(r"\bApplied\s+Bachelor(?:'s)?\s+degree(?:\s+\([^)]*\))?\s+in\s+([^.;]+)", re.IGNORECASE),
    re.compile(r"\bCollaborative\s+Bachelor(?:'s)?\s+degree\s+in\s+([^.;]+)", re.IGNORECASE),
    re.compile(r"\bJoint\s+Bachelor(?:'s)?\s+degree\s+programmes?\s+in\s+([^.;]+)", re.IGNORECASE),
    re.compile(r"\bBakalaurs?\s+programmes?\s+in\s+([^.;]+)", re.IGNORECASE),
    re.compile(r"\bfollowing\s+Bachelor\s+\(([^)]+)\)", re.IGNORECASE),
    re.compile(r"\bR\.?N\.?\s+to\s+Bachelor\s+of\s+Science\s+in\s+([^.;]+)", re.IGNORECASE),
    re.compile(r"\bBachelor\s+of\s+[^/;,.]+", re.IGNORECASE),
)

PROGRAM_PREFIX_PATTERNS = (
    re.compile(r"^(?:joint|concurrent|collaborative|applied|advanced|honours?|professional)\s+", re.IGNORECASE),
    re.compile(r"^also\s+", re.IGNORECASE),
    re.compile(r"^r\.?n\.?\s+to\s+", re.IGNORECASE),
    re.compile(r"^bachelor(?:'s)?(?:\s+degree)?(?:\s+\([^)]*\))?\s+(?:program(?:me)?s?\s+)?in\s+", re.IGNORECASE),
    re.compile(r"^bachelor\s+with\s+honours\s+in\s+", re.IGNORECASE),
    re.compile(r"^bachelor(?:'s)?(?:\s+degree)?(?:\s+\([^)]*\))?\s+", re.IGNORECASE),
    re.compile(r"^bachelor\s+of\s+", re.IGNORECASE),
)

PROGRAM_TRAILING_PATTERNS = (
    re.compile(r",?\s*\d+(?:[.,]\d+)?\s*(?:yr|yrs|year|years)\b.*$", re.IGNORECASE),
    re.compile(r"\s*\([^)]*(?:honours?|professional|degree|programme|program)[^)]*\)", re.IGNORECASE),
    re.compile(r"\s+\b(?:degree|degrees|programme|program|programmes|courses)\b.*$", re.IGNORECASE),
)


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description="Backfill blank Bachelor's Degree cells by parsing multilingual degree titles."
    )
    parser.add_argument(
        "--workbook",
        default=str(WORKBOOK_PATH),
        help="Workbook to update.",
    )
    parser.add_argument(
        "--apply",
        action="store_true",
        help="Write updates back to the workbook. Without this flag the script only previews changes.",
    )
    parser.add_argument(
        "--limit",
        type=int,
        default=20,
        help="Preview at most this many updated rows.",
    )
    return parser


def normalize_space(value: object) -> str:
    return re.sub(r"\s+", " ", str(value or "").replace("\xa0", " ")).strip()


def ascii_fold(value: object) -> str:
    return "".join(
        char
        for char in unicodedata.normalize("NFKD", str(value or ""))
        if not unicodedata.combining(char)
    )


def folded_key(value: object) -> str:
    return normalize_space(ascii_fold(value)).casefold()


def split_lines(value: object) -> list[str]:
    return [normalize_space(line) for line in str(value or "").splitlines() if normalize_space(line)]


def split_key_value(line: str) -> tuple[str | None, str | None]:
    if ":" not in line:
        return None, None
    key, value = line.split(":", 1)
    key = normalize_space(key)
    value = normalize_space(value)
    if not key:
        return None, None
    return key, value


def unique_preserve_order(values: list[str]) -> list[str]:
    result: list[str] = []
    seen: set[str] = set()
    for value in values:
        key = folded_key(value)
        if not key or key in seen:
            continue
        seen.add(key)
        result.append(value)
    return result


def backup_workbook(path: Path) -> Path:
    backup_path = path.with_name(f"{path.stem}{BACKUP_SUFFIX}{path.suffix}")
    shutil.copy2(path, backup_path)
    return backup_path


def looks_like_degree_heading(line: str) -> bool:
    normalized = normalize_space(line)
    if not normalized or normalized.endswith(":"):
        return False
    if split_key_value(normalized)[0] is not None:
        return False

    lowered = folded_key(normalized)
    if lowered in NOTE_LABELS | FIELD_LABELS | META_LABELS:
        return False

    if normalized.count(",") > 2:
        return False
    if len(normalized.split()) > 18:
        return False
    return True


def classify_degree_title(title: object) -> str:
    normalized = normalize_space(title)
    if not looks_like_degree_heading(normalized):
        return ""

    folded_normalized = folded_key(normalized)
    if any(token in folded_normalized for token in ("sub-bachelor", "sub bachelor", "post-baccalaureate", "post bachelor", "post-bachelor")):
        return "other"
    if any(pattern.search(folded_normalized) for pattern in BACHELOR_TITLE_PATTERNS):
        return "bachelor"
    if any(pattern.search(folded_normalized) for pattern in MASTER_TITLE_PATTERNS):
        return "master"
    if any(pattern.search(folded_normalized) for pattern in DOCTOR_TITLE_PATTERNS):
        return "doctor"
    if any(pattern.search(folded_normalized) for pattern in OTHER_DEGREE_TITLE_PATTERNS):
        return "other"
    return ""


def strip_subject_tail(value: str) -> str:
    return re.split(r"\b(?:Note|Main Press|More details|Duration|Remarks?)\b", value, maxsplit=1)[0].strip()


def split_program_candidates(value: str) -> list[str]:
    normalized = strip_subject_tail(normalize_space(value))
    if not normalized:
        return []
    return [normalize_space(part) for part in re.split(r"\s*[;,]\s*", normalized) if normalize_space(part)]


def clean_candidate_title(candidate: str, *, source_title: str = "") -> str:
    value = normalize_space(candidate).strip(" -/|:")
    if not value:
        return ""

    for pattern in PROGRAM_PREFIX_PATTERNS:
        value = pattern.sub("", value)
    for pattern in PROGRAM_TRAILING_PATTERNS:
        value = pattern.sub("", value)

    value = normalize_space(value).strip(" -/|:")
    value = re.sub(r"^(Arts|Science|Education|Fine Arts|Physical Education)\s+(?!and\b)(.+)$", r"\2", value)
    value = clean_program_title(value)
    value = normalize_space(value).strip(" -/|:")
    if not value:
        return ""

    lowered = folded_key(value)
    source_key = folded_key(source_title)
    if source_key and lowered and lowered != source_key and lowered in source_key:
        return ""
    if any(token in lowered for token in ("master", "doctor", "phd", "doktor", "magister", "degree course")):
        return ""
    if lowered in {"degree", "degrees", "llb"}:
        return ""
    if re.fullmatch(r"[A-Z]{2,6}", value):
        return ""

    if any(pattern.match(value) for pattern in NOISE_ITEM_PATTERNS):
        return ""
    if is_noise_program_name(value):
        return ""
    return value


def extract_programs_from_subject_block(value: str, *, source_title: str = "") -> list[str]:
    programs: list[str] = []
    for raw_item in split_program_candidates(value):
        cleaned = clean_candidate_title(raw_item, source_title=source_title)
        if cleaned:
            programs.append(cleaned)
    return unique_preserve_order(programs)


def extract_bachelor_programs_from_degree_text(value: object) -> list[str]:
    programs: list[str] = []
    pending_titles: list[tuple[str, str]] = []
    collecting_fields = False
    current_fields: list[str] = []
    note_lines: list[str] = []
    in_note = False

    def flush_current() -> None:
        nonlocal current_fields, collecting_fields
        bachelor_titles = [title for kind, title in pending_titles if kind == "bachelor"]
        if bachelor_titles and current_fields:
            programs.extend(
                extract_programs_from_subject_block(
                    " ".join(current_fields),
                    source_title=" | ".join(bachelor_titles),
                )
            )
        current_fields = []
        collecting_fields = False
        pending_titles.clear()

    for line in split_lines(value):
        key, key_value = split_key_value(line)
        lowered_key = folded_key(key or line)

        if lowered_key in NOTE_LABELS:
            flush_current()
            in_note = True
            continue

        if in_note:
            note_lines.append(line)
            continue

        degree_kind = classify_degree_title(line)
        if degree_kind:
            if collecting_fields:
                flush_current()
            pending_titles.append((degree_kind, line))
            continue

        if key is not None:
            if folded_key(key) in FIELD_LABELS:
                collecting_fields = True
                current_fields = [key_value] if key_value else []
                continue
            if collecting_fields:
                flush_current()
            continue

        if collecting_fields:
            current_fields.append(line)

    flush_current()
    programs.extend(extract_bachelor_programs_from_notes(" ".join(note_lines)))
    return unique_preserve_order(programs)


def extract_bachelor_programs_from_notes(note_text: str) -> list[str]:
    normalized_note = normalize_space(note_text)
    if not normalized_note:
        return []

    programs: list[str] = []
    for pattern in BACHELOR_NOTE_PATTERNS:
        for match in pattern.finditer(normalized_note):
            if match.groups():
                values = [group for group in match.groups() if group]
            else:
                values = [match.group(0)]
            for raw_value in values:
                programs.extend(extract_programs_from_bachelor_phrase(raw_value))
    return unique_preserve_order(programs)


def extract_programs_from_bachelor_phrase(phrase: str) -> list[str]:
    normalized = normalize_space(phrase)
    if not normalized:
        return []

    programs: list[str] = []

    nested_bachelor_matches = list(re.finditer(r"\bBachelor\s+of\s+[^/;]+", normalized, flags=re.IGNORECASE))
    if len(nested_bachelor_matches) > 1:
        for match in nested_bachelor_matches:
            programs.extend(extract_programs_from_bachelor_phrase(match.group(0)))
        return unique_preserve_order(programs)

    if normalized.lower().startswith("following bachelor"):
        inner_match = re.search(r"\(([^)]+)\)", normalized)
        if inner_match:
            return extract_programs_from_subject_block(inner_match.group(1), source_title=normalized)
        return []

    if re.search(r"\bBakalaurs?\s+programmes?\s+in\b", normalized, flags=re.IGNORECASE):
        normalized = re.sub(r"^\bBakalaurs?\s+programmes?\s+in\s+", "", normalized, flags=re.IGNORECASE)

    cleaned = clean_candidate_title(normalized)
    if cleaned and " and " not in cleaned:
        return [cleaned]

    split_attempts: list[str] = []
    if "/" in normalized:
        split_attempts.extend(part for part in normalized.split("/") if normalize_space(part))
    if "," in normalized:
        split_attempts.extend(part for part in normalized.split(",") if normalize_space(part))

    if split_attempts:
        for part in split_attempts:
            cleaned_part = clean_candidate_title(part)
            if cleaned_part:
                programs.append(cleaned_part)
        if programs:
            return unique_preserve_order(programs)

    cleaned = clean_candidate_title(normalized)
    return [cleaned] if cleaned else []


def extract_bachelor_programs_from_structured_fields(row: tuple[object, ...], index_by_name: dict[str, int]) -> list[str]:
    programs: list[str] = []
    for degree_field_number in range(1, 8):
        type_value = row[index_by_name[f"Degree Fields {degree_field_number} Type"]]
        title_value = row[index_by_name[f"Degree Fields {degree_field_number} Title"]]
        subjects_value = row[index_by_name[f"Degree Fields {degree_field_number} Subjects"]]
        if not normalize_space(subjects_value):
            continue

        title_kind = classify_degree_title(title_value)
        type_kind = classify_degree_title(type_value)
        if title_kind != "bachelor" and type_kind != "bachelor":
            continue

        programs.extend(
            extract_programs_from_subject_block(
                str(subjects_value),
                source_title=str(title_value or type_value or ""),
            )
        )
    return unique_preserve_order(programs)


def extract_degree_section_from_raw_text(raw_text: object) -> str:
    lines = split_lines(raw_text)
    if not lines:
        return ""

    in_degrees = False
    degree_lines: list[str] = []
    for line in lines:
        if line == "Degrees":
            in_degrees = True
            continue
        if in_degrees and line in SECTION_TITLES - {"Degrees"}:
            break
        if in_degrees:
            degree_lines.append(line)
    return "\n".join(degree_lines)


def find_bachelor_programs(row: tuple[object, ...], index_by_name: dict[str, int]) -> list[str]:
    programs = extract_bachelor_programs_from_structured_fields(row, index_by_name)

    degree_text = row[index_by_name[DEGREES_COLUMN]]
    if normalize_space(degree_text):
        programs.extend(extract_bachelor_programs_from_degree_text(degree_text))

    if not programs:
        raw_degree_text = extract_degree_section_from_raw_text(row[index_by_name[RAW_TEXT_COLUMN]])
        if raw_degree_text:
            programs.extend(extract_bachelor_programs_from_degree_text(raw_degree_text))

    return unique_preserve_order(programs)


def is_blank_bachelors(value: object) -> bool:
    normalized = normalize_space(value)
    return not normalized or normalized.casefold() == "not available"


def run(workbook_path: Path, *, apply_changes: bool, preview_limit: int) -> dict[str, object]:
    workbook = load_workbook(workbook_path)
    sheet = workbook["Institutions"]
    header = [str(cell.value) if cell.value is not None else "" for cell in sheet[1]]
    index_by_name = {name: position for position, name in enumerate(header)}
    bachelors_column_index = index_by_name[BACHELORS_COLUMN] + 1

    blank_rows = 0
    updated_rows = 0
    previews: list[dict[str, object]] = []

    for row_number in range(2, sheet.max_row + 1):
        row_values = [sheet.cell(row=row_number, column=column_number).value for column_number in range(1, len(header) + 1)]
        if not is_blank_bachelors(row_values[index_by_name[BACHELORS_COLUMN]]):
            continue

        blank_rows += 1
        programs = find_bachelor_programs(tuple(row_values), index_by_name)
        if not programs:
            continue

        new_value = ", ".join(programs)
        updated_rows += 1
        if len(previews) < preview_limit:
            previews.append(
                {
                    "row": row_number,
                    "iau_code": normalize_space(row_values[index_by_name[IAU_CODE_COLUMN]]),
                    "university_name": normalize_space(row_values[index_by_name[UNIVERSITY_NAME_COLUMN]]),
                    "value": new_value,
                }
            )

        if apply_changes:
            sheet.cell(row=row_number, column=bachelors_column_index).value = new_value

    backup_path: Path | None = None
    if apply_changes and updated_rows:
        backup_path = backup_workbook(workbook_path)
        workbook.save(workbook_path)

    return {
        "blank_rows": blank_rows,
        "updated_rows": updated_rows,
        "remaining_blank_rows": blank_rows - updated_rows,
        "preview": previews,
        "backup_path": str(backup_path) if backup_path else "",
    }


def main() -> None:
    if hasattr(sys.stdout, "reconfigure"):
        sys.stdout.reconfigure(encoding="utf-8")
    args = build_parser().parse_args()
    result = run(Path(args.workbook), apply_changes=args.apply, preview_limit=args.limit)

    mode = "apply" if args.apply else "dry-run"
    print(f"[{mode}] Blank Bachelor's rows: {result['blank_rows']}")
    print(f"[{mode}] Rows with inferred Bachelor's programs: {result['updated_rows']}")
    print(f"[{mode}] Remaining blank rows: {result['remaining_blank_rows']}")
    if result["backup_path"]:
        print(f"[apply] Backup created: {result['backup_path']}")

    preview = result["preview"]
    if preview:
        print("[preview]")
        for item in preview:
            print(f"- row {item['row']} | {item['iau_code']} | {item['university_name']}")
            print(f"  {item['value']}")


if __name__ == "__main__":
    main()
