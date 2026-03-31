import argparse
import bisect
import json
import re
import shutil
from collections import defaultdict
from concurrent.futures import ThreadPoolExecutor, as_completed
from pathlib import Path
from statistics import mean
from typing import Any

import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook

from add_student_costs import (
    GeocodeCache,
    LIVINGCOST_HEADERS,
    LocationKey,
    PageRef,
    ROW_VALUE_KEYS,
    build_page_index,
    compute_cost_columns,
    fetch_with_retry,
    normalize_text,
    page_cache_path,
    resolve_country_slug,
    resolve_page_for_location,
    split_country_value,
)


OUTPUT_COLUMNS = [
    "City Universities Count",
    "City Known Students",
    "Livingcost Quality of Life Index",
    "Livingcost Monthly Salary After Tax (USD)",
    "Student Cost / Salary Ratio",
    "Livingcost Lunch Menu (USD)",
    "Livingcost Fast Food Meal (USD)",
    "Livingcost Cappuccino (USD)",
    "Livingcost Cinema Ticket (USD)",
    "Livingcost Internet Plan (USD)",
    "Livingcost Monthly Transport Pass (USD)",
    "Livingcost Local Transport Ticket (USD)",
    "Livingcost Closest Airport Distance (km)",
    "Livingcost Air Quality",
    "Livingcost PM2.5 (ug/m3)",
    "Livingcost Best University Rank",
    "Student Friendliness Data Level",
    "Student Friendliness Data Completeness (%)",
    "Student Friendliness Affordability Score",
    "Student Friendliness Daily Life Score",
    "Student Friendliness Mobility Score",
    "Student Friendliness Environment Score",
    "Student Friendliness Academic Ecosystem Score",
    "Student Friendliness Score",
    "Student Friendliness Verdict",
    "Student Friendliness Summary",
    "Student Friendliness Source URL",
]

AIR_QUALITY_LABEL_SCORES = {
    "excellent": 95.0,
    "good": 82.0,
    "moderate": 60.0,
    "fair": 55.0,
    "poor": 35.0,
    "unhealthy for sensitive groups": 30.0,
    "unhealthy": 18.0,
    "very unhealthy": 8.0,
    "hazardous": 2.0,
}

RAW_NUMERIC_KEYS = [
    "shared_annual_cost_usd",
    "monthly_salary_after_tax_usd",
    "cost_to_salary_ratio",
    "quality_of_life_index",
    "lunch_menu_usd",
    "fast_food_meal_usd",
    "cappuccino_usd",
    "cinema_ticket_usd",
    "internet_plan_monthly_usd",
    "monthly_transport_pass_usd",
    "local_transport_ticket_usd",
    "closest_airport_distance_km",
    "air_quality_pm25",
    "best_university_rank",
    "city_university_count",
    "city_known_students",
]

METRIC_DIRECTIONS = {
    "shared_annual_cost_usd": "lower",
    "cost_to_salary_ratio": "lower",
    "quality_of_life_index": "higher",
    "lunch_menu_usd": "lower",
    "fast_food_meal_usd": "lower",
    "cappuccino_usd": "lower",
    "cinema_ticket_usd": "lower",
    "internet_plan_monthly_usd": "lower",
    "monthly_transport_pass_usd": "lower",
    "local_transport_ticket_usd": "lower",
    "closest_airport_distance_km": "lower",
    "air_quality_pm25": "lower",
    "best_university_rank": "lower",
    "city_university_count": "higher",
    "city_known_students": "higher",
}

CATEGORY_WEIGHTS = {
    "affordability": 0.30,
    "daily_life": 0.25,
    "mobility": 0.15,
    "environment": 0.10,
    "academic": 0.20,
}

CATEGORY_METRICS = {
    "affordability": [
        "shared_annual_cost_usd",
        "cost_to_salary_ratio",
        "fast_food_meal_usd",
        "cappuccino_usd",
    ],
    "daily_life": [
        "quality_of_life_index",
        "internet_plan_monthly_usd",
        "lunch_menu_usd",
        "cinema_ticket_usd",
    ],
    "mobility": [
        "monthly_transport_pass_usd",
        "local_transport_ticket_usd",
        "closest_airport_distance_km",
    ],
    "environment": [
        "air_quality_pm25",
        "air_quality_label_score",
    ],
    "academic": [
        "best_university_rank",
        "city_university_count",
        "city_known_students",
    ],
}


def backup_workbook(path: Path) -> Path:
    backup_path = path.with_name(f"{path.stem}.student_friendliness_backup{path.suffix}")
    shutil.copy2(path, backup_path)
    return backup_path


def ensure_output_columns(worksheet) -> dict[str, int]:
    header_map = {
        worksheet.cell(row=1, column=column_index).value: column_index
        for column_index in range(1, worksheet.max_column + 1)
    }
    next_column = worksheet.max_column + 1
    for header in OUTPUT_COLUMNS:
        if header not in header_map:
            worksheet.cell(row=1, column=next_column, value=header)
            header_map[header] = next_column
            next_column += 1
    return {header: header_map[header] for header in OUTPUT_COLUMNS}


def parse_number(value: Any) -> float | None:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)

    text = str(value).strip().replace(",", "")
    match = re.search(r"([0-9]+(?:\.[0-9]+)?)", text)
    if not match:
        return None
    number = float(match.group(1))
    suffix = text[match.end() :].strip().lower()
    if suffix in {"b", "bn"}:
        number *= 1_000_000_000
    elif suffix == "m":
        number *= 1_000_000
    elif suffix == "k":
        number *= 1_000
    return number


def parse_int_like(value: Any) -> int | None:
    number = parse_number(value)
    if number is None:
        return None
    return int(round(number))


def build_location_maps(worksheet) -> tuple[list[LocationKey], dict[int, LocationKey], dict[LocationKey, dict[str, int | None]]]:
    rows = worksheet.iter_rows(values_only=True)
    headers = list(next(rows))
    header_map = {header: idx for idx, header in enumerate(headers)}

    country_idx = header_map["Country"]
    province_idx = header_map["Province"]
    city_idx = header_map["City"]
    total_student_idx = header_map.get("Total Student")

    unique_locations: dict[LocationKey, None] = {}
    row_locations: dict[int, LocationKey] = {}
    location_stats: dict[LocationKey, dict[str, int]] = defaultdict(lambda: {"university_count": 0, "known_students": 0, "known_student_rows": 0})

    for row_number, row in enumerate(rows, start=2):
        country = str(row[country_idx] or "").strip()
        province = str(row[province_idx] or "").strip()
        city = str(row[city_idx] or "").strip()
        base_country, _ = split_country_value(country)
        location = LocationKey(country=country, base_country=base_country, province=province, city=city)
        unique_locations.setdefault(location, None)
        row_locations[row_number] = location
        location_stats[location]["university_count"] += 1

        if total_student_idx is not None:
            total_students = parse_int_like(row[total_student_idx])
            if total_students is not None and total_students > 0:
                location_stats[location]["known_students"] += total_students
                location_stats[location]["known_student_rows"] += 1

    normalized_stats: dict[LocationKey, dict[str, int | None]] = {}
    for location, stats in location_stats.items():
        normalized_stats[location] = {
            "university_count": stats["university_count"],
            "known_students": stats["known_students"] if stats["known_student_rows"] > 0 else None,
        }

    return list(unique_locations), row_locations, normalized_stats


def parse_air_quality_score(label: str | None) -> float | None:
    if not label:
        return None
    return AIR_QUALITY_LABEL_SCORES.get(normalize_text(label))


def parse_friendliness_page(
    page_ref: PageRef,
    session: requests.Session | None = None,
    cache_dir: Path | None = None,
) -> dict[str, Any]:
    if cache_dir is not None:
        cache_dir.mkdir(parents=True, exist_ok=True)
        cache_file = page_cache_path(cache_dir, page_ref.url)
        if cache_file.exists():
            payload = json.loads(cache_file.read_text(encoding="utf-8"))
            if payload.get("schema_version") == 1:
                return payload

    http = session or requests.Session()
    response = fetch_with_retry(http, page_ref.url, headers=LIVINGCOST_HEADERS, timeout=30)
    soup = BeautifulSoup(response.text, "html.parser")

    metrics: dict[str, Any] = {}
    for table in soup.find_all("table"):
        for row in table.find_all("tr"):
            cells = [cell.get_text(" ", strip=True) for cell in row.find_all(["th", "td"])]
            if len(cells) < 2:
                continue

            label = normalize_text(cells[0])
            row_value_key = ROW_VALUE_KEYS.get(label)
            if row_value_key:
                metrics[row_value_key] = parse_number(cells[1])

            if label == "quality of life":
                metrics["quality_of_life_index"] = parse_number(cells[1])
            elif label == "monthly salary after tax":
                metrics["monthly_salary_after_tax_usd"] = parse_number(cells[1])
            elif label == "population":
                metrics["population_estimate"] = parse_number(cells[1])
            elif label == "lunch menu":
                metrics["lunch_menu_usd"] = parse_number(cells[1])
            elif label == "fast food meal equiv mcdonald s":
                metrics["fast_food_meal_usd"] = parse_number(cells[1])
            elif label == "cappuccino":
                metrics["cappuccino_usd"] = parse_number(cells[1])
            elif label == "cinema ticket 1 person":
                metrics["cinema_ticket_usd"] = parse_number(cells[1])
            elif label == "internet plan 50 mbps 1 month unlimited":
                metrics["internet_plan_monthly_usd"] = parse_number(cells[1])
            elif label == "monthly ticket local transport":
                metrics["monthly_transport_pass_usd"] = parse_number(cells[1])
            elif label == "local transport ticket":
                metrics["local_transport_ticket_usd"] = parse_number(cells[1])
            elif label == "closest airport":
                metrics["closest_airport_distance_km"] = parse_number(cells[1])
            elif label == "air quality":
                metrics["air_quality_label"] = cells[1]
                if len(cells) >= 3:
                    metrics["air_quality_pm25"] = parse_number(cells[2])
            elif label == "best university rank":
                metrics["best_university_rank"] = parse_number(cells[1])

    payload = {
        "schema_version": 1,
        "url": page_ref.url,
        "source_level": page_ref.source_level,
        "country_slug": page_ref.country_slug,
        "region_slug": page_ref.region_slug,
        "city_slug": page_ref.city_slug,
        "metrics": metrics,
    }

    if cache_dir is not None:
        cache_file.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    return payload


def fetch_all_page_payloads(page_refs: list[PageRef], max_workers: int, page_cache_dir: Path) -> dict[str, dict[str, Any]]:
    payloads: dict[str, dict[str, Any]] = {}
    unique_page_refs = {page_ref.url: page_ref for page_ref in page_refs}
    print(f"[info] Fetching {len(unique_page_refs)} student-friendliness pages...", flush=True)

    def worker(page_ref: PageRef) -> dict[str, Any]:
        session = requests.Session()
        try:
            return parse_friendliness_page(page_ref, session=session, cache_dir=page_cache_dir)
        finally:
            session.close()

    with ThreadPoolExecutor(max_workers=max_workers) as executor:
        future_map = {
            executor.submit(worker, page_ref): page_ref.url
            for page_ref in unique_page_refs.values()
        }
        for future in as_completed(future_map):
            payloads[future_map[future]] = future.result()
    return payloads


def average(values: list[float | None]) -> float | None:
    filtered = [value for value in values if value is not None]
    if not filtered:
        return None
    return round(mean(filtered), 2)


def metric_score(value: float | None, values: list[float], direction: str) -> float | None:
    if value is None or not values:
        return None
    if len(values) == 1:
        return 50.0

    right = bisect.bisect_right(values, value)
    percentile = (right - 1) / (len(values) - 1)
    score = percentile * 100.0 if direction == "higher" else (1.0 - percentile) * 100.0
    return round(max(0.0, min(100.0, score)), 2)


def build_metric_distributions(records: dict[LocationKey, dict[str, Any]]) -> dict[str, list[float]]:
    distributions: dict[str, list[float]] = {}
    for metric_key in RAW_NUMERIC_KEYS:
        values = [
            float(record[metric_key])
            for record in records.values()
            if isinstance(record.get(metric_key), (int, float))
        ]
        distributions[metric_key] = sorted(values)

    air_quality_scores = [
        float(record["air_quality_label_score"])
        for record in records.values()
        if isinstance(record.get("air_quality_label_score"), (int, float))
    ]
    distributions["air_quality_label_score"] = sorted(air_quality_scores)
    return distributions


def build_scores(record: dict[str, Any], distributions: dict[str, list[float]]) -> dict[str, float | None]:
    metric_scores: dict[str, float | None] = {}
    for metric_key, direction in METRIC_DIRECTIONS.items():
        metric_scores[metric_key] = metric_score(
            record.get(metric_key),
            distributions[metric_key],
            direction,
        )

    metric_scores["air_quality_label_score"] = metric_score(
        record.get("air_quality_label_score"),
        distributions["air_quality_label_score"],
        "higher",
    )

    category_scores = {
        category: average([metric_scores.get(metric_key) for metric_key in metric_keys])
        for category, metric_keys in CATEGORY_METRICS.items()
    }

    weighted_parts: list[tuple[float, float]] = []
    for category, weight in CATEGORY_WEIGHTS.items():
        category_score = category_scores.get(category)
        if category_score is not None:
            weighted_parts.append((category_score, weight))

    final_score = None
    if weighted_parts:
        total_weight = sum(weight for _, weight in weighted_parts)
        final_score = round(sum(score * weight for score, weight in weighted_parts) / total_weight, 2)

    return {
        "affordability": category_scores["affordability"],
        "daily_life": category_scores["daily_life"],
        "mobility": category_scores["mobility"],
        "environment": category_scores["environment"],
        "academic": category_scores["academic"],
        "final": final_score,
    }


def data_completeness(record: dict[str, Any]) -> float:
    total_signals = 15
    available_signals = 0
    for metric_key in [
        "shared_annual_cost_usd",
        "cost_to_salary_ratio",
        "quality_of_life_index",
        "lunch_menu_usd",
        "fast_food_meal_usd",
        "cappuccino_usd",
        "cinema_ticket_usd",
        "internet_plan_monthly_usd",
        "monthly_transport_pass_usd",
        "local_transport_ticket_usd",
        "closest_airport_distance_km",
        "best_university_rank",
        "city_university_count",
        "city_known_students",
    ]:
        if record.get(metric_key) is not None:
            available_signals += 1

    if record.get("air_quality_pm25") is not None or record.get("air_quality_label_score") is not None:
        available_signals += 1

    return round(available_signals / total_signals * 100.0, 2)


def verdict_for_score(score: float | None) -> str | None:
    if score is None:
        return None
    if score >= 80:
        return "Very student-friendly"
    if score >= 60:
        return "Student-friendly"
    if score >= 45:
        return "Mixed"
    if score >= 35:
        return "Challenging"
    return "Not student-friendly"


def summary_for_record(record: dict[str, Any], scores: dict[str, float | None]) -> str | None:
    labels = {
        "affordability": "affordability",
        "daily_life": "day-to-day living",
        "mobility": "mobility",
        "environment": "air quality/environment",
        "academic": "academic ecosystem",
    }
    ranked = [
        (category, score)
        for category, score in scores.items()
        if category in labels and score is not None
    ]
    if not ranked:
        return None

    ranked.sort(key=lambda item: item[1], reverse=True)
    strengths = [labels[category] for category, score in ranked if score >= 60][:2]
    weaknesses = [labels[category] for category, score in ranked[::-1] if score <= 40][:1]

    parts: list[str] = []
    if strengths:
        parts.append(f"Strengths: {', '.join(strengths)}")
    if weaknesses:
        parts.append(f"Weakness: {', '.join(weaknesses)}")
    if record.get("source_level") != "city":
        parts.append(f"Uses {record.get('source_level')}-level fallback data")
    return "; ".join(parts) if parts else None


def location_record(payload: dict[str, Any], location_stats: dict[str, int | None]) -> dict[str, Any]:
    metrics = payload["metrics"]
    cost_values = compute_cost_columns(metrics)
    shared_annual_cost = cost_values.get("Estimated Annual Student Cost (Shared Housing, USD)")
    monthly_salary = metrics.get("monthly_salary_after_tax_usd")

    cost_to_salary_ratio = None
    if isinstance(shared_annual_cost, (int, float)) and isinstance(monthly_salary, (int, float)) and monthly_salary > 0:
        cost_to_salary_ratio = round(shared_annual_cost / (monthly_salary * 12.0), 4)

    return {
        "city_university_count": location_stats["university_count"],
        "city_known_students": location_stats["known_students"],
        "quality_of_life_index": metrics.get("quality_of_life_index"),
        "monthly_salary_after_tax_usd": monthly_salary,
        "shared_annual_cost_usd": shared_annual_cost,
        "cost_to_salary_ratio": cost_to_salary_ratio,
        "lunch_menu_usd": metrics.get("lunch_menu_usd"),
        "fast_food_meal_usd": metrics.get("fast_food_meal_usd"),
        "cappuccino_usd": metrics.get("cappuccino_usd"),
        "cinema_ticket_usd": metrics.get("cinema_ticket_usd"),
        "internet_plan_monthly_usd": metrics.get("internet_plan_monthly_usd"),
        "monthly_transport_pass_usd": metrics.get("monthly_transport_pass_usd"),
        "local_transport_ticket_usd": metrics.get("local_transport_ticket_usd"),
        "closest_airport_distance_km": metrics.get("closest_airport_distance_km"),
        "air_quality_label": metrics.get("air_quality_label"),
        "air_quality_pm25": metrics.get("air_quality_pm25"),
        "air_quality_label_score": parse_air_quality_score(metrics.get("air_quality_label")),
        "best_university_rank": metrics.get("best_university_rank"),
        "source_level": payload["source_level"],
        "source_url": payload["url"],
    }


def enrich_workbook(
    workbook_path: Path,
    sheet_name: str,
    max_workers: int,
    geocode_cache_path: Path,
    page_cache_dir: Path,
) -> dict[str, Any]:
    index = build_page_index()
    geocode_cache = GeocodeCache(geocode_cache_path)

    wb = load_workbook(workbook_path)
    ws = wb[sheet_name]
    unique_locations, row_locations, location_stats = build_location_maps(ws)
    print(f"[info] Unique locations: {len(unique_locations)}", flush=True)

    location_to_page: dict[LocationKey, PageRef] = {}
    page_refs: list[PageRef] = []
    source_counts: defaultdict[str, int] = defaultdict(int)
    for location in unique_locations:
        country_slug = resolve_country_slug(location.country, index)
        page_ref = resolve_page_for_location(location, country_slug, index, geocode_cache)
        location_to_page[location] = page_ref
        page_refs.append(page_ref)
        source_counts[page_ref.source_level] += 1

    payloads = fetch_all_page_payloads(page_refs, max_workers=max_workers, page_cache_dir=page_cache_dir)

    records: dict[LocationKey, dict[str, Any]] = {}
    for location in unique_locations:
        page_ref = location_to_page[location]
        records[location] = location_record(payloads[page_ref.url], location_stats[location])

    distributions = build_metric_distributions(records)
    location_scores: dict[LocationKey, dict[str, float | None]] = {}
    for location, record in records.items():
        location_scores[location] = build_scores(record, distributions)
        record["data_completeness"] = data_completeness(record)
        record["verdict"] = verdict_for_score(location_scores[location]["final"])
        record["summary"] = summary_for_record(record, location_scores[location])

    column_map = ensure_output_columns(ws)
    for row_number, location in row_locations.items():
        record = records[location]
        scores = location_scores[location]

        write_values = {
            "City Universities Count": record["city_university_count"],
            "City Known Students": record["city_known_students"],
            "Livingcost Quality of Life Index": record["quality_of_life_index"],
            "Livingcost Monthly Salary After Tax (USD)": record["monthly_salary_after_tax_usd"],
            "Student Cost / Salary Ratio": record["cost_to_salary_ratio"],
            "Livingcost Lunch Menu (USD)": record["lunch_menu_usd"],
            "Livingcost Fast Food Meal (USD)": record["fast_food_meal_usd"],
            "Livingcost Cappuccino (USD)": record["cappuccino_usd"],
            "Livingcost Cinema Ticket (USD)": record["cinema_ticket_usd"],
            "Livingcost Internet Plan (USD)": record["internet_plan_monthly_usd"],
            "Livingcost Monthly Transport Pass (USD)": record["monthly_transport_pass_usd"],
            "Livingcost Local Transport Ticket (USD)": record["local_transport_ticket_usd"],
            "Livingcost Closest Airport Distance (km)": record["closest_airport_distance_km"],
            "Livingcost Air Quality": record["air_quality_label"],
            "Livingcost PM2.5 (ug/m3)": record["air_quality_pm25"],
            "Livingcost Best University Rank": record["best_university_rank"],
            "Student Friendliness Data Level": record["source_level"],
            "Student Friendliness Data Completeness (%)": record["data_completeness"],
            "Student Friendliness Affordability Score": scores["affordability"],
            "Student Friendliness Daily Life Score": scores["daily_life"],
            "Student Friendliness Mobility Score": scores["mobility"],
            "Student Friendliness Environment Score": scores["environment"],
            "Student Friendliness Academic Ecosystem Score": scores["academic"],
            "Student Friendliness Score": scores["final"],
            "Student Friendliness Verdict": record["verdict"],
            "Student Friendliness Summary": record["summary"],
            "Student Friendliness Source URL": record["source_url"],
        }

        for header, value in write_values.items():
            ws.cell(row=row_number, column=column_map[header], value=value)

    backup_path = backup_workbook(workbook_path)
    wb.save(workbook_path)
    return {
        "rows": ws.max_row - 1,
        "page_count": len({page_ref.url for page_ref in page_refs}),
        "source_counts": dict(sorted(source_counts.items())),
        "backup_path": str(backup_path),
    }


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Add student-friendliness columns and a final score to the WHED Institutions sheet.",
    )
    parser.add_argument("--workbook", default="whed_data.xlsx", help="Workbook to enrich in place.")
    parser.add_argument("--sheet", default="Institutions", help="Worksheet name that contains university rows.")
    parser.add_argument("--max-workers", type=int, default=2, help="Maximum number of concurrent page fetches.")
    parser.add_argument(
        "--geocode-cache",
        default=".cache/nominatim_city_cache.json",
        help="Path to the Nominatim fallback cache file.",
    )
    parser.add_argument(
        "--page-cache-dir",
        default=".cache/livingcost_student_friendliness_pages",
        help="Directory used to cache parsed livingcost student-friendliness payloads.",
    )
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    results = enrich_workbook(
        workbook_path=Path(args.workbook),
        sheet_name=args.sheet,
        max_workers=args.max_workers,
        geocode_cache_path=Path(args.geocode_cache),
        page_cache_dir=Path(args.page_cache_dir),
    )
    print(json.dumps(results, ensure_ascii=False, indent=2), flush=True)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
