import argparse
import hashlib
import json
import re
import shutil
import time
import unicodedata
from collections import defaultdict
from concurrent.futures import ThreadPoolExecutor, as_completed
from dataclasses import dataclass
from difflib import SequenceMatcher, get_close_matches
from pathlib import Path
from typing import Any
from xml.etree import ElementTree as ET

import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook


SITEMAP_URLS = [f"https://livingcost.org/sitemap{i}.xml" for i in range(1, 7)]
LIVINGCOST_HEADERS = {
    "User-Agent": "WHED-Scrapping/1.0 (student cost enrichment)",
}
NOMINATIM_HEADERS = {
    "User-Agent": "WHED-Scrapping/1.0 (student cost enrichment)",
}
OUTPUT_COLUMNS = [
    "Estimated Annual Student Cost (Shared Housing, USD)",
    "Estimated Annual Student Cost (Solo Apartment, USD)",
    "Estimated Annual Base Living Cost Without Rent (USD)",
    "Estimated Annual Housing Cost Shared (USD)",
    "Estimated Annual Housing Cost Solo (USD)",
    "Estimated Annual Food Cost (USD)",
    "Estimated Annual Transport Cost (USD)",
    "Estimated Cost Source Level",
    "Estimated Cost Source URL",
]
COUNTRY_ALIASES = {
    "czechia": "czech-republic",
    "slovak republic": "slovakia",
    "united states of america": "united-states",
}
US_STATE_CODES = {
    "alabama": "al",
    "alaska": "ak",
    "arizona": "az",
    "arkansas": "ar",
    "california": "ca",
    "colorado": "co",
    "connecticut": "ct",
    "delaware": "de",
    "district of columbia": "dc",
    "florida": "fl",
    "georgia": "ga",
    "hawaii": "hi",
    "idaho": "id",
    "illinois": "il",
    "indiana": "in",
    "iowa": "ia",
    "kansas": "ks",
    "kentucky": "ky",
    "louisiana": "la",
    "maine": "me",
    "maryland": "md",
    "massachusetts": "ma",
    "michigan": "mi",
    "minnesota": "mn",
    "mississippi": "ms",
    "missouri": "mo",
    "montana": "mt",
    "nebraska": "ne",
    "nevada": "nv",
    "new hampshire": "nh",
    "new jersey": "nj",
    "new mexico": "nm",
    "new york": "ny",
    "north carolina": "nc",
    "north dakota": "nd",
    "ohio": "oh",
    "oklahoma": "ok",
    "oregon": "or",
    "pennsylvania": "pa",
    "rhode island": "ri",
    "south carolina": "sc",
    "south dakota": "sd",
    "tennessee": "tn",
    "texas": "tx",
    "utah": "ut",
    "vermont": "vt",
    "virginia": "va",
    "washington": "wa",
    "west virginia": "wv",
    "wisconsin": "wi",
    "wyoming": "wy",
}
CANADA_PROVINCE_CODES = {
    "alberta": "ab",
    "british columbia": "bc",
    "manitoba": "mb",
    "new brunswick": "nb",
    "newfoundland and labrador": "nl",
    "newfoundland labrador": "nl",
    "nova scotia": "ns",
    "ontario": "on",
    "prince edward island": "pe",
    "quebec": "qc",
    "saskatchewan": "sk",
    "yukon": "yt",
    "northwest territories": "nt",
    "nunavut": "nu",
}
PRIMARY_CITY_ALIASES = {
    ("austria", "wien"): ["vienna"],
    ("germany", "koln"): ["cologne"],
    ("germany", "muenchen"): ["munich"],
    ("germany", "munchen"): ["munich"],
    ("germany", "nurnberg"): ["nuremberg"],
    ("germany", "nuernberg"): ["nuremberg"],
    ("germany", "frankfurt am main"): ["frankfurt"],
    ("hungary", "budapest"): ["budapest"],
    ("italy", "firenze"): ["florence"],
    ("italy", "genova"): ["genoa"],
    ("italy", "milano"): ["milan"],
    ("italy", "napoli"): ["naples"],
    ("italy", "roma"): ["rome"],
    ("italy", "torino"): ["turin"],
    ("romania", "bucuresti"): ["bucharest"],
    ("romania", "bucurești"): ["bucharest"],
    ("spain", "sevilla"): ["seville"],
    ("czech-republic", "praha"): ["prague"],
    ("slovakia", "bratislava"): ["bratislava"],
}
GENERIC_CITY_SUFFIXES = (" city", " district", " municipality", " county", " region")
ROW_VALUE_KEYS = {
    "total with rent": "total_with_rent_monthly",
    "without rent": "without_rent_monthly",
    "rent and utilities": "rent_utilities_monthly",
    "food": "food_monthly",
    "transport": "transport_monthly",
    "1 bedroom apartment in city center 40 m2 or 430 ft2": "apartment_center_1br_monthly",
    "cheap 1 bedroom apartment 40 m2 or 430 ft2": "apartment_cheap_1br_monthly",
    "3 bedroom apartment in city center 80 m2 or 860 ft2": "apartment_center_3br_monthly",
    "cheap 3 bedroom apartment 80 m2 or 860 ft2": "apartment_cheap_3br_monthly",
    "utility bill one person electricity heating water etc": "utilities_one_person_monthly",
    "utility bill for a family electricity heating water etc": "utilities_family_monthly",
    "internet plan 50 mbps 1 month unlimited": "internet_monthly",
    "monthly ticket local transport": "monthly_transport_pass",
}


@dataclass(frozen=True)
class PageRef:
    url: str
    country_slug: str
    region_slug: str | None
    city_slug: str | None
    source_level: str


@dataclass(frozen=True)
class LocationKey:
    country: str
    base_country: str
    province: str
    city: str


def normalize_text(value: str | None) -> str:
    if not value:
        return ""
    folded = unicodedata.normalize("NFKD", value).encode("ascii", "ignore").decode("ascii")
    lowered = folded.lower().replace("&", " and ")
    lowered = lowered.replace("'", " ")
    lowered = re.sub(r"[^a-z0-9]+", " ", lowered)
    return " ".join(lowered.split())


def slugify(value: str | None) -> str:
    return "-".join(normalize_text(value).split())


def first_non_empty(*values: str | None) -> str:
    for value in values:
        if value and str(value).strip():
            return str(value).strip()
    return ""


def split_country_value(country_value: str) -> tuple[str, str]:
    if " - " not in country_value:
        return country_value.strip(), ""
    base_country, region = country_value.split(" - ", 1)
    return base_country.strip(), region.strip()


def ordered_unique(values: list[str]) -> list[str]:
    seen: set[str] = set()
    out: list[str] = []
    for value in values:
        if value and value not in seen:
            seen.add(value)
            out.append(value)
    return out


def city_slug_candidates(base_country_slug: str, city_value: str, province_value: str) -> list[str]:
    candidates: list[str] = []
    raw = first_non_empty(city_value, province_value)
    if not raw:
        return candidates

    raw_variants = [
        raw,
        re.sub(r"\(.*?\)", "", raw).strip(),
        re.sub(r"\s+\d+[a-zA-Z-]*$", "", raw).strip(),
        re.sub(r"\s+[A-Za-z]$", "", raw).strip(),
    ]
    if "," in raw:
        raw_variants.append(raw.split(",", 1)[0].strip())
    for separator in ("/", ";"):
        if separator in raw:
            raw_variants.extend(part.strip() for part in raw.split(separator))

    normalized_raw = normalize_text(raw)
    if normalized_raw.startswith("city of "):
        raw_variants.append(raw[8:].strip())
    if normalized_raw.startswith("st "):
        raw_variants.append(re.sub(r"(?i)^st\.?\s+", "Saint ", raw).strip())
    if normalized_raw.startswith("saint "):
        raw_variants.append(re.sub(r"(?i)^saint\s+", "St ", raw).strip())

    raw_variants = [value for value in raw_variants if value]
    for value in list(raw_variants):
        norm_value = normalize_text(value)
        for suffix in GENERIC_CITY_SUFFIXES:
            if norm_value.endswith(suffix):
                raw_variants.append(value[: -len(suffix)].strip())
        if norm_value.endswith(" city"):
            raw_variants.append(value[: -5].strip())

    candidates.extend(slugify(value) for value in raw_variants if value)
    alias_key = (base_country_slug, normalize_text(raw))
    for alias in PRIMARY_CITY_ALIASES.get(alias_key, []):
        candidates.append(slugify(alias))

    return ordered_unique([candidate for candidate in candidates if candidate])


def region_slug_candidates(base_country_slug: str, country_value: str, province_value: str) -> list[str]:
    raw_values: list[str] = []
    _, country_region = split_country_value(country_value)
    raw_values.extend(part.strip() for part in country_region.split("/") if part.strip())
    raw_values.extend(part.strip() for part in province_value.split("/") if part.strip())

    candidates: list[str] = []
    for value in raw_values:
        normalized = normalize_text(value)
        if not normalized:
            continue
        if base_country_slug == "united-states":
            code = US_STATE_CODES.get(normalized)
            if code:
                candidates.append(code)
        elif base_country_slug == "canada":
            code = CANADA_PROVINCE_CODES.get(normalized)
            if code:
                candidates.append(code)
        else:
            candidates.append(slugify(value))
    return ordered_unique(candidates)


class PageIndex:
    def __init__(self) -> None:
        self.country_pages: dict[str, PageRef] = {}
        self.region_pages: dict[str, dict[str, PageRef]] = defaultdict(dict)
        self.country_city_pages: dict[str, dict[str, list[PageRef]]] = defaultdict(lambda: defaultdict(list))
        self.region_city_pages: dict[str, dict[str, dict[str, PageRef]]] = defaultdict(lambda: defaultdict(dict))
        self.country_city_slugs: dict[str, list[str]] = defaultdict(list)
        self.region_city_slugs: dict[str, dict[str, list[str]]] = defaultdict(lambda: defaultdict(list))

    def finalize(self) -> None:
        for country_slug, city_map in self.country_city_pages.items():
            self.country_city_slugs[country_slug] = sorted(city_map)
        for country_slug, region_map in self.region_city_pages.items():
            for region_slug, city_map in region_map.items():
                self.region_city_slugs[country_slug][region_slug] = sorted(city_map)


def build_page_index() -> PageIndex:
    all_entries: list[tuple[list[str], str]] = []
    deep_prefixes: set[tuple[str, str]] = set()

    for sitemap_url in SITEMAP_URLS:
        response = requests.get(sitemap_url, headers=LIVINGCOST_HEADERS, timeout=60)
        response.raise_for_status()
        root = ET.fromstring(response.text)
        ns = {"sm": "http://www.sitemaps.org/schemas/sitemap/0.9"}
        for loc in root.findall(".//sm:loc", ns):
            url = loc.text or ""
            if "/cost/" not in url:
                continue
            path = url.split("/cost/", 1)[1].strip("/")
            if not path:
                continue
            parts = path.split("/")
            all_entries.append((parts, url))
            if len(parts) == 3:
                deep_prefixes.add((parts[0], parts[1]))

    index = PageIndex()
    for parts, url in all_entries:
        if len(parts) == 1:
            country_slug = parts[0]
            index.country_pages[country_slug] = PageRef(
                url=url,
                country_slug=country_slug,
                region_slug=None,
                city_slug=None,
                source_level="country",
            )
        elif len(parts) == 2:
            country_slug, second = parts
            if (country_slug, second) in deep_prefixes:
                index.region_pages[country_slug][second] = PageRef(
                    url=url,
                    country_slug=country_slug,
                    region_slug=second,
                    city_slug=None,
                    source_level="region",
                )
            else:
                page = PageRef(
                    url=url,
                    country_slug=country_slug,
                    region_slug=None,
                    city_slug=second,
                    source_level="city",
                )
                index.country_city_pages[country_slug][second].append(page)
        elif len(parts) == 3:
            country_slug, region_slug, city_slug = parts
            page = PageRef(
                url=url,
                country_slug=country_slug,
                region_slug=region_slug,
                city_slug=city_slug,
                source_level="city",
            )
            index.country_city_pages[country_slug][city_slug].append(page)
            index.region_city_pages[country_slug][region_slug][city_slug] = page

    index.finalize()
    return index


def resolve_country_slug(country_value: str, index: PageIndex) -> str:
    base_country, _ = split_country_value(country_value)
    normalized = normalize_text(base_country)
    alias = COUNTRY_ALIASES.get(normalized)
    if alias:
        return alias
    candidate = slugify(base_country)
    if candidate in index.country_pages:
        return candidate

    exact_matches = [
        country_slug
        for country_slug in index.country_pages
        if normalize_text(country_slug.replace("-", " ")) == normalized
    ]
    if exact_matches:
        return exact_matches[0]

    scored_matches = sorted(
        (
            SequenceMatcher(None, normalized, normalize_text(country_slug.replace("-", " "))).ratio(),
            country_slug,
        )
        for country_slug in index.country_pages
    )
    best_score, best_country = scored_matches[-1]
    if best_score >= 0.9:
        return best_country
    raise KeyError(f"Could not map country '{country_value}' to a livingcost country page.")


def parse_money(value: str) -> float | None:
    cleaned = value.replace(",", "")
    match = re.search(r"([0-9]+(?:\.[0-9]+)?)", cleaned)
    return float(match.group(1)) if match else None


def simplify_row_label(value: str) -> str:
    normalized = normalize_text(value)
    normalized = normalized.replace(" m 2 ", " m2 ").replace(" ft 2 ", " ft2 ")
    normalized = normalized.replace(" mbps ", " mbps ")
    return normalized


def fetch_with_retry(http: requests.Session, url: str, headers: dict[str, str], timeout: int) -> requests.Response:
    last_error: Exception | None = None
    for attempt in range(1, 8):
        try:
            response = http.get(url, headers=headers, timeout=timeout)
            if response.status_code == 429:
                retry_after = response.headers.get("Retry-After")
                if retry_after:
                    sleep_seconds = max(5.0, float(retry_after))
                else:
                    sleep_seconds = min(60.0, 5.0 * attempt)
                time.sleep(sleep_seconds)
                continue
            response.raise_for_status()
            return response
        except requests.RequestException as exc:
            last_error = exc
            time.sleep(min(30.0, 2.0 * attempt))
    if last_error is None:
        raise RuntimeError(f"Failed to fetch {url}")
    raise last_error


def page_cache_path(cache_dir: Path, url: str) -> Path:
    digest = hashlib.sha1(url.encode("utf-8")).hexdigest()
    return cache_dir / f"{digest}.json"


def parse_cost_page(
    page_ref: PageRef,
    session: requests.Session | None = None,
    cache_dir: Path | None = None,
) -> dict[str, Any]:
    if cache_dir is not None:
        cache_dir.mkdir(parents=True, exist_ok=True)
        cache_file = page_cache_path(cache_dir, page_ref.url)
        if cache_file.exists():
            return json.loads(cache_file.read_text(encoding="utf-8"))

    http = session or requests.Session()
    response = fetch_with_retry(http, page_ref.url, headers=LIVINGCOST_HEADERS, timeout=30)

    soup = BeautifulSoup(response.text, "html.parser")
    metrics: dict[str, float] = {}
    for table in soup.find_all("table"):
        for row in table.find_all("tr"):
            cells = [cell.get_text(" ", strip=True) for cell in row.find_all(["th", "td"])]
            if len(cells) < 2:
                continue
            label = simplify_row_label(cells[0])
            value_key = ROW_VALUE_KEYS.get(label)
            if not value_key:
                continue
            metric_value = parse_money(cells[1])
            if metric_value is not None:
                metrics[value_key] = metric_value

    payload = {
        "url": page_ref.url,
        "source_level": page_ref.source_level,
        "country_slug": page_ref.country_slug,
        "region_slug": page_ref.region_slug,
        "city_slug": page_ref.city_slug,
        "metrics": metrics,
    }
    if cache_dir is not None:
        cache_file.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    return payload


class GeocodeCache:
    def __init__(self, path: Path) -> None:
        self.path = path
        self.path.parent.mkdir(parents=True, exist_ok=True)
        if path.exists():
            self.data: dict[str, dict[str, Any]] = json.loads(path.read_text(encoding="utf-8"))
        else:
            self.data = {}
        self._last_request_ts = 0.0

    def lookup(self, query_key: str) -> dict[str, Any] | None:
        return self.data.get(query_key)

    def store(self, query_key: str, result: dict[str, Any] | None) -> dict[str, Any] | None:
        self.data[query_key] = result or {}
        self.path.write_text(json.dumps(self.data, ensure_ascii=False, indent=2), encoding="utf-8")
        return result

    def geocode(self, city: str, province: str, country: str) -> dict[str, Any] | None:
        query_key = json.dumps({"city": city, "province": province, "country": country}, ensure_ascii=False)
        cached = self.lookup(query_key)
        if cached is not None:
            return cached or None

        elapsed = time.time() - self._last_request_ts
        if elapsed < 1.1:
            time.sleep(1.1 - elapsed)

        params = {
            "q": ", ".join(part for part in [city, province, country] if part),
            "format": "jsonv2",
            "limit": 1,
            "accept-language": "en",
            "addressdetails": 1,
        }
        response = requests.get(
            "https://nominatim.openstreetmap.org/search",
            params=params,
            headers=NOMINATIM_HEADERS,
            timeout=30,
        )
        response.raise_for_status()
        self._last_request_ts = time.time()
        items = response.json()
        if not items:
            return self.store(query_key, None)

        item = items[0]
        address = item.get("address") or {}
        result = {
            "name": item.get("name") or "",
            "display_name": item.get("display_name") or "",
            "city": first_non_empty(
                address.get("city"),
                address.get("town"),
                address.get("village"),
                address.get("municipality"),
                address.get("county"),
                item.get("name"),
            ),
            "state": first_non_empty(
                address.get("state"),
                address.get("region"),
                address.get("province"),
                address.get("county"),
            ),
        }
        return self.store(query_key, result)


def best_page_for_city(
    country_slug: str,
    city_candidates: list[str],
    region_candidates: list[str],
    index: PageIndex,
) -> PageRef | None:
    country_pages = index.country_city_pages.get(country_slug) or {}
    if not country_pages:
        return None

    for city_slug in city_candidates:
        exact_matches = country_pages.get(city_slug) or []
        if not exact_matches:
            continue
        if len(exact_matches) == 1:
            return exact_matches[0]
        for region_slug in region_candidates:
            for page in exact_matches:
                if page.region_slug == region_slug:
                    return page
        direct_matches = [page for page in exact_matches if page.region_slug is None]
        if len(direct_matches) == 1:
            return direct_matches[0]

    region_match_pool: list[str] = []
    for region_slug in region_candidates:
        region_match_pool.extend(index.region_city_slugs.get(country_slug, {}).get(region_slug, []))

    for city_slug in city_candidates:
        if region_match_pool:
            close = get_close_matches(city_slug, region_match_pool, n=1, cutoff=0.84)
            if close:
                for region_slug in region_candidates:
                    page = index.region_city_pages.get(country_slug, {}).get(region_slug, {}).get(close[0])
                    if page:
                        return page

        all_city_slugs = index.country_city_slugs.get(country_slug, [])
        close = get_close_matches(city_slug, all_city_slugs, n=3, cutoff=0.86)
        if not close:
            continue

        best_slug = close[0]
        best_matches = country_pages.get(best_slug) or []
        if len(best_matches) == 1:
            return best_matches[0]
        for region_slug in region_candidates:
            for page in best_matches:
                if page.region_slug == region_slug:
                    return page
        if best_matches:
            return best_matches[0]

    return None


def resolve_page_for_location(
    location: LocationKey,
    country_slug: str,
    index: PageIndex,
    geocode_cache: GeocodeCache,
) -> PageRef:
    region_candidates = region_slug_candidates(country_slug, location.country, location.province)
    city_candidates = city_slug_candidates(country_slug, location.city, location.province)
    page = best_page_for_city(country_slug, city_candidates, region_candidates, index)
    if page:
        return page

    geocoded = geocode_cache.geocode(location.city, location.province, location.base_country)
    if geocoded:
        geocoded_region_candidates = region_slug_candidates(
            country_slug,
            location.country,
            first_non_empty(geocoded.get("state"), location.province),
        )
        geocoded_city_candidates = city_slug_candidates(
            country_slug,
            first_non_empty(geocoded.get("city"), geocoded.get("name")),
            first_non_empty(geocoded.get("state"), location.province),
        )
        page = best_page_for_city(
            country_slug,
            geocoded_city_candidates + city_candidates,
            geocoded_region_candidates + region_candidates,
            index,
        )
        if page:
            return page

    for region_slug in region_candidates:
        page = index.region_pages.get(country_slug, {}).get(region_slug)
        if page:
            return page

    country_page = index.country_pages.get(country_slug)
    if not country_page:
        raise KeyError(f"Missing livingcost country page for '{location.base_country}'.")
    return country_page


def compute_cost_columns(metrics: dict[str, float]) -> dict[str, float | str | None]:
    without_rent = metrics.get("without_rent_monthly")
    food = metrics.get("food_monthly")
    transport = metrics.get("transport_monthly")
    apartment_cheap_1br = metrics.get("apartment_cheap_1br_monthly")
    apartment_cheap_3br = metrics.get("apartment_cheap_3br_monthly")
    utilities_one_person = metrics.get("utilities_one_person_monthly")
    utilities_family = metrics.get("utilities_family_monthly")
    internet = metrics.get("internet_monthly")

    housing_solo_monthly = None
    if apartment_cheap_1br is not None and utilities_one_person is not None and internet is not None:
        housing_solo_monthly = apartment_cheap_1br + utilities_one_person + internet

    housing_shared_monthly = None
    if apartment_cheap_3br is not None and utilities_family is not None and internet is not None:
        housing_shared_monthly = apartment_cheap_3br / 3 + utilities_family / 3 + internet / 3

    student_total_shared_monthly = None
    if without_rent is not None and housing_shared_monthly is not None:
        student_total_shared_monthly = without_rent + housing_shared_monthly

    student_total_solo_monthly = None
    if without_rent is not None and housing_solo_monthly is not None:
        student_total_solo_monthly = without_rent + housing_solo_monthly

    def annualize(value: float | None) -> float | None:
        if value is None:
            return None
        return round(value * 12, 2)

    return {
        "Estimated Annual Student Cost (Shared Housing, USD)": annualize(student_total_shared_monthly),
        "Estimated Annual Student Cost (Solo Apartment, USD)": annualize(student_total_solo_monthly),
        "Estimated Annual Base Living Cost Without Rent (USD)": annualize(without_rent),
        "Estimated Annual Housing Cost Shared (USD)": annualize(housing_shared_monthly),
        "Estimated Annual Housing Cost Solo (USD)": annualize(housing_solo_monthly),
        "Estimated Annual Food Cost (USD)": annualize(food),
        "Estimated Annual Transport Cost (USD)": annualize(transport),
    }


def ensure_output_columns(worksheet) -> dict[str, int]:
    header_map = {
        worksheet.cell(row=1, column=column_index).value: column_index
        for column_index in range(1, worksheet.max_column + 1)
    }
    next_column = worksheet.max_column + 1
    for header in OUTPUT_COLUMNS:
        if header not in header_map:
            worksheet.cell(row=1, column=next_column, value=header)
            header_map[header] = next_column
            next_column += 1
    return {header: header_map[header] for header in OUTPUT_COLUMNS}


def fetch_all_page_payloads(
    page_refs: list[PageRef],
    max_workers: int,
    page_cache_dir: Path,
) -> dict[str, dict[str, Any]]:
    payloads: dict[str, dict[str, Any]] = {}
    unique_page_refs = {page_ref.url: page_ref for page_ref in page_refs}
    print(f"[info] Fetching {len(unique_page_refs)} livingcost pages...", flush=True)

    def worker(page_ref: PageRef) -> dict[str, Any]:
        session = requests.Session()
        try:
            return parse_cost_page(page_ref, session=session, cache_dir=page_cache_dir)
        finally:
            session.close()

    with ThreadPoolExecutor(max_workers=max_workers) as executor:
        future_map = {
            executor.submit(worker, page_ref): page_ref.url
            for page_ref in unique_page_refs.values()
        }
        for future in as_completed(future_map):
            url = future_map[future]
            payloads[url] = future.result()
    return payloads


def backup_workbook(path: Path) -> Path:
    backup_path = path.with_name(f"{path.stem}.student_cost_backup{path.suffix}")
    shutil.copy2(path, backup_path)
    return backup_path


def build_locations(worksheet) -> tuple[list[LocationKey], dict[int, LocationKey]]:
    rows = worksheet.iter_rows(values_only=True)
    headers = list(next(rows))
    country_idx = headers.index("Country")
    province_idx = headers.index("Province")
    city_idx = headers.index("City")

    row_locations: dict[int, LocationKey] = {}
    unique_locations: dict[LocationKey, None] = {}
    for row_number, row in enumerate(rows, start=2):
        country = str(row[country_idx] or "").strip()
        province = str(row[province_idx] or "").strip()
        city = str(row[city_idx] or "").strip()
        base_country, _ = split_country_value(country)
        location = LocationKey(country=country, base_country=base_country, province=province, city=city)
        row_locations[row_number] = location
        unique_locations.setdefault(location, None)
    return list(unique_locations), row_locations


def enrich_workbook(
    workbook_path: Path,
    sheet_name: str,
    max_workers: int,
    geocode_cache_path: Path,
    page_cache_dir: Path,
) -> dict[str, Any]:
    index = build_page_index()
    geocode_cache = GeocodeCache(geocode_cache_path)

    wb = load_workbook(workbook_path)
    ws = wb[sheet_name]
    unique_locations, row_locations = build_locations(ws)
    print(f"[info] Unique locations: {len(unique_locations)}", flush=True)

    location_to_page: dict[LocationKey, PageRef] = {}
    page_refs: list[PageRef] = []
    source_counts: defaultdict[str, int] = defaultdict(int)
    country_counts: defaultdict[str, int] = defaultdict(int)
    for location in unique_locations:
        country_slug = resolve_country_slug(location.country, index)
        page_ref = resolve_page_for_location(location, country_slug, index, geocode_cache)
        location_to_page[location] = page_ref
        page_refs.append(page_ref)
        source_counts[page_ref.source_level] += 1
        country_counts[country_slug] += 1

    payloads = fetch_all_page_payloads(page_refs, max_workers=max_workers, page_cache_dir=page_cache_dir)
    column_map = ensure_output_columns(ws)

    filled_rows = 0
    for row_number, location in row_locations.items():
        page_ref = location_to_page[location]
        payload = payloads[page_ref.url]
        cost_values = compute_cost_columns(payload["metrics"])
        for header, value in cost_values.items():
            ws.cell(row=row_number, column=column_map[header], value=value)
        ws.cell(row=row_number, column=column_map["Estimated Cost Source Level"], value=payload["source_level"])
        ws.cell(row=row_number, column=column_map["Estimated Cost Source URL"], value=payload["url"])
        if cost_values["Estimated Annual Student Cost (Shared Housing, USD)"] is not None:
            filled_rows += 1

    backup_path = backup_workbook(workbook_path)
    wb.save(workbook_path)
    return {
        "rows": ws.max_row - 1,
        "filled_rows": filled_rows,
        "backup_path": str(backup_path),
        "source_counts": dict(sorted(source_counts.items())),
        "country_count": len(country_counts),
        "page_count": len({page_ref.url for page_ref in page_refs}),
    }


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Add estimated annual student living cost columns to the WHED Institutions sheet.",
    )
    parser.add_argument(
        "--workbook",
        default="whed_data.xlsx",
        help="Workbook to enrich in place.",
    )
    parser.add_argument(
        "--sheet",
        default="Institutions",
        help="Worksheet name that contains the university rows.",
    )
    parser.add_argument(
        "--max-workers",
        type=int,
        default=10,
        help="Maximum number of concurrent livingcost page fetches.",
    )
    parser.add_argument(
        "--geocode-cache",
        default=".cache/nominatim_city_cache.json",
        help="Path to the Nominatim fallback cache file.",
    )
    parser.add_argument(
        "--page-cache-dir",
        default=".cache/livingcost_pages",
        help="Directory used to cache parsed livingcost page payloads.",
    )
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    results = enrich_workbook(
        workbook_path=Path(args.workbook),
        sheet_name=args.sheet,
        max_workers=args.max_workers,
        geocode_cache_path=Path(args.geocode_cache),
        page_cache_dir=Path(args.page_cache_dir),
    )
    print(
        json.dumps(
            {
                "rows": results["rows"],
                "filled_rows": results["filled_rows"],
                "page_count": results["page_count"],
                "countries_used": results["country_count"],
                "source_counts": results["source_counts"],
                "backup_path": results["backup_path"],
            },
            ensure_ascii=False,
            indent=2,
        ),
        flush=True,
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
