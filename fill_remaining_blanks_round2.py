import argparse
import concurrent.futures
import io
import json
import re
import shutil
import time
from dataclasses import dataclass
from pathlib import Path
from typing import Any

import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook

try:
    from pypdf import PdfReader
except Exception:  # pragma: no cover
    from PyPDF2 import PdfReader  # type: ignore[assignment]

from add_extended_metrics import USER_AGENT, normalize_country, parse_float
from add_student_costs import normalize_text, split_country_value


WORK_RIGHTS_URL = "https://erudera.com/resources/best-countries-to-work-and-study-for-international-students/"
USCIS_STUDENT_WORK_URL = "https://www.uscis.gov/working-in-the-united-states/students-and-exchange-visitors/students-and-employment"
SLOVAKIA_STUDENT_WORK_URL = "https://home-affairs.ec.europa.eu/policies/migration-and-asylum/eu-immigration-portal/student-slovakia_en"
VISA_REJECTION_URL = "https://gradpilot.com/news/student-visa-rejection-rates-by-country-data"
NAI_WORLD_PDF_URL = "https://academyofinventors.org/wp-content/uploads/2026/02/2025-Top-100-Worldwide-List.pdf"
NAI_US_PDF_URL = "https://academyofinventors.org/wp-content/uploads/2026/03/2025-Top-100-US-List.pdf"

ROUND2_CACHE_DIR = Path(".cache/round2_remaining")
EDURANK_ALUMNI_CACHE_DIR = Path(".cache/edurank/alumni")

ROUND2_NEW_COLUMNS = [
    "Patent Source URL",
    "Work Rights Source URL",
    "Visa Difficulty Source URL",
    "Employment Source URL",
    "Startup Alumni Source URL",
]

COUNTRY_ALIASES = {
    normalize_text("USA (F-1)"): normalize_country("United States of America"),
    normalize_text("UK"): normalize_country("United Kingdom"),
    normalize_text("The Netherlands"): normalize_country("Netherlands"),
    normalize_text("Bosnia & Herzegovina"): normalize_country("Bosnia and Herzegovina"),
    normalize_text("Costa Rica,"): normalize_country("Costa Rica"),
    normalize_text("Slovakia"): normalize_country("Slovak Republic"),
    normalize_text("Czechia"): normalize_country("Czech Republic"),
}

PATENT_NAME_ALIASES = {
    normalize_text("the regents of the university of california"): "University of California",
    normalize_text("georgia tech research corporation"): "Georgia Institute of Technology",
    normalize_text("university of florida research foundation inc"): "University of Florida",
    normalize_text("wisconsin alumni research foundation (university of wisconsin)"): "University of Wisconsin-Madison",
    normalize_text("the florida international university board of trustees"): "Florida International University",
    normalize_text("university of central florida research foundation inc"): "University of Central Florida",
    normalize_text("the trustees of princeton university"): "Princeton University",
    normalize_text("trustees of boston university"): "Boston University",
    normalize_text("trustees of dartmouth college"): "Dartmouth College",
    normalize_text("regents of the university of minnesota"): "University of Minnesota Twin Cities",
    normalize_text("board of regents of the university of nebraska"): "University of Nebraska-Lincoln",
    normalize_text("virginia tech intellectual properties inc"): "Virginia Polytechnic Institute and State University",
    normalize_text("korea university research and business foundation"): "Korea University",
    normalize_text("iowa state university research foundation inc"): "Iowa State University",
    normalize_text("the university of utah research foundation"): "The University of Utah",
    normalize_text("the university of utah research foudnation"): "The University of Utah",
    normalize_text("florida state university research foundation, inc"): "Florida State University",
    normalize_text("university of houston system"): "University of Houston",
    normalize_text("william marsh rice university"): "Rice University",
    normalize_text("the uab research foundation (the university of alabama at birmingham)"): "University of Alabama at Birmingham",
    normalize_text("texas tech university system"): "Texas Tech University",
    normalize_text("university of kentucky research foundation"): "University of Kentucky",
    normalize_text("university of louisville research foundation, inc"): "University of Louisville",
    normalize_text("colorado state university research foundation"): "Colorado State University",
    normalize_text("kansas state university research foundation"): "Kansas State University",
    normalize_text("the board of trustees of the university of arkansas"): "University of Arkansas",
    normalize_text("the university of chicago"): "University of Chicago",
    normalize_text("the university of connecticut"): "The University of Connecticut",
    normalize_text("the university of alabama"): "University of Alabama",
    normalize_text("university of notre dame du lac"): "University of Notre Dame",
    normalize_text("state university of new york (suny)"): "State University of New York-System",
    normalize_text("the state university of new york (suny)"): "State University of New York-System",
    normalize_text("the university of north carolina at chapel hill"): "University of North Carolina at Chapel Hill",
    normalize_text("the university of southern california"): "University of Southern California",
    normalize_text("columbia university"): "Columbia University in the City of New York",
    normalize_text("the pennsylvania state university"): "The Pennsylvania State University-Main Campus",
    normalize_text("indiana university"): "Indiana University Bloomington",
    normalize_text("board of supervisors of louisiana state university and agricultural and mechanical college"): "Louisiana State University and Agricultural and Mechanical College",
    normalize_text("oregon health and science university"): "Oregon Health and Science University",
    normalize_text("oregon health & science university"): "Oregon Health and Science University",
    normalize_text("the george washington university"): "The George Washington University",
    normalize_text("the university of tennessee research foundation"): "The University of Tennessee-Knoxville",
    normalize_text("vanderbilt university"): "Vanderbilt University",
    normalize_text("virginia commonwealth university"): "Virginia Commonwealth University",
}

NAI_SKIP_NAMES = {
    normalize_text("the university of texas system"),
    normalize_text("university system of maryland"),
    normalize_text("university of colorado"),
    normalize_text("university of illinois"),
    normalize_text("university of massachusetts"),
    normalize_text("texas a&m university system"),
    normalize_text("the curators of the university of missouri"),
    normalize_text("research foundation of the city university of new york"),
}

ENTREPRENEUR_KEYWORDS = (
    "entrepreneur",
    "founder",
    "co-founder",
    "startup",
    "businessperson",
    "venture capitalist",
    "investor",
    "industrialist",
)

HYPERLINK_HEADERS = set(ROUND2_NEW_COLUMNS)
PDF_ENTRY_RE = re.compile(r"(?ms)\b(\d{1,3})\s+(.+?)\.{3,}\s*(\d{1,4})\s*(?=(?:\n\s*\d{1,3}\s+)|(?:\n\s*top 100)|$)")
OCR_REPAIRS = {
    "vanderbil t": "vanderbilt",
    "commonweal th": "commonwealth",
    "heal th": "health",
    "agricul tural": "agricultural",
    "l td": "ltd",
    "foudnation": "foundation",
}


@dataclass(frozen=True)
class RowData:
    row_number: int
    university_name: str
    country: str
    base_country_key: str
    graduate_employability: float | None
    industry_placement: float | None
    english_program_ratio: float | None
    international_student_ratio: float | None
    edurank_source_url: str
    salary_source_url: str
    existing_notes: str
    existing_coverage: str


def fetch_text(url: str, cache_name: str) -> str:
    path = ROUND2_CACHE_DIR / cache_name
    path.parent.mkdir(parents=True, exist_ok=True)
    if path.exists():
        return path.read_text(encoding="utf-8", errors="replace")
    last_error: Exception | None = None
    for _ in range(3):
        try:
            response = requests.get(url, timeout=(20, 120), headers={"User-Agent": USER_AGENT})
            response.raise_for_status()
            path.write_text(response.text, encoding="utf-8")
            return response.text
        except Exception as exc:
            last_error = exc
            time.sleep(1.0)
    raise last_error or RuntimeError(f"Failed to fetch text: {url}")


def fetch_bytes(url: str, cache_name: str) -> bytes:
    path = ROUND2_CACHE_DIR / cache_name
    path.parent.mkdir(parents=True, exist_ok=True)
    if path.exists():
        return path.read_bytes()
    last_error: Exception | None = None
    for _ in range(3):
        try:
            response = requests.get(url, timeout=(20, 180), headers={"User-Agent": USER_AGENT})
            response.raise_for_status()
            path.write_bytes(response.content)
            return response.content
        except Exception as exc:
            last_error = exc
            time.sleep(1.0)
    raise last_error or RuntimeError(f"Failed to fetch bytes: {url}")


def country_key(raw_country: str) -> str:
    normalized = normalize_text(raw_country)
    if normalized in COUNTRY_ALIASES:
        return COUNTRY_ALIASES[normalized]
    return normalize_country(raw_country)


def base_country_key(raw_country: str) -> str:
    base_country, _ = split_country_value(raw_country)
    return country_key(base_country)


def backup_workbook(path: Path) -> Path:
    backup_path = path.with_name(f"{path.stem}.round2_backup{path.suffix}")
    shutil.copy2(path, backup_path)
    return backup_path


def ensure_columns(worksheet) -> dict[str, int]:
    header_map = {
        worksheet.cell(row=1, column=column_index).value: column_index
        for column_index in range(1, worksheet.max_column + 1)
    }
    next_column = worksheet.max_column + 1
    for header in ROUND2_NEW_COLUMNS:
        if header not in header_map:
            worksheet.cell(row=1, column=next_column, value=header)
            header_map[header] = next_column
            next_column += 1
    return header_map


def append_note(existing: str, note: str) -> str:
    if not note:
        return existing
    pieces = [part.strip() for part in str(existing or "").split(";") if part.strip()]
    lowered = {part.lower() for part in pieces}
    if note.lower() not in lowered:
        pieces.append(note)
    return "; ".join(pieces)


def merged_coverage(existing: str, additions: list[str]) -> str:
    pieces = [part.strip() for part in str(existing or "").split(";") if part.strip()]
    lowered = {part.lower() for part in pieces}
    for addition in additions:
        if addition and addition.lower() not in lowered:
            pieces.append(addition)
            lowered.add(addition.lower())
    return "; ".join(pieces)


def clamp(value: float, lower: float, upper: float) -> float:
    return max(lower, min(upper, value))


def parse_percent_score(text: str | None) -> float | None:
    if not text:
        return None
    numbers = [float(value) for value in re.findall(r"(\d+(?:\.\d+)?)\s*%", text)]
    if len(numbers) >= 2 and "-" in text:
        return round(sum(numbers[:2]) / 2.0, 2)
    if numbers:
        return round(numbers[0], 2)
    return None


def normalize_patent_name(name: str) -> str:
    text = name.lower().replace("\xa0", " ").strip()
    for original, replacement in OCR_REPAIRS.items():
        text = text.replace(original, replacement)
    text = text.replace("&", "and")
    text = re.sub(r"\s+\.", ".", text)
    text = re.sub(r"\s+", " ", text).strip(" .")
    return normalize_text(text)


def patent_name_variants(name: str) -> list[str]:
    variants = [name.strip()]
    no_paren = re.sub(r"\s*\([^)]*\)", "", name).strip()
    if no_paren:
        variants.append(no_paren)
    if name.lower().startswith("the "):
        variants.append(name[4:].strip())
    if no_paren.lower().startswith("the "):
        variants.append(no_paren[4:].strip())
    seen: set[str] = set()
    ordered: list[str] = []
    for variant in variants:
        key = normalize_patent_name(variant)
        if key and key not in seen:
            seen.add(key)
            ordered.append(key)
    return ordered


def load_rows(workbook_path: Path, sheet_name: str) -> tuple[Any, Any, dict[str, int], list[RowData]]:
    workbook = load_workbook(workbook_path)
    worksheet = workbook[sheet_name]
    rows = list(worksheet.iter_rows(values_only=True))
    headers = [str(value) if value is not None else "" for value in rows[0]]
    header_map = {header: idx for idx, header in enumerate(headers)}

    def value_for(row: tuple[Any, ...], header: str) -> Any:
        index = header_map.get(header)
        return row[index] if index is not None and index < len(row) else None

    institutions: list[RowData] = []
    for row_number, row in enumerate(rows[1:], start=2):
        country = str(value_for(row, "Country") or "").strip()
        institutions.append(
            RowData(
                row_number=row_number,
                university_name=str(value_for(row, "University Name") or "").strip(),
                country=country,
                base_country_key=base_country_key(country),
                graduate_employability=parse_float(value_for(row, "Graduate Employability Rate (%)")),
                industry_placement=parse_float(value_for(row, "Industry Placement Rate")),
                english_program_ratio=parse_float(value_for(row, "English Program Ratio (%)")),
                international_student_ratio=parse_float(value_for(row, "International Student Ratio (%)")),
                edurank_source_url=str(value_for(row, "EduRank Source URL") or "").strip(),
                salary_source_url=str(value_for(row, "Salary Source URL") or "").strip(),
                existing_notes=str(value_for(row, "Unofficial Metrics Notes") or "").strip(),
                existing_coverage=str(value_for(row, "Extended Metrics Coverage") or "").strip(),
            )
        )
    return workbook, worksheet, header_map, institutions


def parse_work_rights_table() -> dict[str, dict[str, str]]:
    html = fetch_text(WORK_RIGHTS_URL, "work_rights_table.html")
    soup = BeautifulSoup(html, "html.parser")
    mapping: dict[str, dict[str, str]] = {}
    for row in soup.find_all("tr"):
        cells = [cell.get_text(" ", strip=True) for cell in row.find_all(["th", "td"])]
        if len(cells) < 4:
            continue
        if normalize_text(cells[0]) in {"country", "countries"}:
            continue
        allowed = cells[1].strip().title()
        if allowed not in {"Yes", "No"}:
            continue
        key = country_key(cells[0].strip())
        mapping[key] = {
            "allowed": allowed,
            "permit_text": cells[2].strip(),
            "hours_text": cells[3].strip(),
            "source_url": WORK_RIGHTS_URL,
        }
    mapping[normalize_country("United States of America")] = {
        "allowed": "Yes",
        "permit_text": "F-1 students may work on campus without separate employment authorization; other employment paths include CPT, OPT, and severe economic hardship authorization.",
        "hours_text": "Up to 20 hours per week while school is in session for on-campus employment.",
        "source_url": USCIS_STUDENT_WORK_URL,
    }
    mapping[normalize_country("Slovak Republic")] = {
        "allowed": "Yes",
        "permit_text": "University students with temporary residence for study can work without an additional work permit while keeping study as their main purpose of stay.",
        "hours_text": "Maximum 20 hours per week for university students.",
        "source_url": SLOVAKIA_STUDENT_WORK_URL,
    }
    return mapping


def parse_visa_difficulty_table() -> dict[str, dict[str, Any]]:
    html = fetch_text(VISA_REJECTION_URL, "visa_difficulty_table.html")
    soup = BeautifulSoup(html, "html.parser")
    mapping: dict[str, dict[str, Any]] = {}
    for row in soup.find_all("tr"):
        cells = row.find_all("td")
        if len(cells) < 5:
            continue
        country_text = cells[0].get_text(" ", strip=True)
        if not country_text:
            continue
        key = country_key(country_text)
        score = parse_percent_score(cells[1].get_text(" ", strip=True))
        if score is None:
            continue
        source_link = cells[4].find("a", href=True)
        mapping[key] = {
            "score": score,
            "source_url": source_link["href"].strip() if source_link else VISA_REJECTION_URL,
        }
    return mapping


def work_permit_adjustment(permit_text: str) -> float:
    normalized = normalize_text(permit_text)
    if not normalized:
        return 0.0
    no_permit_markers = (
        "dont need to apply for a work permit",
        "do not need to apply for a work permit",
        "dont need a work permit",
        "do not need a work permit",
        "do not need an additional working permit",
        "without separate employment authorization",
        "they dont need",
        "they do not need",
        "no additional working permit",
    )
    if any(marker in normalized for marker in no_permit_markers):
        return -2.0
    if "must be enrolled" in normalized or "eligible to work" in normalized:
        return 0.75
    if "work permit" in normalized or "employment authorization" in normalized or "authorities" in normalized:
        return 2.0
    return 0.0


def work_hours_adjustment(hours_text: str) -> float:
    normalized = normalize_text(hours_text)
    if not normalized:
        return 0.0
    if "unlimited" in normalized or "no limitation" in normalized:
        return -1.5
    if "full time" in normalized and "20 hours" not in normalized:
        return -1.0
    numbers = [float(value) for value in re.findall(r"(\d+(?:\.\d+)?)", normalized)]
    if not numbers:
        return 0.0
    smallest = min(numbers)
    if smallest <= 20:
        return 1.0
    if smallest <= 25:
        return 0.5
    if smallest <= 30:
        return 0.25
    return -0.5


def visa_proxy_score(row: RowData, work_info: dict[str, str] | None) -> float | None:
    score = 12.0
    if work_info:
        if work_info.get("allowed") == "Yes":
            score -= 1.5
        else:
            score += 15.0
        score += work_permit_adjustment(work_info.get("permit_text", ""))
        score += work_hours_adjustment(work_info.get("hours_text", ""))
    if row.english_program_ratio is not None:
        if row.english_program_ratio >= 100:
            score -= 2.0
        elif row.english_program_ratio >= 50:
            score -= 1.0
        elif row.english_program_ratio <= 10:
            score += 1.0
        elif row.english_program_ratio <= 1:
            score += 1.5
    if row.international_student_ratio is not None:
        if row.international_student_ratio >= 25:
            score -= 2.0
        elif row.international_student_ratio >= 15:
            score -= 1.0
        elif row.international_student_ratio < 5:
            score += 1.0
    return round(clamp(score, 2.0, 30.0), 2)


def parse_pdf_patent_rankings(url: str, cache_name: str) -> list[tuple[int, str, int]]:
    reader = PdfReader(io.BytesIO(fetch_bytes(url, cache_name)))
    text = "\n".join((page.extract_text() or "") for page in reader.pages)
    records: list[tuple[int, str, int]] = []
    seen: set[tuple[str, int]] = set()
    for match in PDF_ENTRY_RE.finditer(text.lower().replace("\xa0", " ")):
        rank = int(match.group(1))
        institution_name = " ".join(match.group(2).split())
        normalized_name = normalize_patent_name(institution_name)
        if not normalized_name:
            continue
        if any(marker in normalized_name for marker in ("top 100", "universities granted", "institution ranking", "of patents")):
            continue
        count = int(match.group(3))
        token = (normalized_name, count)
        if token in seen:
            continue
        seen.add(token)
        records.append((rank, normalized_name, count))
    return records


def build_patent_counts(rows: list[RowData]) -> dict[str, dict[str, Any]]:
    name_by_variant: dict[str, str] = {}
    all_names = {row.university_name for row in rows if row.university_name}
    for name in all_names:
        for variant in patent_name_variants(name):
            name_by_variant.setdefault(variant, name)

    matched: dict[str, dict[str, Any]] = {}
    world_records = parse_pdf_patent_rankings(NAI_WORLD_PDF_URL, "n_ai_world_2025.pdf")
    us_records = parse_pdf_patent_rankings(NAI_US_PDF_URL, "n_ai_us_2025.pdf")
    world_names = {record[1] for record in world_records}
    records = world_records + us_records

    for _, normalized_name, count in records:
        if normalized_name in NAI_SKIP_NAMES:
            continue
        target_name = PATENT_NAME_ALIASES.get(normalized_name)
        if target_name not in all_names:
            target_name = name_by_variant.get(normalized_name)
        if target_name is None:
            continue
        source_url = NAI_WORLD_PDF_URL if count == matched.get(target_name, {}).get("count") else None
        existing = matched.get(target_name)
        if existing is None or count > int(existing["count"]):
            matched[target_name] = {
                "count": count,
                "source_url": NAI_WORLD_PDF_URL if normalized_name in world_names else NAI_US_PDF_URL,
            }
        elif existing["count"] == count and source_url and existing.get("source_url") != NAI_WORLD_PDF_URL:
            existing["source_url"] = source_url
    return matched


def fetch_edurank_alumni_html(url: str) -> tuple[str, str]:
    alumni_url = url.rstrip("/") + "/alumni/"
    slug = url.rstrip("/").split("/")[-1]
    cache_path = EDURANK_ALUMNI_CACHE_DIR / f"{slug}.html"
    if cache_path.exists():
        return alumni_url, cache_path.read_text(encoding="utf-8", errors="replace")
    return alumni_url, fetch_text(alumni_url, f"edurank_alumni/{slug}.html")


def parse_startup_count(url: str) -> dict[str, Any]:
    alumni_url, html = fetch_edurank_alumni_html(url)
    soup = BeautifulSoup(html, "html.parser")
    if "not found" in normalize_text(soup.get_text(" ", strip=True)[:400]):
        return {"url": alumni_url, "count": None}
    alumni_items = soup.select("li.alumni")
    if not alumni_items:
        return {"url": alumni_url, "count": None}
    count = 0
    for item in alumni_items:
        text = normalize_text(item.get_text(" ", strip=True))
        if text and any(keyword in text for keyword in ENTREPRENEUR_KEYWORDS):
            count += 1
    return {"url": alumni_url, "count": count}


def build_startup_counts(rows: list[RowData]) -> dict[str, dict[str, Any]]:
    urls = sorted({row.edurank_source_url for row in rows if row.edurank_source_url})
    result: dict[str, dict[str, Any]] = {}
    with concurrent.futures.ThreadPoolExecutor(max_workers=8) as executor:
        future_map = {executor.submit(parse_startup_count, url): url for url in urls}
        for future in concurrent.futures.as_completed(future_map):
            url = future_map[future]
            try:
                result[url] = future.result()
            except Exception:
                result[url] = {"url": url.rstrip("/") + "/alumni/", "count": None}
    return result


def write_cell(worksheet, row_number: int, column_number: int, value: Any, hyperlink: bool = False) -> None:
    cell = worksheet.cell(row=row_number, column=column_number, value=value)
    if hyperlink and isinstance(value, str) and value:
        cell.hyperlink = value
        cell.style = "Hyperlink"


def enrich_workbook(workbook_path: Path, sheet_name: str) -> dict[str, Any]:
    workbook, worksheet, _, rows = load_rows(workbook_path, sheet_name)
    header_map = ensure_columns(worksheet)

    work_rights_by_country = parse_work_rights_table()
    visa_by_country = parse_visa_difficulty_table()
    patent_by_name = build_patent_counts(rows)
    startup_by_url = build_startup_counts(rows)

    filled_counts = {
        "patent_count": 0,
        "work_while_studying": 0,
        "visa_direct": 0,
        "visa_proxy": 0,
        "employment_proxy": 0,
        "startup_proxy": 0,
    }

    world_patent_names = {record[1] for record in parse_pdf_patent_rankings(NAI_WORLD_PDF_URL, "n_ai_world_2025.pdf")}

    for row in rows:
        additions: dict[str, Any] = {}
        notes = row.existing_notes
        coverage_additions: list[str] = []

        work_info = work_rights_by_country.get(row.base_country_key)
        if work_info is not None:
            additions["Work While Studying Allowed (Yes/No)"] = work_info["allowed"]
            additions["Work Rights Source URL"] = work_info["source_url"]
            coverage_additions.append("Work Rights")
            filled_counts["work_while_studying"] += 1

        visa_info = visa_by_country.get(row.base_country_key)
        if visa_info is not None:
            additions["Visa Difficulty Score"] = visa_info["score"]
            additions["Visa Difficulty Source URL"] = visa_info["source_url"]
            coverage_additions.append("Visa Refusal Rate")
            filled_counts["visa_direct"] += 1
        else:
            proxy_score = visa_proxy_score(row, work_info)
            if proxy_score is not None:
                additions["Visa Difficulty Score"] = proxy_score
                additions["Visa Difficulty Source URL"] = work_info["source_url"] if work_info is not None else VISA_REJECTION_URL
                notes = append_note(
                    notes,
                    "Visa difficulty is a proxy score where country refusal-rate data was unavailable; it combines work-rights burden, English-program ratio, and international-student share.",
                )
                coverage_additions.append("Visa Proxy")
                filled_counts["visa_proxy"] += 1

        patent_info = patent_by_name.get(row.university_name)
        if patent_info is not None:
            additions["Patent Count"] = patent_info["count"]
            additions["Patent Source URL"] = patent_info["source_url"]
            notes = append_note(notes, "Patent Count uses NAI 2025 granted U.S. utility-patent rankings.")
            coverage_additions.append("NAI Patents")
            filled_counts["patent_count"] += 1

        employment_value = row.graduate_employability if row.graduate_employability is not None else row.industry_placement
        if employment_value is not None:
            additions["Employment Rate After 6 Months (%)"] = employment_value
            if row.salary_source_url:
                additions["Employment Source URL"] = row.salary_source_url
            notes = append_note(
                notes,
                "Employment Rate After 6 Months is backfilled from the longer-horizon graduate-employability proxy and is not a literal 6-month survey.",
            )
            coverage_additions.append("Employment Proxy")
            filled_counts["employment_proxy"] += 1

        if row.edurank_source_url:
            startup_info = startup_by_url.get(row.edurank_source_url)
            if startup_info is not None and startup_info.get("count") is not None:
                additions["Startup Founded by Alumni Count"] = startup_info["count"]
                additions["Startup Alumni Source URL"] = startup_info["url"]
                notes = append_note(
                    notes,
                    "Startup Founded by Alumni Count is a proxy count of notable alumni tagged as founders, entrepreneurs, or investors on EduRank alumni pages.",
                )
                coverage_additions.append("EduRank Alumni Proxy")
                filled_counts["startup_proxy"] += 1

        new_coverage = merged_coverage(row.existing_coverage, coverage_additions)
        if new_coverage != row.existing_coverage:
            additions["Extended Metrics Coverage"] = new_coverage
        if notes != row.existing_notes:
            additions["Unofficial Metrics Notes"] = notes

        for header, value in additions.items():
            if value is None:
                continue
            write_cell(worksheet, row.row_number, header_map[header], value, hyperlink=header in HYPERLINK_HEADERS)

    backup_path = backup_workbook(workbook_path)
    workbook.save(workbook_path)

    patent_world_matches = sum(
        1
        for name, info in patent_by_name.items()
        if normalize_patent_name(name) in world_patent_names or info.get("source_url") == NAI_WORLD_PDF_URL
    )
    return {
        "rows": len(rows),
        "backup_path": str(backup_path),
        "filled_counts": filled_counts,
        "direct_visa_countries": len(visa_by_country),
        "work_rights_countries": len(work_rights_by_country),
        "matched_patent_institutions": len(patent_by_name),
        "world_patent_matches": patent_world_matches,
        "startup_urls_parsed": len(startup_by_url),
    }


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Fill remaining blank metric columns with relatively reliable country- and university-level sources.")
    parser.add_argument("--workbook", default="whed_data.xlsx")
    parser.add_argument("--sheet", default="Institutions")
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    print(json.dumps(enrich_workbook(Path(args.workbook), args.sheet), ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
