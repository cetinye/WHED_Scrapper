from __future__ import annotations

import json
import re
from pathlib import Path
from typing import Dict, Iterable, List

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

from isced_f import (
    classify_bachelor_program,
    classify_bachelors_cell,
    clean_program_title,
    split_bachelor_programs,
    write_bachelor_program_map,
)


ROOT_DIR = Path(__file__).resolve().parent

SECTION_TITLES = [
    "General Information",
    "Officers",
    "Divisions",
    "Degrees",
    "Academic Periodicals",
    "Student & Staff Numbers",
]

SECTION_SET = set(SECTION_TITLES)
IAU_ID_RE = re.compile(r"IAU-\d+")
MULTISPACE_RE = re.compile(r"\s+")
PROGRAM_DURATION_PREFIX_RE = re.compile(
    r"^\d[\d\s./+-]*(?:yr|yrs|year|years|sem|sems|semo|semester|semesters|month|months)\b",
    re.IGNORECASE,
)

ALLOWED_EXACT_COUNTRIES = {
    "Austria",
    "Belgium",
    "Bulgaria",
    "Canada",
    "Croatia",
    "Cyprus",
    "Czechia",
    "Denmark",
    "Estonia",
    "Finland",
    "France",
    "Germany",
    "Greece",
    "Hungary",
    "Ireland",
    "Italy",
    "Latvia",
    "Lithuania",
    "Luxembourg",
    "Malta",
    "Netherlands",
    "Poland",
    "Portugal",
    "Romania",
    "Slovak Republic",
    "Slovenia",
    "Spain",
    "Sweden",
}

ALLOWED_US_JURISDICTIONS = {
    "Alabama",
    "Alaska",
    "Arizona",
    "Arkansas",
    "California",
    "Colorado",
    "Connecticut",
    "Delaware",
    "District of Columbia",
    "Florida",
    "Georgia",
    "Hawaii",
    "Idaho",
    "Illinois",
    "Indiana",
    "Iowa",
    "Kansas",
    "Kentucky",
    "Louisiana",
    "Maine",
    "Maryland",
    "Massachusetts",
    "Michigan",
    "Minnesota",
    "Mississippi",
    "Missouri",
    "Montana",
    "Nebraska",
    "Nevada",
    "New Hampshire",
    "New Jersey",
    "New Mexico",
    "New York",
    "North Carolina",
    "North Dakota",
    "Ohio",
    "Oklahoma",
    "Oregon",
    "Pennsylvania",
    "Rhode Island",
    "South Carolina",
    "South Dakota",
    "Tennessee",
    "Texas",
    "Utah",
    "Vermont",
    "Virginia",
    "Washington",
    "West Virginia",
    "Wisconsin",
    "Wyoming",
}

OUTPUT_COLUMNS = [
    "Whed Link",
    "University Name",
    "IAU Code",
    "Native Name",
    "Country",
    "Street",
    "City",
    "Province",
    "Post Code",
    "Website",
    "Statistics Year",
    "Total Staff",
    "Total Student",
    "Institution Funding",
    "History",
    "Academic Year",
    "Admission Requirements",
    "Admission Requirement IDs",
    "Admission Requirements (Enriched)",
    "Annual Tuition / Cost",
    "Language(s)",
    "Accrediting Agency",
    "Student Body",
    "Permanent URL",
    "Student Statistics Year",
    "Staff Statistics Year",
    "Staff Full Time Total",
    "Staff Part Time Total",
    "Bachelor's Degree",
    "ISCED-F",
    "Master's Degree",
    "Doctor's Degree",
    "Diploma/Certificate",
    "General Information",
    "Officers",
    "Divisions",
    "Degrees",
    "Academic Periodicals",
    "Student & Staff Numbers",
    "Updated On",
    # "TXT File",
    "Raw Text",
]

PROGRAM_SOURCE_CHOICES = {"bachelors", "all-degree-fields"}

UNIVERSITY_TABLE_COLUMNS = [
    "university_key",
    "iau_code",
    "university_name",
    "native_name",
    "country",
    "street",
    "city",
    "province",
    "post_code",
    "website",
    "statistics_year",
    "total_staff",
    "total_student",
    "institution_funding",
    "history",
    "academic_year",
    "admission_requirements",
    "admission_requirement_ids",
    "admission_requirements_enriched",
    "annual_tuition_cost",
    "languages",
    "accrediting_agency",
    "student_body",
    "permanent_url",
    "student_statistics_year",
    "staff_statistics_year",
    "staff_full_time_total",
    "staff_part_time_total",
    "updated_on",
    "whed_link",
]

PROGRAM_TABLE_COLUMNS = [
    "program_key",
    "program_name",
    "isced_code",
]

UNIVERSITY_PROGRAM_TABLE_COLUMNS = [
    "university_program_key",
    "university_key",
    "program_key",
    "iau_code",
    "university_name",
    "program_name",
    "isced_code",
    "degree_types",
]

ADDRESS_FIELD_MAP = {
    "street": "Street",
    "city": "City",
    "province": "Province",
    "post code": "Post Code",
    "post-code": "Post Code",
    "www": "Website",
}

GENERAL_FIELD_MAP = {
    "institution funding": "Institution Funding",
    "history": "History",
    "academic year": "Academic Year",
    "admission requirements": "Admission Requirements",
    "tuition fees": "Annual Tuition / Cost",
    "language(s)": "Language(s)",
    "languages": "Language(s)",
    "accrediting agency": "Accrediting Agency",
    "student body": "Student Body",
}

IGNORED_GENERAL_HEADINGS = {
    "address",
    "other site",
    "other sites",
}

DEFAULT_ENRICHMENT_FILE = "whed_enrichment.jsonl"
DEFAULT_ADMISSION_REQUIREMENT_ID_FILE = ROOT_DIR / "References" / "Codes" / "admission_requirement_condition_ids.json"
ADMISSION_REQUIREMENT_MAPPING_SHEET = "Admission Requirement IDs"
INTERNAL_ADMISSION_REQUIREMENT_LABELS_KEY = "__admission_requirement_labels"
INTERNAL_PROGRAM_ITEMS_KEY = "__program_items"
INTERNAL_PROGRAM_SOURCE_KEY = "__program_source"
INTERNAL_UNIVERSITY_KEY = "__university_key"
GENERIC_REQUIREMENT_HINTS = (
    "certificate",
    "diploma",
    "degree",
    "qualification",
    "exam",
    "examination",
    "test",
    "entrance",
    "admission",
    "matura",
    "baccala",
    "bacalaureat",
    "high school",
    "secondary school",
    "university studies",
    "undergraduate",
    "college",
    "portfolio",
    "audition",
    "essay",
    "recommendation",
    "recommendations",
    "transcript",
    "gpa",
    "semester hours",
    "credit hours",
    "experience",
    "internship",
    "training",
    "language",
    "english",
    "german",
    "toefl",
    "ielts",
    "sat",
    "act",
    "gmat",
    "gre",
    "ged",
    "interview",
    "finances",
    "finance",
    "translation",
    "curriculum",
    "letter of recommendation",
)
GENERIC_LANGUAGE_ONLY_FRAGMENTS = {
    "english",
    "german",
    "french",
    "italian",
    "spanish",
    "portuguese",
    "arabic",
    "turkish",
    "polish",
    "romanian",
    "greek",
    "czech",
    "slovak",
    "hungarian",
    "russian",
}

STUDENT_FIELD_MAP = {
    "statistics year": "Student Statistics Year",
    "total": "Total Student",
}

STAFF_FIELD_MAP = {
    "statistics year": "Staff Statistics Year",
    "total": "Total Staff",
    "full time total": "Staff Full Time Total",
    "full-time total": "Staff Full Time Total",
    "part time total": "Staff Part Time Total",
    "part-time total": "Staff Part Time Total",
}

DEGREE_PATTERNS = {
    "Bachelor's Degree": (
        "bachelor",
        "licenc",
        "licence",
        "licenciat",
        "bacharel",
        "bakalavr",
        "undergraduate",
    ),
    "Master's Degree": (
        "master",
        "magister",
        "maestr",
        "msc",
        "mba",
        "graduate",
    ),
    "Doctor's Degree": (
        "doctor",
        "doktor",
        "doctorat",
        "doctorado",
        "doctorate",
        "phd",
        "ph.d",
    ),
    "Diploma/Certificate": (
        "diploma",
        "certificate",
        "certificat",
        "sertifika",
        "post-bachelor",
        "post bachelor",
    ),
}

DEGREE_NON_SUBJECT_LABELS = {
    "note",
    "notes",
    "duration",
    "remarks",
    "remark",
}

DEGREE_FIELD_LABEL = "fields of study"
DEGREE_DYNAMIC_INSERT_AFTER = "Diploma/Certificate"
INTERNAL_DEGREE_FIELDS_KEY = "__degree_fields_entries"


def normalize_space(value: str) -> str:
    return MULTISPACE_RE.sub(" ", value or "").strip()


def is_allowed_country(country: str) -> bool:
    country = normalize_space(country)
    if not country:
        return False

    if country in ALLOWED_EXACT_COUNTRIES:
        return True

    if country.startswith("Canada -"):
        return True

    if country.startswith("Belgium -"):
        return True

    if country.startswith("United States of America - "):
        jurisdiction = country.removeprefix("United States of America - ").strip()
        return jurisdiction in ALLOWED_US_JURISDICTIONS

    return False


def append_value(record: Dict[str, str], key: str, value: str) -> None:
    value = normalize_space(value)
    if not value:
        return

    current = normalize_space(record.get(key, ""))
    if not current:
        record[key] = value
        return

    existing_parts = {part.strip() for part in current.split(" | ") if part.strip()}
    if value in existing_parts:
        return

    record[key] = f"{current} | {value}"


def split_subject_items(value: str) -> List[str]:
    items: List[str] = []
    for part in re.split(r"\s*[;,]\s*", normalize_space(value)):
        clean = normalize_space(part)
        if clean:
            items.append(clean)
    return items


def merge_subject_values(current: str, value: str) -> str:
    merged_items = split_subject_items(current)
    seen = {item.casefold() for item in merged_items}

    for item in split_subject_items(value):
        lowered = item.casefold()
        if lowered not in seen:
            merged_items.append(item)
            seen.add(lowered)

    return ", ".join(merged_items)


def append_subject_values(record: Dict[str, str], key: str, value: str) -> None:
    merged = merge_subject_values(record.get(key, ""), value)
    if not merged:
        return
    record[key] = merged


def is_navigation_line(line: str) -> bool:
    return sum(1 for title in SECTION_TITLES if title in line) >= 2


def split_lines(text: str) -> List[str]:
    result: List[str] = []
    for raw_line in text.splitlines():
        line = normalize_space(raw_line)
        if line:
            result.append(line)
    return result


def split_key_value(line: str) -> tuple[str, str] | tuple[None, None]:
    if ":" not in line:
        return None, None
    key, value = line.split(":", 1)
    key = normalize_space(key)
    value = normalize_space(value)
    if not key:
        return None, None
    return key, value


def first_non_empty(values: Iterable[str]) -> str:
    for value in values:
        value = normalize_space(value)
        if value:
            return value
    return ""


def normalize_condition_label(value: str) -> str:
    return normalize_space(value).strip(" ,;.")


def unique_preserve_order(values: Iterable[str]) -> List[str]:
    result: List[str] = []
    seen: set[str] = set()
    for value in values:
        clean = normalize_condition_label(value)
        if not clean:
            continue
        lowered = clean.casefold()
        if lowered in seen:
            continue
        seen.add(lowered)
        result.append(clean)
    return result


def next_non_space_char(value: str, start_index: int) -> str:
    for index in range(start_index, len(value)):
        if not value[index].isspace():
            return value[index]
    return ""


def split_admission_requirement_clauses(value: str) -> List[str]:
    normalized = normalize_space(value)
    if not normalized:
        return []

    clauses: List[str] = []
    buffer: List[str] = []
    depth = 0

    def flush_buffer() -> None:
        clause = normalize_condition_label("".join(buffer))
        if clause:
            clauses.append(clause)
        buffer.clear()

    for index, char in enumerate(normalized):
        if char == "(":
            depth += 1
        elif char == ")" and depth > 0:
            depth -= 1

        should_split = False
        if depth == 0 and char == ";":
            should_split = True
        elif depth == 0 and char == ".":
            next_char = next_non_space_char(normalized, index + 1)
            should_split = not next_char or next_char.isupper()
        if should_split:
            flush_buffer()
            continue

        buffer.append(char)

    flush_buffer()
    return clauses


def clean_condition_fragment(value: str) -> str:
    cleaned = normalize_condition_label(value)
    lowered = cleaned.casefold()
    prefixes = (
        "and ",
        "or ",
        "also ",
        "plus ",
        "additionally ",
        "additionally, ",
    )
    for prefix in prefixes:
        if lowered.startswith(prefix):
            cleaned = normalize_condition_label(cleaned[len(prefix) :])
            lowered = cleaned.casefold()
    return cleaned


def split_fragment_on_commas(value: str) -> List[str]:
    normalized = clean_condition_fragment(value)
    if not normalized:
        return []

    parts: List[str] = []
    buffer: List[str] = []
    depth = 0

    def flush_buffer() -> None:
        fragment = clean_condition_fragment("".join(buffer))
        if fragment:
            parts.append(fragment)
        buffer.clear()

    for index, char in enumerate(normalized):
        if char == "(":
            depth += 1
        elif char == ")" and depth > 0:
            depth -= 1

        next_char = next_non_space_char(normalized, index + 1)
        prev_char = buffer[-1] if buffer else ""
        should_split = (
            depth == 0
            and char == ","
            and not (prev_char.isdigit() and next_char.isdigit())
        )

        if should_split:
            flush_buffer()
            continue

        buffer.append(char)

    flush_buffer()
    return parts or [normalized]


def looks_like_requirement_fragment(value: str) -> bool:
    lowered = clean_condition_fragment(value).casefold()
    if not lowered:
        return False
    return any(hint in lowered for hint in GENERIC_REQUIREMENT_HINTS)


def split_fragment_on_and(value: str) -> List[str]:
    normalized = clean_condition_fragment(value)
    if not normalized:
        return []

    parts: List[str] = []
    buffer: List[str] = []
    depth = 0
    index = 0
    separator = " and "

    def flush_buffer() -> None:
        fragment = clean_condition_fragment("".join(buffer))
        if fragment:
            parts.append(fragment)
        buffer.clear()

    while index < len(normalized):
        char = normalized[index]
        if char == "(":
            depth += 1
        elif char == ")" and depth > 0:
            depth -= 1

        if depth == 0 and normalized[index : index + len(separator)].casefold() == separator:
            flush_buffer()
            index += len(separator)
            continue

        buffer.append(char)
        index += 1

    flush_buffer()
    if len(parts) <= 1:
        return parts or [normalized]

    if any(part.casefold() in GENERIC_LANGUAGE_ONLY_FRAGMENTS for part in parts):
        return [normalized]

    if not all(looks_like_requirement_fragment(part) or len(part.split()) >= 3 for part in parts):
        return [normalized]

    return parts


def extract_generic_admission_requirement_conditions(value: str) -> List[str]:
    clauses = split_admission_requirement_clauses(value)
    if not clauses:
        return []

    conditions: List[str] = []
    for clause in clauses:
        comma_fragments = split_fragment_on_commas(clause)
        for comma_fragment in comma_fragments:
            and_fragments = split_fragment_on_and(comma_fragment)
            if and_fragments:
                conditions.extend(and_fragments)
            else:
                clean_fragment = clean_condition_fragment(comma_fragment)
                if clean_fragment:
                    conditions.append(clean_fragment)

    return unique_preserve_order(conditions)


def classify_germany_admission_requirement_clause(clause: str) -> List[str]:
    lowered = normalize_condition_label(clause).casefold()
    if not lowered:
        return []

    labels: List[str] = []

    def add(label: str) -> None:
        if label not in labels:
            labels.append(label)

    has_specific_secondary_certificate = False

    if "reifezeugn" in lowered:
        add("Secondary school certificate (Reifezeugnis)")
        has_specific_secondary_certificate = True

    if any(token in lowered for token in ("abitur", "allgemeine hochschulreife", "allgemeine hochschulsreife")):
        add("Secondary school certificate (Abitur / Allgemeine Hochschulreife)")
        has_specific_secondary_certificate = True

    if any(
        token in lowered
        for token in (
            "fachhochschulreife",
            "fachhochschulereife",
            "fachhochschulreifezeugnis",
            "fachabitur",
            "advanced technical college entrance qualification",
        )
    ):
        add("Secondary school certificate (Fachhochschulreife / Fachabitur)")
        has_specific_secondary_certificate = True

    if "fachgebundene" in lowered:
        add("Secondary school certificate (Fachgebundene Hochschulreife)")
        has_specific_secondary_certificate = True

    if "hochschulzugangsberechtigung" in lowered:
        add("School leaving certificate (Hochschulzugangsberechtigung)")
        has_specific_secondary_certificate = True

    if not has_specific_secondary_certificate and any(
        token in lowered
        for token in (
            "secondary school certificate",
            "secondary school leaving certificate",
            "school leaving certificate",
            "school-leaving certificate",
            "final secondary school examination",
            "high school diploma",
            "graduation from high school",
            "graduation diploma",
        )
    ):
        add("Secondary school certificate")

    if any(
        token in lowered
        for token in (
            "university entry qualification",
            "university entrance certificate",
            "higher education entry qualification",
            "higher education entrance qualification",
            "general qualification for university entrance",
            "general or subject-specific university entrance qualification",
            "school-leaving certificate qualifying for university entrance",
            "academic standard required for university entrance",
            "entrance diploma",
            "a-level",
            "a-levels",
            "matura",
            "heeq",
        )
    ):
        add("University entry qualification")

    if "technical training" in lowered:
        add("Technical training")

    if "2 years' university studies" in lowered or "2 years university studies" in lowered:
        add("Previous university studies")

    if "bachelor" in lowered or "equivalence to ba" in lowered:
        add("Bachelor's degree or equivalent")

    if (
        "staatsexamen" in lowered
        or "magister artium" in lowered
        or "advanced academic degree" in lowered
        or re.search(r"\bdiplom\b", lowered)
    ):
        add("Advanced academic degree (Diplom / Staatsexamen / Magister Artium)")

    if "meisterprü" in lowered or "master craftman" in lowered:
        add("Master craftman's diploma (Meisterprüfung)")

    if any(token in lowered for token in ("zulassungsprü", "zalassungsprü", "special course of studies")):
        add("Special admission examination (Zulassungsprüfung)")

    german_language_tokens = (
        "german language",
        "knowledge of german",
        "knowledge of the german language",
        "good command of the german language",
        "proficiency in german",
        "sufficient knowledge of german",
        "certificate of excellent knowledge of german",
        "german language level",
        "german language proficiency",
        "german language test",
        "testdaf",
        "test daf",
        " dsh",
        "dsh ",
        "(dsh",
        "pnds",
        "zentrale mittelstufenprüfung",
        "(zmp)",
        "daf or dsh",
        "certified knowledge of the german language",
        " c1",
        " c1 ",
    )
    has_german_language_requirement = any(token in lowered for token in german_language_tokens)
    if has_german_language_requirement:
        add("German language proficiency")

    if not has_german_language_requirement and any(
        token in lowered
        for token in (
            "language examination",
            "language competence",
            "language requirements",
            "language test for foreign students",
            "language certificate",
        )
    ):
        add("Language proficiency / examination")

    has_english_test = any(token in lowered for token in ("toefl", "ielts"))
    if not has_english_test and any(
        token in lowered
        for token in (
            "proficiency in german and english",
            "proficiency in english",
            "good command of english",
            "english language",
            "programmes taught in english",
            "programmes in english",
            "taught in english",
        )
    ):
        add("English language proficiency")

    if "toefl" in lowered:
        add("TOEFL")

    if "ielts" in lowered:
        add("IELTS")

    if "sat" in lowered:
        add("SAT")

    if "gmat" in lowered:
        add("GMAT")

    if "working/training contract" in lowered:
        add("Working / training contract")

    if "work experience" in lowered or "practical experience" in lowered:
        add("Work experience")

    practical_training_tokens = (
        "practical training",
        "praktikum",
        "internship",
        "placement of 6 weeks",
        "pre-study internship",
        "pre-practical",
        "pre-practicals",
        "vorpraktikum",
        "3 months practical",
        "relevant practical experience",
        "practical in subject of study",
        "basic three-month internship",
    )
    if any(token in lowered for token in practical_training_tokens):
        add("Practical training / internship")

    if any(token in lowered for token in ("selection process", "auswahlverfahren", "qualification examination", "special entrance qualification")):
        add("Selection / qualification procedure")

    if "entrance examination" in lowered or "entrance exam" in lowered:
        add("Entrance examination")

    if "aptitude test" in lowered or "aptitude tests" in lowered or "physical aptitude test" in lowered:
        add("Aptitude test")

    if "portfolio" in lowered:
        add("Portfolio")

    if "audition" in lowered:
        add("Audition")

    if "artistic" in lowered:
        add("Artistic aptitude evaluation")

    if "armed forces officer examination" in lowered or "armed forces officers examination" in lowered:
        add("Armed Forces Officer examination")

    if any(token in lowered for token in ("12-yr engagement", "12 year engagement", "12-year engagement")):
        add("Military service commitment")

    if "personality" in lowered or "intelligence structure test" in lowered:
        add("Personality / intelligence test")

    if any(token in lowered for token in ("depends on the chosen study programme", "admission procedure")):
        add("Programme-specific admission procedure")

    return labels


def extract_germany_admission_requirement_conditions(value: str) -> List[str]:
    clauses = split_admission_requirement_clauses(value)
    if not clauses:
        return []

    conditions: List[str] = []
    for clause in clauses:
        clause_conditions = classify_germany_admission_requirement_clause(clause)
        if clause_conditions:
            conditions.extend(clause_conditions)
            continue
        conditions.append(clause)

    return unique_preserve_order(conditions)


def extract_country_admission_requirement_conditions(country: str, value: str) -> List[str]:
    if normalize_space(country) == "Germany":
        return extract_germany_admission_requirement_conditions(value)
    return extract_generic_admission_requirement_conditions(value)


def extract_primary_url(value: str) -> str:
    match = re.search(r"https?://\S+", value or "", flags=re.IGNORECASE)
    return match.group(0).rstrip(").,;") if match else normalize_space(value)


def parse_general_information(lines: List[str], record: Dict[str, str]) -> None:
    active_field = ""
    buffer: List[str] = []

    def flush_active() -> None:
        nonlocal active_field, buffer
        if active_field and buffer:
            value = " ".join(buffer)
            append_value(record, active_field, value)
            if active_field == "Staff Full Time Total" and not normalize_space(record.get("Total Staff", "")):
                record["Total Staff"] = value
        active_field = ""
        buffer = []

    for line in lines:
        if line == "* * *":
            flush_active()
            continue

        lowered_line = line.casefold()
        if active_field and lowered_line.startswith(("http://", "https://")):
            buffer.append(line)
            continue

        if lowered_line in GENERAL_FIELD_MAP:
            flush_active()
            active_field = GENERAL_FIELD_MAP[lowered_line]
            continue

        if lowered_line in IGNORED_GENERAL_HEADINGS:
            flush_active()
            continue

        key, value = split_key_value(line)
        if key is not None:
            key_lower = key.casefold()
            mapped_address = ADDRESS_FIELD_MAP.get(key_lower)
            mapped_general = GENERAL_FIELD_MAP.get(key_lower)

            if active_field and not mapped_address and not mapped_general:
                buffer.append(f"{key}: {value}" if value else key)
                continue

            flush_active()
            target_field = mapped_address or mapped_general or key
            if mapped_address and value:
                append_value(record, mapped_address, value)
            elif mapped_general and value:
                append_value(record, mapped_general, value)
            elif value:
                append_value(record, key, value)
            else:
                active_field = target_field
            continue

        if active_field:
            buffer.append(line)

    flush_active()


def parse_student_staff(lines: List[str], record: Dict[str, str]) -> None:
    current_group = ""
    active_field = ""
    buffer: List[str] = []

    def flush_active() -> None:
        nonlocal active_field, buffer
        if active_field and buffer:
            append_value(record, active_field, " ".join(buffer))
        active_field = ""
        buffer = []

    for line in lines:
        if line == "* * *":
            flush_active()
            continue

        if line in {"Students", "Staff"}:
            flush_active()
            current_group = line
            continue

        key, value = split_key_value(line)
        if key is None:
            if active_field:
                buffer.append(line)
            continue

        flush_active()

        if current_group == "Students":
            mapped = STUDENT_FIELD_MAP.get(key.casefold())
            target_field = mapped or f"Students {key}"
            if mapped and value:
                append_value(record, mapped, value)
            elif value:
                append_value(record, f"Students {key}", value)
            else:
                active_field = target_field
        elif current_group == "Staff":
            mapped = STAFF_FIELD_MAP.get(key.casefold())
            target_field = mapped or f"Staff {key}"
            if mapped and value:
                append_value(record, mapped, value)
                if mapped == "Staff Full Time Total" and not normalize_space(record.get("Total Staff", "")):
                    record["Total Staff"] = value
            elif value:
                append_value(record, f"Staff {key}", value)
            else:
                active_field = target_field

    flush_active()


def categorize_degree(title: str) -> str:
    lowered = title.casefold()
    for column, patterns in DEGREE_PATTERNS.items():
        if any(pattern in lowered for pattern in patterns):
            return column
    return ""


def is_degree_heading(line: str) -> bool:
    normalized = normalize_space(line)
    if not normalized or normalized.endswith(":"):
        return False

    key, _ = split_key_value(normalized)
    if key is not None:
        return False

    if not categorize_degree(normalized):
        return False

    lowered = normalized.casefold()
    if lowered in DEGREE_NON_SUBJECT_LABELS or lowered == DEGREE_FIELD_LABEL:
        return False

    has_parenthetical_degree = bool(
        re.search(
            r"\((?:[^)]*(degree|doctorate|ph\.?d|bachelor|master|diploma|certificate)[^)]*)\)",
            normalized,
            flags=re.IGNORECASE,
        )
    )
    if has_parenthetical_degree:
        return True

    if ";" in normalized or normalized.endswith("."):
        return False

    return len(normalized.split()) <= 12


def extract_inline_degree_subjects(title: str) -> str:
    cleaned = normalize_space(title)
    patterns = [
        r"^(?:Also\s+)?Diploma(?:/Certificate)?(?:\s+in)?\s+(.+)$",
        r"^(?:Post-bachelor'?s\s+)?Diploma/Certificate(?:\s+in)?\s+(.+)$",
        r"^Doctor of\s+(.+)$",
        r"^Master of\s+(.+)$",
        r"^Master in\s+(.+)$",
        r"^Bachelor of\s+(.+)$",
    ]
    for pattern in patterns:
        match = re.match(pattern, cleaned, flags=re.IGNORECASE)
        if match:
            return normalize_space(match.group(1))
    return ""


def parse_degree_entries(lines: List[str]) -> List[Dict[str, str]]:
    degree_entries: List[Dict[str, str]] = []
    current_entry: Dict[str, str] | None = None
    pending_label = ""
    pending_buffer: List[str] = []

    def flush_pending() -> None:
        nonlocal pending_label, pending_buffer
        if current_entry and pending_label == DEGREE_FIELD_LABEL and pending_buffer:
            current_entry["fields_of_study"] = merge_subject_values(
                current_entry.get("fields_of_study", ""),
                " ".join(pending_buffer),
            )
        pending_label = ""
        pending_buffer = []

    for line in lines:
        if line == "* * *":
            flush_pending()
            continue

        if is_degree_heading(line):
            flush_pending()
            current_entry = {
                "type": categorize_degree(line),
                "title": normalize_space(line),
                "fields_of_study": "",
            }
            inline_subjects = extract_inline_degree_subjects(line)
            if inline_subjects:
                current_entry["fields_of_study"] = merge_subject_values("", inline_subjects)
            degree_entries.append(current_entry)
            continue

        key, value = split_key_value(line)
        if key is not None:
            flush_pending()
            if current_entry and key.casefold() == DEGREE_FIELD_LABEL:
                if value:
                    current_entry["fields_of_study"] = merge_subject_values(
                        current_entry.get("fields_of_study", ""),
                        value,
                    )
                else:
                    pending_label = DEGREE_FIELD_LABEL
            continue

        if current_entry and pending_label == DEGREE_FIELD_LABEL:
            pending_buffer.append(line)

    flush_pending()

    return [
        entry
        for entry in degree_entries
        if normalize_space(entry.get("type", "")) and normalize_space(entry.get("fields_of_study", ""))
    ]


def parse_degrees(lines: List[str], record: Dict[str, str]) -> None:
    degree_entries = parse_degree_entries(lines)
    record[INTERNAL_DEGREE_FIELDS_KEY] = degree_entries

    for entry in degree_entries:
        append_subject_values(
            record,
            entry.get("type", ""),
            entry.get("fields_of_study", ""),
        )


def collect_sections(lines: List[str]) -> tuple[Dict[str, List[str]], str]:
    sections = {title: [] for title in SECTION_TITLES}
    updated_on = ""
    current_section = ""

    for line in lines:
        if is_navigation_line(line):
            continue

        if line in SECTION_SET:
            current_section = line
            continue

        if line in {"Students", "Staff"} and current_section in {"Degrees", "Academic Periodicals"}:
            current_section = "Student & Staff Numbers"
            sections[current_section].append(line)
            continue

        if line.startswith("Updated on "):
            updated_on = line.removeprefix("Updated on ").strip()
            current_section = ""
            continue

        if current_section:
            sections[current_section].append(line)

    return sections, updated_on


def parse_metadata(lines: List[str], record: Dict[str, str]) -> None:
    front_lines: List[str] = []

    for line in lines:
        if line.startswith("Source URL:"):
            record["Whed Link"] = line.split(":", 1)[1].strip()
            continue

        if line.startswith("Permanent URL:"):
            record["Permanent URL"] = line.split(":", 1)[1].strip()
            continue

        if line in SECTION_SET:
            break

        if is_navigation_line(line):
            continue

        front_lines.append(line)

    iau_id = ""
    for line in front_lines:
        match = IAU_ID_RE.search(line)
        if match:
            iau_id = match.group(0)
            break

    record["IAU Code"] = iau_id

    core_lines = [
        line
        for line in front_lines
        if not line.startswith("Source URL:")
        and not line.startswith("Permanent URL:")
        and not line.startswith("http://")
        and not line.startswith("https://")
    ]

    if core_lines:
        record["University Name"] = core_lines[0]

    if iau_id and iau_id in core_lines:
        id_index = core_lines.index(iau_id)
        between = [line for line in core_lines[1:id_index] if line != iau_id]
        if between:
            record["Native Name"] = " | ".join(between)

        if id_index + 1 < len(core_lines):
            record["Country"] = core_lines[id_index + 1]
    else:
        tail = core_lines[1:]
        if tail:
            record["Country"] = tail[-1]
            if len(tail) > 1:
                record["Native Name"] = " | ".join(tail[:-1])


def parse_txt_file(path: Path) -> Dict[str, str]:
    text = path.read_text(encoding="utf-8", errors="replace")
    lines = split_lines(text)

    record: Dict[str, str] = {column: "" for column in OUTPUT_COLUMNS}
    record["TXT File"] = str(path.resolve())
    record["Raw Text"] = text.strip()

    parse_metadata(lines, record)
    sections, updated_on = collect_sections(lines)
    record["Updated On"] = updated_on

    parse_general_information(sections["General Information"], record)
    parse_student_staff(sections["Student & Staff Numbers"], record)
    parse_degrees(sections["Degrees"], record)
    record["ISCED-F"] = classify_bachelors_cell(record.get("Bachelor's Degree", ""))

    for title in SECTION_TITLES:
        record[title] = "\n".join(sections[title]).strip()

    if not record["Website"]:
        for candidate in lines:
            if candidate.startswith("WWW:"):
                record["Website"] = candidate.split(":", 1)[1].strip()
                break

    record["Website"] = extract_primary_url(record.get("Website", ""))

    if not record["IAU Code"]:
        match = IAU_ID_RE.search(path.name)
        if match:
            record["IAU Code"] = match.group(0)

    if not record["University Name"]:
        record["University Name"] = path.stem

    student_year = normalize_space(record.get("Student Statistics Year", ""))
    staff_year = normalize_space(record.get("Staff Statistics Year", ""))
    if student_year and staff_year and student_year != staff_year:
        record["Statistics Year"] = f"{student_year} | {staff_year}"
    else:
        record["Statistics Year"] = student_year or staff_year

    return record


def load_enrichment_records(path: Path | None) -> Dict[str, Dict[str, str]]:
    if path is None:
        return {}

    path = Path(path)
    if not path.exists():
        return {}

    enrichment_records: Dict[str, Dict[str, str]] = {}
    for raw_line in path.read_text(encoding="utf-8", errors="replace").splitlines():
        line = raw_line.strip()
        if not line:
            continue
        try:
            payload = json.loads(line)
        except json.JSONDecodeError:
            continue

        iau_code = normalize_space(str(payload.get("iau_code", "")))
        if not iau_code:
            continue

        enrichment_records[iau_code] = {
            "Admission Requirements (Enriched)": normalize_space(
                str(payload.get("admission_requirements", ""))
            ),
            "Annual Tuition / Cost": normalize_space(str(payload.get("annual_tuition_cost", ""))),
        }

    return enrichment_records


def load_admission_requirement_id_maps(
    path: Path | None = DEFAULT_ADMISSION_REQUIREMENT_ID_FILE,
) -> Dict[str, Dict[str, int]]:
    if path is None:
        return {}

    path = Path(path)
    if not path.exists():
        return {}

    try:
        payload = json.loads(path.read_text(encoding="utf-8", errors="replace"))
    except json.JSONDecodeError:
        return {}

    if not isinstance(payload, dict):
        return {}

    result: Dict[str, Dict[str, int]] = {}
    for country, country_mapping in payload.items():
        if not isinstance(country_mapping, dict):
            continue

        normalized_country = normalize_space(str(country))
        if not normalized_country:
            continue

        parsed_mapping: Dict[str, int] = {}
        for label, condition_id in country_mapping.items():
            normalized_label = normalize_condition_label(str(label))
            try:
                parsed_id = int(condition_id)
            except (TypeError, ValueError):
                continue
            if normalized_label and parsed_id > 0:
                parsed_mapping[normalized_label] = parsed_id

        if parsed_mapping:
            result[normalized_country] = parsed_mapping

    return result


def write_admission_requirement_id_maps(
    country_maps: Dict[str, Dict[str, int]],
    path: Path | None = DEFAULT_ADMISSION_REQUIREMENT_ID_FILE,
) -> Path | None:
    if path is None:
        return None

    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)

    serializable: Dict[str, Dict[str, int]] = {}
    for country in sorted(country_maps):
        entries = country_maps[country]
        serializable[country] = {
            label: condition_id
            for label, condition_id in sorted(entries.items(), key=lambda item: (item[1], item[0].casefold()))
        }

    path.write_text(json.dumps(serializable, indent=2, ensure_ascii=True), encoding="utf-8")
    return path


def assign_admission_requirement_ids(
    records: List[Dict[str, str]],
    id_map_path: Path | None = DEFAULT_ADMISSION_REQUIREMENT_ID_FILE,
) -> tuple[Dict[str, Dict[str, int]], Dict[tuple[str, str], int]]:
    country_maps = load_admission_requirement_id_maps(id_map_path)
    usage_counts: Dict[tuple[str, str], int] = {}

    for record in records:
        country = normalize_space(record.get("Country", ""))
        conditions = extract_country_admission_requirement_conditions(country, record.get("Admission Requirements", ""))
        record[INTERNAL_ADMISSION_REQUIREMENT_LABELS_KEY] = conditions

        if not conditions:
            record["Admission Requirement IDs"] = ""
            continue

        country_map = country_maps.setdefault(country, {})
        next_id = max(country_map.values(), default=0) + 1
        row_ids: List[str] = []

        for condition in conditions:
            if condition not in country_map:
                country_map[condition] = next_id
                next_id += 1

            row_ids.append(str(country_map[condition]))
            usage_counts[(country, condition)] = usage_counts.get((country, condition), 0) + 1

        record["Admission Requirement IDs"] = ", ".join(row_ids)

    write_admission_requirement_id_maps(country_maps, id_map_path)
    return country_maps, usage_counts


def add_admission_requirement_mapping_sheet(
    workbook: Workbook,
    country_maps: Dict[str, Dict[str, int]],
    usage_counts: Dict[tuple[str, str], int],
) -> None:
    if ADMISSION_REQUIREMENT_MAPPING_SHEET in workbook.sheetnames:
        workbook.remove(workbook[ADMISSION_REQUIREMENT_MAPPING_SHEET])

    sheet = workbook.create_sheet(ADMISSION_REQUIREMENT_MAPPING_SHEET)
    sheet.append(["Country", "Condition ID", "Condition", "Usage Count"])

    header_font = Font(bold=True)
    wrap_alignment = Alignment(vertical="top", wrap_text=True)

    for cell in sheet[1]:
        cell.font = header_font
        cell.alignment = wrap_alignment

    for country in sorted(country_maps):
        for condition, condition_id in sorted(country_maps[country].items(), key=lambda item: (item[1], item[0].casefold())):
            sheet.append(
                [
                    country,
                    condition_id,
                    condition,
                    usage_counts.get((country, condition), 0),
                ]
            )

    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = wrap_alignment

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions


def build_degree_field_columns(max_degree_field_count: int) -> List[str]:
    columns: List[str] = []
    for index in range(1, max_degree_field_count + 1):
        columns.extend(
            [
                f"Degree Fields {index} Type",
                f"Degree Fields {index} Title",
                f"Degree Fields {index} Subjects",
            ]
        )
    return columns


def build_output_columns(max_degree_field_count: int) -> List[str]:
    dynamic_columns = build_degree_field_columns(max_degree_field_count)
    if not dynamic_columns:
        return list(OUTPUT_COLUMNS)

    columns: List[str] = []
    inserted = False
    for column in OUTPUT_COLUMNS:
        columns.append(column)
        if column == DEGREE_DYNAMIC_INSERT_AFTER:
            columns.extend(dynamic_columns)
            inserted = True

    if not inserted:
        columns.extend(dynamic_columns)

    return columns


def build_output_row(record: Dict[str, str], max_degree_field_count: int) -> List[str]:
    row: List[str] = []
    degree_entries = record.get(INTERNAL_DEGREE_FIELDS_KEY, [])
    dynamic_values: List[str] = []

    for index in range(max_degree_field_count):
        entry = degree_entries[index] if index < len(degree_entries) else {}
        dynamic_values.extend(
            [
                entry.get("type", ""),
                entry.get("title", ""),
                entry.get("fields_of_study", ""),
            ]
        )

    inserted = False
    for column in OUTPUT_COLUMNS:
        row.append(record.get(column, ""))
        if column == DEGREE_DYNAMIC_INSERT_AFTER:
            row.extend(dynamic_values)
            inserted = True

    if not inserted:
        row.extend(dynamic_values)

    return row


def resolve_enrichment_file(enrichment_file: Path | None, preferred_dir: Path) -> Path | None:
    if enrichment_file is not None:
        return Path(enrichment_file)

    preferred_dir = Path(preferred_dir)
    default_candidate = preferred_dir / DEFAULT_ENRICHMENT_FILE
    return default_candidate if default_candidate.exists() else None


def collect_txt_records(
    input_dir: Path,
    enrichment_file: Path | None = None,
    include_all_countries: bool = False,
) -> tuple[List[Dict[str, str]], int, set[str], Dict[str, Dict[str, int]], Dict[tuple[str, str], int]]:
    input_dir = Path(input_dir)
    enrichment_records = load_enrichment_records(enrichment_file)

    txt_files = sorted(input_dir.glob("*.txt"), key=lambda item: item.name.casefold())
    bachelor_programs: set[str] = set()
    records: List[Dict[str, str]] = []
    max_degree_field_count = 0

    for txt_file in txt_files:
        record = parse_txt_file(txt_file)
        if not include_all_countries and not is_allowed_country(record.get("Country", "")):
            continue

        enrichment = enrichment_records.get(record.get("IAU Code", ""), {})
        record["Admission Requirements (Enriched)"] = first_non_empty(
            (
                enrichment.get("Admission Requirements (Enriched)", ""),
                record.get("Admission Requirements", ""),
            )
        )
        record["Annual Tuition / Cost"] = first_non_empty(
            (
                enrichment.get("Annual Tuition / Cost", ""),
                record.get("Annual Tuition / Cost", ""),
            )
        )
        bachelor_programs.update(split_bachelor_programs(record.get("Bachelor's Degree", "")))
        max_degree_field_count = max(max_degree_field_count, len(record.get(INTERNAL_DEGREE_FIELDS_KEY, [])))
        records.append(record)

    admission_requirement_id_maps, admission_requirement_usage_counts = assign_admission_requirement_ids(records)
    return (
        records,
        max_degree_field_count,
        bachelor_programs,
        admission_requirement_id_maps,
        admission_requirement_usage_counts,
    )


def normalize_program_source(program_source: str) -> str:
    normalized = normalize_space(program_source or "bachelors").casefold()
    if normalized not in PROGRAM_SOURCE_CHOICES:
        allowed = ", ".join(sorted(PROGRAM_SOURCE_CHOICES))
        raise ValueError(f"Unsupported program source '{program_source}'. Expected one of: {allowed}")
    return normalized


def is_noise_program_name(program_name: str) -> bool:
    normalized = normalize_space(program_name)
    if not normalized:
        return True

    lowered = normalized.casefold()
    if len(normalized) <= 1:
        return True

    if normalized[:1].islower():
        return True

    if "complete a set of curricular units" in lowered:
        return True

    if PROGRAM_DURATION_PREFIX_RE.match(normalized):
        return True

    if normalized[:1].isdigit() and (
        "%" in normalized
        or "+" in normalized
        or any(token in lowered for token in ("online", "programme", "program", "part-time", "full-time", " w/"))
    ):
        return True

    meta_patterns = (
        r"\bprogramme[s]?\b",
        r"\bprogram[s]?\b",
        r"\bcourse[s]?\b",
        r"\boffered\b",
        r"\bcollaboration\b",
        r"\bpartnership\b",
        r"\bpartner institution\b",
        r"\bdual degree\b",
        r"\bdual-degree\b",
        r"\baccelerated\b",
        r"\bpathway\b",
        r"\bresidency\b",
        r"\benrollment\b",
        r"\benrolment\b",
        r"\bsupplement\b",
        r"\bpreparatory\b",
        r"\bpostgraduate\b",
        r"\bdoctoral\b",
        r"\bpostdoctoral\b",
        r"\bexecutive\b",
        r"\bbachelor'?s degree\b",
        r"\bmasters? degree\b",
        r"\bpre-med\b",
        r"\bpre professional\b",
        r"\bpre-professional\b",
        r"\bfree courses\b",
        r"\bshort courses\b",
        r"\blicensure\b",
        r"\bhonors college\b",
    )
    if any(re.search(pattern, lowered) for pattern in meta_patterns):
        return True

    if re.search(r"\b(?:university|college|institute|seminary)\b", lowered):
        return True

    suffix_tokens = (" note", " lekarz", " magister", " mestre", " yrkesexamen", " lizentiat", " abschlusspr")
    if any(token in lowered for token in suffix_tokens):
        return True

    return False


def get_record_program_items(record: Dict[str, str], program_source: str = "bachelors") -> List[Dict[str, str]]:
    normalized_source = normalize_program_source(program_source)
    cached_source = record.get(INTERNAL_PROGRAM_SOURCE_KEY, "")
    cached_items = record.get(INTERNAL_PROGRAM_ITEMS_KEY, [])
    if cached_source == normalized_source and isinstance(cached_items, list):
        return cached_items

    items_by_key: Dict[str, Dict[str, object]] = {}

    def add_program(raw_name: str, degree_type: str = "") -> None:
        program_name = clean_program_title(raw_name)
        if is_noise_program_name(program_name):
            return

        key = program_name.casefold()
        program_item = items_by_key.setdefault(
            key,
            {
                "program_name": program_name,
                "isced_code": "",
                "degree_types": set(),
            },
        )

        isced_code = classify_bachelor_program(raw_name)
        if isced_code and not program_item["isced_code"]:
            program_item["isced_code"] = isced_code
        if degree_type:
            program_item["degree_types"].add(degree_type)

    if normalized_source == "bachelors":
        for raw_name in split_bachelor_programs(record.get("Bachelor's Degree", "")):
            add_program(raw_name, "Bachelor's Degree")
    else:
        for entry in record.get(INTERNAL_DEGREE_FIELDS_KEY, []):
            degree_type = normalize_space(entry.get("type", ""))
            for raw_name in split_subject_items(entry.get("fields_of_study", "")):
                add_program(raw_name, degree_type)

        if not items_by_key:
            for raw_name in split_bachelor_programs(record.get("Bachelor's Degree", "")):
                add_program(raw_name, "Bachelor's Degree")

    result: List[Dict[str, str]] = []
    for item in sorted(items_by_key.values(), key=lambda value: str(value["program_name"]).casefold()):
        degree_types = sorted((str(value) for value in item["degree_types"]), key=str.casefold)
        result.append(
            {
                "program_name": str(item["program_name"]),
                "isced_code": str(item["isced_code"]),
                "degree_types": " | ".join(degree_types),
            }
        )

    record[INTERNAL_PROGRAM_SOURCE_KEY] = normalized_source
    record[INTERNAL_PROGRAM_ITEMS_KEY] = result
    return result


def build_university_table_rows(records: List[Dict[str, str]]) -> List[Dict[str, object]]:
    rows: List[Dict[str, object]] = []

    for university_key, record in enumerate(records, start=1):
        record[INTERNAL_UNIVERSITY_KEY] = university_key
        rows.append(
            {
                "university_key": university_key,
                "iau_code": record.get("IAU Code", ""),
                "university_name": record.get("University Name", ""),
                "native_name": record.get("Native Name", ""),
                "country": record.get("Country", ""),
                "street": record.get("Street", ""),
                "city": record.get("City", ""),
                "province": record.get("Province", ""),
                "post_code": record.get("Post Code", ""),
                "website": record.get("Website", ""),
                "statistics_year": record.get("Statistics Year", ""),
                "total_staff": record.get("Total Staff", ""),
                "total_student": record.get("Total Student", ""),
                "institution_funding": record.get("Institution Funding", ""),
                "history": record.get("History", ""),
                "academic_year": record.get("Academic Year", ""),
                "admission_requirements": record.get("Admission Requirements", ""),
                "admission_requirement_ids": record.get("Admission Requirement IDs", ""),
                "admission_requirements_enriched": record.get("Admission Requirements (Enriched)", ""),
                "annual_tuition_cost": record.get("Annual Tuition / Cost", ""),
                "languages": record.get("Language(s)", ""),
                "accrediting_agency": record.get("Accrediting Agency", ""),
                "student_body": record.get("Student Body", ""),
                "permanent_url": record.get("Permanent URL", ""),
                "student_statistics_year": record.get("Student Statistics Year", ""),
                "staff_statistics_year": record.get("Staff Statistics Year", ""),
                "staff_full_time_total": record.get("Staff Full Time Total", ""),
                "staff_part_time_total": record.get("Staff Part Time Total", ""),
                "updated_on": record.get("Updated On", ""),
                "whed_link": record.get("Whed Link", ""),
            }
        )

    return rows


def build_program_table_rows(
    records: List[Dict[str, str]],
    program_source: str = "bachelors",
) -> tuple[List[Dict[str, object]], Dict[str, Dict[str, object]]]:
    normalized_source = normalize_program_source(program_source)
    programs_by_name: Dict[str, Dict[str, object]] = {}

    for record in records:
        for item in get_record_program_items(record, normalized_source):
            key = item["program_name"].casefold()
            program_row = programs_by_name.setdefault(
                key,
                {
                    "program_key": 0,
                    "program_name": item["program_name"],
                    "isced_code": item["isced_code"],
                },
            )
            if item["isced_code"] and not program_row["isced_code"]:
                program_row["isced_code"] = item["isced_code"]

    rows: List[Dict[str, object]] = []
    for program_key, program_row in enumerate(
        sorted(programs_by_name.values(), key=lambda value: str(value["program_name"]).casefold()),
        start=1,
    ):
        program_row["program_key"] = program_key
        rows.append(dict(program_row))

    return rows, programs_by_name


def build_university_program_table_rows(
    records: List[Dict[str, str]],
    program_lookup: Dict[str, Dict[str, object]],
    program_source: str = "bachelors",
) -> List[Dict[str, object]]:
    normalized_source = normalize_program_source(program_source)
    rows: List[Dict[str, object]] = []

    for university_program_key, record in enumerate(records, start=1):
        record[INTERNAL_UNIVERSITY_KEY] = university_program_key

    row_id = 1
    for record in records:
        university_key = int(record.get(INTERNAL_UNIVERSITY_KEY, 0) or 0)
        for item in get_record_program_items(record, normalized_source):
            program_row = program_lookup.get(item["program_name"].casefold())
            if not program_row:
                continue

            rows.append(
                {
                    "university_program_key": row_id,
                    "university_key": university_key,
                    "program_key": program_row["program_key"],
                    "iau_code": record.get("IAU Code", ""),
                    "university_name": record.get("University Name", ""),
                    "program_name": item["program_name"],
                    "isced_code": item["isced_code"] or program_row.get("isced_code", ""),
                    "degree_types": item["degree_types"],
                }
            )
            row_id += 1

    return rows


def write_table_workbook(
    output_file: Path,
    sheet_name: str,
    columns: List[str],
    rows: List[Dict[str, object]],
) -> int:
    output_file = Path(output_file)
    output_file.parent.mkdir(parents=True, exist_ok=True)

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = sheet_name
    sheet.append(columns)

    header_font = Font(bold=True)
    wrap_alignment = Alignment(vertical="top", wrap_text=True)

    for cell in sheet[1]:
        cell.font = header_font
        cell.alignment = wrap_alignment

    for row in rows:
        sheet.append([row.get(column, "") for column in columns])

    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = wrap_alignment

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    autofit_worksheet(sheet, columns)
    workbook.save(output_file)
    return max(sheet.max_row - 1, 0)


def write_full_workbook(
    records: List[Dict[str, str]],
    output_file: Path,
    max_degree_field_count: int,
    bachelor_programs: Iterable[str],
    admission_requirement_id_maps: Dict[str, Dict[str, int]],
    admission_requirement_usage_counts: Dict[tuple[str, str], int],
) -> int:
    output_file = Path(output_file)
    output_file.parent.mkdir(parents=True, exist_ok=True)

    output_columns = build_output_columns(max_degree_field_count)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Institutions"
    sheet.append(output_columns)

    header_font = Font(bold=True)
    wrap_alignment = Alignment(vertical="top", wrap_text=True)

    for cell in sheet[1]:
        cell.font = header_font
        cell.alignment = wrap_alignment

    for record in records:
        sheet.append(build_output_row(record, max_degree_field_count))

    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = wrap_alignment

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    add_admission_requirement_mapping_sheet(
        workbook,
        admission_requirement_id_maps,
        admission_requirement_usage_counts,
    )
    workbook.save(output_file)
    write_bachelor_program_map(bachelor_programs)
    return max(sheet.max_row - 1, 0)


def write_relational_workbooks(
    records: List[Dict[str, str]],
    universities_output_file: Path,
    programs_output_file: Path,
    university_programs_output_file: Path,
    program_source: str = "bachelors",
) -> Dict[str, int]:
    university_rows = build_university_table_rows(records)
    program_rows, program_lookup = build_program_table_rows(records, program_source=program_source)
    university_program_rows = build_university_program_table_rows(
        records,
        program_lookup,
        program_source=program_source,
    )

    return {
        "universities": write_table_workbook(
            Path(universities_output_file),
            "universities",
            UNIVERSITY_TABLE_COLUMNS,
            university_rows,
        ),
        "programs": write_table_workbook(
            Path(programs_output_file),
            "programs",
            PROGRAM_TABLE_COLUMNS,
            program_rows,
        ),
        "university_programs": write_table_workbook(
            Path(university_programs_output_file),
            "university_programs",
            UNIVERSITY_PROGRAM_TABLE_COLUMNS,
            university_program_rows,
        ),
    }


def export_relational_tables_to_excel(
    input_dir: Path,
    universities_output_file: Path,
    programs_output_file: Path,
    university_programs_output_file: Path,
    enrichment_file: Path | None = None,
    include_all_countries: bool = False,
    program_source: str = "bachelors",
) -> Dict[str, int]:
    reference_output = Path(universities_output_file)
    resolved_enrichment_file = resolve_enrichment_file(enrichment_file, reference_output.parent)
    (
        records,
        _max_degree_field_count,
        _bachelor_programs,
        _admission_requirement_id_maps,
        _admission_requirement_usage_counts,
    ) = collect_txt_records(
        Path(input_dir),
        enrichment_file=resolved_enrichment_file,
        include_all_countries=include_all_countries,
    )

    counts = write_relational_workbooks(
        records,
        universities_output_file=Path(universities_output_file),
        programs_output_file=Path(programs_output_file),
        university_programs_output_file=Path(university_programs_output_file),
        program_source=program_source,
    )
    return counts


def export_txt_directory_outputs(
    input_dir: Path,
    output_file: Path | None = None,
    universities_output_file: Path | None = None,
    programs_output_file: Path | None = None,
    university_programs_output_file: Path | None = None,
    enrichment_file: Path | None = None,
    include_all_countries: bool = False,
    program_source: str = "bachelors",
) -> Dict[str, int]:
    preferred_dir = (
        Path(output_file).parent
        if output_file is not None
        else Path(universities_output_file).parent
        if universities_output_file is not None
        else Path.cwd()
    )
    resolved_enrichment_file = resolve_enrichment_file(enrichment_file, preferred_dir)
    (
        records,
        max_degree_field_count,
        bachelor_programs,
        admission_requirement_id_maps,
        admission_requirement_usage_counts,
    ) = collect_txt_records(
        Path(input_dir),
        enrichment_file=resolved_enrichment_file,
        include_all_countries=include_all_countries,
    )

    results: Dict[str, int] = {}
    if output_file is not None:
        results["full_workbook"] = write_full_workbook(
            records,
            Path(output_file),
            max_degree_field_count,
            bachelor_programs,
            admission_requirement_id_maps,
            admission_requirement_usage_counts,
        )

    if (
        universities_output_file is not None
        and programs_output_file is not None
        and university_programs_output_file is not None
    ):
        results.update(
            write_relational_workbooks(
                records,
                universities_output_file=Path(universities_output_file),
                programs_output_file=Path(programs_output_file),
                university_programs_output_file=Path(university_programs_output_file),
                program_source=program_source,
            )
        )

    return results


def autofit_worksheet(worksheet, column_names: List[str]) -> None:
    width_overrides = {
        "Whed Link": 42,
        "University Name": 34,
        "Native Name": 34,
        "university_name": 34,
        "program_name": 34,
        "Street": 28,
        "street": 28,
        "Website": 28,
        "website": 28,
        "General Information": 40,
        "Divisions": 40,
        "Degrees": 40,
        "ISCED-F": 18,
        "degree_types": 24,
        "Raw Text": 60,
    }

    for index, column_name in enumerate(column_names, start=1):
        letter = get_column_letter(index)
        if column_name in width_overrides:
            worksheet.column_dimensions[letter].width = width_overrides[column_name]
            continue

        max_length = len(column_name)
        for cell in worksheet[letter][: min(worksheet.max_row, 200)]:
            if cell.value is None:
                continue
            max_length = max(max_length, len(str(cell.value)))

        worksheet.column_dimensions[letter].width = min(max(max_length + 2, 12), 60)


def export_txt_directory_to_excel(
    input_dir: Path,
    output_file: Path,
    enrichment_file: Path | None = None,
    include_all_countries: bool = False,
) -> int:
    results = export_txt_directory_outputs(
        input_dir=Path(input_dir),
        output_file=Path(output_file),
        enrichment_file=enrichment_file,
        include_all_countries=include_all_countries,
    )
    return results.get("full_workbook", 0)
