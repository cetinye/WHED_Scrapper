import argparse
import concurrent.futures
import csv
import json
import re
import shutil
from pathlib import Path
from typing import Any

import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook

from whed_enrich import InstitutionRecord, find_scorecard_match


SCORECARD_CSV = Path(".cache/Most-Recent-Cohorts-Institution.csv")
NCES_CACHE_DIR = Path(".cache/nces_collegenavigator_pages")
SOURCE_NAME = "U.S. Department of Education - College Scorecard"
COUNT_SOURCE_NAME = "NCES College Navigator"
OUTPUT_COLUMNS = [
    "Acceptance Rate (%)",
    "Graduation Rate (%)",
    "Open Admission Policy",
    "Applicants Count",
    "Accepted Students Count",
    "Graduates Count",
    "Admission Difficulty Score (0=Easy, 100=Hard)",
    "Admission Difficulty Comment",
    "Graduation Difficulty Score (0=Easy, 100=Hard)",
    "Graduation Difficulty Comment",
    "Admission & Graduation Data Source",
    "Admission & Graduation Reference URL",
    "Count Data Source",
    "Count Data Notes",
    "Count Reference URL",
    "College Scorecard UNITID",
]


def backup_workbook(path: Path) -> Path:
    backup_path = path.with_name(f"{path.stem}.admission_outcomes_backup{path.suffix}")
    shutil.copy2(path, backup_path)
    return backup_path


def ensure_output_columns(worksheet) -> dict[str, int]:
    header_map = {
        worksheet.cell(row=1, column=column_index).value: column_index
        for column_index in range(1, worksheet.max_column + 1)
    }
    next_column = worksheet.max_column + 1
    for header in OUTPUT_COLUMNS:
        if header not in header_map:
            worksheet.cell(row=1, column=next_column, value=header)
            header_map[header] = next_column
            next_column += 1
    return {header: header_map[header] for header in OUTPUT_COLUMNS}


def clean_value(value: str | None) -> str:
    return (value or "").strip()


def parse_float(value: Any) -> float | None:
    if value in (None, "", "NA", "NULL", "PrivacySuppressed", "PS"):
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def load_scorecard_records(csv_path: Path) -> list[dict[str, object]]:
    records: list[dict[str, object]] = []
    with csv_path.open("r", encoding="utf-8", errors="replace", newline="") as handle:
        reader = csv.DictReader(handle)
        for row in reader:
            records.append(
                {
                    "id": row.get("UNITID", ""),
                    "school.name": row.get("INSTNM", ""),
                    "school.state": row.get("STABBR", ""),
                    "school.city": row.get("CITY", ""),
                    "school.school_url": row.get("INSTURL", ""),
                    "admission_rate": row.get("ADM_RATE_SUPP", "") or row.get("ADM_RATE_ALL", "") or row.get("ADM_RATE", ""),
                    "admission_rate_raw": row.get("ADM_RATE", ""),
                    "open_admission_policy": row.get("OPENADMP", ""),
                    "graduation_rate_150": row.get("C150_4_POOLED", "") or row.get("C150_4", ""),
                    "graduation_rate_lt150": row.get("C150_L4_POOLED", "") or row.get("C150_L4", ""),
                    "preddeg": row.get("PREDDEG", ""),
                }
            )
    return records


def scorecard_url(unitid: str) -> str:
    return f"https://collegescorecard.ed.gov/school/?{unitid}" if unitid else ""


def nces_url(unitid: str) -> str:
    return f"https://nces.ed.gov/collegenavigator/?id={unitid}" if unitid else ""


def open_admission_comment(raw_value: str) -> str | None:
    if raw_value == "1":
        return "Yes"
    if raw_value == "2":
        return "No"
    return None


def graduation_rate_from_candidate(candidate: dict[str, object]) -> float | None:
    preferred = parse_float(candidate.get("graduation_rate_150"))
    if preferred is not None:
        return preferred
    return parse_float(candidate.get("graduation_rate_lt150"))


def admission_difficulty(rate: float | None, open_admission_raw: str) -> tuple[float | None, str | None]:
    if open_admission_raw == "1":
        return 5.0, "Open admission / very easy to enter"
    if rate is None:
        return None, None

    score = round((1.0 - rate) * 100.0, 2)
    if rate <= 0.10:
        comment = "Extremely selective"
    elif rate <= 0.25:
        comment = "Highly selective"
    elif rate <= 0.50:
        comment = "Selective"
    elif rate <= 0.75:
        comment = "Moderately selective"
    elif rate <= 0.90:
        comment = "Accessible"
    else:
        comment = "Very accessible"
    return score, comment


def graduation_difficulty(rate: float | None) -> tuple[float | None, str | None]:
    if rate is None:
        return None, None

    score = round((1.0 - rate) * 100.0, 2)
    if rate >= 0.80:
        comment = "Low completion difficulty / high graduation success"
    elif rate >= 0.60:
        comment = "Moderate completion difficulty"
    elif rate >= 0.40:
        comment = "Challenging to finish"
    elif rate >= 0.20:
        comment = "Hard to finish / low graduation rate"
    else:
        comment = "Very hard to finish / very low graduation rate"
    return score, comment


def parse_int(value: str | None) -> int | None:
    digits = re.sub(r"[^0-9]", "", value or "")
    return int(digits) if digits else None


def parse_percent(value: str | None) -> float | None:
    match = re.search(r"(\d+(?:\.\d+)?)\s*%", value or "")
    if not match:
        return None
    return float(match.group(1)) / 100.0


def load_nces_html(unitid: str) -> str:
    cache_path = NCES_CACHE_DIR / f"{unitid}.html"
    if cache_path.exists():
        return cache_path.read_text(encoding="utf-8", errors="replace")

    response = requests.get(
        nces_url(unitid),
        headers={"User-Agent": "Mozilla/5.0 (compatible; WHED-Scrapping/1.0)"},
        timeout=30,
    )
    response.raise_for_status()

    NCES_CACHE_DIR.mkdir(parents=True, exist_ok=True)
    cache_path.write_text(response.text, encoding="utf-8")
    return response.text


def parse_nces_counts(html: str) -> dict[str, Any]:
    soup = BeautifulSoup(html, "html.parser")
    applicants_count = None
    accepted_students_count = None
    graduates_count = None

    for table in soup.find_all("table"):
        row_map: dict[str, list[str]] = {}
        for row in table.find_all("tr"):
            cells = [cell.get_text(" ", strip=True) for cell in row.find_all(["th", "td"])]
            if len(cells) >= 2:
                row_map[cells[0]] = cells[1:]

        if "Number of applicants" in row_map and "Percent admitted" in row_map:
            applicants_count = parse_int(row_map["Number of applicants"][0])
            admitted_percent = parse_percent(row_map["Percent admitted"][0])
            if applicants_count is not None and admitted_percent is not None:
                accepted_students_count = int(round(applicants_count * admitted_percent))
            break

    for table in soup.find_all("table"):
        text = table.get_text(" ", strip=True)
        if "Grand total" not in text or "Program" not in text:
            continue

        for row in table.find_all("tr"):
            cells = [cell.get_text(" ", strip=True) for cell in row.find_all(["th", "td"])]
            if not cells or cells[0] != "Grand total":
                continue

            parsed_values = [parse_int(cell_value) for cell_value in cells[1:]]
            if any(value is not None for value in parsed_values):
                graduates_count = sum(value or 0 for value in parsed_values)
            break

        if graduates_count is not None:
            break

    return {
        "applicants_count": applicants_count,
        "accepted_students_count": accepted_students_count,
        "graduates_count": graduates_count,
    }


def fetch_nces_counts(unitid: str) -> tuple[str, dict[str, Any]]:
    try:
        html = load_nces_html(unitid)
        parsed = parse_nces_counts(html)
    except Exception:
        parsed = {
            "applicants_count": None,
            "accepted_students_count": None,
            "graduates_count": None,
        }

    has_any_count = any(parsed.get(key) is not None for key in parsed)
    parsed.update(
        {
            "source": (
                f"{COUNT_SOURCE_NAME} (accepted count = applicants x percent admitted; "
                "graduates count = completions grand total)"
            )
            if has_any_count
            else None,
            "notes": "; ".join(
                note
                for note in [
                    "Accepted Students Count = Applicants Count x Percent Admitted"
                    if parsed.get("accepted_students_count") is not None
                    else None,
                    "Graduates Count = IPEDS Completions Grand Total"
                    if parsed.get("graduates_count") is not None
                    else None,
                ]
                if note
            )
            or None,
            "reference_url": nces_url(unitid) if has_any_count else None,
        }
    )
    return unitid, parsed


def load_nces_counts_for_unitids(unitids: set[str]) -> dict[str, dict[str, Any]]:
    if not unitids:
        return {}

    counts_by_unitid: dict[str, dict[str, Any]] = {}
    with concurrent.futures.ThreadPoolExecutor(max_workers=8) as executor:
        futures = [executor.submit(fetch_nces_counts, unitid) for unitid in sorted(unitids)]
        for future in concurrent.futures.as_completed(futures):
            unitid, result = future.result()
            counts_by_unitid[unitid] = result
    return counts_by_unitid


def workbook_record(row: tuple[Any, ...], headers: list[str]) -> InstitutionRecord:
    header_map = {header: idx for idx, header in enumerate(headers)}
    university_name = clean_value(str(row[header_map["University Name"]] or ""))
    iau_code = clean_value(str(row[header_map["IAU Code"]] or ""))
    country = clean_value(str(row[header_map["Country"]] or ""))
    city = clean_value(str(row[header_map["City"]] or ""))
    province = clean_value(str(row[header_map["Province"]] or ""))
    website = clean_value(str(row[header_map["Website"]] or ""))
    admission_requirements = clean_value(str(row[header_map["Admission Requirements"]] or ""))
    annual_tuition_cost = clean_value(str(row[header_map["Annual Tuition / Cost"]] or ""))

    return InstitutionRecord(
        txt_file="",
        university_name=university_name,
        iau_code=iau_code,
        country=country,
        city=city,
        province=province,
        website=website,
        admission_requirements=admission_requirements,
        annual_tuition_cost=annual_tuition_cost,
    )


def enrich_workbook(workbook_path: Path, sheet_name: str) -> dict[str, Any]:
    scorecard_records = load_scorecard_records(SCORECARD_CSV)

    wb = load_workbook(workbook_path)
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    headers = list(rows[0])
    column_map = ensure_output_columns(ws)
    matched_candidates: list[tuple[int, dict[str, object] | None]] = []
    unitids: set[str] = set()

    for row_number, row in enumerate(rows[1:], start=2):
        record = workbook_record(row, headers)
        candidate = find_scorecard_match(record, scorecard_records)
        matched_candidates.append((row_number, candidate))
        if candidate is not None:
            unitid = clean_value(str(candidate.get("id", "")))
            if unitid:
                unitids.add(unitid)

    nces_counts = load_nces_counts_for_unitids(unitids)

    matched_rows = 0
    filled_acceptance = 0
    filled_graduation = 0
    filled_applicants = 0
    filled_accepted_count = 0
    filled_graduates_count = 0

    for row_number, candidate in matched_candidates:
        write_values = {
            "Acceptance Rate (%)": None,
            "Graduation Rate (%)": None,
            "Open Admission Policy": None,
            "Applicants Count": None,
            "Accepted Students Count": None,
            "Graduates Count": None,
            "Admission Difficulty Score (0=Easy, 100=Hard)": None,
            "Admission Difficulty Comment": None,
            "Graduation Difficulty Score (0=Easy, 100=Hard)": None,
            "Graduation Difficulty Comment": None,
            "Admission & Graduation Data Source": None,
            "Admission & Graduation Reference URL": None,
            "Count Data Source": None,
            "Count Data Notes": None,
            "Count Reference URL": None,
            "College Scorecard UNITID": None,
        }

        if candidate is not None:
            matched_rows += 1
            unitid = clean_value(str(candidate.get("id", "")))
            count_data = nces_counts.get(unitid, {})
            admission_rate = parse_float(candidate.get("admission_rate"))
            graduation_rate = graduation_rate_from_candidate(candidate)
            open_admission_raw = clean_value(str(candidate.get("open_admission_policy", "")))

            admission_score, admission_comment = admission_difficulty(admission_rate, open_admission_raw)
            graduation_score, graduation_comment = graduation_difficulty(graduation_rate)

            acceptance_percent = round(admission_rate * 100.0, 2) if admission_rate is not None else None
            graduation_percent = round(graduation_rate * 100.0, 2) if graduation_rate is not None else None

            write_values.update(
                {
                    "Acceptance Rate (%)": acceptance_percent,
                    "Graduation Rate (%)": graduation_percent,
                    "Open Admission Policy": open_admission_comment(open_admission_raw),
                    "Applicants Count": count_data.get("applicants_count"),
                    "Accepted Students Count": count_data.get("accepted_students_count"),
                    "Graduates Count": count_data.get("graduates_count"),
                    "Admission Difficulty Score (0=Easy, 100=Hard)": admission_score,
                    "Admission Difficulty Comment": admission_comment,
                    "Graduation Difficulty Score (0=Easy, 100=Hard)": graduation_score,
                    "Graduation Difficulty Comment": graduation_comment,
                    "Admission & Graduation Data Source": SOURCE_NAME,
                    "Admission & Graduation Reference URL": scorecard_url(unitid),
                    "Count Data Source": count_data.get("source"),
                    "Count Data Notes": count_data.get("notes"),
                    "Count Reference URL": count_data.get("reference_url"),
                    "College Scorecard UNITID": unitid or None,
                }
            )

            if acceptance_percent is not None:
                filled_acceptance += 1
            if graduation_percent is not None:
                filled_graduation += 1
            if count_data.get("applicants_count") is not None:
                filled_applicants += 1
            if count_data.get("accepted_students_count") is not None:
                filled_accepted_count += 1
            if count_data.get("graduates_count") is not None:
                filled_graduates_count += 1

        for header, value in write_values.items():
            cell = ws.cell(row=row_number, column=column_map[header], value=value)
            if header in {"Admission & Graduation Reference URL", "Count Reference URL"} and value:
                cell.hyperlink = value
                cell.style = "Hyperlink"

    backup_path = backup_workbook(workbook_path)
    wb.save(workbook_path)
    return {
        "rows": ws.max_row - 1,
        "matched_rows": matched_rows,
        "filled_acceptance_rows": filled_acceptance,
        "filled_graduation_rows": filled_graduation,
        "filled_applicants_rows": filled_applicants,
        "filled_accepted_count_rows": filled_accepted_count,
        "filled_graduates_count_rows": filled_graduates_count,
        "backup_path": str(backup_path),
    }


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Add admission difficulty and graduation difficulty columns using official College Scorecard data where available.",
    )
    parser.add_argument("--workbook", default="whed_data.xlsx", help="Workbook to enrich in place.")
    parser.add_argument("--sheet", default="Institutions", help="Worksheet name that contains university rows.")
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    results = enrich_workbook(Path(args.workbook), args.sheet)
    print(json.dumps(results, ensure_ascii=False, indent=2), flush=True)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
