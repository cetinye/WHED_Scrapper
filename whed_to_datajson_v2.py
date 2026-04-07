from __future__ import annotations

import argparse
import json
from collections import Counter, defaultdict
from pathlib import Path
import re
from typing import Iterable

from openpyxl import load_workbook

from txt_to_excel import (
    build_city_indexes,
    build_country_lookup,
    build_district_indexes,
    match_city_id,
    normalize_text,
    parse_country_name,
    parse_region_value,
    slugify_identifier,
)
from whed_to_datajson import (
    build_germany_state_resolution_context,
    parse_us_state_name,
    resolve_germany_state_name,
)


DEGREE_FIELD_COUNT = 7


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description="Export WHED workbook data into the DataJSONv2 university-details format."
    )
    parser.add_argument(
        "--whed-file",
        default="whed_data.xlsx",
        help="WHED workbook that contains the Institutions sheet.",
    )
    parser.add_argument(
        "--countries-file",
        default=r"References/TercihAnalizi Database Tables/countries.xlsx",
        help="Countries reference workbook.",
    )
    parser.add_argument(
        "--cities-file",
        default=r"References/TercihAnalizi Database Tables/cities.xlsx",
        help="Cities reference workbook.",
    )
    parser.add_argument(
        "--districts-file",
        default=r"References/TercihAnalizi Database Tables/districts.xlsx",
        help="Districts reference workbook.",
    )
    parser.add_argument(
        "--output-dir",
        default="DataJSONv2",
        help="Directory that will receive the generated JSON files.",
    )
    return parser


def format_scalar(value: object) -> str | None:
    if value is None:
        return None

    if isinstance(value, bool):
        return "Yes" if value else "No"

    if isinstance(value, int):
        return str(value)

    if isinstance(value, float):
        if value.is_integer():
            return str(int(value))
        text = format(value, ".15g")
        return text.replace(".", ",")

    text = str(value).strip()
    return text or None


def compose_full_address(*parts: object) -> str | None:
    normalized_parts: list[str] = []
    for part in parts:
        value = format_scalar(part)
        if value:
            normalized_parts.append(value)
    if not normalized_parts:
        return None
    return ", ".join(normalized_parts)


def parse_officers(text: object) -> list[dict[str, str | None]]:
    if not text:
        return []

    lines = [str(line).strip() for line in str(text).splitlines() if str(line).strip()]
    officers: list[dict[str, str | None]] = []
    index = 0

    while index < len(lines):
        line = lines[index]
        if ":" not in line or line.casefold() == "job title:":
            index += 1
            continue

        role, name = [part.strip() for part in line.split(":", 1)]
        job_title: str | None = None
        index += 1

        if index < len(lines) and lines[index].casefold() == "job title:":
            index += 1
            if index < len(lines):
                job_title = lines[index].strip() or None
                index += 1

        if not role and not name and not job_title:
            continue

        officers.append(
            {
                "role": role or None,
                "name": name or None,
                "job_title": job_title,
            }
        )

    return officers


DIVISION_HEADER_RE = re.compile(r"^\s*(.+?)\s*:\s*(.+?)\s*$")


def parse_divisions(text: object) -> list[dict[str, str | None]]:
    if not text:
        return []

    lines = [str(line).strip() for line in str(text).splitlines() if str(line).strip()]
    divisions: list[dict[str, str | None]] = []
    index = 0

    while index < len(lines):
        line = lines[index]
        if line.casefold() in {"fields of study:", "more details:"}:
            index += 1
            continue

        header_match = DIVISION_HEADER_RE.match(line)
        if header_match is None:
            index += 1
            continue

        division_type = header_match.group(1).strip() or None
        division_name = header_match.group(2).strip() or None
        fields_of_study: str | None = None
        more_details: str | None = None
        index += 1

        while index < len(lines):
            marker = lines[index]
            if marker.casefold() == "fields of study:":
                index += 1
                if index < len(lines):
                    fields_of_study = lines[index].strip() or None
                    index += 1
                continue
            if marker.casefold() == "more details:":
                index += 1
                if index < len(lines):
                    more_details = lines[index].strip() or None
                    index += 1
                continue

            next_header = DIVISION_HEADER_RE.match(marker)
            if next_header is not None and marker.casefold() not in {"fields of study:", "more details:"}:
                break

            index += 1

        divisions.append(
            {
                "division_type": division_type,
                "division_name": division_name,
                "fields_of_study": fields_of_study,
                "more_details": more_details,
            }
        )

    return divisions


def extract_degree_fields(row: tuple[object, ...], header_index: dict[str, int]) -> list[dict[str, str | None]]:
    degree_fields: list[dict[str, str | None]] = []
    for number in range(1, DEGREE_FIELD_COUNT + 1):
        degree_type = format_scalar(row[header_index[f"Degree Fields {number} Type"]])
        degree_title = format_scalar(row[header_index[f"Degree Fields {number} Title"]])
        degree_subjects = format_scalar(row[header_index[f"Degree Fields {number} Subjects"]])
        if not any((degree_type, degree_title, degree_subjects)):
            continue
        degree_fields.append(
            {
                "degree_field_type": degree_type,
                "degree_field_title": degree_title,
                "degree_field_subjects": degree_subjects,
            }
        )
    return degree_fields


def build_row_payload(
    row: tuple[object, ...],
    header_index: dict[str, int],
) -> dict[str, object]:
    country_value = row[header_index["Country"]]
    street = row[header_index["Street"]]
    city = row[header_index["City"]]
    province = row[header_index["Province"]]
    post_code = row[header_index["Post Code"]]
    full_address = row[header_index["University Contact Address"]]

    full_address_value = format_scalar(full_address) or compose_full_address(
        street,
        city,
        province,
        post_code,
        country_value,
    )

    code = format_scalar(row[header_index["IAU Code"]])

    return {
        "code": code,
        "data": {
            "general_information": {
                "university_name": format_scalar(row[header_index["University Name"]]),
                "native_name": format_scalar(row[header_index["Native Name"]]),
                "country": format_scalar(country_value),
                "institution_funding": format_scalar(row[header_index["Institution Funding"]]),
                "history": format_scalar(row[header_index["History"]]),
                "academic_year": format_scalar(row[header_index["Academic Year"]]),
                "admission_requirements": format_scalar(row[header_index["Admission Requirements"]]),
                "admission_requirements_enriched": format_scalar(
                    row[header_index["Admission Requirements (Enriched)"]]
                ),
                "annual_tuition_cost": format_scalar(row[header_index["Annual Tuition / Cost"]]),
                "languages": format_scalar(row[header_index["Language(s)"]]),
                "accrediting_agency": format_scalar(row[header_index["Accrediting Agency"]]),
                "student_body": format_scalar(row[header_index["Student Body"]]),
                "statistics_year": format_scalar(row[header_index["Statistics Year"]]),
                "student_statistics_year": format_scalar(row[header_index["Student Statistics Year"]]),
                "staff_statistics_year": format_scalar(row[header_index["Staff Statistics Year"]]),
                "updated_on": format_scalar(row[header_index["Updated On"]]),
            },
            "location_information": {
                "street": format_scalar(street),
                "city": format_scalar(city),
                "province": format_scalar(province),
                "post_code": format_scalar(post_code),
                "full_address": full_address_value,
            },
            "contact_information": {
                "website": format_scalar(row[header_index["University Contact Website"]])
                or format_scalar(row[header_index["Website"]]),
                "contact_page": format_scalar(row[header_index["University Contact Page"]]),
                "email": format_scalar(row[header_index["University Contact Email"]]),
                "phone": format_scalar(row[header_index["University Contact Phone"]]),
                "phone_standardized": format_scalar(
                    row[header_index["University Contact Phone Standardized"]]
                ),
                "key_contacts": format_scalar(row[header_index["University Key Contacts"]]),
            },
            "officers": parse_officers(row[header_index["Officers"]]),
            "divisions": parse_divisions(row[header_index["Divisions"]]),
            "degree_fields": extract_degree_fields(row, header_index),
            "student_and_staff_numbers": {
                "total_students": format_scalar(row[header_index["Total Student"]]),
                "total_staff": format_scalar(row[header_index["Total Staff"]]),
                "staff_full_time_total": format_scalar(row[header_index["Staff Full Time Total"]]),
                "staff_part_time_total": format_scalar(row[header_index["Staff Part Time Total"]]),
            },
            "classification_information": {
                "bachelors_degree": format_scalar(row[header_index["Bachelor's Degree"]]),
                "masters_degree": format_scalar(row[header_index["Master's Degree"]]),
                "doctors_degree": format_scalar(row[header_index["Doctor's Degree"]]),
                "diploma_certificate": format_scalar(row[header_index["Diploma/Certificate"]]),
                "isced_f": format_scalar(row[header_index["ISCED-F"]]),
            },
            "living_cost_information": {
                "estimated_annual_student_cost_shared_housing_usd": format_scalar(
                    row[header_index["Estimated Annual Student Cost (Shared Housing, USD)"]]
                ),
                "estimated_annual_student_cost_solo_apartment_usd": format_scalar(
                    row[header_index["Estimated Annual Student Cost (Solo Apartment, USD)"]]
                ),
                "estimated_annual_base_living_cost_without_rent_usd": format_scalar(
                    row[header_index["Estimated Annual Base Living Cost Without Rent (USD)"]]
                ),
                "estimated_annual_housing_cost_shared_usd": format_scalar(
                    row[header_index["Estimated Annual Housing Cost Shared (USD)"]]
                ),
                "estimated_annual_housing_cost_solo_usd": format_scalar(
                    row[header_index["Estimated Annual Housing Cost Solo (USD)"]]
                ),
                "estimated_annual_food_cost_usd": format_scalar(
                    row[header_index["Estimated Annual Food Cost (USD)"]]
                ),
                "estimated_annual_transport_cost_usd": format_scalar(
                    row[header_index["Estimated Annual Transport Cost (USD)"]]
                ),
                "estimated_cost_source_level": format_scalar(
                    row[header_index["Estimated Cost Source Level"]]
                ),
            },
            "student_friendliness": {
                "data_level": format_scalar(row[header_index["Student Friendliness Data Level"]]),
                "data_completeness_percent": format_scalar(
                    row[header_index["Student Friendliness Data Completeness (%)"]]
                ),
                "affordability_score": format_scalar(
                    row[header_index["Student Friendliness Affordability Score"]]
                ),
                "daily_life_score": format_scalar(
                    row[header_index["Student Friendliness Daily Life Score"]]
                ),
                "mobility_score": format_scalar(row[header_index["Student Friendliness Mobility Score"]]),
                "environment_score": format_scalar(
                    row[header_index["Student Friendliness Environment Score"]]
                ),
                "academic_ecosystem_score": format_scalar(
                    row[header_index["Student Friendliness Academic Ecosystem Score"]]
                ),
                "score": format_scalar(row[header_index["Student Friendliness Score"]]),
                "verdict": format_scalar(row[header_index["Student Friendliness Verdict"]]),
                "summary": format_scalar(row[header_index["Student Friendliness Summary"]]),
            },
            "ranking_and_outcome_information": {
                "acceptance_rate_percent": format_scalar(row[header_index["Acceptance Rate (%)"]]),
                "graduation_rate_percent": format_scalar(row[header_index["Graduation Rate (%)"]]),
                "accepted_students_count": format_scalar(row[header_index["Accepted Students Count"]]),
                "graduates_count": format_scalar(row[header_index["Graduates Count"]]),
                "college_scorecard_unitid": format_scalar(row[header_index["College Scorecard UNITID"]]),
                "citation_per_faculty": format_scalar(row[header_index["Citation per Faculty"]]),
                "average_graduate_salary_usd": format_scalar(
                    row[header_index["Average Graduate Salary (USD)"]]
                ),
                "international_student_ratio_percent": format_scalar(
                    row[header_index["International Student Ratio (%)"]]
                ),
                "international_staff_ratio_percent": format_scalar(
                    row[header_index["International Staff Ratio (%)"]]
                ),
                "number_of_partner_universities": format_scalar(
                    row[header_index["Number of Partner Universities"]]
                ),
                "visa_difficulty_score": format_scalar(row[header_index["Visa Difficulty Score"]]),
                "work_while_studying_allowed": format_scalar(
                    row[header_index["Work While Studying Allowed (Yes/No)"]]
                ),
                "part_time_job_availability_score": format_scalar(
                    row[header_index["Part-time Job Availability Score"]]
                ),
            },
            "environment_and_lifestyle": {
                "cost_of_living_trend_annual_inflation_percent": format_scalar(
                    row[header_index["Cost of Living Trend (Annual Inflation %)"]]
                ),
                "inflation_adjusted_cost_index": format_scalar(
                    row[header_index["Inflation Adjusted Cost Index"]]
                ),
                "dorm_capacity": format_scalar(row[header_index["Dorm Capacity"]]),
                "climate_type": format_scalar(row[header_index["Climate Type"]]),
                "cultural_activity_score": format_scalar(row[header_index["Cultural Activity Score"]]),
                "nightlife_score": format_scalar(row[header_index["Nightlife Score"]]),
                "family_friendliness_score": format_scalar(
                    row[header_index["Family Friendliness Score"]]
                ),
                "digital_infrastructure_score": format_scalar(
                    row[header_index["Digital Infrastructure Score (5G vs)"]]
                ),
            },
        },
    }


def ensure_codes(
    records: Iterable[dict[str, object]],
    *,
    country_id: int,
) -> list[dict[str, object]]:
    fallback_counter = 0
    normalized_records: list[dict[str, object]] = []

    for record in records:
        code = record.get("code")
        if not code:
            fallback_counter += 1
            record["code"] = f"WHED-{country_id}-{fallback_counter}"
        normalized_records.append(record)

    normalized_records.sort(key=lambda item: str(item.get("code") or ""))
    return normalized_records


def write_country_payloads(
    *,
    output_dir: Path,
    grouped_records: dict[int, list[dict[str, object]]],
    country_names_by_id: dict[int, str],
) -> dict[str, int]:
    output_dir.mkdir(parents=True, exist_ok=True)

    all_countries: list[dict[str, object]] = []
    country_file_count = 0

    for country_id in sorted(grouped_records):
        payload = {
            "country_id": country_id,
            "locale": "en",
            "university_details": ensure_codes(grouped_records[country_id], country_id=country_id),
        }
        all_countries.append(payload)

        country_name = country_names_by_id.get(country_id, str(country_id))
        file_name = f"{country_id}-{slugify_identifier(country_name) or country_id}.json"
        output_path = output_dir / file_name
        output_path.write_text(
            json.dumps(payload, ensure_ascii=False, indent=2),
            encoding="utf-8",
        )
        country_file_count += 1

    (output_dir / "all_countries.json").write_text(
        json.dumps(all_countries, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )

    return {
        "country_json_files": country_file_count,
        "all_countries_entries": len(all_countries),
    }


def write_partition_payloads(
    *,
    output_dir: Path,
    country_id: int,
    grouped_records: dict[str, list[dict[str, object]]],
    file_prefix: str,
    partition_label: str,
) -> dict[str, int]:
    output_dir.mkdir(parents=True, exist_ok=True)

    manifest: list[dict[str, object]] = []
    total_universities = 0

    for partition_name in sorted(grouped_records):
        universities = ensure_codes(grouped_records[partition_name], country_id=country_id)
        payload = {
            "country_id": country_id,
            "locale": "en",
            "university_details": universities,
        }
        file_name = f"{file_prefix}-{slugify_identifier(partition_name)}.json"
        output_path = output_dir / file_name
        output_path.write_text(
            json.dumps(payload, ensure_ascii=False, indent=2),
            encoding="utf-8",
        )
        manifest.append(
            {
                partition_label: partition_name,
                "file": file_name,
                "universities": len(universities),
            }
        )
        total_universities += len(universities)

    (output_dir / "manifest.json").write_text(
        json.dumps(manifest, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )

    return {
        "partition_json_files": len(manifest),
        "partition_universities": total_universities,
    }


def export_datajson_v2(
    *,
    whed_file: Path,
    countries_file: Path,
    cities_file: Path,
    districts_file: Path,
    output_dir: Path,
) -> dict[str, int]:
    country_lookup = build_country_lookup(countries_file)
    country_names_by_id = {int(country_id): country_name for country_id, country_name in country_lookup.values()}
    city_exact_index, city_simplified_index, city_iso2_index = build_city_indexes(cities_file)
    district_exact_index, district_simplified_index = build_district_indexes(districts_file)
    germany_state_context = build_germany_state_resolution_context(cities_file, districts_file)

    workbook = load_workbook(whed_file, read_only=True, data_only=True)
    worksheet = workbook["Institutions"]
    rows = worksheet.iter_rows(values_only=True)
    headers = [str(cell) if cell is not None else "" for cell in next(rows)]
    header_index = {header: index for index, header in enumerate(headers)}

    grouped_records: dict[int, list[dict[str, object]]] = defaultdict(list)
    us_records_by_state: dict[str, list[dict[str, object]]] = defaultdict(list)
    germany_records_by_state: dict[str, list[dict[str, object]]] = defaultdict(list)
    stats = {
        "institution_rows_scanned": 0,
        "institution_rows_exported": 0,
        "country_misses": 0,
    }
    country_miss_samples: Counter[str] = Counter()
    germany_state_unresolved_names: list[str] = []

    for row in rows:
        stats["institution_rows_scanned"] += 1
        raw_country = row[header_index["Country"]]
        country_name = parse_country_name(raw_country)
        country_match = country_lookup.get(normalize_text(country_name))

        if country_match is None:
            stats["country_misses"] += 1
            country_miss_samples[str(raw_country or "").strip()] += 1
            continue

        country_id = int(country_match[0])
        matched_country_name = str(country_match[1] or "").strip()
        payload = build_row_payload(row, header_index)
        grouped_records[country_id].append(payload)
        stats["institution_rows_exported"] += 1

        us_state_name = parse_us_state_name(raw_country)
        if us_state_name:
            us_records_by_state[us_state_name].append(payload)

        if country_id == 78:
            region_value = parse_region_value(raw_country, row[header_index["Province"]])
            city_id = match_city_id(
                country_id=country_id,
                country_name=matched_country_name,
                city_value=row[header_index["City"]],
                region_value=region_value,
                city_exact_index=city_exact_index,
                city_simplified_index=city_simplified_index,
                city_iso2_index=city_iso2_index,
                district_exact_index=district_exact_index,
                district_simplified_index=district_simplified_index,
            )
            germany_state_name = resolve_germany_state_name(
                city_id=city_id,
                context=germany_state_context,
            )
            if germany_state_name:
                germany_records_by_state[germany_state_name].append(payload)
            else:
                germany_state_unresolved_names.append(
                    format_scalar(row[header_index["University Name"]]) or "<unknown>"
                )

    write_stats = write_country_payloads(
        output_dir=output_dir,
        grouped_records=grouped_records,
        country_names_by_id=country_names_by_id,
    )
    us_state_output_dir = output_dir / "UnitedStatesStates"
    us_state_write_stats = write_partition_payloads(
        output_dir=us_state_output_dir,
        country_id=229,
        grouped_records=us_records_by_state,
        file_prefix="229-united-states",
        partition_label="state",
    )
    germany_state_output_dir = output_dir / "GermanyStates"
    germany_state_write_stats = write_partition_payloads(
        output_dir=germany_state_output_dir,
        country_id=78,
        grouped_records=germany_records_by_state,
        file_prefix="78-germany",
        partition_label="state",
    )

    summary = {
        "source_file": str(whed_file.resolve()),
        "output_dir": str(output_dir.resolve()),
        **stats,
        **write_stats,
        "us_state_json_files": us_state_write_stats["partition_json_files"],
        "us_state_universities": us_state_write_stats["partition_universities"],
        "germany_state_json_files": germany_state_write_stats["partition_json_files"],
        "germany_state_universities": germany_state_write_stats["partition_universities"],
        "germany_state_unresolved": len(germany_state_unresolved_names),
        "germany_state_unresolved_names": germany_state_unresolved_names,
        "country_miss_samples": dict(country_miss_samples.most_common(20)),
    }
    (output_dir / "summary.json").write_text(
        json.dumps(summary, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )

    return summary


def main() -> None:
    args = build_parser().parse_args()
    summary = export_datajson_v2(
        whed_file=Path(args.whed_file),
        countries_file=Path(args.countries_file),
        cities_file=Path(args.cities_file),
        districts_file=Path(args.districts_file),
        output_dir=Path(args.output_dir),
    )
    print(f"[done] Institution rows scanned: {summary['institution_rows_scanned']}", flush=True)
    print(f"[done] Institution rows exported: {summary['institution_rows_exported']}", flush=True)
    print(f"[done] Country JSON files written: {summary['country_json_files']}", flush=True)
    print(f"[done] US state JSON files written: {summary['us_state_json_files']}", flush=True)
    print(f"[done] Germany state JSON files written: {summary['germany_state_json_files']}", flush=True)
    print(f"[warn] Country misses: {summary['country_misses']}", flush=True)
    print(f"[warn] Germany state unresolved: {summary['germany_state_unresolved']}", flush=True)


if __name__ == "__main__":
    main()
