from __future__ import annotations

import argparse
import html
import hashlib
import json
import re
import shutil
import threading
import unicodedata
from concurrent.futures import ThreadPoolExecutor, as_completed
from dataclasses import dataclass
from functools import lru_cache
from pathlib import Path
from typing import Any
from urllib.parse import unquote, urljoin, urlparse

import phonenumbers
import pycountry
import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook


WORKBOOK_PATH = Path("whed_data.xlsx")
ENRICHMENT_PATH = Path("whed_enrichment.jsonl")
CACHE_DIR = Path(".cache/contact_language_enrichment")
BACKUP_SUFFIX = ".language_contact_backup"
USER_AGENT = (
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
    "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/133.0 Safari/537.36"
)

NEW_COLUMNS = [
    "Admission Language Requirement Level",
    "University Contact Information",
    "University Contact Website",
    "University Contact Email",
    "University Contact Phone",
    "University Contact Phone Standardized",
    "University Contact Page",
    "University Contact Address",
    "University Key Contacts",
]

CONTACT_LINK_HINTS = (
    "contact",
    "contact-us",
    "contactus",
    "kontakt",
    "kontakte",
    "kontakti",
    "kontaktiere",
    "contat",
    "contacto",
    "contacte",
    "contatti",
    "contatti-us",
    "contato",
    "impressum",
    "directory",
    "about/contact",
)

EMAIL_RE = re.compile(r"(?i)\b[a-z0-9._%+-]+@[a-z0-9.-]+\.[a-z]{2,}\b")
EMAIL_SEARCH_RE = re.compile(
    r"(?<![A-Z0-9._%+-])([A-Z0-9._%+-]+@[A-Z0-9.-]+\.[A-Z]{2,})(?![A-Z0-9._%+-])",
    re.IGNORECASE,
)
PHONE_RE = re.compile(
    r"(?:(?:\+\d{1,3}[\s()./-]*)?(?:\(?\d{2,4}\)?[\s()./-]*){2,5}\d{2,4})"
)
WHITESPACE_RE = re.compile(r"\s+")
CEFR_RE = re.compile(
    r"\b(?:level\s*)?((?:A1|A2|B1|B2|C1|C2)(?:\s*[-/]\s*(?:A1|A2|B1|B2|C1|C2))?)\b",
    re.IGNORECASE,
)
IELTS_RE = re.compile(r"IELTS[^\d]{0,25}(\d(?:\.\d)?)", re.IGNORECASE)
TOEFL_NEAR_RE = re.compile(r"TOEFL.{0,80}", re.IGNORECASE)
PTE_RE = re.compile(r"\bPTE(?:\s+Academic)?[^\d]{0,20}(\d{2,3})", re.IGNORECASE)
DUOLINGO_RE = re.compile(r"(?:Duolingo|DET)[^\d]{0,20}(\d{2,3})", re.IGNORECASE)
CAMBRIDGE_RE = re.compile(r"\b(CAE|CPE|FCE|Cambridge English)\b", re.IGNORECASE)
LANGUAGE_RE = re.compile(
    r"\b("
    r"English|French|German|Spanish|Italian|Portuguese|Russian|Arabic|Turkish|Dutch|"
    r"Swedish|Norwegian|Danish|Finnish|Polish|Czech|Slovak|Hungarian|Romanian|Bulgarian|"
    r"Greek|Croatian|Serbian|Slovenian|Ukrainian|Chinese|Japanese|Korean|Thai|Vietnamese|"
    r"Indonesian|Malay|Hebrew|Persian|Urdu|Hindi"
    r")\b",
    re.IGNORECASE,
)
GENERIC_LANGUAGE_HINTS = (
    "toefl",
    "ielts",
    "pte",
    "duolingo",
    "cambridge",
    "cefr",
    "language proficiency",
    "english proficiency",
    "proof of english",
    "knowledge of english",
    "knowledge of french",
    "knowledge of german",
    "language certificate",
    "language skills",
    "language requirement",
)
CONTACT_ROLE_PRIORITY = (
    "admissions",
    "admission",
    "enrollment",
    "enrolment",
    "registrar",
    "international",
    "student",
    "contact",
    "recruit",
    "provost",
    "president",
    "rector",
    "head",
)
CONTACT_PAGE_SKIP_HINTS = (
    "faq",
    "news",
    "press",
    "alumni",
    "privacy",
    "cookies",
    "events",
)
NOISE_EMAIL_DOMAIN_HINTS = (
    "sentry",
    "wixpress",
    "example.com",
)
NOISE_EMAIL_LOCAL_HINTS = (
    "noreply",
    "no-reply",
    "do-not-reply",
    "donotreply",
    "mailer-daemon",
)
COUNTRY_REGION_ALIASES = {
    "United States of America": "US",
    "Russia": "RU",
    "Viet Nam": "VN",
    "Türkiye": "TR",
    "Czechia": "CZ",
    "Taiwan, China": "TW",
    "Republic of Moldova": "MD",
    "Palestine": "PS",
    "Bolivia": "BO",
    "Tanzania": "TZ",
    "Venezuela": "VE",
    "Iran, Islamic Republic of": "IR",
    "Syrian Arab Republic": "SY",
    "Laos": "LA",
    "Brunei Darussalam": "BN",
    "Micronesia": "FM",
    "Korea, Republic of": "KR",
    "Korea, Democratic People's Republic of": "KP",
}

thread_local = threading.local()


@dataclass(frozen=True)
class RowInput:
    row_number: int
    iau_code: str
    university_name: str
    country: str
    website: str
    street: str
    city: str
    province: str
    post_code: str
    languages: str
    officers: str
    admission_requirements: str
    admission_requirements_enriched: str


@dataclass(frozen=True)
class RowResult:
    row_number: int
    admission_language_value: str
    contact_value: str
    contact_website_value: str
    contact_email_value: str
    contact_phone_value: str
    contact_phone_standardized_value: str
    contact_page_value: str
    contact_address_value: str
    key_contacts_value: str
    contact_email_count: int
    contact_phone_count: int
    used_contact_page: bool


def normalize_text(value: Any) -> str:
    return WHITESPACE_RE.sub(" ", str(value or "").replace("\xa0", " ")).strip()


def normalize_multiline_text(value: Any) -> str:
    text = str(value or "").replace("\r\n", "\n").replace("\r", "\n").replace("\xa0", " ")
    lines = [WHITESPACE_RE.sub(" ", line).strip() for line in text.splitlines()]
    return "\n".join(line for line in lines if line).strip()


def strip_invisible_chars(value: str) -> str:
    return "".join(
        char
        for char in value
        if unicodedata.category(char) != "Cf" and char not in {"\u00ad", "\ufeff"}
    )


def normalize_phone_text(value: str) -> str:
    value = html.unescape(str(value or ""))
    value = strip_invisible_chars(value.replace("\xa0", " "))
    for _ in range(3):
        decoded = unquote(value)
        if decoded == value:
            break
        value = decoded
    return normalize_text(value)


def clean_url(value: str) -> str:
    value = normalize_text(value)
    if not value:
        return ""
    if value.startswith("www."):
        value = f"https://{value}"
    if not value.startswith(("http://", "https://")):
        value = f"https://{value}"
    return value


def clean_domain(value: str) -> str:
    value = clean_url(value)
    if not value:
        return ""
    parsed = urlparse(value)
    return parsed.netloc.casefold().removeprefix("www.")


def domains_related(left: str, right: str) -> bool:
    left_domain = clean_domain(left)
    right_domain = clean_domain(right)
    if not left_domain or not right_domain:
        return False
    return (
        left_domain == right_domain
        or left_domain.endswith(f".{right_domain}")
        or right_domain.endswith(f".{left_domain}")
    )


def homepage_url(value: str) -> str:
    value = clean_url(value)
    if not value:
        return ""
    parsed = urlparse(value)
    scheme = parsed.scheme or "https"
    netloc = parsed.netloc or parsed.path
    return f"{scheme}://{netloc.strip('/')}/"


def base_country_name(value: str) -> str:
    return normalize_text(value).split(" - ", 1)[0].strip()


@lru_cache(maxsize=128)
def country_to_region_code(country_value: str) -> str:
    base_country = base_country_name(country_value)
    if not base_country:
        return ""
    alias = COUNTRY_REGION_ALIASES.get(base_country)
    if alias:
        return alias
    try:
        return pycountry.countries.search_fuzzy(base_country)[0].alpha_2
    except Exception:
        return ""


def get_session() -> requests.Session:
    session = getattr(thread_local, "session", None)
    if session is None:
        session = requests.Session()
        session.headers.update({"User-Agent": USER_AGENT})
        thread_local.session = session
    return session


def cache_key_for_url(url: str) -> str:
    return hashlib.sha1(url.encode("utf-8")).hexdigest()


def fetch_html(url: str, timeout: tuple[int, int] = (8, 15)) -> str:
    CACHE_DIR.mkdir(parents=True, exist_ok=True)
    cache_path = CACHE_DIR / f"{cache_key_for_url(url)}.html"
    if cache_path.exists():
        return cache_path.read_text(encoding="utf-8", errors="replace")

    session = get_session()
    response = session.get(url, timeout=timeout, allow_redirects=True)
    response.raise_for_status()
    content_type = response.headers.get("content-type", "")
    if "html" not in content_type and "<html" not in response.text[:500].casefold():
        raise ValueError(f"Non-HTML response for {url}")
    cache_path.write_text(response.text, encoding="utf-8")
    return response.text


def load_enrichment_index(path: Path) -> dict[str, dict[str, str]]:
    if not path.exists():
        return {}
    index: dict[str, dict[str, str]] = {}
    for raw_line in path.read_text(encoding="utf-8", errors="replace").splitlines():
        line = raw_line.strip()
        if not line:
            continue
        try:
            payload = json.loads(line)
        except json.JSONDecodeError:
            continue
        iau_code = normalize_text(payload.get("iau_code"))
        if iau_code:
            index[iau_code] = payload
    return index


def ensure_output_columns(worksheet) -> dict[str, int]:
    header_map = {
        worksheet.cell(row=1, column=column_index).value: column_index
        for column_index in range(1, worksheet.max_column + 1)
    }
    next_column = worksheet.max_column + 1
    for header in NEW_COLUMNS:
        if header not in header_map:
            worksheet.cell(row=1, column=next_column, value=header)
            header_map[header] = next_column
            next_column += 1
    return {header: header_map[header] for header in NEW_COLUMNS}


def backup_workbook(path: Path) -> Path:
    backup_path = path.with_name(f"{path.stem}{BACKUP_SUFFIX}{path.suffix}")
    shutil.copy2(path, backup_path)
    return backup_path


def build_row_inputs(worksheet) -> tuple[list[RowInput], dict[str, int]]:
    headers = [cell.value for cell in next(worksheet.iter_rows(min_row=1, max_row=1))]
    header_map = {header: index for index, header in enumerate(headers)}

    def cell_value(values: tuple[Any, ...], header: str, multiline: bool = False) -> str:
        column_index = header_map.get(header)
        if column_index is None or column_index >= len(values):
            return ""
        if multiline:
            return normalize_multiline_text(values[column_index])
        return normalize_text(values[column_index])

    rows: list[RowInput] = []
    for row_number, values in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
        rows.append(
            RowInput(
                row_number=row_number,
                iau_code=cell_value(values, "IAU Code"),
                university_name=cell_value(values, "University Name"),
                country=cell_value(values, "Country"),
                website=cell_value(values, "Website"),
                street=cell_value(values, "Street"),
                city=cell_value(values, "City"),
                province=cell_value(values, "Province"),
                post_code=cell_value(values, "Post Code"),
                languages=cell_value(values, "Language(s)"),
                officers=cell_value(values, "Officers", multiline=True),
                admission_requirements=cell_value(values, "Admission Requirements"),
                admission_requirements_enriched=cell_value(values, "Admission Requirements (Enriched)"),
            )
        )
    return rows, header_map


def split_admission_sources(row: RowInput, enrichment_index: dict[str, dict[str, str]]) -> list[str]:
    values = [
        row.admission_requirements_enriched,
        normalize_text(enrichment_index.get(row.iau_code, {}).get("admission_requirements")),
        row.admission_requirements,
    ]
    unique_values: list[str] = []
    seen: set[str] = set()
    for value in values:
        normalized = normalize_text(value)
        if normalized and normalized.casefold() not in seen:
            unique_values.append(normalized)
            seen.add(normalized.casefold())
    return unique_values


def find_language_name(text: str, instruction_languages: str) -> str:
    match = LANGUAGE_RE.search(text or "")
    if match:
        return match.group(1).title()
    if instruction_languages:
        first = normalize_text(instruction_languages).split(";")[0]
        if first:
            return first
    return "Instruction language"


def extract_toefl_details(text: str) -> list[str]:
    details: list[str] = []
    for snippet in TOEFL_NEAR_RE.findall(text):
        lowered = snippet.casefold()
        for value in re.findall(r"\d{2,3}", snippet):
            numeric = int(value)
            detail = ""
            if ("ibt" in lowered or "internet" in lowered) and 30 <= numeric <= 130:
                detail = f"iBT {numeric}"
            elif ("computer" in lowered or "cbt" in lowered) and 30 <= numeric <= 300:
                detail = f"CBT {numeric}"
            elif ("paper" in lowered or "pbt" in lowered) and 300 <= numeric <= 677:
                detail = f"PBT {numeric}"
            elif 30 <= numeric <= 130:
                detail = f"iBT {numeric}"
            elif 300 <= numeric <= 677:
                detail = f"PBT {numeric}"
            elif 30 <= numeric <= 300:
                detail = f"CBT {numeric}"
            if detail and detail not in details:
                details.append(detail)
    return details


def classify_language_requirement(text: str, instruction_languages: str) -> tuple[int, str]:
    normalized = normalize_text(text)
    if not normalized:
        if instruction_languages:
            return (
                1,
                f"Instruction language(s): {instruction_languages}. No explicit admission language threshold found.",
            )
        return (0, "No explicit admission language threshold found in available source text.")

    language_name = find_language_name(normalized, instruction_languages)
    cefr_values = list(dict.fromkeys(value.upper().replace(" ", "") for value in CEFR_RE.findall(normalized)))
    ielts_values = [
        value
        for value in dict.fromkeys(IELTS_RE.findall(normalized))
        if 0.0 <= float(value) <= 9.0
    ]
    pte_values = [
        value
        for value in dict.fromkeys(PTE_RE.findall(normalized))
        if 10 <= int(value) <= 90
    ]
    duolingo_values = [
        value
        for value in dict.fromkeys(DUOLINGO_RE.findall(normalized))
        if 10 <= int(value) <= 160
    ]
    toefl_values = extract_toefl_details(normalized)
    cambridge_values = list(dict.fromkeys(value.upper() for value in CAMBRIDGE_RE.findall(normalized)))

    detail_parts: list[str] = []
    if cefr_values:
        detail_parts.append("CEFR " + "/".join(cefr_values))
    if ielts_values:
        detail_parts.append("IELTS " + "/".join(ielts_values))
    if toefl_values:
        detail_parts.append("TOEFL " + "/".join(toefl_values))
    if pte_values:
        detail_parts.append("PTE " + "/".join(pte_values))
    if duolingo_values:
        detail_parts.append("Duolingo " + "/".join(duolingo_values))
    if cambridge_values:
        detail_parts.append("Cambridge " + "/".join(cambridge_values))

    lowered = normalized.casefold()
    if detail_parts:
        return (3, f"{language_name} proficiency required: " + "; ".join(detail_parts) + ".")

    if any(hint in lowered for hint in GENERIC_LANGUAGE_HINTS):
        return (2, f"{language_name} proficiency required; score/level not specified.")

    if instruction_languages:
        return (
            1,
            f"Instruction language(s): {instruction_languages}. No explicit admission language threshold found.",
        )
    return (0, "No explicit admission language threshold found in available source text.")


def best_admission_language_value(row: RowInput, enrichment_index: dict[str, dict[str, str]]) -> str:
    candidates = split_admission_sources(row, enrichment_index)
    best_score = -1
    best_value = ""
    for candidate in candidates:
        score, value = classify_language_requirement(candidate, row.languages)
        if score > best_score:
            best_score = score
            best_value = value
    if best_value:
        return best_value
    _, fallback_value = classify_language_requirement("", row.languages)
    return fallback_value


def clean_phone(value: str) -> str:
    value = normalize_phone_text(value).strip(" .;,:|")
    if not value:
        return ""
    value = re.sub(r"^(?:tel|telephone|telefon|phone|tlf|ph)[\s:./-]*", "", value, flags=re.IGNORECASE)
    value = value.strip(" .;,:|")
    digits = re.sub(r"\D", "", value)
    if len(digits) < 7 or len(digits) > 15:
        return ""
    if len(set(digits)) == 1:
        return ""
    if not any(symbol in value for symbol in ("+", "(", ")", "-", ".", "/", " ")) and not value.startswith("+"):
        if len(digits) in {10, 11, 12}:
            return digits
        return ""
    return value


def looks_like_invalid_phone_text(value: str) -> bool:
    normalized = normalize_phone_text(value)
    if not normalized:
        return True
    if re.search(r"[A-Za-z]{2,}", normalized):
        return True
    compact = re.sub(r"\s*([./-])\s*", r"\1", normalized)
    if re.search(r"\b\d{1,2}[./-]\d{1,2}[./-]\d{2,4}\b", compact):
        return True
    invalid_patterns = (
        r"^\(?\d{4}\)?\s*[-/]\s*\d{2,4}$",
        r"^\d{4}[./-]\d{1,2}[./-]\d{1,2}(?:\s+\d{1,2}(?::\d{2})?)?$",
        r"^\d{1,2}[./-]\d{1,2}[./-]\d{2,4}(?:\s*[-–]\s*\d{1,2}[./-]\d{1,2}[./-]\d{2,4})?(?:\s+\d{1,2}(?::\d{2})?)?$",
        r"^\d{1,2}[./-]\d{1,2}(?:\s*[-–/]\s*\d{1,2}[./-]\d{1,2}[./-]\d{2,4})+$",
        r"^\d{1,2}[.:]\d{2}\s*[-–]\s*\d{1,2}[.:]\d{2}$",
        r"^\d{1,4}\s*/\s*\d{4}$",
        r"^\d{4}\s+\d{1,2}[./-]\d{1,2}[./-]\d{2,4}$",
        r"^(?:\d{4}\s+){2,}\d{4}$",
        r"^\d{4}\s+\d{1,2}\s+\d{1,2}\s*[-/]\s*\d{4}\s+\d{1,2}\s+\d{1,2}$",
        r"^\d{1,3}\.\d{6,}$",
        r"^\d{5}\s+\d{3}[.\-]\d{3}[.\-]\d{4}$",
        r"^\d{5}-\d{2,4}$",
    )
    return any(re.match(pattern, normalized) or re.match(pattern, compact) for pattern in invalid_patterns)


def unique_keep_order(values: list[str]) -> list[str]:
    seen: set[str] = set()
    result: list[str] = []
    for value in values:
        normalized = normalize_text(value)
        if not normalized:
            continue
        key = normalized.casefold()
        if key in seen:
            continue
        seen.add(key)
        result.append(normalized)
    return result


def unique_phone_values(values: list[str]) -> list[str]:
    seen_keys: set[str] = set()
    result: list[str] = []
    for raw_value in values:
        normalized_raw = normalize_phone_text(raw_value)
        candidates: list[str] = []
        for part in re.split(r"[,;|\n]+", normalized_raw):
            part = part.strip()
            if not part:
                continue
            formatted_candidates = re.findall(r"\+?\d[\d()./-]{5,}\d", part)
            valid_formatted = [
                candidate
                for candidate in formatted_candidates
                if not looks_like_invalid_phone_text(candidate) and clean_phone(candidate)
            ]
            if valid_formatted:
                candidates.extend(valid_formatted)
                continue

            if not looks_like_invalid_phone_text(part) and clean_phone(part):
                candidates.append(part)
                continue

            digit_only_tokens = re.findall(r"\b\d{7,12}\b", part)
            candidates.extend(digit_only_tokens)
        if not candidates and normalized_raw:
            candidates = [normalized_raw]

        for candidate in candidates:
            if looks_like_invalid_phone_text(candidate):
                continue
            cleaned = clean_phone(candidate)
            if not cleaned:
                continue
            digits = re.sub(r"\D", "", cleaned)
            if len(digits) == 11 and digits.startswith("1"):
                digits = digits[1:]
            if digits in seen_keys:
                continue
            seen_keys.add(digits)
            result.append(cleaned)
    return result


def standardize_phone_values(values: list[str], country_value: str) -> list[str]:
    region_code = country_to_region_code(country_value)
    seen: set[str] = set()
    result: list[str] = []
    for raw_phone in unique_phone_values(values):
        try:
            parsed = phonenumbers.parse(raw_phone, region_code or None)
        except phonenumbers.NumberParseException:
            continue
        if not phonenumbers.is_valid_number(parsed):
            continue
        standardized = phonenumbers.format_number(parsed, phonenumbers.PhoneNumberFormat.E164)
        if standardized in seen:
            continue
        seen.add(standardized)
        result.append(standardized)
    return result


def normalize_email_text(value: str) -> str:
    value = html.unescape(str(value or ""))
    value = strip_invisible_chars(value.replace("\xa0", " "))
    for _ in range(3):
        decoded = unquote(value)
        if decoded == value:
            break
        value = decoded
    replacements = {
        "(at)": "@",
        "[at]": "@",
        "{at}": "@",
        "(AT)": "@",
        "[AT]": "@",
        "{AT}": "@",
        "(dot)": ".",
        "[dot]": ".",
        "{dot}": ".",
        "(DOT)": ".",
        "[DOT]": ".",
        "{DOT}": ".",
    }
    for source, target in replacements.items():
        value = value.replace(source, target)
    value = re.sub(r"^\s*(?:mailto:|email\s+|e-mail\s+)", "", value, flags=re.IGNORECASE)
    value = value.lstrip("/")
    value = value.replace('"', " ").replace("'", " ").replace("<", " ").replace(">", " ")
    return normalize_text(value)


def is_valid_email(value: str) -> bool:
    if not value or " " in value or value.count("@") != 1:
        return False
    if ".." in value:
        return False
    local_part, domain_part = value.split("@", 1)
    if not local_part or not domain_part:
        return False
    if local_part.startswith(".") or local_part.endswith("."):
        return False
    if domain_part.startswith(".") or domain_part.endswith("."):
        return False
    if "." not in domain_part:
        return False
    if not re.fullmatch(r"[A-Z0-9._%+-]+", local_part, re.IGNORECASE):
        return False
    if not re.fullmatch(r"[A-Z0-9.-]+\.[A-Z]{2,}", domain_part, re.IGNORECASE):
        return False
    return True


def is_noise_email(value: str) -> bool:
    lowered = value.casefold()
    if "@" not in lowered:
        return True
    local_part, domain_part = lowered.split("@", 1)
    if any(hint in local_part for hint in NOISE_EMAIL_LOCAL_HINTS):
        return True
    if any(hint in domain_part for hint in NOISE_EMAIL_DOMAIN_HINTS):
        return True
    if domain_part == "example.com":
        return True
    return False


def extract_valid_emails(value: str) -> list[str]:
    normalized = normalize_email_text(value)
    found: list[str] = []
    for match in EMAIL_SEARCH_RE.findall(normalized):
        email = match.strip(" .,:;|/\\")
        if is_valid_email(email) and not is_noise_email(email):
            found.append(email.lower())
    for token in re.split(r"[\s,;]+", normalized):
        email = token.strip(" .,:;|/\\()[]{}<>\"'")
        if is_valid_email(email) and not is_noise_email(email):
            found.append(email.lower())
    return found


def unique_email_values(values: list[str]) -> list[str]:
    seen: set[str] = set()
    result: list[str] = []
    for raw_value in values:
        for email in extract_valid_emails(raw_value):
            key = email.casefold()
            if key in seen:
                continue
            seen.add(key)
            result.append(email)
    return result


def extract_contact_data_from_html(page_url: str, html: str, homepage: str) -> dict[str, Any]:
    soup = BeautifulSoup(html, "html.parser")
    emails: list[str] = []
    phones: list[str] = []
    contact_links: list[str] = []

    for tag in soup.select("a[href^='mailto:']"):
        href = tag.get("href", "")
        emails.extend(extract_valid_emails(href.split(":", 1)[1].split("?", 1)[0]))

    for tag in soup.select("a[href^='tel:']"):
        href = tag.get("href", "")
        phone = clean_phone(href.split(":", 1)[1])
        if phone:
            phones.append(phone)

    visible_text = soup.get_text(" ", strip=True)
    emails.extend(extract_valid_emails(visible_text))
    emails.extend(extract_valid_emails(html))
    if not phones:
        phones.extend(clean_phone(match) for match in PHONE_RE.findall(visible_text))

    for tag in soup.select("a[href]"):
        href = urljoin(page_url, tag.get("href", ""))
        if not href.startswith(("http://", "https://")):
            continue
        if not domains_related(href, homepage):
            continue
        candidate = href.split("#", 1)[0]
        text_blob = f"{tag.get_text(' ', strip=True)} {candidate}".casefold()
        if any(hint in text_blob for hint in CONTACT_LINK_HINTS):
            contact_links.append(candidate)

    return {
        "emails": unique_email_values(emails)[:3],
        "phones": unique_phone_values(phones)[:3],
        "contact_links": unique_keep_order(contact_links)[:8],
    }


def score_contact_link(url: str, anchor_blob: str) -> float:
    url_lower = url.casefold()
    blob_lower = anchor_blob.casefold()
    score = 0.0
    if "/contact" in url_lower or "contact-us" in url_lower or "contactus" in url_lower:
        score += 5.0
    if any(hint in url_lower for hint in CONTACT_LINK_HINTS):
        score += 2.5
    if any(hint in blob_lower for hint in CONTACT_LINK_HINTS):
        score += 2.0
    if any(skip in url_lower for skip in CONTACT_PAGE_SKIP_HINTS):
        score -= 1.5
    score -= min(len(url), 140) / 250.0
    return score


def discover_contact_page(homepage: str, homepage_html: str) -> str:
    soup = BeautifulSoup(homepage_html, "html.parser")
    candidates: list[tuple[float, str]] = []
    for tag in soup.select("a[href]"):
        href = urljoin(homepage, tag.get("href", ""))
        if not href.startswith(("http://", "https://")):
            continue
        if not domains_related(href, homepage):
            continue
        href = href.split("#", 1)[0]
        if href.rstrip("/") == homepage.rstrip("/"):
            continue
        anchor_blob = f"{tag.get_text(' ', strip=True)} {href}"
        score = score_contact_link(href, anchor_blob)
        if score > 1.5:
            candidates.append((score, href))

    if not candidates:
        fallback_candidates = [
            urljoin(homepage, "/contact"),
            urljoin(homepage, "/contact-us"),
            urljoin(homepage, "/about/contact"),
        ]
        candidates = [(score_contact_link(url, url), url) for url in fallback_candidates]

    ordered = sorted(candidates, reverse=True)
    for _, candidate in ordered:
        return candidate
    return ""


def parse_officer_entries(officers_text: str) -> list[str]:
    if not officers_text:
        return []
    entries: list[tuple[int, str]] = []
    lines = [line.strip() for line in officers_text.splitlines()]
    index = 0
    while index < len(lines):
        line = lines[index]
        if ":" not in line or line.lower() == "job title:":
            index += 1
            continue
        role, name = [part.strip() for part in line.split(":", 1)]
        title = ""
        if index + 2 < len(lines) and lines[index + 1].lower() == "job title:":
            title = lines[index + 2].strip()
            index += 3
        else:
            index += 1
        display = ""
        if name and title:
            display = f"{name} ({title})"
        elif name:
            display = f"{name} ({role})"
        if not display:
            continue
        blob = f"{role} {title}".casefold()
        priority = len(CONTACT_ROLE_PRIORITY) + 5
        for rank, token in enumerate(CONTACT_ROLE_PRIORITY):
            if token in blob:
                priority = rank
                break
        entries.append((priority, display))
    ordered = [item for _, item in sorted(entries, key=lambda item: (item[0], item[1].casefold()))]
    return unique_keep_order(ordered)[:2]


def compose_address(row: RowInput) -> str:
    parts = [row.street, row.city, row.province, row.post_code, row.country]
    return ", ".join(part for part in parts if part)


def fetch_contact_details(homepage: str) -> dict[str, Any]:
    if not homepage:
        return {"emails": [], "phones": [], "contact_page": ""}

    try:
        homepage_html = fetch_html(homepage)
        homepage_data = extract_contact_data_from_html(homepage, homepage_html, homepage)
    except Exception:
        return {"emails": [], "phones": [], "contact_page": ""}

    emails = list(homepage_data["emails"])
    phones = list(homepage_data["phones"])
    contact_page = ""
    should_follow_contact_page = not emails or not phones
    if should_follow_contact_page:
        candidate_page = discover_contact_page(homepage, homepage_html)
        if candidate_page:
            try:
                candidate_html = fetch_html(candidate_page)
                candidate_data = extract_contact_data_from_html(candidate_page, candidate_html, homepage)
                emails = unique_keep_order(emails + candidate_data["emails"])[:3]
                phones = unique_keep_order(phones + candidate_data["phones"])[:3]
                contact_page = candidate_page
            except Exception:
                contact_page = ""

    return {
        "emails": emails,
        "phones": phones,
        "contact_page": contact_page,
    }


def contact_cache_path(iau_code: str) -> Path:
    safe_code = normalize_text(iau_code) or "unknown"
    return CACHE_DIR / f"{safe_code}.contact.json"


def get_contact_details(row: RowInput, fetch_online_contact: bool) -> dict[str, Any]:
    if not fetch_online_contact:
        return {"emails": [], "phones": [], "contact_page": ""}
    CACHE_DIR.mkdir(parents=True, exist_ok=True)
    cache_path = contact_cache_path(row.iau_code)
    if cache_path.exists():
        try:
            return json.loads(cache_path.read_text(encoding="utf-8"))
        except json.JSONDecodeError:
            pass

    details = fetch_contact_details(homepage_url(row.website))
    cache_path.write_text(json.dumps(details, ensure_ascii=False, indent=2), encoding="utf-8")
    return details


def build_contact_value(row: RowInput, contact_details: dict[str, Any]) -> str:
    lines: list[str] = []
    website = build_contact_website_value(row)
    if website:
        lines.append(f"Website: {website}")
    emails = unique_email_values(list(contact_details.get("emails", [])))[:2]
    if emails:
        lines.append(f"Email: {', '.join(emails)}")
    phones = unique_phone_values(list(contact_details.get("phones", [])))[:2]
    if phones:
        lines.append(f"Phone: {', '.join(phones)}")
    contact_page = build_contact_page_value(contact_details)
    if contact_page:
        lines.append(f"Contact page: {contact_page}")
    address = build_contact_address_value(row)
    if address:
        lines.append(f"Address: {address}")
    key_contacts = build_key_contacts_value(row)
    if key_contacts:
        lines.append(f"Key contact(s): {key_contacts}")
    if not lines:
        lines.append("No contact information found in available source data.")
    return "\n".join(lines)


def build_contact_website_value(row: RowInput) -> str:
    return clean_url(row.website)


def build_contact_email_value(contact_details: dict[str, Any]) -> str:
    emails = unique_email_values(list(contact_details.get("emails", [])))[:2]
    return ", ".join(emails)


def build_contact_phone_value(contact_details: dict[str, Any]) -> str:
    phones = unique_phone_values(list(contact_details.get("phones", [])))[:2]
    return ", ".join(phones)


def build_contact_phone_standardized_value(row: RowInput, contact_details: dict[str, Any]) -> str:
    phones = standardize_phone_values(list(contact_details.get("phones", [])), row.country)[:2]
    return ", ".join(phones)


def build_contact_page_value(contact_details: dict[str, Any]) -> str:
    return normalize_text(contact_details.get("contact_page"))


def build_contact_address_value(row: RowInput) -> str:
    return compose_address(row)


def build_key_contacts_value(row: RowInput) -> str:
    officers = parse_officer_entries(row.officers)
    if officers:
        return "; ".join(officers)
    return ""


def process_row(
    row: RowInput,
    enrichment_index: dict[str, dict[str, str]],
    fetch_online_contact: bool,
) -> RowResult:
    admission_language_value = best_admission_language_value(row, enrichment_index)
    contact_details = (
        get_contact_details(row, fetch_online_contact)
        if row.website
        else {"emails": [], "phones": [], "contact_page": ""}
    )
    contact_value = build_contact_value(row, contact_details)
    contact_website_value = build_contact_website_value(row)
    contact_email_value = build_contact_email_value(contact_details)
    contact_phone_value = build_contact_phone_value(contact_details)
    contact_phone_standardized_value = build_contact_phone_standardized_value(row, contact_details)
    contact_page_value = build_contact_page_value(contact_details)
    contact_address_value = build_contact_address_value(row)
    key_contacts_value = build_key_contacts_value(row)
    return RowResult(
        row_number=row.row_number,
        admission_language_value=admission_language_value,
        contact_value=contact_value,
        contact_website_value=contact_website_value,
        contact_email_value=contact_email_value,
        contact_phone_value=contact_phone_value,
        contact_phone_standardized_value=contact_phone_standardized_value,
        contact_page_value=contact_page_value,
        contact_address_value=contact_address_value,
        key_contacts_value=key_contacts_value,
        contact_email_count=len(unique_email_values(list(contact_details.get("emails", [])))),
        contact_phone_count=len(unique_phone_values(list(contact_details.get("phones", [])))),
        used_contact_page=bool(contact_details.get("contact_page")),
    )


def run_enrichment(
    workbook_path: Path,
    enrichment_path: Path,
    workers: int,
    fetch_online_contact: bool,
) -> dict[str, Any]:
    workbook = load_workbook(workbook_path)
    worksheet = workbook.active
    output_columns = ensure_output_columns(worksheet)
    rows, _ = build_row_inputs(worksheet)
    enrichment_index = load_enrichment_index(enrichment_path)

    results: dict[int, RowResult] = {}
    completed = 0
    with ThreadPoolExecutor(max_workers=max(1, workers)) as executor:
        future_map = {
            executor.submit(process_row, row, enrichment_index, fetch_online_contact): row
            for row in rows
        }
        for future in as_completed(future_map):
            row = future_map[future]
            try:
                result = future.result()
            except Exception:
                result = RowResult(
                    row_number=row.row_number,
                    admission_language_value=best_admission_language_value(row, enrichment_index),
                    contact_value=build_contact_value(
                        row,
                        {"emails": [], "phones": [], "contact_page": ""},
                    ),
                    contact_website_value=build_contact_website_value(row),
                    contact_email_value="",
                    contact_phone_value="",
                    contact_phone_standardized_value="",
                    contact_page_value="",
                    contact_address_value=build_contact_address_value(row),
                    key_contacts_value=build_key_contacts_value(row),
                    contact_email_count=0,
                    contact_phone_count=0,
                    used_contact_page=False,
                )
            results[result.row_number] = result
            completed += 1
            if completed % 100 == 0 or completed == len(rows):
                print(f"[progress] {completed}/{len(rows)} rows processed", flush=True)

    explicit_language_count = 0
    fallback_language_count = 0
    contact_email_rows = 0
    contact_phone_rows = 0
    standardized_phone_rows = 0
    contact_page_rows = 0

    for row in rows:
        result = results[row.row_number]
        worksheet.cell(
            row=row.row_number,
            column=output_columns["Admission Language Requirement Level"],
            value=result.admission_language_value,
        )
        worksheet.cell(
            row=row.row_number,
            column=output_columns["University Contact Information"],
            value=result.contact_value,
        )
        worksheet.cell(
            row=row.row_number,
            column=output_columns["University Contact Website"],
            value=result.contact_website_value,
        )
        worksheet.cell(
            row=row.row_number,
            column=output_columns["University Contact Email"],
            value=result.contact_email_value,
        )
        worksheet.cell(
            row=row.row_number,
            column=output_columns["University Contact Phone"],
            value=result.contact_phone_value,
        )
        worksheet.cell(
            row=row.row_number,
            column=output_columns["University Contact Phone Standardized"],
            value=result.contact_phone_standardized_value,
        )
        worksheet.cell(
            row=row.row_number,
            column=output_columns["University Contact Page"],
            value=result.contact_page_value,
        )
        worksheet.cell(
            row=row.row_number,
            column=output_columns["University Contact Address"],
            value=result.contact_address_value,
        )
        worksheet.cell(
            row=row.row_number,
            column=output_columns["University Key Contacts"],
            value=result.key_contacts_value,
        )
        if "No explicit admission language threshold found" in result.admission_language_value:
            fallback_language_count += 1
        else:
            explicit_language_count += 1
        if result.contact_email_count:
            contact_email_rows += 1
        if result.contact_phone_count:
            contact_phone_rows += 1
        if result.contact_phone_standardized_value:
            standardized_phone_rows += 1
        if result.used_contact_page:
            contact_page_rows += 1

    workbook.save(workbook_path)
    return {
        "rows": len(rows),
        "explicit_or_generic_language_rows": explicit_language_count,
        "instruction_language_fallback_rows": fallback_language_count,
        "contact_email_rows": contact_email_rows,
        "contact_phone_rows": contact_phone_rows,
        "contact_phone_standardized_rows": standardized_phone_rows,
        "contact_page_rows": contact_page_rows,
    }


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description="Add admission language requirement and university contact information columns to whed_data.xlsx."
    )
    parser.add_argument("--workbook", type=Path, default=WORKBOOK_PATH)
    parser.add_argument("--enrichment", type=Path, default=ENRICHMENT_PATH)
    parser.add_argument("--workers", type=int, default=16)
    parser.add_argument(
        "--fetch-online-contact",
        action="store_true",
        help="Try to extract email/phone/contact page from official websites. Slower but can add richer contact details.",
    )
    return parser


def main() -> int:
    args = build_parser().parse_args()
    backup_path = backup_workbook(args.workbook)
    print(f"[info] Backup created: {backup_path}", flush=True)
    stats = run_enrichment(
        args.workbook,
        args.enrichment,
        args.workers,
        args.fetch_online_contact,
    )
    print(json.dumps(stats, ensure_ascii=False, indent=2), flush=True)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
