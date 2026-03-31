import argparse
import concurrent.futures
import csv
import io
import json
import re
import shutil
import subprocess
import zipfile
from dataclasses import dataclass
from difflib import SequenceMatcher
from pathlib import Path
from statistics import median
from typing import Any
from urllib.parse import quote

import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook

from add_student_costs import normalize_text, split_country_value
from whed_enrich import clean_domain, domains_related, find_scorecard_match


THE_YEARS = [2026, 2025, 2024, 2023, 2022]
ARWU_YEARS = [2025, 2024, 2023, 2022, 2021]
SCORECARD_CSV = Path(".cache/Most-Recent-Cohorts-Institution.csv")
ROR_ZIP = Path("References/Zenodo-ROR/v2.4-2026-03-12-ror-data.zip")
THE_CACHE_DIR = Path(".cache/the_rankings")
ARWU_CACHE_DIR = Path(".cache/arwu_rankings")
OPENALEX_CACHE_DIR = Path(".cache/openalex_institutions")
WORLDBANK_CACHE_DIR = Path(".cache/worldbank")
NUMBEO_CACHE_DIR = Path(".cache/numbeo")
USER_AGENT = "Mozilla/5.0 (compatible; WHED-Scrapping/1.0)"
OPENALEX_USER_AGENT = "mailto:test@example.com"

OUTPUT_COLUMNS = [
    "Global Ranking - QS",
    "Global Ranking - THE",
    "Global Ranking - ARWU",
    "Ranking Trend (5Y) - THE",
    "Ranking Trend (5Y) - ARWU",
    "Citation per Faculty",
    "Research Output Count",
    "Patent Count",
    "Industry Collaboration Score",
    "Graduate Employability Rate (%)",
    "Average Graduate Salary (USD)",
    "Dropout Rate (%)",
    "Average Time to Graduate",
    "International Student Ratio (%)",
    "International Staff Ratio (%)",
    "Number of Partner Universities",
    "Erasmus Participation Score",
    "Visa Difficulty Score",
    "English Program Ratio (%)",
    "Scholarship Availability (%)",
    "Average Scholarship Amount",
    "Work While Studying Allowed (Yes/No)",
    "Part-time Job Availability Score",
    "Cost of Living Trend (Annual Inflation %)",
    "Inflation Adjusted Cost Index",
    "Campus Size (m²)",
    "Dorm Capacity",
    "Library Size (books/digital)",
    "Lab Count",
    "Sports Facilities Score",
    "Campus Safety Score",
    "Green Campus Score",
    "Employment Rate After 6 Months (%)",
    "Top Hiring Companies",
    "Alumni Network Strength Score",
    "Startup Founded by Alumni Count",
    "Industry Placement Rate",
    "Climate Type",
    "Crime Rate Index",
    "Cultural Activity Score",
    "Nightlife Score",
    "Family Friendliness Score",
    "Digital Infrastructure Score (5G vs)",
    "Extended Metrics Coverage",
    "THE Source URL",
    "ARWU Source URL",
    "OpenAlex Source URL",
    "Salary Source URL",
    "Inflation Source URL",
    "Crime Source URL",
]

COUNTRY_ALIASES = {
    normalize_text("United States of America"): normalize_text("United States"),
    normalize_text("United Kingdom of Great Britain and Northern Ireland"): normalize_text("United Kingdom"),
    normalize_text("Russian Federation"): normalize_text("Russia"),
    normalize_text("Iran (Islamic Republic of)"): normalize_text("Iran"),
    normalize_text("Turkey"): normalize_text("Turkiye"),
    normalize_text("Republic of Korea"): normalize_text("South Korea"),
    normalize_text("Korea, Republic of"): normalize_text("South Korea"),
    normalize_text("Korea, Democratic People's Republic of"): normalize_text("North Korea"),
    normalize_text("Hong Kong SAR, China"): normalize_text("Hong Kong China"),
    normalize_text("Hong Kong"): normalize_text("Hong Kong China"),
    normalize_text("Hong Kong, China"): normalize_text("Hong Kong China"),
    normalize_text("Macau SAR, China"): normalize_text("Macau China"),
    normalize_text("Macau"): normalize_text("Macau China"),
    normalize_text("Macau, China"): normalize_text("Macau China"),
    normalize_text("Taiwan"): normalize_text("Taiwan China"),
    normalize_text("Taiwan, China"): normalize_text("Taiwan China"),
    normalize_text("Macao, China"): normalize_text("Macau China"),
    normalize_text("Venezuela, RB"): normalize_text("Venezuela"),
    normalize_text("Czechia"): normalize_text("Czech Republic"),
}


@dataclass(frozen=True)
class WorkbookInstitution:
    row_number: int
    university_name: str
    country: str
    base_country: str
    city: str
    province: str
    website: str
    annual_cost_usd: float | None


@dataclass(frozen=True)
class RorRecord:
    ror_id: str
    name: str
    aliases: tuple[str, ...]
    country: str
    city: str
    website: str
    domains: tuple[str, ...]


def backup_workbook(path: Path) -> Path:
    backup_path = path.with_name(f"{path.stem}.extended_metrics_backup{path.suffix}")
    shutil.copy2(path, backup_path)
    return backup_path


def ensure_output_columns(worksheet) -> dict[str, int]:
    header_map = {
        worksheet.cell(row=1, column=column_index).value: column_index
        for column_index in range(1, worksheet.max_column + 1)
    }
    next_column = worksheet.max_column + 1
    for header in OUTPUT_COLUMNS:
        if header not in header_map:
            worksheet.cell(row=1, column=next_column, value=header)
            header_map[header] = next_column
            next_column += 1
    return {header: header_map[header] for header in OUTPUT_COLUMNS}


def parse_float(value: Any) -> float | None:
    if value in (None, "", "NA", "NULL", "PrivacySuppressed", "PS"):
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def parse_int_like(value: Any) -> int | None:
    number = parse_float(value)
    if number is None:
        return None
    return int(round(number))


def normalize_country(value: str | None) -> str:
    normalized = normalize_text(value)
    return COUNTRY_ALIASES.get(normalized, normalized)


def parse_rank_number(rank_value: str | None) -> int | None:
    if not rank_value:
        return None
    text = str(rank_value).strip().lstrip("=")
    match = re.search(r"(\d+)", text)
    return int(match.group(1)) if match else None


def trend_delta(current_rank: str | None, old_rank: str | None) -> int | None:
    current = parse_rank_number(current_rank)
    old = parse_rank_number(old_rank)
    if current is None or old is None:
        return None
    return old - current


def load_workbook_rows(workbook_path: Path, sheet_name: str) -> tuple[Any, list[WorkbookInstitution]]:
    wb = load_workbook(workbook_path)
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    headers = [str(value) if value is not None else "" for value in rows[0]]
    header_map = {header: idx for idx, header in enumerate(headers)}
    annual_cost_idx = header_map.get("Estimated Annual Student Cost (Shared Housing, USD)")
    institutions: list[WorkbookInstitution] = []

    for row_number, row in enumerate(rows[1:], start=2):
        country = str(row[header_map["Country"]] or "").strip()
        base_country, _ = split_country_value(country)
        annual_cost = parse_float(row[annual_cost_idx]) if annual_cost_idx is not None else None
        institutions.append(
            WorkbookInstitution(
                row_number=row_number,
                university_name=str(row[header_map["University Name"]] or "").strip(),
                country=country,
                base_country=base_country,
                city=str(row[header_map["City"]] or "").strip(),
                province=str(row[header_map["Province"]] or "").strip(),
                website=str(row[header_map["Website"]] or "").strip(),
                annual_cost_usd=annual_cost,
            )
        )
    return wb, institutions


def fetch_json(url: str, cache_path: Path, headers: dict[str, str] | None = None) -> Any:
    cache_path.parent.mkdir(parents=True, exist_ok=True)
    if cache_path.exists():
        return json.loads(cache_path.read_text(encoding="utf-8"))
    last_error: Exception | None = None
    for _ in range(3):
        try:
            response = requests.get(url, timeout=60, headers=headers or {"User-Agent": USER_AGENT})
            response.raise_for_status()
            cache_path.write_text(response.text, encoding="utf-8")
            return response.json()
        except Exception as exc:  # pragma: no cover - network retry path
            last_error = exc
    raise last_error or RuntimeError(f"Failed to fetch JSON: {url}")


def fetch_text(url: str, cache_path: Path, headers: dict[str, str] | None = None) -> str:
    cache_path.parent.mkdir(parents=True, exist_ok=True)
    if cache_path.exists():
        return cache_path.read_text(encoding="utf-8", errors="replace")
    last_error: Exception | None = None
    for _ in range(3):
        try:
            response = requests.get(url, timeout=(20, 120), headers=headers or {"User-Agent": USER_AGENT})
            response.raise_for_status()
            cache_path.write_text(response.text, encoding="utf-8")
            return response.text
        except Exception as exc:  # pragma: no cover - network retry path
            last_error = exc
    raise last_error or RuntimeError(f"Failed to fetch text: {url}")


def load_the_rankings(year: int) -> list[dict[str, Any]]:
    url = f"https://www.timeshighereducation.com/json/ranking_tables/world_university_rankings/{year}"
    return fetch_json(url, THE_CACHE_DIR / f"the_{year}.json").get("data", [])


def load_arwu_rankings(year: int) -> list[dict[str, Any]]:
    url = f"https://www.shanghairanking.com/rankings/arwu/{year}.html"
    html = fetch_text(url, ARWU_CACHE_DIR / f"arwu_{year}.html")
    parsed_cache = ARWU_CACHE_DIR / f"arwu_{year}_payload.json"
    if parsed_cache.exists():
        try:
            payload = json.loads(parsed_cache.read_text(encoding="utf-8"))
        except UnicodeDecodeError:
            payload = json.loads(parsed_cache.read_text(encoding="utf-16"))
    else:
        payload_path_match = re.search(r'href="(?P<path>/_nuxt/static/\d+/rankings/arwu/' + str(year) + r'/payload\.js)"', html)
        if payload_path_match is None:
            payload = None
        else:
            payload_url = "https://www.shanghairanking.com" + payload_path_match.group("path")
            payload_js = fetch_text(payload_url, ARWU_CACHE_DIR / f"arwu_{year}_payload.js")
            node_script = (
                "const fs=require('fs');"
                "const input=fs.readFileSync(0,'utf8');"
                "let captured=null;"
                "global.__NUXT_JSONP__=(path,data)=>{captured=data;};"
                "eval(input);"
                "process.stdout.write(JSON.stringify(captured));"
            )
            completed = subprocess.run(
                ["node", "-e", node_script],
                input=payload_js,
                text=True,
                encoding="utf-8",
                capture_output=True,
                check=True,
            )
            payload = json.loads(completed.stdout)
            parsed_cache.write_text(json.dumps(payload), encoding="utf-8")

    if payload and payload.get("data"):
        return [
            {
                "rank": item.get("ranking"),
                "name": item.get("univNameEn"),
                "country": item.get("region"),
                "source_url": f"https://www.shanghairanking.com/institution/{item.get('univUp')}",
            }
            for item in payload["data"][0].get("univList", [])
            if item.get("ranking") and item.get("univNameEn")
        ]

    soup = BeautifulSoup(html, "html.parser")
    results: list[dict[str, Any]] = []
    for row in soup.select("tbody tr"):
        cells = row.find_all("td")
        link = row.select_one('a[href*="/institution/"]')
        country_node = row.select_one(".location")
        if len(cells) < 2 or link is None:
            continue
        rank = cells[0].get_text(" ", strip=True)
        if not rank or not re.search(r"\d", rank):
            continue
        href = link.get("href", "").strip()
        results.append(
            {
                "rank": rank,
                "name": link.get_text(" ", strip=True),
                "country": country_node.get_text(" ", strip=True) if country_node else "",
                "source_url": f"https://www.shanghairanking.com{href}" if href.startswith("/") else href,
            }
        )
    return results


def build_rank_index(records: list[dict[str, Any]]) -> dict[str, list[dict[str, Any]]]:
    bucketed: dict[str, list[dict[str, Any]]] = {}
    for record in records:
        country_key = normalize_country(record.get("location") or record.get("country") or "")
        bucketed.setdefault(country_key, []).append(record)
    return bucketed


def candidate_names_for_match(record: dict[str, Any], alias_field: str | None = None) -> list[str]:
    values = [str(record.get("name", "")).strip()]
    if alias_field:
        alias_value = str(record.get(alias_field, "") or "").strip()
        if alias_value:
            values.extend(part.strip() for part in re.split(r"[|;,/]", alias_value) if part.strip())
    unique: list[str] = []
    seen: set[str] = set()
    for value in values:
        normalized = normalize_text(value)
        if normalized and normalized not in seen:
            seen.add(normalized)
            unique.append(value)
    return unique


def best_name_ratio(source_names: list[str], candidate_names: list[str]) -> float:
    best = 0.0
    for left in source_names:
        left_norm = normalize_text(left)
        for right in candidate_names:
            ratio = SequenceMatcher(None, left_norm, normalize_text(right)).ratio()
            if ratio > best:
                best = ratio
    return best


def find_best_rank_match(
    names: list[str],
    country: str,
    bucketed_records: dict[str, list[dict[str, Any]]],
    alias_field: str | None = None,
    threshold: float = 0.93,
) -> dict[str, Any] | None:
    candidates = bucketed_records.get(normalize_country(country), [])
    if not candidates:
        return None
    exact_names = {normalize_text(name) for name in names if normalize_text(name)}
    for candidate in candidates:
        candidate_names = candidate_names_for_match(candidate, alias_field)
        if any(normalize_text(name) in exact_names for name in candidate_names):
            return candidate

    best_candidate = None
    best_ratio = 0.0
    for candidate in candidates:
        ratio = best_name_ratio(names, candidate_names_for_match(candidate, alias_field))
        if ratio > best_ratio:
            best_ratio = ratio
            best_candidate = candidate
    return best_candidate if best_candidate is not None and best_ratio >= threshold else None


def load_ror_records() -> tuple[dict[str, list[RorRecord]], dict[str, list[RorRecord]]]:
    with zipfile.ZipFile(ROR_ZIP) as archive:
        csv_name = next(name for name in archive.namelist() if name.endswith(".csv"))
        with archive.open(csv_name) as handle:
            reader = csv.DictReader(io.TextIOWrapper(handle, encoding="utf-8", errors="replace", newline=""))
            domain_index: dict[str, list[RorRecord]] = {}
            country_index: dict[str, list[RorRecord]] = {}
            for row in reader:
                if "education" not in str(row.get("types", "")):
                    continue
                website = str(row.get("links.type.website", "") or "").strip()
                aliases: list[str] = []
                for raw_aliases in [row.get("names.types.alias", ""), row.get("names.types.label", ""), row.get("names.types.acronym", "")]:
                    for raw_alias in str(raw_aliases or "").split(";"):
                        alias = re.sub(r"^[a-z]{2}:\s*", "", raw_alias.strip(), flags=re.IGNORECASE)
                        if alias:
                            aliases.append(alias)
                domains: list[str] = []
                for raw_domain in [row.get("domains", ""), website]:
                    clean = clean_domain(str(raw_domain or ""))
                    if clean:
                        domains.append(clean)
                record = RorRecord(
                    ror_id=str(row.get("id", "") or "").strip(),
                    name=str(row.get("names.types.ror_display", "") or "").strip(),
                    aliases=tuple(dict.fromkeys(aliases)),
                    country=str(row.get("locations.geonames_details.country_name", "") or "").strip(),
                    city=str(row.get("locations.geonames_details.name", "") or "").strip(),
                    website=website,
                    domains=tuple(dict.fromkeys(domains)),
                )
                country_index.setdefault(normalize_country(record.country), []).append(record)
                for domain in record.domains:
                    domain_index.setdefault(domain, []).append(record)
    return domain_index, country_index


def find_ror_match(institution: WorkbookInstitution, domain_index: dict[str, list[RorRecord]], country_index: dict[str, list[RorRecord]]) -> RorRecord | None:
    record_domain = clean_domain(institution.website)
    candidates = domain_index.get(record_domain, []) if record_domain else []
    if not candidates:
        candidates = country_index.get(normalize_country(institution.base_country), [])
    if not candidates:
        return None

    best_candidate = None
    best_score = 0.0
    city_exact = False
    for candidate in candidates:
        ratio = best_name_ratio([institution.university_name], [candidate.name, *candidate.aliases])
        score = ratio
        domain_match = record_domain and any(domains_related(record_domain, domain) for domain in candidate.domains)
        current_city_exact = bool(
            institution.city and candidate.city and normalize_text(institution.city) == normalize_text(candidate.city)
        )
        if domain_match:
            score += 1.5
        if current_city_exact:
            score += 0.3
        if score > best_score:
            best_score = score
            best_candidate = candidate
            city_exact = current_city_exact

    if best_candidate is None:
        return None
    base_ratio = best_score
    if record_domain and any(domains_related(record_domain, domain) for domain in best_candidate.domains):
        base_ratio -= 1.5
    if city_exact:
        base_ratio -= 0.3
    if record_domain and any(domains_related(record_domain, domain) for domain in best_candidate.domains) and base_ratio >= 0.72:
        return best_candidate
    if base_ratio >= 0.97:
        return best_candidate
    if base_ratio >= 0.90 and city_exact:
        return best_candidate
    return None


def fetch_openalex_for_ror(ror_id: str) -> dict[str, Any]:
    cache_path = OPENALEX_CACHE_DIR / (ror_id.rstrip("/").split("/")[-1] + ".json")
    if cache_path.exists():
        return json.loads(cache_path.read_text(encoding="utf-8"))
    url = "https://api.openalex.org/institutions?filter=ror:" + quote(ror_id, safe=":/")
    response = requests.get(url, timeout=60, headers={"User-Agent": OPENALEX_USER_AGENT})
    response.raise_for_status()
    payload = response.json()
    result = payload.get("results", [{}])[0] if payload.get("results") else {}
    cache_path.parent.mkdir(parents=True, exist_ok=True)
    cache_path.write_text(json.dumps(result), encoding="utf-8")
    return result


def load_scorecard_salary_records() -> list[dict[str, Any]]:
    records: list[dict[str, Any]] = []
    with SCORECARD_CSV.open("r", encoding="utf-8", errors="replace", newline="") as handle:
        reader = csv.DictReader(handle)
        for row in reader:
            records.append(
                {
                    "id": row.get("UNITID", ""),
                    "school.name": row.get("INSTNM", ""),
                    "school.state": row.get("STABBR", ""),
                    "school.city": row.get("CITY", ""),
                    "school.school_url": row.get("INSTURL", ""),
                    "earnings_1yr": row.get("MD_EARN_WNE_1YR", ""),
                    "dropout_6yr": row.get("WDRAW_ORIG_YR6_RT", ""),
                }
            )
    return records


def build_scorecard_record(institution: WorkbookInstitution):
    class _Record:
        university_name = institution.university_name
        country = institution.country
        city = institution.city
        website = institution.website

    return _Record


def load_worldbank_country_map() -> dict[str, dict[str, str]]:
    payload = fetch_json("https://api.worldbank.org/v2/country?format=json&per_page=400", WORLDBANK_CACHE_DIR / "countries.json")
    mapping: dict[str, dict[str, str]] = {}
    for item in payload[1]:
        iso3 = str(item.get("id", "") or "").strip()
        name = str(item.get("name", "") or "").strip()
        if not iso3 or not name:
            continue
        aliases = {
            name,
            "United States of America" if iso3 == "USA" else "",
            "Türkiye" if iso3 == "TUR" else "",
            "Russia" if iso3 == "RUS" else "",
            "Iran" if iso3 == "IRN" else "",
            "South Korea" if iso3 == "KOR" else "",
            "Hong Kong, China" if iso3 == "HKG" else "",
            "Macau, China" if iso3 == "MAC" else "",
            "Taiwan, China" if iso3 == "TWN" else "",
        }
        for alias in aliases:
            if alias:
                mapping[normalize_country(alias)] = {"iso3": iso3, "name": name}
    return mapping


def fetch_worldbank_inflation(iso3: str) -> dict[str, Any]:
    payload = fetch_json(
        f"https://api.worldbank.org/v2/country/{iso3}/indicator/FP.CPI.TOTL.ZG?format=json&per_page=10",
        WORLDBANK_CACHE_DIR / f"{iso3}_inflation.json",
    )
    latest = next((row for row in payload[1] if row.get("value") is not None), None)
    return {
        "value": float(latest["value"]) if latest else None,
        "source_url": f"https://api.worldbank.org/v2/country/{iso3}/indicator/FP.CPI.TOTL.ZG?format=json&per_page=10",
    }


def load_numbeo_crime_index() -> dict[tuple[str, str], dict[str, Any]]:
    html = fetch_text("https://www.numbeo.com/crime/rankings_current.jsp", NUMBEO_CACHE_DIR / "crime_rankings_current.html")
    soup = BeautifulSoup(html, "html.parser")
    table = soup.find("table", {"id": "t2"})
    mapping: dict[tuple[str, str], dict[str, Any]] = {}
    if table is None:
        return mapping
    for row in table.find_all("tr")[1:]:
        cells = row.find_all("td")
        link = row.find("a", href=True)
        if len(cells) < 4 or link is None:
            continue
        city_country = cells[1].get_text(" ", strip=True)
        if "," not in city_country:
            continue
        city, country = [part.strip() for part in city_country.rsplit(",", 1)]
        crime_index = parse_float(cells[2].get_text(" ", strip=True))
        if crime_index is None:
            continue
        mapping[(normalize_text(city), normalize_country(country))] = {
            "crime_index": round(crime_index, 2),
            "source_url": link.get("href", "").strip(),
        }
    return mapping


def hyperlink_columns() -> set[str]:
    return {"THE Source URL", "ARWU Source URL", "OpenAlex Source URL", "Salary Source URL", "Inflation Source URL", "Crime Source URL"}


def build_name_candidates(institution: WorkbookInstitution, ror_match: RorRecord | None) -> list[str]:
    candidates = [institution.university_name]
    if ror_match is not None:
        candidates.append(ror_match.name)
        candidates.extend(ror_match.aliases[:3])
    ordered: list[str] = []
    seen: set[str] = set()
    for candidate in candidates:
        key = normalize_text(candidate)
        if key and key not in seen:
            seen.add(key)
            ordered.append(candidate)
    return ordered


def enrich_workbook(workbook_path: Path, sheet_name: str) -> dict[str, Any]:
    wb, institutions = load_workbook_rows(workbook_path, sheet_name)
    ws = wb[sheet_name]
    column_map = ensure_output_columns(ws)

    the_by_year = {year: load_the_rankings(year) for year in THE_YEARS}
    arwu_by_year = {year: load_arwu_rankings(year) for year in ARWU_YEARS}
    the_current_bucket = build_rank_index(the_by_year[THE_YEARS[0]])
    arwu_current_bucket = build_rank_index(arwu_by_year[ARWU_YEARS[0]])
    the_previous_by_iid = {year: {str(item.get("iid", "")): item for item in records if item.get("iid")} for year, records in the_by_year.items()}
    arwu_previous_by_url = {year: {str(item.get("source_url", "")): item for item in records if item.get("source_url")} for year, records in arwu_by_year.items()}

    ror_domain_index, ror_country_index = load_ror_records()
    ror_matches = {institution.row_number: find_ror_match(institution, ror_domain_index, ror_country_index) for institution in institutions}

    unique_ror_ids = sorted({match.ror_id for match in ror_matches.values() if match is not None})
    openalex_by_ror: dict[str, dict[str, Any]] = {}
    with concurrent.futures.ThreadPoolExecutor(max_workers=8) as executor:
        futures = {executor.submit(fetch_openalex_for_ror, ror_id): ror_id for ror_id in unique_ror_ids}
        for future in concurrent.futures.as_completed(futures):
            ror_id = futures[future]
            try:
                openalex_by_ror[ror_id] = future.result()
            except Exception:
                openalex_by_ror[ror_id] = {}

    scorecard_records = load_scorecard_salary_records()
    country_map = load_worldbank_country_map()
    unique_iso3 = sorted({country_map[normalize_country(i.base_country)]["iso3"] for i in institutions if normalize_country(i.base_country) in country_map})
    inflation_by_iso3: dict[str, dict[str, Any]] = {}
    with concurrent.futures.ThreadPoolExecutor(max_workers=8) as executor:
        futures = {executor.submit(fetch_worldbank_inflation, iso3): iso3 for iso3 in unique_iso3}
        for future in concurrent.futures.as_completed(futures):
            iso3 = futures[future]
            try:
                inflation_by_iso3[iso3] = future.result()
            except Exception:
                inflation_by_iso3[iso3] = {"value": None, "source_url": None}

    numbeo_crime = load_numbeo_crime_index()
    row_results: dict[int, dict[str, Any]] = {}
    adjusted_values: list[float] = []
    filled_counts = {"the_rank": 0, "arwu_rank": 0, "research_output": 0, "industry_score": 0, "salary": 0, "dropout": 0, "intl_students": 0, "inflation": 0, "crime": 0}

    for institution in institutions:
        ror_match = ror_matches[institution.row_number]
        name_candidates = build_name_candidates(institution, ror_match)
        the_match = find_best_rank_match(name_candidates, institution.base_country, the_current_bucket, alias_field="aliases", threshold=0.92)
        arwu_match = find_best_rank_match(name_candidates, institution.base_country, arwu_current_bucket, threshold=0.93)
        openalex_match = openalex_by_ror.get(ror_match.ror_id, {}) if ror_match is not None else {}

        if institution.country.startswith("United States of America"):
            scorecard_match = find_scorecard_match(build_scorecard_record(institution), scorecard_records)  # type: ignore[arg-type]
        else:
            scorecard_match = None

        inflation_value = None
        inflation_source_url = None
        country_key = normalize_country(institution.base_country)
        if country_key in country_map:
            iso3 = country_map[country_key]["iso3"]
            inflation_payload = inflation_by_iso3.get(iso3, {})
            inflation_value = parse_float(inflation_payload.get("value"))
            inflation_source_url = inflation_payload.get("source_url")

        crime_match = numbeo_crime.get((normalize_text(institution.city), normalize_country(institution.base_country)))
        the_rank = str(the_match.get("rank", "")).strip() if the_match else None
        the_iid = str(the_match.get("iid", "")).strip() if the_match else ""
        the_old = the_previous_by_iid[THE_YEARS[-1]].get(the_iid) if the_iid else None
        the_url = f"https://www.timeshighereducation.com{str(the_match.get('url', '')).strip()}" if the_match and str(the_match.get("url", "")).startswith("/") else None
        arwu_rank = str(arwu_match.get("rank", "")).strip() if arwu_match else None
        arwu_url = str(arwu_match.get("source_url", "")).strip() if arwu_match else None
        arwu_old = arwu_previous_by_url[ARWU_YEARS[-1]].get(arwu_url) if arwu_url else None
        research_output = parse_int_like(openalex_match.get("works_count")) if openalex_match else None
        openalex_url = str(openalex_match.get("id", "")).strip() if openalex_match else None
        industry_score = parse_float(the_match.get("scores_industry_income")) if the_match else None
        intl_student_ratio = parse_float(str(the_match.get("stats_pc_intl_students", "")).replace("%", "")) if the_match and the_match.get("stats_pc_intl_students") else None

        average_salary = None
        dropout_rate = None
        salary_source_url = None
        if scorecard_match is not None:
            average_salary = parse_int_like(scorecard_match.get("earnings_1yr"))
            dropout = parse_float(scorecard_match.get("dropout_6yr"))
            dropout_rate = round(dropout * 100.0, 2) if dropout is not None else None
            unitid = str(scorecard_match.get("id", "")).strip()
            salary_source_url = f"https://collegescorecard.ed.gov/school/?{unitid}" if unitid else None

        adjusted_raw = None
        if institution.annual_cost_usd is not None and inflation_value is not None:
            adjusted_raw = institution.annual_cost_usd * (1.0 + inflation_value / 100.0)
            adjusted_values.append(adjusted_raw)

        coverage = []
        if the_rank or industry_score is not None or intl_student_ratio is not None:
            coverage.append("THE")
        if arwu_rank:
            coverage.append("ARWU")
        if research_output is not None:
            coverage.append("OpenAlex")
        if average_salary is not None or dropout_rate is not None:
            coverage.append("College Scorecard")
        if inflation_value is not None:
            coverage.append("World Bank")
        if crime_match is not None:
            coverage.append("Numbeo")

        row_results[institution.row_number] = {
            "values": {
                "Global Ranking - QS": None,
                "Global Ranking - THE": the_rank,
                "Global Ranking - ARWU": arwu_rank,
                "Ranking Trend (5Y) - THE": trend_delta(the_rank, str(the_old.get("rank", "")).strip() if the_old else None),
                "Ranking Trend (5Y) - ARWU": trend_delta(arwu_rank, str(arwu_old.get("rank", "")).strip() if arwu_old else None),
                "Citation per Faculty": None,
                "Research Output Count": research_output,
                "Patent Count": None,
                "Industry Collaboration Score": round(industry_score, 2) if industry_score is not None else None,
                "Graduate Employability Rate (%)": None,
                "Average Graduate Salary (USD)": average_salary,
                "Dropout Rate (%)": dropout_rate,
                "Average Time to Graduate": None,
                "International Student Ratio (%)": round(intl_student_ratio, 2) if intl_student_ratio is not None else None,
                "International Staff Ratio (%)": None,
                "Number of Partner Universities": None,
                "Erasmus Participation Score": None,
                "Visa Difficulty Score": None,
                "English Program Ratio (%)": None,
                "Scholarship Availability (%)": None,
                "Average Scholarship Amount": None,
                "Work While Studying Allowed (Yes/No)": None,
                "Part-time Job Availability Score": None,
                "Cost of Living Trend (Annual Inflation %)": round(inflation_value, 2) if inflation_value is not None else None,
                "Inflation Adjusted Cost Index": None,
                "Campus Size (m²)": None,
                "Dorm Capacity": None,
                "Library Size (books/digital)": None,
                "Lab Count": None,
                "Sports Facilities Score": None,
                "Campus Safety Score": None,
                "Green Campus Score": None,
                "Employment Rate After 6 Months (%)": None,
                "Top Hiring Companies": None,
                "Alumni Network Strength Score": None,
                "Startup Founded by Alumni Count": None,
                "Industry Placement Rate": None,
                "Climate Type": None,
                "Crime Rate Index": crime_match.get("crime_index") if crime_match else None,
                "Cultural Activity Score": None,
                "Nightlife Score": None,
                "Family Friendliness Score": None,
                "Digital Infrastructure Score (5G vs)": None,
                "Extended Metrics Coverage": "; ".join(coverage) if coverage else None,
                "THE Source URL": the_url,
                "ARWU Source URL": arwu_url,
                "OpenAlex Source URL": openalex_url,
                "Salary Source URL": salary_source_url,
                "Inflation Source URL": inflation_source_url,
                "Crime Source URL": crime_match.get("source_url") if crime_match else None,
            },
            "adjusted_raw": adjusted_raw,
        }

        if the_rank:
            filled_counts["the_rank"] += 1
        if arwu_rank:
            filled_counts["arwu_rank"] += 1
        if research_output is not None:
            filled_counts["research_output"] += 1
        if industry_score is not None:
            filled_counts["industry_score"] += 1
        if average_salary is not None:
            filled_counts["salary"] += 1
        if dropout_rate is not None:
            filled_counts["dropout"] += 1
        if intl_student_ratio is not None:
            filled_counts["intl_students"] += 1
        if inflation_value is not None:
            filled_counts["inflation"] += 1
        if crime_match is not None:
            filled_counts["crime"] += 1

    adjusted_median = median(adjusted_values) if adjusted_values else None
    for payload in row_results.values():
        if payload["adjusted_raw"] is not None and adjusted_median:
            payload["values"]["Inflation Adjusted Cost Index"] = round((payload["adjusted_raw"] / adjusted_median) * 100.0, 2)

    for row_number, payload in row_results.items():
        for header, value in payload["values"].items():
            cell = ws.cell(row=row_number, column=column_map[header], value=value)
            if header in hyperlink_columns() and value:
                cell.hyperlink = value
                cell.style = "Hyperlink"

    backup_path = backup_workbook(workbook_path)
    wb.save(workbook_path)
    return {"rows": len(institutions), "filled_counts": filled_counts, "inflation_adjusted_rows": len(adjusted_values), "backup_path": str(backup_path)}


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Append extended ranking, research, economic, and city metrics where reliable public sources are available.")
    parser.add_argument("--workbook", default="whed_data.xlsx")
    parser.add_argument("--sheet", default="Institutions")
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    print(json.dumps(enrich_workbook(Path(args.workbook), args.sheet), ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
