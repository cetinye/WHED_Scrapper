import argparse
import concurrent.futures
import csv
import json
import math
import re
import shutil
import time
import zipfile
from dataclasses import dataclass
from pathlib import Path
from typing import Any

import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook

from add_extended_metrics import (
    OPENALEX_CACHE_DIR,
    ROR_ZIP,
    USER_AGENT,
    fetch_openalex_for_ror,
    find_ror_match,
    normalize_country,
    parse_float,
)
from add_student_costs import normalize_text, split_country_value
from whed_enrich import clean_domain, domains_related, find_scorecard_match


WIKIDATA_CACHE_DIR = Path(".cache/wikidata")
EDURANK_CACHE_DIR = Path(".cache/edurank")
MAPRESSO_CACHE_DIR = Path(".cache/mapresso")
NUMBEO_CACHE_DIR = Path(".cache/numbeo")
SCORECARD_CSV = Path(".cache/Most-Recent-Cohorts-Institution.csv")

QS_ITEM_QID = "Q1790510"
EDURANK_WORLD_UNIVERSITIES = 14131

NEW_OUTPUT_COLUMNS = [
    "QS Source URL",
    "EduRank Source URL",
    "Wikidata Source URL",
    "Numbeo QoL Source URL",
    "Climate Source URL",
    "Unofficial Metrics Notes",
]

LANGUAGE_SPLIT_RE = re.compile(r"[;,/]|(?:\band\b)", re.IGNORECASE)
ERASMUS_RE = re.compile(
    r"(?i)\bErasmus(?:\+|\s+(?:Mundus|Coordinator|Officer|Office|Programme|Program|Mobility|Department|Institutional))"
)
LAB_RE = re.compile(r"(?im)^\s*(?:Laboratory|Lab(?:oratory)?)\s*:")
SPORTS_FACILITY_RE = re.compile(
    r"(?im)^\s*(?:Sports (?:Centre|Center|Complex)|Gymnasium|Stadium|Athletic Center|Athletics Center|Fitness (?:Centre|Center)|Recreation Center)\s*:"
)
SPORTS_UNIT_RE = re.compile(
    r"(?im)^\s*(?:School|Faculty|College|Department|Centre|Center|Campus|Institute|Unit)\s*:\s*.*(?:sport|athletic|fitness|physical education|recreation)"
)
PARTNER_COUNT_PATTERNS = [
    re.compile(r"(?i)\b(\d+)\s+partner institutions?\b"),
    re.compile(r"(?i)\b(\d+)\s+partner universities?\b"),
    re.compile(r"(?i)\bwith\s+(\d+)\s+partner institutions?\b"),
    re.compile(r"(?i)\bwith\s+(\d+)\s+partner universities?\b"),
    re.compile(r"(?i)\b(one|two|three|four|five|six|seven|eight|nine|ten)\s+partner institutions?\b"),
    re.compile(r"(?i)\b(one|two|three|four|five|six|seven|eight|nine|ten)\s+partner universities?\b"),
]
NUMBER_WORDS = {
    "one": 1,
    "two": 2,
    "three": 3,
    "four": 4,
    "five": 5,
    "six": 6,
    "seven": 7,
    "eight": 8,
    "nine": 9,
    "ten": 10,
}
AIR_QUALITY_SCORES = {
    "excellent": 100.0,
    "good": 85.0,
    "moderate": 65.0,
    "fair": 55.0,
    "poor": 35.0,
    "unhealthy": 20.0,
    "very unhealthy": 10.0,
    "hazardous": 0.0,
}
AREA_UNIT_MULTIPLIERS = {
    "square metre": 1.0,
    "square meter": 1.0,
    "hectare": 10000.0,
    "acre": 4046.8564224,
    "square kilometre": 1_000_000.0,
    "square kilometer": 1_000_000.0,
}


@dataclass(frozen=True)
class InstitutionRow:
    row_number: int
    university_name: str
    country: str
    base_country: str
    city: str
    province: str
    website: str
    permanent_url: str
    language_text: str
    raw_text: str
    staff_full_time_total: float | None
    annual_cost_usd: float | None
    monthly_salary_usd: float | None
    cinema_usd: float | None
    fast_food_usd: float | None
    cappuccino_usd: float | None
    internet_plan_usd: float | None
    city_universities_count: int | None
    city_known_students: int | None
    quality_of_life_index: float | None
    air_quality_text: str
    pm25: float | None
    student_friendliness_source_url: str
    affordability_score: float | None
    daily_life_score: float | None
    mobility_score: float | None
    environment_score: float | None
    academic_ecosystem_score: float | None
    crime_index: float | None
    existing_coverage: str
    college_scorecard_unitid: str


@dataclass(frozen=True)
class RorGeoRecord:
    ror_id: str
    name: str
    aliases: tuple[str, ...]
    country: str
    city: str
    website: str
    domains: tuple[str, ...]
    latitude: float | None
    longitude: float | None


class _ScorecardInstitution:
    def __init__(self, institution: InstitutionRow):
        self.university_name = institution.university_name
        self.country = institution.country
        self.city = institution.city
        self.website = institution.website


def clamp(value: float | None, lower: float = 0.0, upper: float = 100.0) -> float | None:
    if value is None:
        return None
    return max(lower, min(upper, value))


def average_available(*values: float | None) -> float | None:
    numeric = [float(value) for value in values if value is not None]
    if not numeric:
        return None
    return round(sum(numeric) / len(numeric), 2)


def parse_int_like(value: Any) -> int | None:
    number = parse_float(value)
    if number is None:
        return None
    return int(round(number))


def normalize_row_text(text: str) -> str:
    return text.replace("\r\n", "\n").replace(" | ", "\n")


def backup_workbook(path: Path) -> Path:
    backup_path = path.with_name(f"{path.stem}.unofficial_remaining_backup{path.suffix}")
    shutil.copy2(path, backup_path)
    return backup_path


def ensure_columns(worksheet) -> dict[str, int]:
    header_map = {
        worksheet.cell(row=1, column=column_index).value: column_index
        for column_index in range(1, worksheet.max_column + 1)
    }
    next_column = worksheet.max_column + 1
    for header in NEW_OUTPUT_COLUMNS:
        if header not in header_map:
            worksheet.cell(row=1, column=next_column, value=header)
            header_map[header] = next_column
            next_column += 1
    return header_map


def hyperlink_columns() -> set[str]:
    return {
        "QS Source URL",
        "EduRank Source URL",
        "Wikidata Source URL",
        "Numbeo QoL Source URL",
        "Climate Source URL",
    }


def fetch_text(url: str, cache_path: Path) -> str:
    cache_path.parent.mkdir(parents=True, exist_ok=True)
    if cache_path.exists():
        return cache_path.read_text(encoding="utf-8", errors="replace")
    last_error: Exception | None = None
    for _ in range(3):
        try:
            response = requests.get(url, timeout=(20, 120), headers={"User-Agent": USER_AGENT})
            response.raise_for_status()
            cache_path.write_text(response.text, encoding="utf-8")
            return response.text
        except Exception as exc:
            last_error = exc
            time.sleep(1.0)
    raise last_error or RuntimeError(f"Failed to fetch text: {url}")


def fetch_json(url: str, cache_path: Path) -> Any:
    cache_path.parent.mkdir(parents=True, exist_ok=True)
    if cache_path.exists():
        return json.loads(cache_path.read_text(encoding="utf-8"))
    last_error: Exception | None = None
    for _ in range(3):
        try:
            response = requests.get(url, timeout=(20, 120), headers={"User-Agent": USER_AGENT})
            response.raise_for_status()
            cache_path.write_text(response.text, encoding="utf-8")
            return response.json()
        except Exception as exc:
            last_error = exc
            time.sleep(1.0)
    raise last_error or RuntimeError(f"Failed to fetch JSON: {url}")


def load_institutions(workbook_path: Path, sheet_name: str) -> tuple[Any, Any, dict[str, int], list[InstitutionRow]]:
    wb = load_workbook(workbook_path)
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    headers = [str(value) if value is not None else "" for value in rows[0]]
    header_map = {header: idx for idx, header in enumerate(headers)}
    institutions: list[InstitutionRow] = []

    def get_value(row: tuple[Any, ...], header: str) -> Any:
        idx = header_map.get(header)
        return row[idx] if idx is not None and idx < len(row) else None

    for row_number, row in enumerate(rows[1:], start=2):
        country = str(get_value(row, "Country") or "").strip()
        base_country, _ = split_country_value(country)
        institutions.append(
            InstitutionRow(
                row_number=row_number,
                university_name=str(get_value(row, "University Name") or "").strip(),
                country=country,
                base_country=base_country,
                city=str(get_value(row, "City") or "").strip(),
                province=str(get_value(row, "Province") or "").strip(),
                website=str(get_value(row, "Website") or "").strip(),
                permanent_url=str(get_value(row, "Permanent URL") or "").strip(),
                language_text=str(get_value(row, "Language(s)") or "").strip(),
                raw_text=str(get_value(row, "Raw Text") or ""),
                staff_full_time_total=parse_float(get_value(row, "Staff Full Time Total")),
                annual_cost_usd=parse_float(get_value(row, "Estimated Annual Student Cost (Shared Housing, USD)")),
                monthly_salary_usd=parse_float(get_value(row, "Livingcost Monthly Salary After Tax (USD)")),
                cinema_usd=parse_float(get_value(row, "Livingcost Cinema Ticket (USD)")),
                fast_food_usd=parse_float(get_value(row, "Livingcost Fast Food Meal (USD)")),
                cappuccino_usd=parse_float(get_value(row, "Livingcost Cappuccino (USD)")),
                internet_plan_usd=parse_float(get_value(row, "Livingcost Internet Plan (USD)")),
                city_universities_count=parse_int_like(get_value(row, "City Universities Count")),
                city_known_students=parse_int_like(get_value(row, "City Known Students")),
                quality_of_life_index=parse_float(get_value(row, "Livingcost Quality of Life Index")),
                air_quality_text=str(get_value(row, "Livingcost Air Quality") or "").strip(),
                pm25=parse_float(get_value(row, "Livingcost PM2.5 (ug/m3)")),
                student_friendliness_source_url=str(get_value(row, "Student Friendliness Source URL") or "").strip(),
                affordability_score=parse_float(get_value(row, "Student Friendliness Affordability Score")),
                daily_life_score=parse_float(get_value(row, "Student Friendliness Daily Life Score")),
                mobility_score=parse_float(get_value(row, "Student Friendliness Mobility Score")),
                environment_score=parse_float(get_value(row, "Student Friendliness Environment Score")),
                academic_ecosystem_score=parse_float(get_value(row, "Student Friendliness Academic Ecosystem Score")),
                crime_index=parse_float(get_value(row, "Crime Rate Index")),
                existing_coverage=str(get_value(row, "Extended Metrics Coverage") or "").strip(),
                college_scorecard_unitid=str(get_value(row, "College Scorecard UNITID") or "").strip(),
            )
        )
    return wb, ws, header_map, institutions


def load_ror_records_with_geo() -> tuple[dict[str, list[RorGeoRecord]], dict[str, list[RorGeoRecord]]]:
    with zipfile.ZipFile(ROR_ZIP) as archive:
        csv_name = next(name for name in archive.namelist() if name.endswith(".csv"))
        with archive.open(csv_name) as handle:
            reader = csv.DictReader(handle.read().decode("utf-8", errors="replace").splitlines())
            domain_index: dict[str, list[RorGeoRecord]] = {}
            country_index: dict[str, list[RorGeoRecord]] = {}
            for row in reader:
                if "education" not in str(row.get("types", "")):
                    continue
                website = str(row.get("links.type.website", "") or "").strip()
                aliases: list[str] = []
                for raw_aliases in [row.get("names.types.alias", ""), row.get("names.types.label", ""), row.get("names.types.acronym", "")]:
                    for raw_alias in str(raw_aliases or "").split(";"):
                        alias = re.sub(r"^[a-z]{2}:\s*", "", raw_alias.strip(), flags=re.IGNORECASE)
                        if alias:
                            aliases.append(alias)
                domains: list[str] = []
                for raw_domain in [row.get("domains", ""), website]:
                    clean = clean_domain(str(raw_domain or ""))
                    if clean:
                        domains.append(clean)
                record = RorGeoRecord(
                    ror_id=str(row.get("id", "") or "").strip(),
                    name=str(row.get("names.types.ror_display", "") or "").strip(),
                    aliases=tuple(dict.fromkeys(aliases)),
                    country=str(row.get("locations.geonames_details.country_name", "") or "").strip(),
                    city=str(row.get("locations.geonames_details.name", "") or "").strip(),
                    website=website,
                    domains=tuple(dict.fromkeys(domains)),
                    latitude=parse_float(row.get("locations.geonames_details.lat")),
                    longitude=parse_float(row.get("locations.geonames_details.lng")),
                )
                country_index.setdefault(normalize_country(record.country), []).append(record)
                for domain in record.domains:
                    domain_index.setdefault(domain, []).append(record)
    return domain_index, country_index


def load_openalex_for_matches(ror_matches: dict[int, RorGeoRecord | None]) -> dict[str, dict[str, Any]]:
    unique_ror_ids = sorted({match.ror_id for match in ror_matches.values() if match is not None})
    openalex_by_ror: dict[str, dict[str, Any]] = {}

    def load_one(ror_id: str) -> tuple[str, dict[str, Any]]:
        cache_path = OPENALEX_CACHE_DIR / (ror_id.rstrip("/").split("/")[-1] + ".json")
        if cache_path.exists():
            return ror_id, json.loads(cache_path.read_text(encoding="utf-8"))
        return ror_id, fetch_openalex_for_ror(ror_id)

    with concurrent.futures.ThreadPoolExecutor(max_workers=8) as executor:
        futures = {executor.submit(load_one, ror_id): ror_id for ror_id in unique_ror_ids}
        for future in concurrent.futures.as_completed(futures):
            ror_id = futures[future]
            try:
                _, payload = future.result()
                openalex_by_ror[ror_id] = payload
            except Exception:
                openalex_by_ror[ror_id] = {}
    return openalex_by_ror


def load_scorecard_proxy_records() -> list[dict[str, Any]]:
    records: list[dict[str, Any]] = []
    with SCORECARD_CSV.open("r", encoding="utf-8", errors="replace", newline="") as handle:
        reader = csv.DictReader(handle)
        for row in reader:
            records.append(
                {
                    "id": row.get("UNITID", ""),
                    "school.name": row.get("INSTNM", ""),
                    "school.state": row.get("STABBR", ""),
                    "school.city": row.get("CITY", ""),
                    "school.school_url": row.get("INSTURL", ""),
                    "unemp_rate": row.get("UNEMP_RATE", ""),
                    "count_wne_p6": row.get("COUNT_WNE_P6", ""),
                    "count_nwne_p6": row.get("COUNT_NWNE_P6", ""),
                    "mthcmp6": row.get("MTHCMP6", ""),
                }
            )
    return records


def build_scorecard_proxy_for_institution(institution: InstitutionRow, scorecard_records: list[dict[str, Any]]) -> dict[str, Any] | None:
    if not institution.country.startswith("United States of America"):
        return None
    matched = find_scorecard_match(_ScorecardInstitution(institution), scorecard_records)  # type: ignore[arg-type]
    if matched is None:
        return None
    count_wne = parse_float(matched.get("count_wne_p6"))
    count_nwne = parse_float(matched.get("count_nwne_p6"))
    employed_share = None
    if count_wne is not None and count_nwne is not None and (count_wne + count_nwne) > 0:
        employed_share = round((count_wne / (count_wne + count_nwne)) * 100.0, 2)
    months_to_completion = parse_float(matched.get("mthcmp6"))
    years_to_completion = round(months_to_completion / 12.0, 2) if months_to_completion is not None else None
    unitid = str(matched.get("id", "") or "").strip()
    return {
        "employed_share": employed_share,
        "years_to_completion": years_to_completion,
        "source_url": f"https://collegescorecard.ed.gov/school/?{unitid}" if unitid else None,
    }


def load_edurank_url_map() -> dict[str, list[str]]:
    cache_path = EDURANK_CACHE_DIR / "url_map.json"
    if cache_path.exists():
        return json.loads(cache_path.read_text(encoding="utf-8"))
    mapping: dict[str, list[str]] = {}
    for index in range(1, 25):
        xml = fetch_text(f"https://edurank.org/uni-sitemap{index}.xml", EDURANK_CACHE_DIR / f"sitemap_{index}.xml")
        soup = BeautifulSoup(xml, "xml")
        for loc in soup.find_all("loc"):
            url = loc.get_text(strip=True)
            if "/uni/" not in url or url.rstrip("/").endswith("/alumni") or url.rstrip("/").endswith("/rankings"):
                continue
            slug = url.rstrip("/").split("/")[-1]
            key = normalize_text(slug.replace("-", " "))
            mapping.setdefault(key, []).append(url)
    cache_path.parent.mkdir(parents=True, exist_ok=True)
    cache_path.write_text(json.dumps(mapping), encoding="utf-8")
    return mapping


def name_variants_for_edurank(name: str) -> list[str]:
    variants = [name.strip()]
    no_paren = re.sub(r"\s*\([^)]*\)", "", name).strip()
    if no_paren:
        variants.append(no_paren)
    if name.lower().startswith("the "):
        variants.append(name[4:].strip())
    if no_paren.lower().startswith("the "):
        variants.append(no_paren[4:].strip())
    seen: set[str] = set()
    ordered: list[str] = []
    for variant in variants:
        key = normalize_text(variant)
        if key and key not in seen:
            seen.add(key)
            ordered.append(variant)
    return ordered


def match_edurank_url(institution: InstitutionRow, ror_match: RorGeoRecord | None, openalex_match: dict[str, Any], url_map: dict[str, list[str]]) -> str | None:
    candidates: list[str] = []
    wiki_url = str(openalex_match.get("ids", {}).get("wikipedia", "") or "").strip()
    if wiki_url:
        wiki_title = wiki_url.rstrip("/").rsplit("/", 1)[-1].replace("_", " ")
        candidates.extend(name_variants_for_edurank(wiki_title))
    candidates.extend(name_variants_for_edurank(institution.university_name))
    if ror_match is not None:
        candidates.extend(name_variants_for_edurank(ror_match.name))
        for alias in ror_match.aliases[:5]:
            candidates.extend(name_variants_for_edurank(alias))
    display_name = str(openalex_match.get("display_name", "") or "").strip()
    if display_name:
        candidates.extend(name_variants_for_edurank(display_name))
    for alias in openalex_match.get("display_name_alternatives", [])[:5]:
        alias_text = str(alias or "").strip()
        if alias_text:
            candidates.extend(name_variants_for_edurank(alias_text))

    seen_keys: set[str] = set()
    for candidate in candidates:
        key = normalize_text(candidate)
        if not key or key in seen_keys:
            continue
        seen_keys.add(key)
        urls = url_map.get(key)
        if urls:
            return urls[0]
    return None


def parse_money(value: str | None) -> float | None:
    if not value:
        return None
    text = str(value).strip()
    text = text.replace("$", "").replace("USD", "").replace("/year", "").replace("/y", "").replace(",", "").strip()
    match = re.search(r"(-?\d+(?:\.\d+)?)", text)
    return float(match.group(1)) if match else None


def parse_percent_text(value: str | None) -> float | None:
    if not value:
        return None
    match = re.search(r"(-?\d+(?:\.\d+)?)\s*%", str(value))
    return float(match.group(1)) if match else None


def parse_int_text(value: str | None) -> int | None:
    if not value:
        return None
    cleaned = str(value).replace(",", "").strip()
    match = re.search(r"(-?\d+)", cleaned)
    return int(match.group(1)) if match else None


def parse_edurank_page(url: str) -> dict[str, Any]:
    slug = url.rstrip("/").split("/")[-1]
    html = fetch_text(url, EDURANK_CACHE_DIR / "pages" / f"{slug}.html")
    soup = BeautifulSoup(html, "html.parser")
    field_map: dict[str, str] = {}
    for dt in soup.find_all("dt"):
        label = dt.get_text(" ", strip=True)
        dd = dt.find_next_sibling("dd")
        if label and dd and label not in field_map:
            field_map[label] = dd.get_text(" ", strip=True)
    for row in soup.select("table tr"):
        th = row.find("th")
        td = row.find("td")
        if th is None or td is None:
            continue
        label = th.get_text(" ", strip=True)
        value = td.get_text(" ", strip=True)
        if label and value and label not in field_map:
            field_map[label] = value
    return {
        "url": url,
        "citations": parse_int_text(field_map.get("Citations")),
        "staff_fte": parse_int_text(field_map.get("Total FTE staff")),
        "scholarship_availability": parse_percent_text(
            field_map.get("Students receiving aid") or field_map.get("Aid receiving") or field_map.get("Receiving Aid")
        ),
        "scholarship_amount": parse_money(field_map.get("Average aid awarded") or field_map.get("Average Aid")),
        "dorm_capacity": parse_int_text(field_map.get("Dormitory capacity")),
    }


def parse_edurank_alumni(url: str) -> dict[str, Any]:
    alumni_url = url.rstrip("/") + "/alumni/"
    slug = url.rstrip("/").split("/")[-1]
    html = fetch_text(alumni_url, EDURANK_CACHE_DIR / "alumni" / f"{slug}.html")
    text = BeautifulSoup(html, "html.parser").get_text(" ", strip=True)
    match = re.search(r"\bis\s+(\d+)(?:st|nd|rd|th)\s+in the world\b", text, flags=re.IGNORECASE)
    world_rank = int(match.group(1)) if match else None
    score = None
    if world_rank is not None:
        score = round(((EDURANK_WORLD_UNIVERSITIES - world_rank + 1) / EDURANK_WORLD_UNIVERSITIES) * 100.0, 2)
    return {"url": alumni_url, "alumni_rank": world_rank, "alumni_score": score}


def load_numbeo_quality_of_life() -> dict[tuple[str, str], dict[str, Any]]:
    html = fetch_text(
        "https://www.numbeo.com/quality-of-life/rankings_current.jsp",
        NUMBEO_CACHE_DIR / "quality_of_life_rankings_current.html",
    )
    soup = BeautifulSoup(html, "html.parser")
    table = soup.find("table", {"id": "t2"})
    mapping: dict[tuple[str, str], dict[str, Any]] = {}
    if table is None:
        return mapping
    for row in table.find_all("tr")[1:]:
        cells = row.find_all("td")
        link = row.find("a", href=True)
        if len(cells) < 11 or link is None:
            continue
        city_country = cells[1].get_text(" ", strip=True)
        if "," not in city_country:
            continue
        city, country = [part.strip() for part in city_country.rsplit(",", 1)]
        normalized_city = normalize_text(re.sub(r"\s*\([^)]*\)", "", city))
        normalized_country = normalize_country(country)
        mapping[(normalized_city, normalized_country)] = {
            "quality_of_life": parse_float(cells[2].get_text(" ", strip=True)),
            "purchasing_power": parse_float(cells[3].get_text(" ", strip=True)),
            "safety": parse_float(cells[4].get_text(" ", strip=True)),
            "healthcare": parse_float(cells[5].get_text(" ", strip=True)),
            "traffic": parse_float(cells[8].get_text(" ", strip=True)),
            "pollution": parse_float(cells[9].get_text(" ", strip=True)),
            "climate": parse_float(cells[10].get_text(" ", strip=True)),
            "source_url": link.get("href", "").strip(),
        }
    return mapping


def query_wikidata(query: str, cache_path: Path) -> dict[str, Any]:
    cache_path.parent.mkdir(parents=True, exist_ok=True)
    if cache_path.exists():
        return json.loads(cache_path.read_text(encoding="utf-8"))
    last_error: Exception | None = None
    for _ in range(3):
        try:
            response = requests.get(
                "https://query.wikidata.org/sparql",
                params={"format": "json", "query": query},
                headers={"User-Agent": USER_AGENT},
                timeout=(20, 120),
            )
            response.raise_for_status()
            cache_path.write_text(response.text, encoding="utf-8")
            return response.json()
        except Exception as exc:
            last_error = exc
            time.sleep(1.5)
    raise last_error or RuntimeError("Wikidata query failed")


def load_wikidata_qs_ranks(qids: list[str]) -> dict[str, dict[str, Any]]:
    result: dict[str, dict[str, Any]] = {}
    batch_size = 100
    for start in range(0, len(qids), batch_size):
        batch = qids[start:start + batch_size]
        qid_token = "_".join(batch[:3]) + f"_{len(batch)}"
        query = (
            "SELECT ?item ?rank ?time WHERE { "
            f"VALUES ?item {{ {' '.join('wd:' + qid for qid in batch)} }} "
            "?item p:P1352 ?stmt. "
            "?stmt ps:P1352 ?rank; pq:P459 wd:" + QS_ITEM_QID + ". "
            "OPTIONAL { ?stmt pq:P585 ?time. } "
            "}"
        )
        payload = query_wikidata(query, WIKIDATA_CACHE_DIR / "qs_batches" / f"{qid_token}.json")
        for binding in payload.get("results", {}).get("bindings", []):
            qid = binding["item"]["value"].rsplit("/", 1)[-1]
            rank = parse_int_text(binding["rank"]["value"])
            time_value = binding.get("time", {}).get("value", "")
            if rank is None:
                continue
            previous = result.get(qid)
            if previous is None or time_value > str(previous.get("time", "")):
                result[qid] = {"rank": rank, "time": time_value, "source_url": f"https://www.wikidata.org/wiki/{qid}"}
    return result


def load_wikidata_unit_labels(unit_qids: set[str]) -> dict[str, str]:
    labels: dict[str, str] = {}
    for qid in unit_qids:
        cache_path = WIKIDATA_CACHE_DIR / "units" / f"{qid}.json"
        if cache_path.exists():
            payload = json.loads(cache_path.read_text(encoding="utf-8"))
        else:
            payload = fetch_json(f"https://www.wikidata.org/wiki/Special:EntityData/{qid}.json", cache_path)
        entity = payload.get("entities", {}).get(qid, {})
        label = str(entity.get("labels", {}).get("en", {}).get("value", "") or "").strip().lower()
        if label:
            labels[qid] = label
    return labels


def load_wikidata_areas(qids: list[str]) -> dict[str, dict[str, Any]]:
    raw_results: dict[str, list[tuple[float, str]]] = {}
    batch_size = 100
    for start in range(0, len(qids), batch_size):
        batch = qids[start:start + batch_size]
        qid_token = "_".join(batch[:3]) + f"_{len(batch)}"
        query = (
            "SELECT ?item ?amount ?unit WHERE { "
            f"VALUES ?item {{ {' '.join('wd:' + qid for qid in batch)} }} "
            "?item p:P2046 ?stmt. "
            "?stmt psv:P2046 ?v. "
            "?v wikibase:quantityAmount ?amount; wikibase:quantityUnit ?unit. "
            "}"
        )
        payload = query_wikidata(query, WIKIDATA_CACHE_DIR / "area_batches" / f"{qid_token}.json")
        for binding in payload.get("results", {}).get("bindings", []):
            qid = binding["item"]["value"].rsplit("/", 1)[-1]
            amount = parse_float(binding["amount"]["value"])
            unit_qid = binding["unit"]["value"].rsplit("/", 1)[-1]
            if amount is not None:
                raw_results.setdefault(qid, []).append((amount, unit_qid))

    unit_labels = load_wikidata_unit_labels({unit_qid for pairs in raw_results.values() for _, unit_qid in pairs})
    result: dict[str, dict[str, Any]] = {}
    for qid, pairs in raw_results.items():
        converted_values: list[float] = []
        for amount, unit_qid in pairs:
            label = unit_labels.get(unit_qid, "")
            multiplier = next((value for key, value in AREA_UNIT_MULTIPLIERS.items() if key in label), None)
            if multiplier is not None:
                converted_values.append(amount * multiplier)
        if converted_values:
            result[qid] = {"area_m2": round(max(converted_values), 2), "source_url": f"https://www.wikidata.org/wiki/{qid}"}
    return result


def climate_key(institution: InstitutionRow) -> tuple[str, str]:
    return normalize_text(institution.city), normalize_country(institution.base_country)


def build_city_coordinates(
    institutions: list[InstitutionRow],
    ror_matches: dict[int, RorGeoRecord | None],
    openalex_by_ror: dict[str, dict[str, Any]],
) -> dict[tuple[str, str], tuple[float, float]]:
    coordinates: dict[tuple[str, str], tuple[float, float]] = {}
    for institution in institutions:
        key = climate_key(institution)
        if not key[0] or key in coordinates:
            continue
        ror_match = ror_matches[institution.row_number]
        openalex_match = openalex_by_ror.get(ror_match.ror_id, {}) if ror_match is not None else {}
        geo = openalex_match.get("geo", {}) if openalex_match else {}
        latitude = parse_float(geo.get("latitude"))
        longitude = parse_float(geo.get("longitude"))
        if latitude is None or longitude is None:
            latitude = ror_match.latitude if ror_match is not None else None
            longitude = ror_match.longitude if ror_match is not None else None
        if latitude is not None and longitude is not None:
            coordinates[key] = (latitude, longitude)
    return coordinates


def fetch_climate_type(city_key: tuple[str, str], latitude: float, longitude: float) -> dict[str, Any]:
    slug = f"{city_key[0]}__{city_key[1]}".replace("/", "_")
    url = f"https://climate.mapresso.com/api/koeppen/?lat={latitude:.4f}&lon={longitude:.4f}"
    payload = fetch_json(url, MAPRESSO_CACHE_DIR / f"{slug}.json")
    climate_text = None
    for item in payload.get("data", []):
        if item.get("short") == "KG" and item.get("text"):
            climate_text = item["text"]
            break
    if climate_text is None:
        climate_text = next((item.get("text") for item in payload.get("data", []) if item.get("text")), None)
    return {"climate_type": climate_text, "source_url": url}


def split_languages(value: str) -> list[str]:
    tokens = [part.strip() for part in LANGUAGE_SPLIT_RE.split(value) if part.strip()]
    unique: list[str] = []
    seen: set[str] = set()
    for token in tokens:
        key = normalize_text(token)
        if key and key not in seen:
            seen.add(key)
            unique.append(token)
    return unique


def english_program_ratio(language_text: str) -> tuple[float | None, str | None]:
    languages = split_languages(language_text)
    if not languages:
        return None, None
    normalized = [normalize_text(language) for language in languages]
    if normalize_text("English") not in normalized:
        return 0.0, "WHED instructional language list"
    return round(100.0 / len(languages), 2), "WHED instructional language split estimate"


def extract_partner_count(raw_text: str) -> int | None:
    text = normalize_row_text(raw_text)
    for pattern in PARTNER_COUNT_PATTERNS:
        match = pattern.search(text)
        if match:
            value = match.group(1).strip().lower()
            return int(value) if value.isdigit() else NUMBER_WORDS.get(value)
    return None


def extract_lab_count(raw_text: str) -> int | None:
    count = len(LAB_RE.findall(normalize_row_text(raw_text)))
    return count or None


def sports_facilities_score(raw_text: str) -> float | None:
    text = normalize_row_text(raw_text)
    count = len(SPORTS_FACILITY_RE.findall(text))
    count += len(SPORTS_UNIT_RE.findall(text))
    return round(min(count * 25.0, 100.0), 2) if count > 0 else None


def air_quality_score(air_quality_text: str, pm25: float | None) -> float | None:
    quality_key = normalize_text(air_quality_text)
    mapped = AIR_QUALITY_SCORES.get(quality_key)
    pm25_score = clamp(100.0 - ((pm25 or 0.0) * 2.0)) if pm25 is not None else None
    return average_available(mapped, pm25_score)


def monthly_cost(institution: InstitutionRow) -> float | None:
    return round(institution.annual_cost_usd / 12.0, 2) if institution.annual_cost_usd is not None else None


def salary_cover_score(institution: InstitutionRow) -> float | None:
    monthly_student_cost = monthly_cost(institution)
    if institution.monthly_salary_usd is None or monthly_student_cost in (None, 0):
        return None
    return round(clamp((institution.monthly_salary_usd / monthly_student_cost) * 50.0) or 0.0, 2)


def price_affordability_score(price: float | None, salary: float | None) -> float | None:
    if price is None or salary in (None, 0):
        return None
    return round(clamp(100.0 - ((price / salary) * 2000.0)) or 0.0, 2)


def student_scene_score(institution: InstitutionRow) -> float | None:
    universities_score = clamp((institution.city_universities_count or 0) / 20.0 * 100.0) if institution.city_universities_count is not None else None
    students_score = clamp((math.log1p(institution.city_known_students) / math.log1p(300000)) * 100.0) if institution.city_known_students not in (None, 0) else None
    return average_available(universities_score, students_score)


def normalized_qol_score(value: float | None) -> float | None:
    return round(clamp((value / 240.0) * 100.0) or 0.0, 2) if value is not None else None


def merged_coverage(existing: str, additions: list[str]) -> str | None:
    values = [part.strip() for part in existing.split(";") if part.strip()]
    seen = {part.lower() for part in values}
    for addition in additions:
        if addition and addition.lower() not in seen:
            values.append(addition)
            seen.add(addition.lower())
    return "; ".join(values) if values else None


def enrich_workbook(workbook_path: Path, sheet_name: str) -> dict[str, Any]:
    wb, ws, _, institutions = load_institutions(workbook_path, sheet_name)
    full_header_map = ensure_columns(ws)
    ror_domain_index, ror_country_index = load_ror_records_with_geo()
    ror_matches = {institution.row_number: find_ror_match(institution, ror_domain_index, ror_country_index) for institution in institutions}
    openalex_by_ror = load_openalex_for_matches(ror_matches)
    scorecard_records = load_scorecard_proxy_records()
    edurank_url_map = load_edurank_url_map()
    numbeo_qol = load_numbeo_quality_of_life()

    edurank_urls_by_row: dict[int, str] = {}
    qids: set[str] = set()
    for institution in institutions:
        ror_match = ror_matches[institution.row_number]
        openalex_match = openalex_by_ror.get(ror_match.ror_id, {}) if ror_match is not None else {}
        url = match_edurank_url(institution, ror_match, openalex_match, edurank_url_map)
        if url:
            edurank_urls_by_row[institution.row_number] = url
        qid_url = str(openalex_match.get("ids", {}).get("wikidata", "") or "").strip()
        if qid_url:
            qids.add(qid_url.rsplit("/", 1)[-1])

    qids_sorted = sorted(qids)
    wikidata_qs = load_wikidata_qs_ranks(qids_sorted)
    wikidata_area = load_wikidata_areas(qids_sorted)

    unique_edurank_urls = sorted(set(edurank_urls_by_row.values()))
    edurank_main_by_url: dict[str, dict[str, Any]] = {}
    edurank_alumni_by_url: dict[str, dict[str, Any]] = {}
    with concurrent.futures.ThreadPoolExecutor(max_workers=6) as executor:
        main_futures = {executor.submit(parse_edurank_page, url): url for url in unique_edurank_urls}
        for future in concurrent.futures.as_completed(main_futures):
            url = main_futures[future]
            try:
                payload = future.result()
            except Exception:
                payload = {"url": url}
            edurank_main_by_url[payload["url"]] = payload
        alumni_futures = {executor.submit(parse_edurank_alumni, url): url for url in unique_edurank_urls}
        for future in concurrent.futures.as_completed(alumni_futures):
            url = alumni_futures[future]
            try:
                payload = future.result()
            except Exception:
                payload = {"url": url.rstrip("/") + "/alumni/"}
            edurank_alumni_by_url[payload["url"].replace("/alumni/", "/")] = payload

    city_coordinates = build_city_coordinates(institutions, ror_matches, openalex_by_ror)
    climate_by_city: dict[tuple[str, str], dict[str, Any]] = {}
    with concurrent.futures.ThreadPoolExecutor(max_workers=6) as executor:
        future_map = {executor.submit(fetch_climate_type, city_key, lat, lon): city_key for city_key, (lat, lon) in city_coordinates.items()}
        for future in concurrent.futures.as_completed(future_map):
            city_key = future_map[future]
            try:
                climate_by_city[city_key] = future.result()
            except Exception:
                climate_by_city[city_key] = {}

    filled_counts = {key: 0 for key in [
        "qs_rank", "citation_per_faculty", "scholarship_availability", "scholarship_amount", "dorm_capacity",
        "english_ratio", "campus_safety", "green_campus", "alumni_strength", "climate_type",
        "family_friendliness", "cultural_activity", "nightlife", "digital_infrastructure",
        "part_time_job", "campus_size", "erasmus", "lab_count", "sports_score", "employability_proxy",
    ]}

    for institution in institutions:
        row_number = institution.row_number
        ror_match = ror_matches[row_number]
        openalex_match = openalex_by_ror.get(ror_match.ror_id, {}) if ror_match is not None else {}
        qid_url = str(openalex_match.get("ids", {}).get("wikidata", "") or "").strip()
        qid = qid_url.rsplit("/", 1)[-1] if qid_url else None
        edurank_url = edurank_urls_by_row.get(row_number)
        edurank_main = edurank_main_by_url.get(edurank_url, {}) if edurank_url else {}
        edurank_alumni = edurank_alumni_by_url.get(edurank_url, {}) if edurank_url else {}
        numbeo_match = numbeo_qol.get(climate_key(institution))
        climate_match = climate_by_city.get(climate_key(institution), {})
        scorecard_proxy = build_scorecard_proxy_for_institution(institution, scorecard_records)

        staff_count = institution.staff_full_time_total or parse_float(edurank_main.get("staff_fte"))
        cited_by_count = parse_float(openalex_match.get("cited_by_count")) or parse_float(edurank_main.get("citations"))
        citation_per_faculty = round(cited_by_count / staff_count, 2) if cited_by_count is not None and staff_count not in (None, 0) else None
        scholarship_availability = parse_float(edurank_main.get("scholarship_availability"))
        scholarship_amount = parse_float(edurank_main.get("scholarship_amount"))
        dorm_capacity = parse_int_like(edurank_main.get("dorm_capacity"))
        english_ratio, english_note = english_program_ratio(institution.language_text)
        partner_count = extract_partner_count(institution.raw_text)
        lab_count = extract_lab_count(institution.raw_text)
        sports_score = sports_facilities_score(institution.raw_text)
        erasmus_score = 100.0 if ERASMUS_RE.search(normalize_row_text(institution.raw_text)) else None
        safety_score = parse_float(numbeo_match.get("safety")) if numbeo_match else None
        if safety_score is None and institution.crime_index is not None:
            safety_score = round(clamp(100.0 - institution.crime_index) or 0.0, 2)

        pollution_score = round(clamp(100.0 - parse_float(numbeo_match.get("pollution"))) or 0.0, 2) if numbeo_match and parse_float(numbeo_match.get("pollution")) is not None else None
        climate_index_score = parse_float(numbeo_match.get("climate")) if numbeo_match else None
        healthcare_score = parse_float(numbeo_match.get("healthcare")) if numbeo_match else None
        purchasing_power_score = clamp(parse_float(numbeo_match.get("purchasing_power")) / 2.0) if numbeo_match and parse_float(numbeo_match.get("purchasing_power")) is not None else None
        quality_score = normalized_qol_score(parse_float(numbeo_match.get("quality_of_life"))) if numbeo_match else normalized_qol_score(institution.quality_of_life_index)

        green_campus = average_available(institution.environment_score, pollution_score, climate_index_score, air_quality_score(institution.air_quality_text, institution.pm25))
        family_friendliness = average_available(safety_score, healthcare_score, climate_index_score, pollution_score, institution.mobility_score)
        scene_score = student_scene_score(institution)
        cultural_activity = average_available(institution.daily_life_score, institution.academic_ecosystem_score, scene_score, quality_score)
        nightlife = average_available(institution.daily_life_score, scene_score, price_affordability_score(institution.fast_food_usd, institution.monthly_salary_usd), price_affordability_score(institution.cappuccino_usd, institution.monthly_salary_usd))
        part_time_job = average_available(salary_cover_score(institution), institution.affordability_score, purchasing_power_score)
        digital_infrastructure = average_available(price_affordability_score(institution.internet_plan_usd, institution.monthly_salary_usd), institution.daily_life_score)

        qs_rank = wikidata_qs.get(qid, {}).get("rank") if qid else None
        campus_size = wikidata_area.get(qid, {}).get("area_m2") if qid else None
        alumni_strength = parse_float(edurank_alumni.get("alumni_score"))
        climate_type = climate_match.get("climate_type")
        employability_proxy = scorecard_proxy.get("employed_share") if scorecard_proxy else None
        average_time_to_graduate = scorecard_proxy.get("years_to_completion") if scorecard_proxy else None

        additions = {
            "Global Ranking - QS": qs_rank,
            "Citation per Faculty": citation_per_faculty,
            "Graduate Employability Rate (%)": employability_proxy,
            "Average Time to Graduate": average_time_to_graduate,
            "Number of Partner Universities": partner_count,
            "Erasmus Participation Score": erasmus_score,
            "English Program Ratio (%)": english_ratio,
            "Scholarship Availability (%)": scholarship_availability,
            "Average Scholarship Amount": scholarship_amount,
            "Part-time Job Availability Score": part_time_job,
            "Campus Size (m²)": campus_size,
            "Dorm Capacity": dorm_capacity,
            "Lab Count": lab_count,
            "Sports Facilities Score": sports_score,
            "Campus Safety Score": safety_score,
            "Green Campus Score": green_campus,
            "Alumni Network Strength Score": alumni_strength,
            "Industry Placement Rate": employability_proxy,
            "Climate Type": climate_type,
            "Cultural Activity Score": cultural_activity,
            "Nightlife Score": nightlife,
            "Family Friendliness Score": family_friendliness,
            "Digital Infrastructure Score (5G vs)": digital_infrastructure,
            "QS Source URL": wikidata_qs.get(qid, {}).get("source_url") if qid else None,
            "EduRank Source URL": edurank_url,
            "Wikidata Source URL": wikidata_area.get(qid, {}).get("source_url") if qid else None,
            "Numbeo QoL Source URL": numbeo_match.get("source_url") if numbeo_match else None,
            "Climate Source URL": climate_match.get("source_url"),
        }

        notes = []
        if english_note and english_ratio is not None:
            notes.append(english_note)
        if employability_proxy is not None:
            notes.append("Graduate employability and industry placement use College Scorecard employed-share proxy at 6 years")
        if part_time_job is not None:
            notes.append("Part-time job score is a proxy from salary-cover, affordability, and purchasing power")
        if cultural_activity is not None or nightlife is not None or digital_infrastructure is not None:
            notes.append("City score columns are proxy composites from Livingcost, WHED, and Numbeo inputs")
        if family_friendliness is not None or green_campus is not None:
            notes.append("Family and green scores combine environment, safety, pollution, and climate indicators")
        additions["Unofficial Metrics Notes"] = "; ".join(dict.fromkeys(notes)) if notes else None
        additions["Extended Metrics Coverage"] = merged_coverage(
            institution.existing_coverage,
            [
                "Wikidata" if qs_rank is not None or campus_size is not None else "",
                "EduRank" if edurank_url and any(value is not None for value in [scholarship_availability, scholarship_amount, dorm_capacity, alumni_strength, citation_per_faculty]) else "",
                "Numbeo QoL" if numbeo_match and any(value is not None for value in [safety_score, green_campus, family_friendliness, cultural_activity, nightlife]) else "",
                "Mapresso" if climate_type else "",
                "WHED" if any(value is not None for value in [english_ratio, partner_count, lab_count, sports_score, erasmus_score]) else "",
                "College Scorecard Proxy" if any(value is not None for value in [employability_proxy, average_time_to_graduate]) else "",
            ],
        )

        for header, value in additions.items():
            if value is None:
                continue
            cell = ws.cell(row=row_number, column=full_header_map[header], value=value)
            if header in hyperlink_columns():
                cell.hyperlink = value
                cell.style = "Hyperlink"

        for key, value in {
            "qs_rank": qs_rank, "citation_per_faculty": citation_per_faculty, "scholarship_availability": scholarship_availability,
            "scholarship_amount": scholarship_amount, "dorm_capacity": dorm_capacity, "english_ratio": english_ratio,
            "campus_safety": safety_score, "green_campus": green_campus, "alumni_strength": alumni_strength,
            "climate_type": climate_type, "family_friendliness": family_friendliness, "cultural_activity": cultural_activity,
            "nightlife": nightlife, "digital_infrastructure": digital_infrastructure, "part_time_job": part_time_job,
            "campus_size": campus_size, "erasmus": erasmus_score, "lab_count": lab_count, "sports_score": sports_score,
            "employability_proxy": employability_proxy,
        }.items():
            if value is not None:
                filled_counts[key] += 1

    backup_path = backup_workbook(workbook_path)
    wb.save(workbook_path)
    return {"rows": len(institutions), "backup_path": str(backup_path), "matched_edurank_urls": len(unique_edurank_urls), "matched_qids": len(qids_sorted), "filled_counts": filled_counts}


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Fill remaining extended-metric blanks using reliable unofficial sources and transparent proxy scores.")
    parser.add_argument("--workbook", default="whed_data.xlsx")
    parser.add_argument("--sheet", default="Institutions")
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    print(json.dumps(enrich_workbook(Path(args.workbook), args.sheet), ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
